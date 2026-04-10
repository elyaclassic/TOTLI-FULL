"""Oddiy Excel (.xlsx) ga operatsiyalar yozish va hisobot formulalarini saqlash."""
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from src.config import EXCEL_FILE_PATH
from src.services.parse_operation import parse_operation_text

OPERATIONS_HEADERS = [
    "Sana",
    "Vaqt",
    "Mijoz_ID",
    "Mijoz_nomi",
    "Turi",
    "Valyuta",
    "Summa",
    "Kurs",
    "Summa_uzs",
    "Summa_usd",
    "Izoh",
    "Telegram_user",
    "Matn",
    "Manba",
]

CUSTOMERS_HEADERS = [
    "ID",
    "Nomi",
    "Telefon",
    "Boshlang'ich_qarz_uzs",
    "Boshlang'ich_qarz_usd",
    "Mijoz_to'lagan_uzs",
    "Mijoz_to'lagan_usd",
    "Biz_bergan_uzs",
    "Biz_bergan_usd",
    "Qarz_qoldiq_uzs",
    "Qarz_qoldiq_usd",
]

SUMMARY_HEADERS = ["Ko'rsatkich", "UZS", "USD"]


def _excel_path() -> Path:
    p = Path(EXCEL_FILE_PATH).expanduser()
    p.parent.mkdir(parents=True, exist_ok=True)
    return p


def _bold_row(ws, row_num: int) -> None:
    for cell in ws[row_num]:
        cell.font = Font(bold=True)


def _find_header_row(ws, first_header: str) -> int | None:
    for row in range(1, min(ws.max_row, 20) + 1):
        value = str(ws[f"A{row}"].value or "").strip()
        if value == first_header:
            return row
    return None


def _next_data_row(
    ws,
    first_header: str,
    default_header_row: int = 1,
    data_columns: tuple[int, ...] | None = None,
) -> int:
    header_row = _find_header_row(ws, first_header) or default_header_row
    check_columns = data_columns or tuple(range(1, ws.max_column + 1))
    row = header_row + 1
    while row <= ws.max_row:
        if any(ws.cell(row=row, column=col).value not in (None, "") for col in check_columns):
            row += 1
            continue
        break
    return row


def _ensure_workbook() -> Path:
    path = _excel_path()
    if path.exists():
        return path

    wb = Workbook()

    ws = wb.active
    ws.title = "Operatsiyalar"
    ws.append(OPERATIONS_HEADERS)
    _bold_row(ws, 1)

    customers = wb.create_sheet("Mijozlar")
    customers.append(CUSTOMERS_HEADERS)
    _bold_row(customers, 1)

    summary = wb.create_sheet("Hisobot")
    summary.append(SUMMARY_HEADERS)
    _bold_row(summary, 1)

    wb.save(path)
    return path


def _ensure_headers(ws, headers: list[str]) -> int:
    header_row = _find_header_row(ws, headers[0])
    if header_row:
        for idx, header in enumerate(headers, start=1):
            if ws.cell(row=header_row, column=idx).value in (None, ""):
                ws.cell(row=header_row, column=idx, value=header)
        _bold_row(ws, header_row)
        return header_row

    if not ws["A1"].value:
        ws.append(headers)
        _bold_row(ws, 1)
        return 1

    for idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=header)
    _bold_row(ws, 1)
    return 1


def _migrate_operations_sheet(ws) -> None:
    header_row = _find_header_row(ws, "Sana")
    if not header_row:
        return
    if str(ws[f"F{header_row}"].value or "").strip() == "Valyuta":
        return
    if str(ws[f"F{header_row}"].value or "").strip() != "Summa":
        return

    ws.insert_cols(6, 4)
    for idx, header in enumerate(OPERATIONS_HEADERS, start=1):
        ws.cell(row=header_row, column=idx, value=header)
    _bold_row(ws, header_row)

    for row in range(header_row + 1, ws.max_row + 1):
        old_sum = _num(ws.cell(row=row, column=10).value)
        if not any(ws.cell(row=row, column=col).value not in (None, "") for col in range(1, 15)):
            continue
        ws.cell(row=row, column=6, value="UZS")
        ws.cell(row=row, column=7, value=old_sum if old_sum else "")
        ws.cell(row=row, column=8, value="")
        ws.cell(row=row, column=9, value=old_sum if old_sum else "")
        ws.cell(row=row, column=10, value="")


def _migrate_customers_sheet(ws) -> None:
    header_row = _find_header_row(ws, "ID")
    if not header_row:
        return
    if str(ws[f"D{header_row}"].value or "").strip() == "Boshlang'ich_qarz_uzs":
        return
    if str(ws[f"D{header_row}"].value or "").strip() != "Boshlang'ich_qarz":
        return

    ws.insert_cols(5, 4)
    for idx, header in enumerate(CUSTOMERS_HEADERS, start=1):
        ws.cell(row=header_row, column=idx, value=header)
    _bold_row(ws, header_row)

    for row in range(header_row + 1, ws.max_row + 1):
        if not ws.cell(row=row, column=1).value:
            continue
        old_paid = _num(ws.cell(row=row, column=9).value)
        old_given = _num(ws.cell(row=row, column=10).value)
        old_balance = _num(ws.cell(row=row, column=11).value)
        ws.cell(row=row, column=5, value=0)
        ws.cell(row=row, column=6, value=old_paid if old_paid else "")
        ws.cell(row=row, column=7, value=0)
        ws.cell(row=row, column=8, value=old_given if old_given else "")
        ws.cell(row=row, column=9, value=0)
        ws.cell(row=row, column=10, value=old_balance if old_balance else "")
        ws.cell(row=row, column=11, value=0)


def _ensure_summary_formulas(summary_ws, operations_ws) -> None:
    header_row = _ensure_headers(summary_ws, SUMMARY_HEADERS)
    base_row = header_row + 1

    summary_ws.cell(row=header_row, column=2, value="UZS")
    summary_ws.cell(row=header_row, column=3, value="USD")
    _bold_row(summary_ws, header_row)

    summary_ws[f"A{base_row}"] = "Mijozlar to'lagan"
    summary_ws[f"B{base_row}"] = '=SUMIFS(Operatsiyalar!$I:$I,Operatsiyalar!$E:$E,"kirim")'
    summary_ws[f"C{base_row}"] = '=SUMIFS(Operatsiyalar!$J:$J,Operatsiyalar!$E:$E,"kirim")'

    summary_ws[f"A{base_row + 1}"] = "Biz bergan"
    summary_ws[f"B{base_row + 1}"] = '=SUMIFS(Operatsiyalar!$I:$I,Operatsiyalar!$E:$E,"chiqim")'
    summary_ws[f"C{base_row + 1}"] = '=SUMIFS(Operatsiyalar!$J:$J,Operatsiyalar!$E:$E,"chiqim")'

    summary_ws[f"A{base_row + 2}"] = "Jami qarz qoldiq"
    summary_ws[f"B{base_row + 2}"] = f"=B{base_row + 1}-B{base_row}"
    summary_ws[f"C{base_row + 2}"] = f"=C{base_row + 1}-C{base_row}"

    summary_ws[f"A{base_row + 3}"] = "Operatsiyalar soni"
    summary_ws[f"B{base_row + 3}"] = '=COUNTA(Operatsiyalar!$A:$A)-3'
    summary_ws[f"C{base_row + 3}"] = ""


def _ensure_customer_formulas(customers_ws) -> None:
    header_row = _find_header_row(customers_ws, CUSTOMERS_HEADERS[0]) or 1
    for row in range(header_row + 1, max(customers_ws.max_row, header_row + 1) + 1):
        if not customers_ws[f"A{row}"].value:
            continue
        customers_ws[f"F{row}"] = (
            f'=SUMIFS(Operatsiyalar!$I:$I,Operatsiyalar!$C:$C,A{row},Operatsiyalar!$E:$E,"kirim")'
        )
        customers_ws[f"G{row}"] = (
            f'=SUMIFS(Operatsiyalar!$J:$J,Operatsiyalar!$C:$C,A{row},Operatsiyalar!$E:$E,"kirim")'
        )
        customers_ws[f"H{row}"] = (
            f'=SUMIFS(Operatsiyalar!$I:$I,Operatsiyalar!$C:$C,A{row},Operatsiyalar!$E:$E,"chiqim")'
        )
        customers_ws[f"I{row}"] = (
            f'=SUMIFS(Operatsiyalar!$J:$J,Operatsiyalar!$C:$C,A{row},Operatsiyalar!$E:$E,"chiqim")'
        )
        customers_ws[f"J{row}"] = f"=D{row}+H{row}-F{row}"
        customers_ws[f"K{row}"] = f"=E{row}+I{row}-G{row}"


def _num(value) -> float:
    try:
        if value is None or value == "":
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        return float(str(value).replace(" ", "").replace(",", ""))
    except Exception:
        return 0.0


def _int_or_none(value) -> int | None:
    try:
        if value is None or value == "":
            return None
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        text = str(value).strip()
        if text.isdigit():
            return int(text)
        return None
    except Exception:
        return None


def _match_period(row_date: str, period: str) -> bool:
    if period == "all":
        return True
    try:
        d = datetime.strptime(str(row_date), "%Y-%m-%d").date()
    except Exception:
        return False
    today = date.today()
    if period == "today":
        return d == today
    if period == "this_month":
        return d.year == today.year and d.month == today.month
    return True


def _resolve_currency(currency: str | None) -> str:
    cur = str(currency or "UZS").strip().upper()
    return cur if cur in {"UZS", "USD"} else "UZS"


def _compute_amounts(currency: str, amount: float | None, rate: float | None) -> tuple[float | None, float | None]:
    if amount is None:
        return None, None
    cur = _resolve_currency(currency)
    r = _num(rate)
    if cur == "USD":
        return (amount * r if r > 0 else 0.0), float(amount)
    return float(amount), (float(amount) / r if r > 0 else 0.0)


def _calc_customer_totals(operations_ws, customer_id: int) -> dict[str, float]:
    totals = {
        "kirim_uzs": 0.0,
        "kirim_usd": 0.0,
        "chiqim_uzs": 0.0,
        "chiqim_usd": 0.0,
    }
    header_row = _find_header_row(operations_ws, OPERATIONS_HEADERS[0]) or 1
    for row in range(header_row + 1, operations_ws.max_row + 1):
        cid = _int_or_none(operations_ws[f"C{row}"].value)
        if cid is None or cid != int(customer_id):
            continue
        amount_uzs = _num(operations_ws[f"I{row}"].value)
        amount_usd = _num(operations_ws[f"J{row}"].value)
        op_type = str(operations_ws[f"E{row}"].value or "").strip().lower()
        if op_type == "kirim":
            totals["kirim_uzs"] += amount_uzs
            totals["kirim_usd"] += amount_usd
        elif op_type == "chiqim":
            totals["chiqim_uzs"] += amount_uzs
            totals["chiqim_usd"] += amount_usd
    return totals


def _load_book():
    path = _ensure_workbook()
    wb = load_workbook(path)
    operations = wb["Operatsiyalar"] if "Operatsiyalar" in wb.sheetnames else wb.active
    operations.title = "Operatsiyalar"
    _ensure_headers(operations, OPERATIONS_HEADERS)
    _migrate_operations_sheet(operations)
    _ensure_headers(operations, OPERATIONS_HEADERS)

    if "Mijozlar" not in wb.sheetnames:
        wb.create_sheet("Mijozlar")
    customers = wb["Mijozlar"]
    _ensure_headers(customers, CUSTOMERS_HEADERS)
    _migrate_customers_sheet(customers)
    _ensure_headers(customers, CUSTOMERS_HEADERS)

    if "Hisobot" not in wb.sheetnames:
        wb.create_sheet("Hisobot")
    summary = wb["Hisobot"]
    _ensure_headers(summary, SUMMARY_HEADERS)

    return path, wb, operations, customers, summary


def list_customers() -> list[dict]:
    path, wb, operations, customers, summary = _load_book()
    _ensure_customer_formulas(customers)
    _ensure_summary_formulas(summary, operations)
    wb.save(path)

    items: list[dict] = []
    header_row = _find_header_row(customers, CUSTOMERS_HEADERS[0]) or 1
    for row in range(header_row + 1, customers.max_row + 1):
        cid = _int_or_none(customers[f"A{row}"].value)
        name = customers[f"B{row}"].value
        if cid is None or not name:
            continue
        totals = _calc_customer_totals(operations, int(cid))
        opening_uzs = _num(customers[f"D{row}"].value)
        opening_usd = _num(customers[f"E{row}"].value)
        items.append(
            {
                "id": int(cid),
                "name": str(name),
                "phone": str(customers[f"C{row}"].value or ""),
                "opening_uzs": opening_uzs,
                "opening_usd": opening_usd,
                "kirim_uzs": totals["kirim_uzs"],
                "kirim_usd": totals["kirim_usd"],
                "chiqim_uzs": totals["chiqim_uzs"],
                "chiqim_usd": totals["chiqim_usd"],
                "qoldiq_uzs": opening_uzs + totals["chiqim_uzs"] - totals["kirim_uzs"],
                "qoldiq_usd": opening_usd + totals["chiqim_usd"] - totals["kirim_usd"],
            }
        )
    return items


def get_customer(customer_id: int) -> dict | None:
    for item in list_customers():
        if item["id"] == customer_id:
            return item
    return None


def add_customer(
    name: str,
    phone: str = "",
    opening_balance_uzs: float = 0,
    opening_balance_usd: float = 0,
) -> dict:
    path, wb, _operations, customers, _summary = _load_book()

    max_id = 0
    header_row = _find_header_row(customers, CUSTOMERS_HEADERS[0]) or 1
    for row in range(header_row + 1, customers.max_row + 1):
        cid = _int_or_none(customers[f"A{row}"].value)
        if cid is not None:
            max_id = max(max_id, cid)
    new_id = max_id + 1

    target_row = _next_data_row(
        customers,
        CUSTOMERS_HEADERS[0],
        default_header_row=header_row,
        data_columns=(1, 2, 3, 4, 5),
    )
    customers[f"A{target_row}"] = new_id
    customers[f"B{target_row}"] = name.strip()
    customers[f"C{target_row}"] = phone.strip()
    customers[f"D{target_row}"] = float(opening_balance_uzs or 0)
    customers[f"E{target_row}"] = float(opening_balance_usd or 0)
    _ensure_customer_formulas(customers)
    wb.save(path)
    return get_customer(new_id) or {"id": new_id, "name": name.strip(), "phone": phone.strip()}


def delete_customer(customer_id: int) -> bool:
    path, wb, operations, customers, summary = _load_book()
    deleted = False

    customer_header = _find_header_row(customers, CUSTOMERS_HEADERS[0]) or 1
    for row in range(customers.max_row, customer_header, -1):
        cid = _int_or_none(customers[f"A{row}"].value)
        if cid == int(customer_id):
            customers.delete_rows(row, 1)
            deleted = True
            break

    operations_header = _find_header_row(operations, OPERATIONS_HEADERS[0]) or 1
    for row in range(operations.max_row, operations_header, -1):
        cid = _int_or_none(operations[f"C{row}"].value)
        if cid == int(customer_id):
            operations.delete_rows(row, 1)

    if deleted:
        _ensure_customer_formulas(customers)
        _ensure_summary_formulas(summary, operations)
        wb.save(path)
    return deleted


def customer_history(customer_id: int, limit: int = 20) -> list[dict]:
    _path, _wb, operations, _customers, _summary = _load_book()
    rows: list[dict] = []
    header_row = _find_header_row(operations, OPERATIONS_HEADERS[0]) or 1
    for row in range(header_row + 1, operations.max_row + 1):
        cid = _int_or_none(operations[f"C{row}"].value)
        if cid is None or cid != int(customer_id):
            continue
        rows.append(
            {
                "date": str(operations[f"A{row}"].value or ""),
                "time": str(operations[f"B{row}"].value or ""),
                "customer_id": int(customer_id),
                "customer_name": str(operations[f"D{row}"].value or ""),
                "type": str(operations[f"E{row}"].value or ""),
                "currency": _resolve_currency(operations[f"F{row}"].value),
                "amount": _num(operations[f"G{row}"].value),
                "rate": _num(operations[f"H{row}"].value),
                "amount_uzs": _num(operations[f"I{row}"].value),
                "amount_usd": _num(operations[f"J{row}"].value),
                "note": str(operations[f"K{row}"].value or ""),
                "telegram_user": str(operations[f"L{row}"].value or ""),
                "text": str(operations[f"M{row}"].value or ""),
                "source": str(operations[f"N{row}"].value or ""),
            }
        )
    return rows[-limit:]


def customer_history_by_period(customer_id: int, period: str = "all", limit: int = 50) -> list[dict]:
    rows = customer_history(customer_id, limit=10000)
    filtered = [row for row in rows if _match_period(row["date"], period)]
    return filtered[-limit:]


def customer_operation_count(customer_id: int) -> int:
    return len(customer_history(customer_id, limit=100000))


def summary_report_by_period(period: str = "all") -> dict:
    _path, _wb, operations, _customers, _summary = _load_book()
    jami_kirim_uzs = 0.0
    jami_kirim_usd = 0.0
    jami_chiqim_uzs = 0.0
    jami_chiqim_usd = 0.0
    operatsiyalar_soni = 0
    customer_map: dict[int, dict] = {}
    header_row = _find_header_row(operations, OPERATIONS_HEADERS[0]) or 1
    for row in range(header_row + 1, operations.max_row + 1):
        row_date = str(operations[f"A{row}"].value or "")
        if not _match_period(row_date, period):
            continue
        if not any(operations.cell(row=row, column=col).value not in (None, "") for col in range(1, 15)):
            continue
        operatsiyalar_soni += 1
        amount_uzs = _num(operations[f"I{row}"].value)
        amount_usd = _num(operations[f"J{row}"].value)
        op_type = str(operations[f"E{row}"].value or "").strip().lower()
        customer_id = _int_or_none(operations[f"C{row}"].value)
        customer_name = str(operations[f"D{row}"].value or "").strip()
        if op_type == "kirim":
            jami_kirim_uzs += amount_uzs
            jami_kirim_usd += amount_usd
        elif op_type == "chiqim":
            jami_chiqim_uzs += amount_uzs
            jami_chiqim_usd += amount_usd
        if customer_id is not None:
            item = customer_map.setdefault(
                customer_id,
                {
                    "customer_id": customer_id,
                    "customer_name": customer_name or f"#{customer_id}",
                    "kirim_uzs": 0.0,
                    "kirim_usd": 0.0,
                    "chiqim_uzs": 0.0,
                    "chiqim_usd": 0.0,
                },
            )
            if customer_name:
                item["customer_name"] = customer_name
            if op_type == "kirim":
                item["kirim_uzs"] += amount_uzs
                item["kirim_usd"] += amount_usd
            elif op_type == "chiqim":
                item["chiqim_uzs"] += amount_uzs
                item["chiqim_usd"] += amount_usd
    customers = []
    for customer_id, item in customer_map.items():
        customer = get_customer(customer_id)
        opening_uzs = float(customer.get("opening_uzs", 0) if customer else 0)
        opening_usd = float(customer.get("opening_usd", 0) if customer else 0)
        item["qoldiq_uzs"] = opening_uzs + item["chiqim_uzs"] - item["kirim_uzs"]
        item["qoldiq_usd"] = opening_usd + item["chiqim_usd"] - item["kirim_usd"]
        customers.append(item)
    customers.sort(key=lambda x: x["customer_name"].lower())
    return {
        "jami_kirim_uzs": jami_kirim_uzs,
        "jami_kirim_usd": jami_kirim_usd,
        "jami_chiqim_uzs": jami_chiqim_uzs,
        "jami_chiqim_usd": jami_chiqim_usd,
        "farq_uzs": jami_chiqim_uzs - jami_kirim_uzs,
        "farq_usd": jami_chiqim_usd - jami_kirim_usd,
        "operatsiyalar_soni": operatsiyalar_soni,
        "customers": customers,
    }


def customer_report_by_period(customer_id: int, period: str = "all") -> dict | None:
    customer = get_customer(customer_id)
    if not customer:
        return None
    history = customer_history_by_period(customer_id, period, limit=1000)
    kirim_uzs = sum(float(item["amount_uzs"] or 0) for item in history if item["type"] == "kirim")
    kirim_usd = sum(float(item["amount_usd"] or 0) for item in history if item["type"] == "kirim")
    chiqim_uzs = sum(float(item["amount_uzs"] or 0) for item in history if item["type"] == "chiqim")
    chiqim_usd = sum(float(item["amount_usd"] or 0) for item in history if item["type"] == "chiqim")
    return {
        "customer": customer,
        "history": history[-20:],
        "period": period,
        "kirim_uzs": kirim_uzs,
        "kirim_usd": kirim_usd,
        "chiqim_uzs": chiqim_uzs,
        "chiqim_usd": chiqim_usd,
        "farq_uzs": customer.get("opening_uzs", 0) + chiqim_uzs - kirim_uzs,
        "farq_usd": customer.get("opening_usd", 0) + chiqim_usd - kirim_usd,
    }


def export_report_excel(report_type: str, period: str = "all", customer_id: int | None = None) -> str:
    out_dir = _excel_path().parent
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = out_dir / f"hisobot_{report_type}_{period}_{ts}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Hisobot"

    if report_type == "summary":
        report = summary_report_by_period(period)
        ws.append(["Ko'rsatkich", "UZS", "USD"])
        ws.append(["Mijozlar to'lagan", report["jami_kirim_uzs"], report["jami_kirim_usd"]])
        ws.append(["Biz bergan", report["jami_chiqim_uzs"], report["jami_chiqim_usd"]])
        ws.append(["Qarz qoldiq", report["farq_uzs"], report["farq_usd"]])
        ws.append(["Operatsiyalar soni", report["operatsiyalar_soni"], ""])
        ws.append([])
        ws.append(["Mijoz", "To'lagan_UZS", "To'lagan_USD"])
        ws.append(["", "Bergan_UZS", "Bergan_USD"])
        for item in report["customers"]:
            ws.append([item["customer_name"], item["kirim_uzs"], item["kirim_usd"]])
            ws.append(["", item["chiqim_uzs"], item["chiqim_usd"]])
            ws.append(["", item["qoldiq_uzs"], item["qoldiq_usd"]])
            ws.append([])
    else:
        report = customer_report_by_period(int(customer_id or 0), period)
        if not report:
            raise RuntimeError("Mijoz topilmadi")
        customer = report["customer"]
        ws.append(["Mijoz", customer["name"], ""])
        ws.append(["Telefon", customer.get("phone") or "", ""])
        ws.append(["Ko'rsatkich", "UZS", "USD"])
        ws.append(["Boshlang'ich qarz", customer.get("opening_uzs", 0), customer.get("opening_usd", 0)])
        ws.append(["Mijoz to'lagan", report["kirim_uzs"], report["kirim_usd"]])
        ws.append(["Biz bergan", report["chiqim_uzs"], report["chiqim_usd"]])
        ws.append(["Qarz qoldiq", report["farq_uzs"], report["farq_usd"]])
        ws.append([])
        ws.append(["Sana", "Vaqt", "Turi", "Valyuta", "Summa", "Kurs", "Summa_UZS", "Summa_USD", "Izoh"])
        for item in report["history"]:
            ws.append(
                [
                    item["date"],
                    item["time"],
                    item["type"],
                    item["currency"],
                    item["amount"],
                    item["rate"],
                    item["amount_uzs"],
                    item["amount_usd"],
                    item["note"],
                ]
            )

    for cell in ws[1]:
        cell.font = Font(bold=True)
    wb.save(path)
    return str(path)


def summary_report() -> dict:
    path, wb, operations, customers, summary = _load_book()
    _ensure_customer_formulas(customers)
    _ensure_summary_formulas(summary, operations)
    wb.save(path)
    return summary_report_by_period("all")


def append_operation_row(
    text: str,
    telegram_user_id: int,
    telegram_username: str | None,
    manba: str = "matn",
    customer_id: int | None = None,
    customer_name: str | None = None,
    operation_type: str | None = None,
    amount: float | None = None,
    note: str | None = None,
    currency: str | None = None,
    rate: float | None = None,
) -> str:
    """
    Lokal Excel ga yangi qator qo'shadi.
    Qaytadi: ishlatilgan fayl yo'li.
    """
    path, wb, operations, customers, summary = _load_book()

    turi, summa, izoh = parse_operation_text(text)
    now = datetime.now()
    uname = f"@{telegram_username}" if telegram_username else str(telegram_user_id)
    if operation_type:
        turi = operation_type
    if amount is not None:
        summa = amount
    if note:
        izoh = note
    if customer_id and not customer_name:
        customer = get_customer(customer_id)
        if customer:
            customer_name = customer["name"]

    op_currency = _resolve_currency(currency)
    amount_uzs, amount_usd = _compute_amounts(op_currency, summa, rate)

    target_row = _next_data_row(
        operations,
        OPERATIONS_HEADERS[0],
        data_columns=(1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14),
    )
    values = [
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M:%S"),
        customer_id or "",
        customer_name or "",
        turi or "",
        op_currency,
        summa if summa is not None else "",
        rate if rate not in (None, "") else "",
        amount_uzs if amount_uzs is not None else "",
        amount_usd if amount_usd is not None else "",
        izoh or "",
        uname,
        text[:5000],
        manba,
    ]
    for idx, value in enumerate(values, start=1):
        operations.cell(row=target_row, column=idx, value=value)

    _ensure_customer_formulas(customers)
    _ensure_summary_formulas(summary, operations)
    wb.save(path)
    return str(path)


def append_voice_row(
    text: str,
    telegram_user_id: int,
    telegram_username: str | None,
    customer_id: int | None = None,
    customer_name: str | None = None,
    operation_type: str | None = None,
    amount: float | None = None,
    note: str | None = None,
    currency: str | None = None,
    rate: float | None = None,
) -> str:
    return append_operation_row(
        text,
        telegram_user_id,
        telegram_username,
        manba="ovoz",
        customer_id=customer_id,
        customer_name=customer_name,
        operation_type=operation_type,
        amount=amount,
        note=note,
        currency=currency,
        rate=rate,
    )
