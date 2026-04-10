"""
Excel shablon: Mijozlar + Operatsiyalar + Hisobot (qarzdorlik formulalari).
Ishga tushirish: python scripts/build_excel_template.py
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "templates" / "mijoz_hisob_kitobi.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FILL = PatternFill("solid", fgColor="D9EAF7")
THIN = Side(style="thin", color="D9D9D9")
ALL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_sheet(ws, title: str, subtitle: str | None = None) -> None:
    ws.insert_rows(1, 2)
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(vertical="center")
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(italic=True, color="666666")
    ws.freeze_panes = "A4"


def style_header_row(ws, row_num: int = 3) -> None:
    for cell in ws[row_num]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = ALL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def set_widths(ws, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def style_data_grid(ws, start_row: int, end_row: int, end_col: int) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=end_col):
        for cell in row:
            cell.border = ALL_BORDER
            if cell.column in (1, 2):
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(vertical="center")


def money_format(ws, cells: list[str]) -> None:
    for addr in cells:
        ws[addr].number_format = '#,##0'


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    op = wb.active
    op.title = "Operatsiyalar"
    style_sheet(
        op,
        "Operatsiyalar jurnali",
        "Bot yozadigan asosiy jadval. Mijoz va summa shu yerga tushadi.",
    )
    headers = [
        "Sana",
        "Vaqt",
        "Mijoz_ID",
        "Mijoz_nomi",
        "Turi",
        "Valyuta",
        "Summa",
        "Kurs",
        "Summa_uzs",
        "Summa_usd",
        "Izoh",
        "Telegram_user",
        "Matn",
        "Manba",
    ]
    op.append(headers)
    style_header_row(op, 3)
    set_widths(op, [14, 12, 12, 28, 16, 12, 16, 14, 16, 16, 24, 18, 44, 12])
    style_data_grid(op, 3, 400, len(headers))
    dv_op_type = DataValidation(type="list", formula1='"kirim,chiqim"', allow_blank=True)
    op.add_data_validation(dv_op_type)
    dv_op_type.add("E4:E400")
    dv_currency = DataValidation(type="list", formula1='"UZS,USD"', allow_blank=True)
    op.add_data_validation(dv_currency)
    dv_currency.add("F4:F400")
    for col in ("G", "H", "I", "J"):
        for cell in op[col][3:400]:
            cell.number_format = '#,##0.00'
    for cell in op["G"][3:400]:
        cell.number_format = '#,##0'

    mj = wb.create_sheet("Mijozlar")
    style_sheet(
        mj,
        "Mijozlar bazasi",
        "Mijozlarni shu yerda saqlaysiz. Formulalar avtomatik hisoblaydi.",
    )
    mj.append(
        [
            "ID",
            "Nomi",
            "Telefon",
            "Boshlang'ich_qarz_uzs",
            "Boshlang'ich_qarz_usd",
            "Mijoz_to'lagan_uzs",
            "Mijoz_to'lagan_usd",
            "Biz_bergan_uzs",
            "Biz_bergan_usd",
            "Qarz_qoldiq_uzs",
            "Qarz_qoldiq_usd",
        ]
    )
    style_header_row(mj, 3)
    set_widths(mj, [10, 28, 18, 20, 20, 20, 20, 20, 20, 20, 20])
    # Formulalarni oldindan ko'p qatorga yozib qo'yamiz, Excelda buzilib ketmasin.
    for row in range(4, 304):
        mj[f"F{row}"] = f'=IF(A{row}="","",SUMIFS(Operatsiyalar!$I:$I,Operatsiyalar!$C:$C,A{row},Operatsiyalar!$E:$E,"kirim"))'
        mj[f"G{row}"] = f'=IF(A{row}="","",SUMIFS(Operatsiyalar!$J:$J,Operatsiyalar!$C:$C,A{row},Operatsiyalar!$E:$E,"kirim"))'
        mj[f"H{row}"] = f'=IF(A{row}="","",SUMIFS(Operatsiyalar!$I:$I,Operatsiyalar!$C:$C,A{row},Operatsiyalar!$E:$E,"chiqim"))'
        mj[f"I{row}"] = f'=IF(A{row}="","",SUMIFS(Operatsiyalar!$J:$J,Operatsiyalar!$C:$C,A{row},Operatsiyalar!$E:$E,"chiqim"))'
        mj[f"J{row}"] = f'=IF(A{row}="","",D{row}+H{row}-F{row})'
        mj[f"K{row}"] = f'=IF(A{row}="","",E{row}+I{row}-G{row})'
    style_data_grid(mj, 3, 303, 11)
    for col in ("D", "E", "F", "G", "H", "I", "J", "K"):
        for cell in mj[col][3:303]:
            cell.number_format = '#,##0.00'

    # Qarz mantiqi:
    # kirim = mijoz to'lovi -> qarz kamayadi
    # chiqim = biz berdik -> qarz oshadi

    hs = wb.create_sheet("Hisobot")
    style_sheet(
        hs,
        "Hisobot",
        "Umumiy ko'rsatkichlar. Davr bo'yicha botdan ham alohida hisobot olish mumkin.",
    )
    hs.append(["Ko'rsatkich", "UZS", "USD"])
    style_header_row(hs, 3)
    set_widths(hs, [28, 20, 20])
    hs["A4"] = "Mijozlar to'lagan"
    hs["B4"] = '=SUMIFS(Operatsiyalar!$I:$I,Operatsiyalar!$E:$E,"kirim")'
    hs["C4"] = '=SUMIFS(Operatsiyalar!$J:$J,Operatsiyalar!$E:$E,"kirim")'
    hs["A5"] = "Biz bergan"
    hs["B5"] = '=SUMIFS(Operatsiyalar!$I:$I,Operatsiyalar!$E:$E,"chiqim")'
    hs["C5"] = '=SUMIFS(Operatsiyalar!$J:$J,Operatsiyalar!$E:$E,"chiqim")'
    hs["A6"] = "Jami qarz qoldiq"
    hs["B6"] = "=B5-B4"
    hs["C6"] = "=C5-C4"
    hs["A7"] = "Operatsiyalar soni"
    hs["B7"] = '=COUNTA(Operatsiyalar!$A:$A)-3'
    hs["C7"] = ""
    hs["A9"] = "Izoh"
    hs["B9"] = "Bot: Mijozlar -> amal -> valyuta -> summa -> kurs; Hisobot ikki valyutada chiqadi"
    style_data_grid(hs, 3, 9, 3)
    money_format(hs, ["B4", "B5", "B6", "C4", "C5", "C6"])
    hs.conditional_formatting.add("B6", FormulaRule(formula=["B6>0"], stopIfTrue=False, fill=PatternFill("solid", fgColor="FCE4D6")))
    hs.conditional_formatting.add("B6", FormulaRule(formula=["B6<=0"], stopIfTrue=False, fill=PatternFill("solid", fgColor="E2F0D9")))

    wb.save(OUT)
    print(f"Yaratildi: {OUT}")


if __name__ == "__main__":
    main()
