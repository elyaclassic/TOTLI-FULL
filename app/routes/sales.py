"""
Savdo (sales) — sotuvlar ro'yxati, yangi sotuv, tahrir, tasdiq, revert, o'chirish, POS, qaytarish.
"""
import base64
import io
import json
from datetime import datetime, timedelta
from urllib.parse import quote, unquote
from typing import Optional

import barcode
from barcode.writer import ImageWriter
from fastapi import APIRouter, Request, Depends, Form, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, or_

from app.core import templates
from app.models.database import (
    get_db,
    User,
    Product,
    Partner,
    Warehouse,
    Stock,
    Order,
    OrderItem,
    Payment,
    ProductPrice,
    PriceType,
    Category,
    PosDraft,
    CashRegister,
    CashTransfer,
    ExpenseType,
    StockMovement,
    Employee,
    EmployeeAdvance,
    Driver,
)
from app.deps import require_auth, require_admin


def _check_order_access(order: Order, current_user: User):
    """Admin/manager — hamma buyurtma, boshqalar — faqat o'ziniki."""
    if current_user.role in ("admin", "manager"):
        return
    if order.user_id and order.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Bu buyurtmaga ruxsat yo'q")


from app.services.period_service import is_period_closed
from app.utils.notifications import check_low_stock_and_notify
from app.utils.user_scope import get_warehouses_for_user
from app.utils.audit import log_action
from app.utils.production_order import (
    create_production_from_order,
    get_semi_finished_warehouse,
    get_product_stock_in_warehouse,
    notify_qiyom_operators,
    notify_cutting_packing_operators,
)
from app.utils.db_schema import ensure_orders_payment_due_date_column, ensure_order_item_warehouse_id_column
from app.services.stock_service import create_stock_movement
from app.services.finance_service import sync_cash_balance as _sync_cash_balance
from app.services.pos_helpers import (
    get_pos_price_type as _get_pos_price_type,
    get_pos_warehouses_for_user as _get_pos_warehouses_for_user,
    get_pos_warehouse_for_user as _get_pos_warehouse_for_user,
    get_pos_partner as _get_pos_partner,
    get_pos_cash_register as _get_pos_cash_register,
)
from app.constants import QUERY_LIMIT_DEFAULT, QUERY_LIMIT_HISTORY

router = APIRouter(prefix="/sales", tags=["sales"])


@router.get("", response_class=HTMLResponse)
async def sales_list(
    request: Request,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    warehouse_id: Optional[str] = None,
    status: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_dir: Optional[str] = None,
    page: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    from urllib.parse import unquote
    from app.utils.pagination import paginate, pagination_query_string
    from sqlalchemy.orm import subqueryload
    q = db.query(Order).options(
        subqueryload(Order.partner), subqueryload(Order.warehouse)
    ).filter(Order.type == "sale")
    if date_from and date_from.strip():
        q = q.filter(Order.date >= date_from.strip()[:10] + " 00:00:00")
    if date_to and date_to.strip():
        q = q.filter(Order.date <= date_to.strip()[:10] + " 23:59:59")
    wh_id = None
    if warehouse_id and str(warehouse_id).strip().isdigit():
        try:
            wh_id = int(warehouse_id)
        except (ValueError, TypeError):
            pass
    if wh_id is not None and wh_id > 0:
        q = q.filter(Order.warehouse_id == wh_id)
    status_filter = (status or "").strip()
    if status_filter:
        q = q.filter(Order.status == status_filter)
    sort_col = (sort_by or "date").strip().lower()
    sort_order = (sort_dir or "desc").strip().lower()
    if sort_order not in ("asc", "desc"):
        sort_order = "desc"
    if sort_col == "number":
        q = q.order_by(Order.number.asc() if sort_order == "asc" else Order.number.desc())
    elif sort_col == "date":
        q = q.order_by(Order.date.asc() if sort_order == "asc" else Order.date.desc())
    elif sort_col == "partner":
        q = q.outerjoin(Partner, Order.partner_id == Partner.id).order_by(
            Partner.name.asc() if sort_order == "asc" else Partner.name.desc()
        )
    elif sort_col == "warehouse":
        q = q.outerjoin(Warehouse, Order.warehouse_id == Warehouse.id).order_by(
            Warehouse.name.asc() if sort_order == "asc" else Warehouse.name.desc()
        )
    elif sort_col == "total":
        q = q.order_by(Order.total.asc() if sort_order == "asc" else Order.total.desc())
    elif sort_col == "status":
        q = q.order_by(Order.status.asc() if sort_order == "asc" else Order.status.desc())
    else:
        q = q.order_by(Order.date.desc())

    pg = paginate(q, page or 1, per_page=50)
    orders = pg["items"]

    from sqlalchemy import func as sa_func
    from app.services.sales_metrics import SALE_REALIZED
    stats_row = db.query(
        sa_func.coalesce(sa_func.sum(Order.total), 0),
        sa_func.coalesce(sa_func.sum(Order.debt), 0),
        sa_func.count(Order.id),
    ).filter(
        Order.type == "sale",
        Order.status.in_(SALE_REALIZED),
    )
    if date_from and date_from.strip():
        stats_row = stats_row.filter(Order.date >= date_from.strip()[:10] + " 00:00:00")
    if date_to and date_to.strip():
        stats_row = stats_row.filter(Order.date <= date_to.strip()[:10] + " 23:59:59")
    if wh_id is not None and wh_id > 0:
        stats_row = stats_row.filter(Order.warehouse_id == wh_id)
    total_sum, qarz_sum, completed_count = stats_row.one()
    total_sum = float(total_sum or 0)
    qarz_sum = float(qarz_sum or 0)
    draft_q = db.query(sa_func.count(Order.id)).filter(
        Order.type == "sale", Order.status == "draft"
    )
    if date_from and date_from.strip():
        draft_q = draft_q.filter(Order.date >= date_from.strip()[:10] + " 00:00:00")
    if date_to and date_to.strip():
        draft_q = draft_q.filter(Order.date <= date_to.strip()[:10] + " 23:59:59")
    if wh_id is not None and wh_id > 0:
        draft_q = draft_q.filter(Order.warehouse_id == wh_id)
    draft_count = int(draft_q.scalar() or 0)

    pay_stats = db.query(Payment.payment_type, sa_func.sum(Payment.amount)).join(
        Order, Order.id == Payment.order_id
    ).filter(
        Order.type == "sale",
        Order.status.in_(SALE_REALIZED),
        Payment.type == "income",
        Payment.status == "confirmed",
    )
    if date_from and date_from.strip():
        pay_stats = pay_stats.filter(Order.date >= date_from.strip()[:10] + " 00:00:00")
    if date_to and date_to.strip():
        pay_stats = pay_stats.filter(Order.date <= date_to.strip()[:10] + " 23:59:59")
    if wh_id is not None and wh_id > 0:
        pay_stats = pay_stats.filter(Order.warehouse_id == wh_id)
    pay_stats = pay_stats.group_by(Payment.payment_type).all()
    pay_map = {(pt or "").strip().lower(): float(s or 0) for pt, s in pay_stats}
    naqd_sum = pay_map.get("cash", 0) + pay_map.get("naqd", 0)
    plastik_sum = pay_map.get("card", 0) + pay_map.get("plastik", 0)
    terminal_sum = pay_map.get("terminal", 0)
    click_sum = pay_map.get("click", 0)

    # Chegirma + Tannarx + Foyda (faqat admin/manager/raxbar)
    _role = (getattr(current_user, "role", None) or "").strip().lower()
    show_profit = _role in ("admin", "manager", "rahbar", "raxbar")
    chegirma_sum = 0.0
    tannarx_sum = 0.0
    foyda_sum = 0.0
    foyda_margin_pct = 0.0
    if show_profit:
        # Chegirma — subtotal − total (admin/manager qatori)
        chg_q = db.query(sa_func.coalesce(sa_func.sum(Order.subtotal - Order.total), 0)).filter(
            Order.type == "sale",
            Order.status.in_(SALE_REALIZED),
        )
        if date_from and date_from.strip():
            chg_q = chg_q.filter(Order.date >= date_from.strip()[:10] + " 00:00:00")
        if date_to and date_to.strip():
            chg_q = chg_q.filter(Order.date <= date_to.strip()[:10] + " 23:59:59")
        if wh_id is not None and wh_id > 0:
            chg_q = chg_q.filter(Order.warehouse_id == wh_id)
        chegirma_sum = float(chg_q.scalar() or 0)

        # Tannarx — qty * purchase_price (item-level JOIN)
        tnx_q = db.query(
            sa_func.coalesce(sa_func.sum(OrderItem.quantity * Product.purchase_price), 0)
        ).join(
            Order, Order.id == OrderItem.order_id
        ).join(
            Product, Product.id == OrderItem.product_id
        ).filter(
            Order.type == "sale",
            Order.status.in_(SALE_REALIZED),
        )
        if date_from and date_from.strip():
            tnx_q = tnx_q.filter(Order.date >= date_from.strip()[:10] + " 00:00:00")
        if date_to and date_to.strip():
            tnx_q = tnx_q.filter(Order.date <= date_to.strip()[:10] + " 23:59:59")
        if wh_id is not None and wh_id > 0:
            tnx_q = tnx_q.filter(Order.warehouse_id == wh_id)
        tannarx_sum = float(tnx_q.scalar() or 0)

        foyda_sum = total_sum - tannarx_sum
        foyda_margin_pct = (foyda_sum / total_sum * 100) if total_sum else 0.0

    warehouses = get_warehouses_for_user(db, current_user)
    error = request.query_params.get("error")
    error_detail = unquote(request.query_params.get("detail", "") or "")
    info = request.query_params.get("info")
    info_detail = unquote(request.query_params.get("detail", "") or "") if info else ""
    sort_by_val = sort_col if sort_col in ("number", "date", "partner", "warehouse", "total", "status") else "date"
    filter_params = {
        "date_from": (date_from or "").strip()[:10] or "",
        "date_to": (date_to or "").strip()[:10] or "",
        "warehouse_id": str(wh_id) if wh_id else "",
        "status": status_filter or "",
        "sort_by": sort_by_val,
        "sort_dir": sort_order,
    }
    pq = pagination_query_string(filter_params)
    active_drivers = db.query(Driver).filter(Driver.is_active == True).order_by(Driver.full_name).all()
    today_iso = datetime.now().date().isoformat()
    return templates.TemplateResponse("sales/list.html", {
        "request": request,
        "orders": orders,
        "total_sum": total_sum,
        "naqd_sum": naqd_sum,
        "plastik_sum": plastik_sum,
        "terminal_sum": terminal_sum,
        "click_sum": click_sum,
        "qarz_sum": qarz_sum,
        "warehouses": warehouses,
        "date_from": (date_from or "").strip()[:10] or None,
        "date_to": (date_to or "").strip()[:10] or None,
        "selected_warehouse_id": wh_id,
        "selected_status": status_filter,
        "sort_by": sort_by_val,
        "sort_dir": sort_order,
        "filter_params": "&".join(f"{k}={v}" for k, v in filter_params.items() if v),
        "page": pg["page"],
        "per_page": pg["per_page"],
        "total_count": pg["total_count"],
        "total_pages": pg["total_pages"],
        "items_count": pg["items_count"],
        "base_url": "/sales",
        "pagination_query": pq,
        "completed_count": completed_count,
        "draft_count": draft_count,
        "show_profit": show_profit,
        "chegirma_sum": chegirma_sum,
        "tannarx_sum": tannarx_sum,
        "foyda_sum": foyda_sum,
        "foyda_margin_pct": foyda_margin_pct,
        "page_title": "Sotuvlar",
        "current_user": current_user,
        "error": error,
        "error_detail": error_detail,
        "active_drivers": active_drivers,
        "today_iso": today_iso,
    })


@router.get("/new", response_class=HTMLResponse)
async def sales_new(
    request: Request,
    price_type_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    ensure_order_item_warehouse_id_column(db)
    products = db.query(Product).options(
        joinedload(Product.unit),
    ).filter(
        Product.type.in_(["tayyor", "yarim_tayyor", "hom_ashyo", "material"]),
        Product.is_active == True,
    ).options(joinedload(Product.unit)).order_by(Product.name).all()
    partners = db.query(Partner).filter(Partner.is_active == True).order_by(Partner.name).all()
    warehouses = get_warehouses_for_user(db, current_user)
    price_types = db.query(PriceType).filter(PriceType.is_active == True).order_by(PriceType.name).all()
    # Manager uchun default — "Chakana" (foydalanuvchi xohlasa boshqa narx turini tanlaydi).
    # Boshqa rollar uchun avvalgi xulq (birinchi narx turi).
    current_pt_id = price_type_id
    if not current_pt_id and price_types:
        role_l = (current_user.role or "").strip().lower()
        if role_l in ("manager", "menejer"):
            chakana = next((pt for pt in price_types if (pt.name or "").strip().lower() == "chakana"), None)
            current_pt_id = chakana.id if chakana else price_types[0].id
        else:
            current_pt_id = price_types[0].id
    product_prices_by_type = {}
    if current_pt_id:
        pps = db.query(ProductPrice).filter(ProductPrice.price_type_id == current_pt_id).all()
        product_prices_by_type = {pp.product_id: pp.sale_price for pp in pps}
    warehouse_products = {}
    warehouse_stock_quantities = {}
    role = (current_user.role or "").strip()
    show_all_warehouses = role in ("admin", "manager")
    for wh in warehouses:
        rows = (
            db.query(Stock.product_id)
            .filter(Stock.warehouse_id == wh.id)
            .group_by(Stock.product_id)
            .having(func.sum(Stock.quantity) > 0)
            .all()
        )
        warehouse_products[str(wh.id)] = [r[0] for r in rows]
        qty_rows = (
            db.query(Stock.product_id, func.coalesce(func.sum(Stock.quantity), 0).label("total"))
            .filter(Stock.warehouse_id == wh.id)
            .group_by(Stock.product_id)
            .all()
        )
        warehouse_stock_quantities[str(wh.id)] = {str(r[0]): float(r[1] or 0) for r in qty_rows}
    product_warehouse_quantities = {}
    if show_all_warehouses and warehouses:
        all_pids = set()
        all_qty = {}
        for wh in warehouses:
            for pid in warehouse_products.get(str(wh.id), []):
                all_pids.add(pid)
            qty = warehouse_stock_quantities.get(str(wh.id), {})
            for pid, q in qty.items():
                all_qty[pid] = all_qty.get(pid, 0) + q
                if pid not in product_warehouse_quantities:
                    product_warehouse_quantities[pid] = {}
                product_warehouse_quantities[pid][str(wh.id)] = float(q)
        warehouse_products["all"] = list(all_pids)
        warehouse_stock_quantities["all"] = {str(k): v for k, v in all_qty.items()}
    return templates.TemplateResponse("sales/new.html", {
        "request": request,
        "products": products,
        "partners": partners,
        "warehouses": warehouses,
        "show_all_warehouses": show_all_warehouses,
        "price_types": price_types,
        "current_price_type_id": current_pt_id,
        "product_prices_by_type": product_prices_by_type,
        "warehouse_products": warehouse_products,
        "warehouse_stock_quantities": warehouse_stock_quantities,
        "product_warehouse_quantities": product_warehouse_quantities,
        "current_user": current_user,
        "page_title": "Yangi sotuv",
        "now_iso": datetime.now().strftime("%Y-%m-%dT%H:%M"),
    })


@router.post("/create")
async def sales_create(
    request: Request,
    partner_id: int = Form(...),
    warehouse_id: int = Form(...),
    price_type_id: Optional[int] = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    form = await request.form()
    product_ids = [int(x) for x in form.getlist("product_id") if str(x).strip().isdigit()]
    quantities_raw = form.getlist("quantity")
    prices_raw = form.getlist("price")
    warehouse_ids_raw = form.getlist("warehouse_id")
    quantities = []
    for q in quantities_raw:
        try:
            val = float(q)
            if 0 < val < 1_000_000:
                quantities.append(val)
        except (ValueError, TypeError):
            pass
    prices = []
    for p in prices_raw:
        try:
            val = float(p)
            if 0 <= val < 1_000_000_000:
                prices.append(val)
        except (ValueError, TypeError):
            pass
    warehouse_ids = []
    for w in warehouse_ids_raw:
        try:
            if str(w).strip().isdigit():
                warehouse_ids.append(int(w))
            else:
                warehouse_ids.append(None)
        except (ValueError, TypeError):
            warehouse_ids.append(None)
    # Sana — foydalanuvchi backdate qilishi mumkin (closed period va kelajak guard)
    sale_date_str = (form.get("sale_date") or "").strip()
    order_date = None
    if sale_date_str:
        for fmt in ("%Y-%m-%dT%H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
            try:
                order_date = datetime.strptime(sale_date_str, fmt)
                break
            except ValueError:
                continue
    if order_date is None:
        order_date = datetime.now()
    if order_date > datetime.now():
        order_date = datetime.now()  # kelajak rad — hozirgi vaqtga clamp
    if is_period_closed(db, order_date):
        return RedirectResponse(
            url="/sales/new?error=" + quote("Tanlangan sana yopiq davrga to'g'ri keladi. Boshqa sana tanlang."),
            status_code=303,
        )
    last_order = db.query(Order).filter(Order.type == "sale").order_by(Order.id.desc()).first()
    new_number = f"S-{order_date.strftime('%Y%m%d')}-{(last_order.id + 1) if last_order else 1:04d}"
    order = Order(
        number=new_number,
        type="sale",
        partner_id=partner_id,
        warehouse_id=warehouse_id,
        price_type_id=price_type_id if price_type_id else None,
        status="draft",
        date=order_date,
    )
    db.add(order)
    db.commit()
    db.refresh(order)
    for i in range(min(len(product_ids), len(quantities))):
        pid, qty = product_ids[i], float(quantities[i])
        if pid and qty > 0:
            item_wh_id = warehouse_ids[i] if i < len(warehouse_ids) and warehouse_ids[i] else warehouse_id
            price = prices[i] if i < len(prices) and prices[i] >= 0 else None
            if price is None or price < 0:
                pp = db.query(ProductPrice).filter(
                    ProductPrice.product_id == pid,
                    ProductPrice.price_type_id == order.price_type_id,
                ).first()
                price = pp.sale_price or 0 if pp else 0
                if not price:
                    prod = db.query(Product).filter(Product.id == pid).first()
                    price = (prod.sale_price or prod.purchase_price or 0) if prod else 0
            total_row = qty * price
            db.add(OrderItem(order_id=order.id, product_id=pid, warehouse_id=item_wh_id, quantity=qty, price=price, total=total_row))
            order.subtotal = (order.subtotal or 0) + total_row
            order.total = (order.total or 0) + total_row
    db.commit()
    return RedirectResponse(url=f"/sales/edit/{order.id}", status_code=303)


@router.get("/exchange/{order_id}", response_class=HTMLResponse)
async def sales_exchange_detail(
    request: Request,
    order_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Obmen hujjati ko'rinishi — bir parent (return_sale) va child (sale) ni birga ko'rsatadi."""
    parent = db.query(Order).options(
        joinedload(Order.items).joinedload(OrderItem.product),
        joinedload(Order.partner),
        joinedload(Order.agent),
    ).filter(Order.id == order_id).first()
    if not parent:
        raise HTTPException(status_code=404, detail="Obmen topilmadi")
    if parent.parent_order_id:
        # Child sahifaga kirilsa, parent ga redirect
        return RedirectResponse(url=f"/sales/exchange/{parent.parent_order_id}", status_code=303)
    child = db.query(Order).options(
        joinedload(Order.items).joinedload(OrderItem.product),
    ).filter(Order.parent_order_id == parent.id).first()
    if not child:
        raise HTTPException(status_code=404, detail="Obmen ning sale qismi topilmadi (yarim hujjat)")
    _check_order_access(parent, current_user)
    return templates.TemplateResponse("sales/exchange_detail.html", {
        "request": request,
        "parent": parent,        # qaytgan tovar (return_sale)
        "child": child,          # yangi tovar (sale)
        "current_user": current_user,
        "page_title": f"Обмен: {parent.number} ↔ {child.number}",
    })


@router.get("/edit/{order_id}", response_class=HTMLResponse)
async def sales_edit(
    request: Request,
    order_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    from urllib.parse import unquote
    order = db.query(Order).options(
        joinedload(Order.items).joinedload(OrderItem.product),
    ).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Sotuv topilmadi")
    # Obmen orderlar uchun maxsus sahifaga redirect
    if order.type == "return_sale":
        if order.parent_order_id:
            return RedirectResponse(url=f"/sales/exchange/{order.parent_order_id}", status_code=303)
        has_child = db.query(Order.id).filter(Order.parent_order_id == order.id).first()
        if has_child:
            return RedirectResponse(url=f"/sales/exchange/{order.id}", status_code=303)
        raise HTTPException(status_code=404, detail="Yarim obmen hujjati — qaytadan obmen urilishi kerak")
    if order.parent_order_id and order.type == "sale":
        return RedirectResponse(url=f"/sales/exchange/{order.parent_order_id}", status_code=303)
    if order.type != "sale":
        raise HTTPException(status_code=404, detail="Sotuv topilmadi")
    _check_order_access(order, current_user)
    products = db.query(Product).filter(
        Product.type.in_(["tayyor", "yarim_tayyor"]),
        Product.is_active == True,
    ).order_by(Product.name).all()
    product_prices_by_type = {}
    if order.price_type_id:
        pps = db.query(ProductPrice).filter(ProductPrice.price_type_id == order.price_type_id).all()
        product_prices_by_type = {pp.product_id: pp.sale_price for pp in pps}
    error = request.query_params.get("error")
    error_detail = unquote(request.query_params.get("detail", "") or "")
    info = request.query_params.get("info")
    info_detail = unquote(request.query_params.get("detail", "") or "") if info else ""
    foyda_zarar = 0
    for item in order.items:
        cost = 0
        if item.product:
            # Tannarxni stock.cost_price dan olish (ombordagi haqiqiy tannarx)
            item_wh = getattr(item, "warehouse_id", None) or order.warehouse_id
            st = db.query(Stock).filter(
                Stock.warehouse_id == item_wh,
                Stock.product_id == item.product_id,
            ).first()
            if st and st.cost_price and st.cost_price > 0:
                cost = st.cost_price
            # cost_price topilmasa, cost=0 qoladi (noto'g'ri purchase_price ishlatilmaydi)
        foyda_zarar += (item.quantity or 0) * ((item.price or 0) - cost)
    role = (getattr(current_user, "role", None) or "").strip().lower() if current_user else ""
    # Foyda/Zarar faqat admin yoki rahbar/raxbar ko'radi
    show_foyda_zarar = bool(role in ("admin", "rahbar", "raxbar"))
    # Waiting production sababini ko'rsatish — qaysi mahsulot yetishmayapti
    missing_items = []
    if order.status == "waiting_production":
        from app.services.stock_service import compute_missing_items
        missing_items = compute_missing_items(db, order)
    return templates.TemplateResponse("sales/edit.html", {
        "request": request,
        "order": order,
        "products": products,
        "product_prices_by_type": product_prices_by_type,
        "current_user": current_user,
        "page_title": f"Sotuv: {order.number}",
        "error": error,
        "error_detail": error_detail,
        "info": info,
        "info_detail": info_detail,
        "foyda_zarar": foyda_zarar,
        "show_foyda_zarar": show_foyda_zarar,
        "missing_items": missing_items,
        "now_iso": datetime.now().strftime("%Y-%m-%dT%H:%M"),
    })


@router.post("/{order_id}/add-item")
async def sales_add_item(
    order_id: int,
    product_id: int = Form(...),
    quantity: float = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    # return_sale (obmen qaytarish) ham dispatch → driver oqimidan o'tadi.
    order = db.query(Order).filter(
        Order.id == order_id, Order.type.in_(("sale", "return_sale"))
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Sotuv topilmadi")
    _check_order_access(order, current_user)
    if order.status != "draft":
        return RedirectResponse(url=f"/sales/edit/{order_id}", status_code=303)
    price = 0
    pp = db.query(ProductPrice).filter(
        ProductPrice.product_id == product_id,
        ProductPrice.price_type_id == order.price_type_id,
    ).first()
    if pp:
        price = pp.sale_price or 0
    if not price:
        prod = db.query(Product).filter(Product.id == product_id).first()
        price = (prod.sale_price or prod.purchase_price or 0) if prod else 0
    total_row = quantity * price
    # Bir xil mahsulot+narx allaqachon bo'lsa — yangi qator emas, soniga qo'shish
    existing = db.query(OrderItem).filter(
        OrderItem.order_id == order_id,
        OrderItem.product_id == product_id,
        OrderItem.price == price,
    ).first()
    if existing:
        existing.quantity = (existing.quantity or 0) + quantity
        existing.total = (existing.total or 0) + total_row
    else:
        db.add(OrderItem(order_id=order_id, product_id=product_id, quantity=quantity, price=price, total=total_row))
    order.subtotal = (order.subtotal or 0) + total_row
    order.total = (order.total or 0) + total_row
    db.commit()
    return RedirectResponse(url=f"/sales/edit/{order_id}", status_code=303)


@router.post("/{order_id}/add-items")
async def sales_add_items(
    request: Request,
    order_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    order = db.query(Order).filter(Order.id == order_id, Order.type == "sale").first()
    if not order:
        raise HTTPException(status_code=404, detail="Sotuv topilmadi")
    _check_order_access(order, current_user)
    if order.status != "draft":
        return RedirectResponse(url=f"/sales/edit/{order_id}", status_code=303)
    form = await request.form()
    product_ids = [int(x) for x in form.getlist("product_id") if str(x).strip().isdigit()]
    quantities_raw = form.getlist("quantity")
    prices_raw = form.getlist("price")
    quantities = []
    for q in quantities_raw:
        try:
            val = float(q)
            if 0 < val < 1_000_000:
                quantities.append(val)
        except (ValueError, TypeError):
            pass
    prices = []
    for p in prices_raw:
        try:
            val = float(p)
            if 0 <= val < 1_000_000_000:
                prices.append(val)
        except (ValueError, TypeError):
            pass
    for i in range(min(len(product_ids), len(quantities))):
        pid, qty = product_ids[i], quantities[i]
        if not pid or qty <= 0:
            continue
        price = prices[i] if i < len(prices) and prices[i] >= 0 else None
        if price is None or price < 0:
            pp = db.query(ProductPrice).filter(
                ProductPrice.product_id == pid,
                ProductPrice.price_type_id == order.price_type_id,
            ).first()
            price = pp.sale_price or 0 if pp else 0
            if not price:
                prod = db.query(Product).filter(Product.id == pid).first()
                price = (prod.sale_price or prod.purchase_price or 0) if prod else 0
        total_row = qty * price
        existing = db.query(OrderItem).filter(
            OrderItem.order_id == order_id,
            OrderItem.product_id == pid,
            OrderItem.price == price,
            OrderItem.warehouse_id == order.warehouse_id,
        ).first()
        if existing:
            existing.quantity = (existing.quantity or 0) + qty
            existing.total = (existing.total or 0) + total_row
        else:
            db.add(OrderItem(order_id=order_id, product_id=pid, warehouse_id=order.warehouse_id, quantity=qty, price=price, total=total_row))
        order.subtotal = (order.subtotal or 0) + total_row
        order.total = (order.total or 0) + total_row
    db.commit()
    return RedirectResponse(url=f"/sales/edit/{order_id}", status_code=303)


@router.post("/{order_id}/update-item/{item_id}")
async def sales_update_item(
    order_id: int,
    item_id: int,
    quantity: float = Form(...),
    price: float = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Draft sotuv qatorining miqdor/narxini yangilash. Faqat draft holatda."""
    order = db.query(Order).filter(Order.id == order_id, Order.type == "sale").first()
    if not order or order.status != "draft":
        return RedirectResponse(url=f"/sales/edit/{order_id}", status_code=303)
    _check_order_access(order, current_user)
    item = db.query(OrderItem).filter(OrderItem.id == item_id, OrderItem.order_id == order_id).first()
    if not item:
        return RedirectResponse(url=f"/sales/edit/{order_id}", status_code=303)
    if not (0 < quantity < 1_000_000) or not (0 <= price < 1_000_000_000):
        return RedirectResponse(
            url=f"/sales/edit/{order_id}?error=item&detail=" + quote("Miqdor yoki narx noto'g'ri."),
            status_code=303,
        )
    new_total = quantity * price
    delta = new_total - (item.total or 0)
    item.quantity = quantity
    item.price = price
    item.total = new_total
    order.subtotal = (order.subtotal or 0) + delta
    order.total = (order.total or 0) + delta
    db.commit()
    return RedirectResponse(url=f"/sales/edit/{order_id}?message=item-updated", status_code=303)


@router.post("/{order_id}/set-date")
async def sales_set_date(
    order_id: int,
    sale_date: str = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Draft sotuv sanasini o'zgartirish (backdate). Faqat draft holatda.
    Kelajak sana va closed period rad etiladi."""
    order = db.query(Order).filter(Order.id == order_id, Order.type == "sale").first()
    if not order:
        raise HTTPException(status_code=404, detail="Sotuv topilmadi")
    if order.status != "draft":
        return RedirectResponse(
            url=f"/sales/edit/{order_id}?error=" + quote("Faqat qoralama holatdagi sotuv sanasini o'zgartirib bo'ladi."),
            status_code=303,
        )
    new_date = None
    for fmt in ("%Y-%m-%dT%H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            new_date = datetime.strptime((sale_date or "").strip(), fmt)
            break
        except ValueError:
            continue
    if new_date is None:
        return RedirectResponse(
            url=f"/sales/edit/{order_id}?error=" + quote("Sana formati noto'g'ri."),
            status_code=303,
        )
    if new_date > datetime.now():
        return RedirectResponse(
            url=f"/sales/edit/{order_id}?error=" + quote("Kelajakdagi sana qabul qilinmaydi."),
            status_code=303,
        )
    if is_period_closed(db, new_date):
        return RedirectResponse(
            url=f"/sales/edit/{order_id}?error=" + quote("Tanlangan sana yopiq davrga to'g'ri keladi."),
            status_code=303,
        )
    order.date = new_date
    db.commit()
    return RedirectResponse(url=f"/sales/edit/{order_id}?message=date-updated", status_code=303)


@router.post("/{order_id}/confirm")
async def sales_confirm(
    order_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Buyurtmani tasdiqlash. 2 xil flow:

    - POS (source != 'agent'): mijoz do'konda turibdi, darrov yetkaziladi.
      → status='delivered', stock kamayadi, balance += debt. Yetkazib berish yo'q.
    - Agent (source='agent'): keyinroq haydovchi yetkazib beradi.
      → faqat status='confirmed'. Stock va balance keyingi bosqichlarda.
    """
    order = db.query(Order).filter(Order.id == order_id, Order.type == "sale").first()
    if not order:
        raise HTTPException(status_code=404, detail="Sotuv topilmadi")
    _check_order_access(order, current_user)
    if is_period_closed(db, order.date):
        return RedirectResponse(url=f"/sales/edit/{order_id}?error=period_closed", status_code=303)
    if order.status != "draft":
        return RedirectResponse(url=f"/sales/edit/{order_id}", status_code=303)

    is_agent = (order.source or "").strip().lower() == "agent"
    from sqlalchemy import text as _text

    if is_agent:
        # Agent flow — faqat status (dispatch keyinroq)
        claim = db.execute(
            _text("UPDATE orders SET status='confirmed' WHERE id=:id AND type='sale' AND status='draft'"),
            {"id": order_id},
        )
        if claim.rowcount == 0:
            return RedirectResponse(url=f"/sales/edit/{order_id}?already=1", status_code=303)
        db.commit()
        try:
            from app.bot.services.audit_watchdog import audit_sale
            audit_sale(order.id)
        except Exception:
            pass
        try:
            if order.type == "sale":
                from app.bot.customer_bot.notify import notify_customer, msg_order_confirmed
                notify_customer(order.partner_id, msg_order_confirmed(order))
        except Exception:
            pass
        return RedirectResponse(url=f"/sales/edit/{order_id}?confirmed=1", status_code=303)

    # POS flow — stock yetadimi tekshir, yetsa darrov yetkaziladi
    insufficient = []
    for item in order.items:
        wh_id = item.warehouse_id if item.warehouse_id else order.warehouse_id
        stock = db.query(Stock).filter(
            Stock.warehouse_id == wh_id,
            Stock.product_id == item.product_id,
        ).first()
        have = float(stock.quantity or 0) if stock else 0.0
        need = float(item.quantity or 0)
        if have + 1e-6 < need:
            pname = item.product.name if item.product else f"#{item.product_id}"
            insufficient.append(f"{pname}: kerak {need:g}, bor {have:g}")

    if insufficient:
        # POS sotuvida ishlab chiqarish AVTOMAT yaratilmaydi — manager hal qiladi
        detail = "; ".join(insufficient[:5]) + ("; ..." if len(insufficient) > 5 else "")
        return RedirectResponse(
            url=f"/sales/edit/{order_id}?error=shortage&detail=" + quote(
                f"Ombor yetmaydi: {detail}. Buyurtmani tahrirlang yoki ishlab chiqarishni kuting."
            ),
            status_code=303,
        )

    # Atomik UPDATE — status'ni delivered'ga o'tkazish
    claim = db.execute(
        _text("UPDATE orders SET status='delivered' WHERE id=:id AND type='sale' AND status='draft'"),
        {"id": order_id},
    )
    if claim.rowcount == 0:
        return RedirectResponse(url=f"/sales/edit/{order_id}?already=1", status_code=303)

    # Stock kamaytirish
    for item in order.items:
        wh_id = item.warehouse_id if item.warehouse_id else order.warehouse_id
        if not wh_id or not item.product_id:
            continue
        create_stock_movement(
            db=db, warehouse_id=wh_id, product_id=item.product_id,
            quantity_change=-float(item.quantity or 0),
            operation_type="sale", document_type="Sale",
            document_id=order.id, document_number=order.number,
            user_id=current_user.id if current_user else None,
            note=f"POS sotuvi: {order.number}",
            created_at=order.date or datetime.now(),
        )

    # Qarz va balans
    order.debt = max(0.0, (order.total or 0) - (order.paid or 0))
    if order.partner_id:
        from app.services.partner_balance_service import recompute_partner_balance
        db.flush()
        recompute_partner_balance(db, order.partner_id, reason="sale_confirm",
                                  ref=order.number,
                                  actor=current_user.username if current_user else None)

    db.commit()
    check_low_stock_and_notify(db)

    try:
        from app.bot.services.notifier import notify_new_sale, notify_big_sale
        partner = db.query(Partner).filter(Partner.id == order.partner_id).first() if order.partner_id else None
        p_name = partner.name if partner else "Naqd"
        notify_new_sale(order.number, p_name, order.total or 0, order.paid or 0)
        if (order.total or 0) >= 10_000_000:
            notify_big_sale(order.number, p_name, order.total)
    except Exception:
        pass
    try:
        from app.bot.services.audit_watchdog import audit_sale
        audit_sale(order.id)
    except Exception:
        pass

    return RedirectResponse(url=f"/sales/edit/{order_id}?delivered=1", status_code=303)


@router.post("/{order_id}/dispatch")
async def sales_dispatch(
    request: Request,
    order_id: int,
    delivery_date: str = Form(...),
    driver_id: int = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Buyurtmani yo'lga chiqarish: confirmed -> out_for_delivery (yoki waiting_production).

    Stock yetarli bo'lsa: stock kamayadi, Delivery yaratiladi, status=out_for_delivery.
    Stock yetmasa: Production buyurtmasi yaratiladi, status=waiting_production,
    delivery_date + pending_driver_id saqlanadi (production tugagach auto-dispatch).
    """
    from datetime import date as _date, datetime as _dt
    from sqlalchemy import text as _text
    from app.models.database import Driver, Delivery

    if not current_user or getattr(current_user, "role", None) not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")

    order = db.query(Order).filter(Order.id == order_id, Order.type == "sale").first()
    if not order:
        raise HTTPException(status_code=404, detail="Sotuv topilmadi")
    _check_order_access(order, current_user)
    if is_period_closed(db, order.date):
        return RedirectResponse(url=f"/sales/edit/{order_id}?error=period_closed", status_code=303)
    # Faqat agent buyurtmalari yetkazib berishni talab qiladi (POS — darrov)
    if (order.source or "").strip().lower() != "agent":
        return RedirectResponse(
            url=f"/sales/edit/{order_id}?error=non_agent_dispatch&detail=" + quote(
                "POS sotuvi yetkazib berishni talab qilmaydi. Buyurtmani tasdiqlash darrov yetkazib beradi."
            ),
            status_code=303,
        )
    if order.status != "confirmed":
        return RedirectResponse(url=f"/sales/edit/{order_id}?already=1", status_code=303)

    try:
        delivery_d = _date.fromisoformat(delivery_date.strip())
    except ValueError:
        return RedirectResponse(
            url=f"/sales/edit/{order_id}?error=date_format",
            status_code=303,
        )

    drv = db.query(Driver).filter(Driver.id == driver_id, Driver.is_active == True).first()
    if not drv:
        return RedirectResponse(
            url=f"/sales/edit/{order_id}?error=driver_not_found",
            status_code=303,
        )

    items = list(order.items or [])
    if not items:
        return RedirectResponse(
            url=f"/sales/edit/{order_id}?error=no_items",
            status_code=303,
        )

    # Obmen qaytarish (return_sale): qaytgan tovar — stock chiqim YO'Q, stock
    # yetishmovchilik tekshiruvi YO'Q. Faqat out_for_delivery + Delivery.
    # Qaytgan tovar omborga kirimi haydovchi "Yetkazdim" bosganda (api_driver_ops.py).
    if order.type == "return_sale":
        r = db.execute(
            _text("UPDATE orders SET status='out_for_delivery', delivery_date=:dd, "
                  "dispatched_at=:now, pending_driver_id=:drv "
                  "WHERE id=:id AND status='confirmed'"),
            {"id": order_id, "dd": delivery_d, "now": _dt.now(), "drv": driver_id},
        )
        if r.rowcount == 0:
            return RedirectResponse(url=f"/sales/edit/{order_id}?already=1", status_code=303)
        prefix = f"DLV-{delivery_d.strftime('%Y%m%d')}"
        last = (
            db.query(Delivery)
            .filter(Delivery.number.like(f"{prefix}%"))
            .order_by(Delivery.id.desc())
            .first()
        )
        try:
            seq = int(last.number.split("-")[-1]) + 1 if last and last.number else 1
        except Exception:
            seq = 1
        partner_obj = db.query(Partner).filter(Partner.id == order.partner_id).first() if order.partner_id else None
        db.add(Delivery(
            number=f"{prefix}-{seq:04d}",
            order_id=order.id,
            order_number=order.number,
            driver_id=driver_id,
            delivery_address=(partner_obj.address or "") if partner_obj else "",
            latitude=partner_obj.latitude if partner_obj else None,
            longitude=partner_obj.longitude if partner_obj else None,
            planned_date=delivery_d,
            notes=f"OBMEN qaytarish — Mijoz: {partner_obj.name if partner_obj else ''}, Tel: {partner_obj.phone if partner_obj else ''}",
            status="pending",
        ))
        db.commit()
        referer = request.headers.get("referer", "")
        if "/agents/" in referer:
            return RedirectResponse(url=referer, status_code=303)
        return RedirectResponse(url=f"/sales/edit/{order_id}?dispatched=1", status_code=303)

    # Stock yetarli yoki yo'qligini tekshirish (semi-finished fallback bilan)
    semi_warehouse = get_semi_finished_warehouse(db)
    insufficient_items = []
    for item in items:
        wh_id = item.warehouse_id if item.warehouse_id else order.warehouse_id
        stock = db.query(Stock).filter(
            Stock.warehouse_id == wh_id,
            Stock.product_id == item.product_id,
        ).first()
        available = float(stock.quantity) if stock and stock.quantity else 0.0
        if available + 1e-6 < float(item.quantity):
            # Semi-finished omborda yetarli mahsulot bormi?
            semi_avail = 0.0
            if semi_warehouse:
                semi_avail = get_product_stock_in_warehouse(db, semi_warehouse.id, item.product_id)
            if semi_avail >= 1 and semi_avail >= float(item.quantity):
                notify_cutting_packing_operators(
                    db=db, order_number=order.number, order_id=order.id,
                    product_name=(item.product.name if item.product else "Mahsulot"),
                )
                continue
            notify_qiyom_operators(
                db=db, order_number=order.number, order_id=order.id,
                product_name=(item.product.name if item.product else "Mahsulot"),
            )
            insufficient_items.append({
                "product": item.product,
                "required": float(item.quantity),
                "available": available,
            })

    # Yetishmaydigan mahsulot bo'lsa — Production yaratish, status=waiting_production
    if insufficient_items:
        try:
            productions, missing = create_production_from_order(
                db=db, order=order,
                insufficient_items=insufficient_items, current_user=current_user,
            )
            if not productions:
                db.rollback()
                msg = "Ishlab chiqarish buyurtmasi yaratilmadi."
                if missing:
                    msg += " Retsept topilmadi: " + ", ".join(missing[:10]) + ("…" if len(missing) > 10 else "")
                return RedirectResponse(
                    url=f"/sales/edit/{order_id}?error=production&detail=" + quote(msg),
                    status_code=303,
                )

            # waiting_production'ga atomik o'tkazish, delivery_date va driver saqlanadi
            r = db.execute(
                _text("UPDATE orders SET status='waiting_production', "
                      "delivery_date=:dd, pending_driver_id=:drv "
                      "WHERE id=:id AND status='confirmed'"),
                {"id": order_id, "dd": delivery_d, "drv": driver_id},
            )
            if r.rowcount == 0:
                db.rollback()
                return RedirectResponse(url=f"/sales/edit/{order_id}?already=1", status_code=303)
            db.commit()

            production_numbers = ", ".join([p.number for p in productions])
            parts = []
            for it in insufficient_items:
                p = it.get("product")
                name = (getattr(p, "name", None) or "Mahsulot")
                req = float(it.get("required") or 0)
                avail = float(it.get("available") or 0)
                lack = max(req - avail, 0.0)
                parts.append(f"{name}: kerak {req:g}, mavjud {avail:g}, yetmaydi {lack:g}")
            detail_list = "; ".join(parts[:12]) + ("; ..." if len(parts) > 12 else "")
            return RedirectResponse(
                url=f"/sales/edit/{order_id}?info=production&detail=" + quote(
                    f"Yetmayotganlar: {detail_list}. "
                    f"Ishlab chiqarish: {production_numbers}. "
                    f"Tayyor bo'lgach yo'lga chiqariladi."
                ),
                status_code=303,
            )
        except Exception as e:
            db.rollback()
            import traceback
            traceback.print_exc()
            return RedirectResponse(
                url=f"/sales/edit/{order_id}?error=production&detail=" + quote(f"Ishlab chiqarish xato: {str(e)[:200]}"),
                status_code=303,
            )

    # Stock yetarli — atomik out_for_delivery + stock decrement + Delivery yaratish
    r = db.execute(
        _text("UPDATE orders SET status='out_for_delivery', delivery_date=:dd, "
              "dispatched_at=:now, pending_driver_id=:drv "
              "WHERE id=:id AND status='confirmed'"),
        {"id": order_id, "dd": delivery_d, "now": _dt.now(), "drv": driver_id},
    )
    if r.rowcount == 0:
        return RedirectResponse(url=f"/sales/edit/{order_id}?already=1", status_code=303)

    for item in items:
        wh_id = item.warehouse_id if item.warehouse_id else order.warehouse_id
        if not wh_id or not item.product_id:
            continue
        create_stock_movement(
            db=db,
            warehouse_id=wh_id,
            product_id=item.product_id,
            quantity_change=-float(item.quantity or 0),
            operation_type="sale",
            document_type="Sale",
            document_id=order.id,
            document_number=order.number,
            user_id=current_user.id if current_user else None,
            note=f"Sotuv yo'lga chiqarildi: {order.number}",
            created_at=order.date or _dt.now(),
        )

    # Delivery hujjatini yaratish (DLV-YYYYMMDD-NNNN ketma-ket raqam bilan)
    prefix = f"DLV-{delivery_d.strftime('%Y%m%d')}"
    last = (
        db.query(Delivery)
        .filter(Delivery.number.like(f"{prefix}%"))
        .order_by(Delivery.id.desc())
        .first()
    )
    try:
        seq = int(last.number.split("-")[-1]) + 1 if last and last.number else 1
    except Exception:
        seq = 1
    partner_obj = db.query(Partner).filter(Partner.id == order.partner_id).first() if order.partner_id else None
    delivery = Delivery(
        number=f"{prefix}-{seq:04d}",
        order_id=order.id,
        order_number=order.number,
        driver_id=driver_id,
        delivery_address=(partner_obj.address or "") if partner_obj else "",
        latitude=partner_obj.latitude if partner_obj else None,
        longitude=partner_obj.longitude if partner_obj else None,
        planned_date=delivery_d,
        notes=f"Mijoz: {partner_obj.name if partner_obj else ''}, Tel: {partner_obj.phone if partner_obj else ''}",
        status="pending",
    )
    db.add(delivery)
    db.commit()

    check_low_stock_and_notify(db)

    # Telegram bildirish
    try:
        from app.bot.services.notifier import notify_new_sale, notify_big_sale
        p_name = partner_obj.name if partner_obj else "Naqd"
        notify_new_sale(order.number, p_name, order.total or 0, order.paid or 0)
        if (order.total or 0) >= 10_000_000:
            notify_big_sale(order.number, p_name, order.total)
    except Exception:
        pass

    try:
        from app.bot.services.audit_watchdog import audit_sale
        audit_sale(order.id)
    except Exception:
        pass

    try:
        from app.bot.customer_bot.notify import notify_customer, msg_order_dispatched
        notify_customer(order.partner_id, msg_order_dispatched(order))
    except Exception:
        pass

    # Agar foydalanuvchi agent detail sahifasidan kelgan bo'lsa, o'sha sahifaga qaytarish
    # (browser back tugmasi yaxshi ishlasin — modal qayta ochilmasin).
    referer = request.headers.get("referer", "")
    if "/agents/" in referer:
        return RedirectResponse(url=referer, status_code=303)
    return RedirectResponse(url=f"/sales/edit/{order_id}?dispatched=1", status_code=303)


@router.post("/{order_id}/delete-item/{item_id}")
async def sales_delete_item(
    order_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    order = db.query(Order).filter(Order.id == order_id, Order.type == "sale").first()
    if not order or order.status != "draft":
        return RedirectResponse(url=f"/sales/edit/{order_id}", status_code=303)
    _check_order_access(order, current_user)
    item = db.query(OrderItem).filter(OrderItem.id == item_id, OrderItem.order_id == order_id).first()
    if item:
        order.total = (order.total or 0) - (item.total or 0)
        order.subtotal = (order.subtotal or 0) - (item.total or 0)
        db.delete(item)
        db.commit()
    return RedirectResponse(url=f"/sales/edit/{order_id}", status_code=303)


def _revert_return_sale_exchange(db, order, current_user):
    """Qaytarish/almashtirish (return_sale) tasdiqini bekor qilish — XAVFSIZ minimal.

    Faqat yetkazilmagan (status='confirmed') va to'lovsiz almashtirishni DRAFT holatga
    qaytaradi (tahrirlab qayta tasdiqlash uchun): qaytarish (parent) + juft sotuv (child)
    ikkalasi 'draft' bo'ladi + partner balans recompute. Oddiy sotuv revert kabi.
    Yetkazilgan ('delivered'/'completed') yoki to'lovli almashtirishni RAD etadi (to'liq Sub-2 keyin).
    """
    child = db.query(Order).filter(Order.parent_order_id == order.id, Order.type == "sale").first()
    docs = [order] + ([child] if child else [])
    doc_ids = [d.id for d in docs]
    # Xavfsizlik: yetkazilgan (stock harakatlangan) yoki to'lovli bo'lsa bu yerda bekor qilmaymiz
    if any((d.status or "") in ("delivered", "completed") for d in docs):
        return RedirectResponse(
            url="/sales?error=exchange_revert&detail=" + quote(
                "Yetkazilgan almashtirishni bu yerdan bekor qilib bo'lmaydi. Admin bilan bog'laning."),
            status_code=303)
    if db.query(Payment.id).filter(Payment.order_id.in_(doc_ids)).first():
        return RedirectResponse(
            url="/sales?error=exchange_revert&detail=" + quote(
                "To'lovli almashtirishni bu yerdan bekor qilib bo'lmaydi. Admin bilan bog'laning."),
            status_code=303)
    affected = set()
    for d in docs:
        d.status = "draft"
        if d.partner_id:
            affected.add(d.partner_id)
    db.flush()
    from app.services.partner_balance_service import recompute_partner_balance
    for pid in affected:
        recompute_partner_balance(db, pid, reason="exchange_revert", ref=order.number,
                                  actor=current_user.username if current_user else None)
    db.commit()
    return RedirectResponse(url="/sales?reverted=1&number=" + quote(order.number), status_code=303)


@router.post("/{order_id}/revert")
async def sales_revert(
    order_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Buyurtma topilmadi")
    if order.type == "return_sale":
        return _revert_return_sale_exchange(db, order, current_user)
    if order.type != "sale":
        raise HTTPException(status_code=404, detail="Sotuv topilmadi")
    # Bug 1 — paid > 0 bo'lgan orderni revert qilish taqiqlanadi (refund kerak)
    if (order.paid or 0) > 0:
        return RedirectResponse(
            url=f"/sales/edit/{order_id}?error=paid_block&detail=" + quote(
                f"Bu sotuvga {order.paid:,.0f} so'm to'lov qabul qilingan. Avval refund (qaytarish) hujjati yarating, keyin tasdiqni bekor qilishingiz mumkin."
            ),
            status_code=303,
        )
    status = (getattr(order, "status", None) or "").strip().lower()
    if status in ("waiting_production", "confirmed"):
        # Ombor hisobdan chiqarilmagan — faqat draft ga qaytarish
        order.status = "draft"
        # Confirmed bo'lsa — tegishli yetkazishni ham bekor qilish
        # ('pending', 'in_progress', 'failed' — yetkazilmagan har qanday holat)
        if status == "confirmed":
            from app.models.database import Delivery as DeliveryModel
            for delivery in db.query(DeliveryModel).filter(
                DeliveryModel.order_id == order.id,
                DeliveryModel.status.in_(["pending", "in_progress", "failed"]),
            ).all():
                delivery.status = "cancelled"
        db.commit()
        referer = "/sales/edit/" + str(order_id)
        return RedirectResponse(url=referer, status_code=303)
    if status == "out_for_delivery":
        # /dispatch stock kamaytirgan — qaytarish kerak (balance YO'Q, hali yozilmagan)
        from app.services.stock_service import create_stock_movement
        for item in order.items:
            wh_id = item.warehouse_id if item.warehouse_id else order.warehouse_id
            if not wh_id or not item.product_id:
                continue
            create_stock_movement(
                db=db,
                warehouse_id=wh_id,
                product_id=item.product_id,
                quantity_change=float(item.quantity or 0),
                operation_type="dispatch_revert",
                document_type="Sale",
                document_id=order.id,
                document_number=order.number,
                user_id=current_user.id if current_user else None,
                note=f"Yo'lga chiqarishni bekor qilish: {order.number}",
                created_at=order.date or datetime.now(),
            )
        # Yetkazilmagan delivery'larni cancel qilish
        from app.models.database import Delivery as DeliveryModel
        for delivery in db.query(DeliveryModel).filter(
            DeliveryModel.order_id == order.id,
            DeliveryModel.status.in_(["pending", "in_progress", "failed", "picked_up"]),
        ).all():
            delivery.status = "cancelled"
        # delivery_date va pending_driver_id'ni NULL qilamiz (qayta dispatch uchun toza holat)
        order.delivery_date = None
        order.dispatched_at = None
        order.pending_driver_id = None
        order.status = "draft"
        db.commit()
        return RedirectResponse(url=f"/sales/edit/{order_id}", status_code=303)
    if status not in ("completed", "delivered"):
        return RedirectResponse(
            url=f"/sales/edit/{order_id}?error=revert&detail=" + quote("Faqat yo'lga chiqqan yoki yetkazilgan sotuvning tasdiqini bekor qilish mumkin."),
            status_code=303,
        )
    from app.services.stock_service import create_stock_movement
    for item in order.items:
        wh_id = item.warehouse_id if item.warehouse_id else order.warehouse_id
        if not wh_id or not item.product_id:
            continue
        create_stock_movement(
            db=db,
            warehouse_id=wh_id,
            product_id=item.product_id,
            quantity_change=float(item.quantity or 0),
            operation_type="sale_revert",
            document_type="Sale",
            document_id=order.id,
            document_number=order.number,
            user_id=current_user.id if current_user else None,
            note=f"Sotuv tasdiqini bekor qilish: {order.number}",
            created_at=order.date or datetime.now(),
        )
    order.previous_partner_balance = None
    # Yetkazilmagan delivery larni cancel qilish (cancelled/failed/delivered emasini)
    from app.models.database import Delivery as DeliveryModel
    for delivery in db.query(DeliveryModel).filter(
        DeliveryModel.order_id == order.id,
        DeliveryModel.status.in_(["pending", "in_progress", "failed"]),
    ).all():
        delivery.status = "cancelled"
    order.status = "draft"
    if order.partner_id:
        from app.services.partner_balance_service import recompute_partner_balance
        db.flush()
        recompute_partner_balance(db, order.partner_id, reason="sale_revert",
                                  ref=order.number,
                                  actor=current_user.username if current_user else None)
    db.commit()
    return RedirectResponse(url=f"/sales/edit/{order_id}", status_code=303)


@router.get("/{order_id}/nakladnoy", response_class=HTMLResponse)
async def sales_nakladnoy(
    request: Request,
    order_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Sotuv nakladnoy — tasdiqlangan sotuv uchun chop etish."""
    order = (
        db.query(Order)
        .options(
            joinedload(Order.items).joinedload(OrderItem.product).joinedload(Product.unit),
            joinedload(Order.partner),
            joinedload(Order.warehouse),
            joinedload(Order.user),
            joinedload(Order.price_type),
        )
        .filter(Order.id == order_id, Order.type == "sale")
        .first()
    )
    if not order:
        raise HTTPException(status_code=404, detail="Sotuv topilmadi")
    return templates.TemplateResponse("sales/nakladnoy.html", {
        "request": request,
        "order": order,
        "current_user": current_user,
    })


@router.get("/nakladnoy/excel/bulk")
async def sales_nakladnoy_excel_bulk(
    ids: str = "",
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Bir nechta sotuv yuk xatini bitta Excel da jamlama. ids=1,2,3 formatda."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.worksheet.pagebreak import Break
    from fastapi.responses import StreamingResponse

    try:
        order_ids = [int(x) for x in ids.split(",") if x.strip().isdigit()]
    except Exception:
        order_ids = []
    if not order_ids:
        raise HTTPException(status_code=400, detail="Hech qanday buyurtma tanlanmagan")

    orders = (
        db.query(Order)
        .options(
            joinedload(Order.items).joinedload(OrderItem.product).joinedload(Product.unit),
            joinedload(Order.partner),
            joinedload(Order.warehouse),
            joinedload(Order.agent),
        )
        .filter(Order.id.in_(order_ids), Order.type.in_(("sale", "return_sale")))
        .order_by(Order.id)
        .all()
    )
    if not orders:
        raise HTTPException(status_code=404, detail="Buyurtmalar topilmadi")

    # Obmen parentlari uchun child sale ni avtomatik qo'shish
    parent_return_ids = [o.id for o in orders if o.type == "return_sale" and o.parent_order_id is None]
    exchange_children: dict[int, Order] = {}
    if parent_return_ids:
        children = (
            db.query(Order)
            .options(joinedload(Order.items).joinedload(OrderItem.product).joinedload(Product.unit))
            .filter(Order.parent_order_id.in_(parent_return_ids))
            .all()
        )
        for ch in children:
            exchange_children[ch.parent_order_id] = ch

    # Partner bo'yicha gruhlash (kontragent → orderlar)
    from collections import OrderedDict
    by_partner: OrderedDict = OrderedDict()
    for o in orders:
        # Child sale (parent_order_id) — exchange_children orqali avtomatik kiradi, ro'yxatda alohida ko'rsatmaymiz
        if o.parent_order_id and o.type == "sale":
            continue
        key = o.partner_id or 0
        by_partner.setdefault(key, []).append(o)

    wb = Workbook()
    ws = wb.active
    ws.title = "Yuk xati (jamlama)"

    # A4 portrait (knijniy) + fit to width — jadval bir sahifa kengligiga sig'adi
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # balandligi avtomatik (ko'p sahifa)
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.print_options.horizontalCentered = True

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    section_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    bold_white = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    row = 1
    today_str = datetime.now().strftime("%d.%m.%Y")

    section_send_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    section_ret_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
    section_buyurtma_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    HEADERS = ["№", "Kodi", "Nomi", "O'lchov birligi", "Soni", "Narxi", "Summa"]

    # A4 portrait amalda ~62-65 row sig'adi (banner/title rowlar balandroq).
    # Har yuk xati uchun kerakli qator sonini oldindan hisoblab, joy yetmasa
    # manual page break qo'yamiz — yuk xati butunligicha ko'chadi.
    ROWS_PER_PAGE = 62

    def calc_partner_rows(p_orders: list) -> int:
        # Header bloki: 1 sarlavha + 3 info + 1 spacer = 5
        n = 5
        for o in p_orders:
            if o.type == "sale" and not o.parent_order_id:
                # section: banner(1) + headers(1) + items(N) + jami(1) + spacer(1) = N + 4
                n += len(o.items) + 4
        for o in p_orders:
            if o.type != "return_sale":
                continue
            ch = exchange_children.get(o.id)
            if ch:
                n += len(ch.items) + 4
                n += len(o.items) + 4
            else:
                n += len(o.items) + 4
        n += 2  # imzo + spacer
        return n

    def render_section(start_row: int, title: str, fill, items: list, doc_number: str) -> int:
        # Section banner (1 satr — title)
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=7)
        ban = ws.cell(row=start_row, column=1, value=f"{title} ({doc_number})")
        ban.font = Font(bold=True, size=11)
        ban.fill = fill
        ban.alignment = left
        ban.border = border

        # Headers
        hr = start_row + 1
        for c, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=hr, column=c, value=h)
            cell.font = bold_white
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        r = hr + 1
        total_qty = 0.0
        total_sum = 0.0
        for i, it in enumerate(items, 1):
            prod = it.product
            prod_name = prod.name if prod else f"#{it.product_id}"
            prod_code = (prod.code if prod and prod.code else "") or ""
            unit = prod.unit.name if (prod and prod.unit) else ""
            qty = float(it.quantity or 0)
            price = float(it.price or 0)
            summa = float(it.total or (qty * price))
            total_qty += qty
            total_sum += summa
            ws.cell(row=r, column=1, value=i).alignment = center
            ws.cell(row=r, column=2, value=prod_code).alignment = center
            ws.cell(row=r, column=3, value=prod_name).alignment = left
            ws.cell(row=r, column=4, value=unit).alignment = center
            ws.cell(row=r, column=5, value=qty).alignment = right
            ws.cell(row=r, column=6, value=price).alignment = right
            ws.cell(row=r, column=7, value=summa).alignment = right
            for c in range(1, 8):
                ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=6).number_format = '#,##0'
            ws.cell(row=r, column=7).number_format = '#,##0'
            r += 1
        # Itog
        ws.cell(row=r, column=1, value="Jami").font = Font(bold=True)
        ws.cell(row=r, column=1).alignment = left
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = border
        ws.cell(row=r, column=5, value=total_qty).alignment = right
        ws.cell(row=r, column=5).font = Font(bold=True)
        sum_cell = ws.cell(row=r, column=7, value=total_sum)
        sum_cell.font = Font(bold=True, color="2E7D32")
        sum_cell.alignment = right
        sum_cell.number_format = '#,##0" so\'m"'
        return r + 1

    current_page_used = 0
    is_first_partner = True

    for partner_id, partner_orders in by_partner.items():
        # Yuk xati joriy sahifaga sig'maydigan bo'lsa — page break qo'yamiz,
        # yangi sahifaga butun yuk xati o'tadi
        needed_rows = calc_partner_rows(partner_orders)
        if not is_first_partner and current_page_used + needed_rows > ROWS_PER_PAGE:
            ws.row_breaks.append(Break(id=row - 1, man=True, max=16383))
            current_page_used = 0
        is_first_partner = False
        partner_start_row = row

        # Bir vakil order'dan partner ma'lumotlarini olamiz
        first_order = partner_orders[0]
        partner = first_order.partner
        agent = first_order.agent
        partner_name = partner.name if partner else "Naqd mijoz"
        partner_addr = partner.address if (partner and partner.address) else ""
        partner_phone = partner.phone if (partner and partner.phone) else ""

        # Накладная sarlavhasi
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=1, value=f"Накладная № {first_order.number}    от {first_order.date.strftime('%d.%m.%Y') if first_order.date else today_str}")
        cell.font = Font(bold=True, size=14)
        cell.alignment = center
        row += 1

        # Kontragent va agent ma'lumotlari (2 ustunli)
        ws.cell(row=row, column=1, value="Kimga:").font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row=row, column=2, value=partner_name)
        ws.cell(row=row, column=5, value="ТП:").font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        ws.cell(row=row, column=6, value=agent.full_name if agent else "")
        row += 1
        ws.cell(row=row, column=1, value="Manzil:").font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row=row, column=2, value=partner_addr)
        ws.cell(row=row, column=5, value="Тел(тп):").font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        ws.cell(row=row, column=6, value=agent.phone if agent else "")
        row += 1
        ws.cell(row=row, column=1, value="Telefon:").font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row=row, column=2, value=partner_phone)
        ws.cell(row=row, column=5, value="Kod agenta:").font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        ws.cell(row=row, column=6, value=agent.code if agent else "")
        row += 2

        # Sub-jadvallar tartibi: BUYURTMA → ОБМЕН (отгруз) → ОБМЕН (возврат)
        # Avval oddiy buyurtma (sale, parent_order_id=None) larni chiqaramiz
        for o in partner_orders:
            if o.type == "sale" and not o.parent_order_id:
                row = render_section(row, "BUYURTMA", section_buyurtma_fill, list(o.items), o.number)
                row += 1

        # Keyin obmen lar (return_sale parent + child sale juftligi)
        for o in partner_orders:
            if o.type != "return_sale":
                continue
            child = exchange_children.get(o.id)
            if child:
                row = render_section(row, "ОБМЕН (отгруз)", section_send_fill, list(child.items), child.number)
                row += 1
                row = render_section(row, "ОБМЕН (возврат)", section_ret_fill, list(o.items), o.number)
                row += 1
            else:
                # Yarim/oddiy qaytarish (child yo'q)
                row = render_section(row, "QAYTARISH", section_ret_fill, list(o.items), o.number)
                row += 1

        # Imzo joylari (har Накладная oxirida bir marta)
        ws.cell(row=row, column=1, value="Topshirdi:").font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row=row, column=2, value="_______________________")
        ws.cell(row=row, column=5, value="Qabul qildi:").font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        ws.cell(row=row, column=6, value="_______________________")
        row += 2

        # Joriy sahifa hisobini yangilash
        current_page_used += (row - partner_start_row)

    # Ustun kengligi — 7 ustun
    widths = [12, 12, 38, 14, 10, 14, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # Fayl nomi: 1 ta order bo'lsa o'sha order raqamida, ko'p bo'lsa jamlama
    if len(orders) == 1:
        filename = f"yuk_xati_{orders[0].number}.xlsx"
    else:
        filename = f"yuk_xati_jamlama_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/exchange/{order_id}/excel")
async def sales_exchange_excel(
    order_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Obmen yuk xatini yagona Excel: ichida ОБМЕН (отгруз) + ОБМЕН (возврат) jadvallari."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    parent = db.query(Order).options(
        joinedload(Order.items).joinedload(OrderItem.product).joinedload(Product.unit),
        joinedload(Order.partner),
        joinedload(Order.agent),
    ).filter(Order.id == order_id).first()
    if not parent:
        raise HTTPException(status_code=404, detail="Obmen topilmadi")
    if parent.parent_order_id:
        return RedirectResponse(url=f"/sales/exchange/{parent.parent_order_id}/excel", status_code=303)
    child = db.query(Order).options(
        joinedload(Order.items).joinedload(OrderItem.product).joinedload(Product.unit),
    ).filter(Order.parent_order_id == parent.id).first()
    if not child:
        raise HTTPException(status_code=404, detail="Obmen ning sale qismi topilmadi")
    _check_order_access(parent, current_user)

    wb = Workbook()
    ws = wb.active
    ws.title = "Obmen"

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True)
    bold_lg = Font(bold=True, size=12)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    fill_send = PatternFill("solid", fgColor="C8E6C9")
    fill_ret = PatternFill("solid", fgColor="FFE0B2")
    fill_head_row = PatternFill("solid", fgColor="EEEEEE")

    partner = parent.partner
    agent = parent.agent

    # 1-qator: Накладная № sarlavha
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Накладная № {parent.number}    от {parent.date.strftime('%d.%m.%Y') if parent.date else ''}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    # Kontragent (chap) va agent (o'ng) bloklari
    ws["A2"] = "Кому:"; ws["A2"].font = bold
    ws.merge_cells("B2:D2"); ws["B2"] = partner.name if partner else ""
    ws["E2"] = "ТП:"; ws["E2"].font = bold
    ws.merge_cells("F2:G2"); ws["F2"] = agent.full_name if agent else ""

    ws["A3"] = "Адрес:"; ws["A3"].font = bold
    ws.merge_cells("B3:D3"); ws["B3"] = partner.address if partner else ""
    ws["E3"] = "Тел(тп):"; ws["E3"].font = bold
    ws.merge_cells("F3:G3"); ws["F3"] = agent.phone if agent else ""

    ws["A4"] = "Тел:"; ws["A4"].font = bold
    ws.merge_cells("B4:D4"); ws["B4"] = partner.phone if partner else ""
    ws["E4"] = "Код агента:"; ws["E4"].font = bold
    ws.merge_cells("F4:G4"); ws["F4"] = agent.code if agent else ""

    headers = ["№", "Код", "Наименование", "ЕИ", "Кол-во", "Цена", "Сумма"]

    def render_section(start_row: int, title: str, fill, items, doc_number: str) -> int:
        # Section banner
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=7)
        ban = ws.cell(row=start_row, column=1, value=f"{title} ({doc_number})")
        ban.font = bold_lg
        ban.fill = fill
        ban.alignment = left
        ban.border = border

        # Headers
        hr = start_row + 1
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=hr, column=c, value=h)
            cell.font = bold
            cell.fill = fill_head_row
            cell.alignment = center
            cell.border = border

        # Rows
        r = hr + 1
        total_qty = 0.0
        total_sum = 0.0
        for i, it in enumerate(items, 1):
            prod = it.product
            prod_name = prod.name if prod else f"#{it.product_id}"
            prod_code = (prod.code if prod and prod.code else "") or ""
            unit = prod.unit.name if (prod and prod.unit) else ""
            qty = float(it.quantity or 0)
            price = float(it.price or 0)
            summa = float(it.total or (qty * price))
            total_qty += qty
            total_sum += summa
            ws.cell(row=r, column=1, value=i).alignment = center
            ws.cell(row=r, column=2, value=prod_code).alignment = center
            ws.cell(row=r, column=3, value=prod_name).alignment = left
            ws.cell(row=r, column=4, value=unit).alignment = center
            ws.cell(row=r, column=5, value=qty).alignment = right
            ws.cell(row=r, column=6, value=price).alignment = right
            ws.cell(row=r, column=7, value=summa).alignment = right
            for c in range(1, 8):
                ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=6).number_format = '#,##0'
            ws.cell(row=r, column=7).number_format = '#,##0'
            r += 1

        # Footer (Итог)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        f1 = ws.cell(row=r, column=1, value="Итог")
        f1.font = bold; f1.alignment = right; f1.border = border
        ws.cell(row=r, column=5, value=total_qty).alignment = right
        ws.cell(row=r, column=5).font = bold; ws.cell(row=r, column=5).border = border
        ws.cell(row=r, column=6).border = border
        sf = ws.cell(row=r, column=7, value=total_sum)
        sf.font = bold; sf.alignment = right; sf.number_format = '#,##0'; sf.border = border
        return r + 1

    # Section 1: ОБМЕН (отгруз) — yangi tovar (child sale)
    next_row = render_section(6, "ОБМЕН (отгруз)", fill_send, list(child.items), child.number)
    next_row += 1
    # Section 2: ОБМЕН (возврат) — qaytgan (parent return_sale)
    next_row = render_section(next_row, "ОБМЕН (возврат)", fill_ret, list(parent.items), parent.number)

    # Imzo joylari
    sig_row = next_row + 2
    ws.cell(row=sig_row, column=1, value="Topshirdi:").font = bold
    ws.merge_cells(start_row=sig_row, start_column=2, end_row=sig_row, end_column=3)
    ws.cell(row=sig_row, column=2, value="_______________________").alignment = left
    ws.cell(row=sig_row, column=5, value="Qabul qildi:").font = bold
    ws.merge_cells(start_row=sig_row, start_column=6, end_row=sig_row, end_column=7)
    ws.cell(row=sig_row, column=6, value="_______________________").alignment = left

    widths = [5, 10, 38, 8, 12, 14, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"obmen_{parent.number}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/{order_id}/nakladnoy/excel")
async def sales_nakladnoy_excel(
    order_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Sotuv yuk xatini Excel formatda yuklab olish — bulk endpoint'iga delegate.
    Format izchilligi uchun (Накладная № ..., 7 ustun, ОБМЕН bo'limlari) bulk
    funksiyasini chaqiramiz. Bulk fayl nomi 1 ta order uchun yuk_xati_{number}.xlsx.
    """
    return await sales_nakladnoy_excel_bulk(ids=str(order_id), db=db, current_user=current_user)


@router.get("/{order_id}/nakladnoy/excel/legacy", include_in_schema=False)
async def sales_nakladnoy_excel_legacy(
    order_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Eski single yuk xati (TOTLI HOLVA — Yuk xati formati) — saqlangan emergency uchun."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    order = (
        db.query(Order)
        .options(
            joinedload(Order.items).joinedload(OrderItem.product).joinedload(Product.unit),
            joinedload(Order.partner),
            joinedload(Order.warehouse),
        )
        .filter(Order.id == order_id, Order.type.in_(["sale", "return_sale"]))
        .first()
    )
    if not order:
        raise HTTPException(status_code=404, detail="Sotuv topilmadi")

    wb = Workbook()
    ws = wb.active
    ws.title = "Yuk xati"

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    bold_white = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    # Sarlavha
    ws.merge_cells("A1:F1")
    ws["A1"] = "TOTLI HOLVA — Yuk xati"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = center

    ws.merge_cells("A2:F2")
    ws["A2"] = f"№ {order.number}    Sana: {order.date.strftime('%d.%m.%Y %H:%M') if order.date else '-'}"
    ws["A2"].font = Font(bold=True, size=11)
    ws["A2"].alignment = center

    # Mijoz va ombor
    ws["A4"] = "Mijoz:"
    ws["A4"].font = Font(bold=True)
    ws.merge_cells("B4:F4")
    ws["B4"] = order.partner.name if order.partner else "Naqd mijoz"

    ws["A5"] = "Telefon:"
    ws["A5"].font = Font(bold=True)
    ws.merge_cells("B5:F5")
    ws["B5"] = (order.partner.phone if order.partner and order.partner.phone else "-")

    ws["A6"] = "Manzil:"
    ws["A6"].font = Font(bold=True)
    ws.merge_cells("B6:F6")
    ws["B6"] = (order.partner.address if order.partner and order.partner.address else "-")

    ws["A7"] = "Ombor:"
    ws["A7"].font = Font(bold=True)
    ws.merge_cells("B7:F7")
    ws["B7"] = order.warehouse.name if order.warehouse else "-"

    # Jadval sarlavhasi
    headers = ["№", "Mahsulot", "Birlik", "Miqdor", "Narx (so'm)", "Summa (so'm)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col, value=h)
        cell.font = bold_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Items
    row = 10
    total = 0.0
    for i, item in enumerate(order.items, 1):
        prod_name = item.product.name if item.product else f"#{item.product_id}"
        unit = item.product.unit.name if (item.product and item.product.unit) else ""
        qty = float(item.quantity or 0)
        price = float(item.price or 0)
        summa = float(item.total or (qty * price))
        total += summa

        ws.cell(row=row, column=1, value=i).alignment = center
        ws.cell(row=row, column=2, value=prod_name).alignment = left
        ws.cell(row=row, column=3, value=unit).alignment = center
        ws.cell(row=row, column=4, value=qty).alignment = right
        ws.cell(row=row, column=5, value=price).alignment = right
        ws.cell(row=row, column=6, value=summa).alignment = right

        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
            if col in (5, 6):
                ws.cell(row=row, column=col).number_format = '#,##0'
        row += 1

    # JAMI
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    jami_label = ws.cell(row=row, column=1, value="JAMI:")
    jami_label.font = Font(bold=True, size=12)
    jami_label.alignment = right
    jami_label.border = border
    jami_val = ws.cell(row=row, column=6, value=total)
    jami_val.font = Font(bold=True, size=12, color="2E7D32")
    jami_val.alignment = right
    jami_val.number_format = '#,##0'
    jami_val.border = border

    # Imzo joylari
    sig_row = row + 3
    ws.cell(row=sig_row, column=1, value="Topshirdi (omborchi):").font = Font(bold=True)
    ws.merge_cells(start_row=sig_row, start_column=2, end_row=sig_row, end_column=3)
    ws.cell(row=sig_row, column=2, value="_______________________").alignment = left

    ws.cell(row=sig_row, column=4, value="Qabul qildi:").font = Font(bold=True)
    ws.merge_cells(start_row=sig_row, start_column=5, end_row=sig_row, end_column=6)
    ws.cell(row=sig_row, column=5, value="_______________________").alignment = left

    # Ustun kengligi
    widths = [5, 35, 10, 12, 16, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"yuk_xati_{order.number}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/delete/{order_id}")
async def sales_delete(
    order_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Sotuvni o'chirish (admin). Qoralama — bekor qilingan qiladi; bekor qilingan — bazadan o'chiradi.
    To'lov bog'langan bo'lsa orphan oldini olish uchun rad etiladi."""
    order = db.query(Order).filter(Order.id == order_id, Order.type == "sale").first()
    if not order:
        raise HTTPException(status_code=404, detail="Sotuv topilmadi")

    from app.services.document_service import delete_sale_fully, DocumentError
    try:
        delete_sale_fully(db, order)
    except DocumentError as e:
        return RedirectResponse(
            url=f"/sales?error=delete&detail=" + quote(e.detail),
            status_code=303,
        )
    return RedirectResponse(url="/sales", status_code=303)


# ---------- POS (sotuv oynasi) ----------
@router.get("/pos", response_class=HTMLResponse)
async def sales_pos(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Sotuv oynasi: faqat sotuvchi (yoki admin/menejer). Tovarlar foydalanuvchi bo'limi/omboridan."""
    ensure_orders_payment_due_date_column(db)
    role = (current_user.role or "").strip()
    if role not in ("sotuvchi", "admin", "manager"):
        return RedirectResponse(url="/?error=pos_access", status_code=303)
    pos_user_warehouses = _get_pos_warehouses_for_user(db, current_user)
    sales_warehouse = _get_pos_warehouse_for_user(db, current_user)
    warehouse_id_param = request.query_params.get("warehouse_id")
    if warehouse_id_param and pos_user_warehouses:
        try:
            wid = int(warehouse_id_param)
            chosen = next((w for w in pos_user_warehouses if w.id == wid), None)
            if chosen:
                sales_warehouse = chosen
        except (TypeError, ValueError):
            pass
    from datetime import date as date_type
    today_date = date_type.today()
    pos_today_orders = db.query(Order).filter(
        Order.type == "sale",
        Order.status.in_(("completed", "delivered")),
        func.date(Order.created_at) == today_date,
    ).order_by(Order.created_at.desc()).limit(10).all()
    if not sales_warehouse and role == "sotuvchi":
        err = "no_warehouse"
        detail_msg = "Sizga ombor yoki bo'lim biriktirilmagan. Administrator bilan bog'laning."
        return templates.TemplateResponse("sales/pos.html", {
            "request": request,
            "page_title": "Sotuv oynasi",
            "current_user": current_user,
            "warehouse": None,
            "pos_user_warehouses": pos_user_warehouses,
            "pos_today_orders": pos_today_orders,
            "products": [],
            "product_prices": {},
            "stock_by_product": {},
            "pos_categories": [],
            "pos_all_categories": [],
            "pos_partners": [],
            "default_partner_id": None,
            "success": request.query_params.get("success"),
            "error": err,
            "error_detail": detail_msg,
            "number": request.query_params.get("number", ""),
        })
    # Faqat tanlangan ombordagi mahsulotlar (admin/menejer/sotuvchi — hammasi uchun bir xil)
    if not sales_warehouse:
        stock_rows = []
    else:
        stock_rows = db.query(Stock.product_id, Stock.quantity).filter(
            Stock.warehouse_id == sales_warehouse.id,
            Stock.quantity > 0,
        ).all()
    stock_by_product = {}
    for r in stock_rows:
        pid, qty = r[0], float(r[1] or 0)
        stock_by_product[pid] = stock_by_product.get(pid, 0) + qty
    product_ids_in_warehouse = list(stock_by_product.keys())
    if product_ids_in_warehouse:
        products = db.query(Product).options(joinedload(Product.unit)).filter(
            Product.id.in_(product_ids_in_warehouse),
            Product.is_active == True,
        ).order_by(Product.name).all()
    else:
        products = []
    price_type = _get_pos_price_type(db)
    product_prices = {}
    if price_type and product_ids_in_warehouse:
        pps = db.query(ProductPrice).filter(
            ProductPrice.price_type_id == price_type.id,
            ProductPrice.product_id.in_(product_ids_in_warehouse),
        ).all()
        product_prices = {pp.product_id: float(pp.sale_price or 0) for pp in pps}
    for p in products:
        if p.id not in product_prices or product_prices[p.id] == 0:
            product_prices[p.id] = float(p.sale_price or p.purchase_price or 0)
    pos_categories = []
    if products:
        cat_ids = list({p.category_id for p in products if p.category_id})
        if cat_ids:
            for c in db.query(Category).filter(Category.id.in_(cat_ids)).order_by(Category.name).all():
                pos_categories.append({"id": c.id, "name": c.name or c.code or ""})
    # Dropdown uchun barcha kategoriyalar (omborda qoldiq bo'lmasa ham)
    pos_all_categories = [
        {"id": c.id, "name": c.name or c.code or ""} for c in
        db.query(Category).order_by(Category.name).all()
    ]
    success = request.query_params.get("success")
    error = request.query_params.get("error")
    error_detail = unquote(request.query_params.get("detail", "") or "")
    number = request.query_params.get("number", "")
    user_with_partners = db.query(User).options(joinedload(User.partners_list)).filter(User.id == current_user.id).first()
    user_role = (current_user.role or "").strip()
    assigned_partners = []
    if user_with_partners and getattr(user_with_partners, "partners_list", None):
        assigned_partners = [p for p in user_with_partners.partners_list if getattr(p, "is_active", True)]
    if user_role == "sotuvchi":
        # Sotuvchi uchun cheklov SHART — biriktirilgan mijozlar bilan cheklash
        # Bo'sh bo'lsa, default chakana xaridor qo'shiladi (POS ishlashi uchun)
        if assigned_partners:
            pos_partners = sorted(assigned_partners, key=lambda p: (p.name or ""))
        else:
            chakana = db.query(Partner).filter(
                Partner.is_active == True,
                or_(Partner.code == "chakana", Partner.code == "pos"),
            ).first()
            pos_partners = [chakana] if chakana else []
    else:
        # Admin/manager — agar biriktirilgan bo'lsa shu, aks holda barchasi
        if assigned_partners:
            pos_partners = sorted(assigned_partners, key=lambda p: (p.name or ""))
        else:
            pos_partners = db.query(Partner).filter(Partner.is_active == True).order_by(Partner.name).all()
    default_partner = _get_pos_partner(db)
    default_partner_id = default_partner.id if default_partner else None
    if pos_partners and default_partner_id is not None:
        if not any(p.id == default_partner_id for p in pos_partners):
            default_partner_id = pos_partners[0].id if pos_partners else None
    # Harajat turlari (POS expense modal uchun)
    pos_expense_types = db.query(ExpenseType).order_by(ExpenseType.name).all()
    # Kassalar (inkasatsiya "Yangi yuborish" modali uchun)
    # Sotuvchi faqat naqd pul ko'taradi va Asosiy kassaga (markaziy naqd) keltiradi —
    # boshqa kassalarni (do'kon plastik, click, terminal) ko'rsatish keraksiz.
    _cash_q = db.query(CashRegister).filter(CashRegister.is_active == True)
    _role = (current_user.role or "").strip().lower()
    if _role == "sotuvchi":
        _cash_q = _cash_q.filter(
            CashRegister.payment_type == "naqd",
            CashRegister.name == "Asosiy kassa",
        )
    pos_cash_registers = _cash_q.order_by(CashRegister.name).all()
    return templates.TemplateResponse("sales/pos.html", {
        "request": request,
        "page_title": "Sotuv oynasi",
        "current_user": current_user,
        "warehouse": sales_warehouse,
        "pos_user_warehouses": pos_user_warehouses,
        "pos_today_orders": pos_today_orders,
        "products": products,
        "product_prices": product_prices,
        "stock_by_product": stock_by_product,
        "price_type": price_type,
        "pos_categories": pos_categories,
        "pos_all_categories": pos_all_categories,
        "pos_partners": pos_partners,
        "pos_cash_registers": pos_cash_registers,
        "default_partner_id": default_partner_id,
        "pos_expense_types": pos_expense_types,
        "success": success,
        "error": error,
        "error_detail": error_detail,
        "number": number,
    })


@router.get("/pos/daily-orders")
async def sales_pos_daily_orders(
    request: Request,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    order_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Kunlik / sanadan-sanagacha sotuvlar yoki qaytarishlar ro'yxati (JSON)."""
    from datetime import date as date_type, datetime as dt
    role = (current_user.role or "").strip()
    if role not in ("sotuvchi", "admin", "manager"):
        return []
    today = date_type.today()
    try:
        d_from = dt.strptime(date_from, "%Y-%m-%d").date() if date_from else today
    except (ValueError, TypeError):
        d_from = today
    try:
        d_to = dt.strptime(date_to, "%Y-%m-%d").date() if date_to else today
    except (ValueError, TypeError):
        d_to = today
    if d_from > d_to:
        d_from, d_to = d_to, d_from
    o_type = (order_type or "sale").strip().lower()
    if o_type != "return_sale":
        o_type = "sale"
    q = db.query(Order).filter(
        Order.type == o_type,
        Order.status.in_(("completed", "delivered")),
        func.date(Order.created_at) >= d_from,
        func.date(Order.created_at) <= d_to,
    )
    # Sotuvchi faqat o'z POS ombori sotuvlarini ko'radi (Do'kon 1 ↔ Do'kon 2 ajratish)
    if role == "sotuvchi":
        pos_wh = _get_pos_warehouse_for_user(db, current_user)
        if pos_wh:
            q = q.filter(Order.warehouse_id == pos_wh.id)
        else:
            return []
    orders = q.order_by(Order.created_at.desc()).limit(QUERY_LIMIT_DEFAULT).all()
    out = []
    for o in orders:
        out.append({
            "id": o.id,
            "number": o.number or "",
            "type": o.type or "sale",
            "created_at": o.created_at.strftime("%H:%M") if o.created_at else "-",
            "date": o.created_at.strftime("%d.%m.%Y") if o.created_at else "-",
            "partner_name": o.partner.name if o.partner else "-",
            "warehouse_name": o.warehouse.name if o.warehouse else "-",
            "total": float(o.total or 0),
            "payment_type": o.payment_type or "naqd",
        })
    return out


@router.get("/pos/x-report")
async def sales_pos_x_report(
    request: Request,
    date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """X hisobot — sotuvchi smena yakuni: sotuv/qaytarish/bekor, to'lov turi, kassa balansi, inkasatsiya.

    date — YYYY-MM-DD, default=today, max=today.
    """
    from datetime import date as date_type, datetime as dt
    role = (current_user.role or "").strip()
    if role not in ("sotuvchi", "admin", "manager"):
        return JSONResponse({"error": "Ruxsat yo'q"}, status_code=403)

    target_date = date_type.today()
    if date:
        try:
            parsed = dt.strptime(date[:10], "%Y-%m-%d").date()
            if parsed > date_type.today():
                return JSONResponse({"error": "Kelajak sanasi bo'lishi mumkin emas"}, status_code=400)
            target_date = parsed
        except ValueError:
            return JSONResponse({"error": "Sana formati xato (YYYY-MM-DD)"}, status_code=400)

    base_q = db.query(Order).filter(func.date(Order.created_at) == target_date)
    pos_wh = None
    if role == "sotuvchi":
        pos_wh = _get_pos_warehouse_for_user(db, current_user)
        if not pos_wh:
            return JSONResponse({"error": "Sizga POS ombor biriktirilmagan"}, status_code=400)
        base_q = base_q.filter(Order.warehouse_id == pos_wh.id)

    completed_q = base_q.filter(Order.status.in_(("completed", "delivered")))
    cancelled_q = base_q.filter(Order.status == "cancelled")
    orders = completed_q.all()
    cancelled_orders = cancelled_q.all()

    sales = [o for o in orders if (o.type or "sale") == "sale"]
    returns = [o for o in orders if o.type == "return_sale"]
    sales_total = sum(float(o.total or 0) for o in sales)
    returns_total = sum(float(o.total or 0) for o in returns)
    cancelled_total = sum(float(o.total or 0) for o in cancelled_orders)

    by_type: dict = {}
    sale_with_payments: set = set()
    sale_order_ids = [o.id for o in sales]
    if sale_order_ids:
        try:
            pmt_q = db.query(Payment).filter(
                Payment.type == "income",
                or_(Payment.status == "confirmed", Payment.status.is_(None)),
                Payment.order_id.in_(sale_order_ids),
            ).all()
            cash_pt_map = {c.id: ((c.payment_type or "naqd").strip().lower()) for c in db.query(CashRegister).all()}
            for p in pmt_q:
                sale_with_payments.add(p.order_id)
                pt = cash_pt_map.get(p.cash_register_id, "naqd")
                if pt == "perechisleniye":
                    pt = "bank"
                if pt not in by_type:
                    by_type[pt] = {"count": 0, "sum": 0.0}
                by_type[pt]["count"] += 1
                by_type[pt]["sum"] += float(p.amount or 0)
        except Exception:
            pass

    qarz_orders = [o for o in sales if o.id not in sale_with_payments]
    if qarz_orders:
        by_type["qarz"] = {
            "count": len(qarz_orders),
            "sum": sum(float(o.total or 0) for o in qarz_orders),
        }

    payment_breakdown = [{"type": k, "count": v["count"], "sum": v["sum"]} for k, v in sorted(by_type.items(), key=lambda x: -x[1]["sum"])]

    by_user: list = []
    if role in ("admin", "manager"):
        bu: dict = {}
        for o in sales:
            uid = o.user_id or 0
            if uid not in bu:
                bu[uid] = {"count": 0, "sum": 0.0, "returns": 0.0}
            bu[uid]["count"] += 1
            bu[uid]["sum"] += float(o.total or 0)
        for o in returns:
            uid = o.user_id or 0
            if uid not in bu:
                bu[uid] = {"count": 0, "sum": 0.0, "returns": 0.0}
            bu[uid]["returns"] += float(o.total or 0)
        if bu:
            user_names = {u.id: (u.full_name or u.username) for u in db.query(User).filter(User.id.in_([uid for uid in bu.keys() if uid])).all()}
            by_user = [
                {
                    "user_id": uid,
                    "user": user_names.get(uid, "-") if uid else "—",
                    "count": v["count"],
                    "sum": v["sum"],
                    "returns": v["returns"],
                    "net": v["sum"] - v["returns"],
                }
                for uid, v in sorted(bu.items(), key=lambda x: -x[1]["sum"])
            ]

    cash_balances: list = []
    inkasatsiya_today = {"count": 0, "sum": 0.0}
    inkasatsiya_naqd_today = {"count": 0, "sum": 0.0}
    expense_to_partner = {"count": 0, "sum": 0.0}
    expense_other = {"count": 0, "sum": 0.0}
    expense_non_cash: dict = {}
    try:
        from app.services.finance_service import cash_balance_formula
        from sqlalchemy.orm import joinedload
        user_full = db.query(User).options(joinedload(User.cash_registers_list)).filter(User.id == current_user.id).first()
        user_cashes = list(getattr(user_full, "cash_registers_list", None) or []) if user_full else []
        if role == "sotuvchi" and pos_wh:
            wh_dept_id = getattr(pos_wh, "department_id", None)
            shop_cashes = [c for c in user_cashes if getattr(c, "department_id", None) == wh_dept_id] if wh_dept_id else []
            if not shop_cashes and wh_dept_id:
                shop_cashes = db.query(CashRegister).filter(
                    CashRegister.department_id == wh_dept_id,
                    CashRegister.is_active == True,
                ).all()
            cash_ids = [c.id for c in shop_cashes]
            if cash_ids:
                transfers = db.query(CashTransfer).filter(
                    CashTransfer.from_cash_id.in_(cash_ids),
                    CashTransfer.status.in_(("in_transit", "completed")),
                    func.date(CashTransfer.date) == target_date,
                ).all()
                inkasatsiya_today = {
                    "count": len(transfers),
                    "sum": sum(float(t.amount or 0) for t in transfers),
                }

                naqd_cash_ids = [c.id for c in shop_cashes if (c.payment_type or "").strip().lower() == "naqd"]
                naqd_transfers = [t for t in transfers if t.from_cash_id in naqd_cash_ids]
                inkasatsiya_naqd_today = {
                    "count": len(naqd_transfers),
                    "sum": sum(float(t.amount or 0) for t in naqd_transfers),
                }
                cash_pt_lookup = {c.id: ((c.payment_type or "naqd").strip().lower()) for c in shop_cashes}
                if cash_ids:
                    expenses_q = db.query(Payment).filter(
                        Payment.cash_register_id.in_(cash_ids),
                        Payment.type == "expense",
                        or_(Payment.status == "confirmed", Payment.status.is_(None)),
                        func.date(Payment.created_at) == target_date,
                    ).all()
                    for e in expenses_q:
                        amt = float(e.amount or 0)
                        pt = cash_pt_lookup.get(e.cash_register_id, "naqd")
                        if pt == "perechisleniye":
                            pt = "bank"
                        if pt == "naqd":
                            if e.partner_id:
                                expense_to_partner["count"] += 1
                                expense_to_partner["sum"] += amt
                            else:
                                expense_other["count"] += 1
                                expense_other["sum"] += amt
                        else:
                            if pt not in expense_non_cash:
                                expense_non_cash[pt] = {"count": 0, "sum": 0.0}
                            expense_non_cash[pt]["count"] += 1
                            expense_non_cash[pt]["sum"] += amt
    except Exception:
        pass

    naqd_income_today = float(by_type.get("naqd", {}).get("sum", 0) or 0)
    today_net_naqd = naqd_income_today - expense_to_partner["sum"] - expense_other["sum"] - inkasatsiya_naqd_today["sum"]

    # Boshlang'ich qoldiq (kechagi qoldiq) — kassa balansidan bugungi net naqdni ayirib topiladi.
    # Bugungi X-hisobotda qoldiq = hozirgi haqiqiy CashRegister.balance (real-time).
    # O'tgan sanada balans real-time bo'lgani uchun opening hisoblab bo'lmaydi → eski xulq.
    opening_naqd = None
    qoldiq = today_net_naqd
    try:
        if role == "sotuvchi" and pos_wh and target_date == date_type.today():
            naqd_cash_objs = [c for c in shop_cashes if (c.payment_type or "").strip().lower() == "naqd"]
            current_naqd_balance = sum(float(c.balance or 0) for c in naqd_cash_objs)
            opening_naqd = current_naqd_balance - today_net_naqd
            qoldiq = current_naqd_balance
    except Exception:
        pass

    # Oldingi Z-hisobot bo'lsa farqni hisoblash (majburiy keyingi yopilish uchun).
    # Birinchi (eng eski) Z bilan solishtirish — chunki eng yangi Z hozirgi live state
    # bilan teng bo'ladi (uning ustiga yangi sotuv qo'shilmagan bo'lsa).
    last_z_info = None
    diff_sales_count = None
    diff_sales_total = None

    def _z_snaps_for_date(d):
        """Tanlangan sana uchun current_user + pos_wh bo'yicha Z snapshotlar (eski->yangi)."""
        import os as _xos2
        import json as _xjson2
        folder2 = _xos2.path.join("data", "z_reports", d.strftime("%Y-%m-%d"))
        if not _xos2.path.isdir(folder2):
            return []
        wid = pos_wh.id if pos_wh else None
        snaps = []
        for fname in _xos2.listdir(folder2):
            if not fname.endswith(".json"):
                continue
            try:
                with open(_xos2.path.join(folder2, fname), "r", encoding="utf-8") as f:
                    snap = _xjson2.load(f)
            except (OSError, _xjson2.JSONDecodeError):
                continue
            if int(snap.get("user_id") or 0) != current_user.id:
                continue
            if wid is not None and snap.get("warehouse_id") != wid:
                continue
            snaps.append(snap)
        snaps.sort(key=lambda s: s.get("closed_at") or "")
        return snaps

    try:
        all_snaps = _z_snaps_for_date(target_date)
        if all_snaps:
            first_snap = all_snaps[0]
            prev_count = int(first_snap.get("sales_count") or 0)
            prev_total = float(first_snap.get("sales_total") or 0)
            d_count = len(sales) - prev_count
            d_total = sales_total - prev_total
            if d_count != 0 or abs(d_total) > 0.01:
                last_z_info = {
                    "z_id": first_snap.get("z_id"),
                    "closed_at": first_snap.get("closed_at"),
                    "sales_count": prev_count,
                    "sales_total": prev_total,
                    "total_closes": len(all_snaps),
                }
                diff_sales_count = d_count
                diff_sales_total = d_total
    except Exception:
        pass

    # OLDINGI KUNLARDAN HALI KELMAGAN PUL: oxirgi 7 kun ichida 2+ Z bo'lgan kunlar
    # uchun (voluntary close + forced close) → farq = topshirilmagan naqd.
    # CashTransfer (kassadan kassaga, FROM naqd POS kassa) — bu pulning topshirilgani
    # demakdir. FIFO: eski kun avval qoplanadi.
    pending_prev_days = []
    try:
        from datetime import timedelta as _td
        for offset in range(1, 8):
            prev_d = target_date - _td(days=offset)
            snaps = _z_snaps_for_date(prev_d)
            if len(snaps) < 2:
                continue
            first = snaps[0]
            last = snaps[-1]
            d_c = int(last.get("sales_count") or 0) - int(first.get("sales_count") or 0)
            d_t = float(last.get("sales_total") or 0) - float(first.get("sales_total") or 0)
            if d_c == 0 and abs(d_t) <= 0.01:
                continue
            pending_prev_days.append({
                "date": prev_d.strftime("%Y-%m-%d"),
                "date_display": prev_d.strftime("%d.%m.%Y"),
                "first_z_id": first.get("z_id"),
                "first_closed_at": first.get("closed_at"),
                "first_sales_total": float(first.get("sales_total") or 0),
                "diff_sales_count": d_c,
                "diff_sales_total": d_t,
            })

        # CashTransfer'larni qoplash: pending kunlardan eski→yangi tartibda
        if pending_prev_days:
            naqd_cash_ids_local = []
            try:
                user_full_l = db.query(User).options(joinedload(User.cash_registers_list)).filter(User.id == current_user.id).first()
                user_cashes_l = list(getattr(user_full_l, "cash_registers_list", None) or []) if user_full_l else []
                if role == "sotuvchi" and pos_wh:
                    wh_dept_l = getattr(pos_wh, "department_id", None)
                    cands = [c for c in user_cashes_l if getattr(c, "department_id", None) == wh_dept_l] if wh_dept_l else []
                    if not cands and wh_dept_l:
                        cands = db.query(CashRegister).filter(
                            CashRegister.department_id == wh_dept_l,
                            CashRegister.is_active == True,
                        ).all()
                else:
                    cands = user_cashes_l
                naqd_cash_ids_local = [c.id for c in cands if (c.payment_type or "").strip().lower() == "naqd"]
            except Exception:
                naqd_cash_ids_local = []

            if naqd_cash_ids_local:
                cutoff_dt = datetime.combine(target_date - _td(days=10), datetime.min.time())
                transfers = db.query(CashTransfer).filter(
                    CashTransfer.from_cash_id.in_(naqd_cash_ids_local),
                    CashTransfer.status.in_(("in_transit", "completed")),
                    CashTransfer.date >= cutoff_dt,
                ).order_by(CashTransfer.date.asc()).all()
                transfer_remaining = [float(t.amount or 0) for t in transfers]
                transfer_dates = [t.date for t in transfers]

                pending_prev_days.sort(key=lambda x: x["date"])
                filtered = []
                for pd_item in pending_prev_days:
                    needed = float(pd_item["diff_sales_total"] or 0)
                    if needed <= 0.01:
                        continue
                    first_closed_iso = pd_item.get("first_closed_at") or ""
                    try:
                        first_closed_dt = datetime.fromisoformat(first_closed_iso) if first_closed_iso else None
                    except (ValueError, TypeError):
                        first_closed_dt = None
                    for i, td in enumerate(transfer_dates):
                        if transfer_remaining[i] <= 0:
                            continue
                        if first_closed_dt and td < first_closed_dt:
                            continue
                        consumed = min(transfer_remaining[i], needed)
                        transfer_remaining[i] -= consumed
                        needed -= consumed
                        if needed <= 0.01:
                            break
                    if needed > 0.01:
                        pd_item["diff_sales_total"] = round(needed, 2)
                        filtered.append(pd_item)
                pending_prev_days = filtered
                pending_prev_days.sort(key=lambda x: x["date"], reverse=True)
    except Exception:
        pass

    return JSONResponse({
        "date": target_date.strftime("%d.%m.%Y"),
        "date_iso": target_date.strftime("%Y-%m-%d"),
        "user": current_user.full_name or current_user.username,
        "warehouse": pos_wh.name if pos_wh else "Barcha",
        "sales_count": len(sales),
        "sales_total": sales_total,
        "last_z": last_z_info,
        "diff_sales_count": diff_sales_count,
        "diff_sales_total": diff_sales_total,
        "pending_prev_days": pending_prev_days,
        "returns_count": len(returns),
        "returns_total": returns_total,
        "cancelled_count": len(cancelled_orders),
        "cancelled_total": cancelled_total,
        "net_total": sales_total - returns_total,
        "payment_breakdown": payment_breakdown,
        "by_user": by_user,
        "cash_balances": cash_balances,
        "inkasatsiya_today": inkasatsiya_today,
        "inkasatsiya_naqd_today": inkasatsiya_naqd_today,
        "expense_to_partner": expense_to_partner,
        "expense_other": expense_other,
        "expense_non_cash": [{"type": k, "count": v["count"], "sum": v["sum"]} for k, v in sorted(expense_non_cash.items(), key=lambda x: -x[1]["sum"])],
        "opening_naqd": opening_naqd,
        "qoldiq": qoldiq,
    })


@router.get("/pos/employees-active")
async def sales_pos_employees_active(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """POS uchun aktiv xodimlar ro'yxati (Xodim mahsulot modal)."""
    role = (current_user.role or "").strip()
    if role not in ("sotuvchi", "admin", "manager"):
        return JSONResponse({"ok": False, "error": "Ruxsat yo'q"}, status_code=403)
    rows = db.query(Employee).filter(Employee.is_active == True).order_by(Employee.full_name).all()
    return JSONResponse({
        "ok": True,
        "items": [{
            "id": e.id,
            "name": e.full_name or "",
            "position": e.position or "",
            "department": e.department or "",
            "free_quota": float(getattr(e, "monthly_free_quota", None) or 0),
        } for e in rows],
    })


@router.get("/pos/employee-quota")
async def sales_pos_employee_quota(
    employee_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Tanlangan xodimning shu oydagi mahsulot xaridi va kvota qoldig'i."""
    from datetime import date as date_type
    role = (current_user.role or "").strip()
    if role not in ("sotuvchi", "admin", "manager"):
        return JSONResponse({"ok": False, "error": "Ruxsat yo'q"}, status_code=403)
    emp = db.query(Employee).filter(Employee.id == employee_id, Employee.is_active == True).first()
    if not emp:
        return JSONResponse({"ok": False, "error": "Xodim topilmadi"}, status_code=404)
    today = date_type.today()
    month_start = today.replace(day=1)
    used_q = (
        db.query(func.coalesce(func.sum(EmployeeAdvance.amount), 0))
        .filter(
            EmployeeAdvance.employee_id == employee_id,
            EmployeeAdvance.is_product == True,
            EmployeeAdvance.advance_date >= month_start,
            EmployeeAdvance.advance_date <= today,
        )
        .scalar()
    )
    free_quota = float(getattr(emp, "monthly_free_quota", None) or 0)
    used = float(used_q or 0)
    return JSONResponse({
        "ok": True,
        "employee": {"id": emp.id, "name": emp.full_name or "", "position": emp.position or ""},
        "month": today.strftime("%Y-%m"),
        "free_quota": free_quota,
        "used_this_month": used,
        "free_remaining": max(0.0, free_quota - used),
    })


@router.post("/pos/employee-product")
async def sales_pos_employee_product(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """POS savatdan xodim mahsuloti yozish.

    Body JSON: {employee_id, warehouse_id, items: [{product_id, quantity, price}]}
    Yaratiladi:
      - Order (type=sale, employee_id, payment_type='employee_advance', paid=0, debt=0)
      - OrderItem(lar)
      - StockMovement(lar) — oddiy sotuv kabi (operation_type='sale')
      - EmployeeAdvance(is_product=True, amount=order.total, note=order.number)
    """
    from datetime import date as date_type
    role = (current_user.role or "").strip()
    if role not in ("sotuvchi", "admin", "manager"):
        return JSONResponse({"ok": False, "error": "Ruxsat yo'q"}, status_code=403)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"ok": False, "error": "Body JSON xato"}, status_code=400)

    employee_id = body.get("employee_id")
    warehouse_id = body.get("warehouse_id")
    items_in = body.get("items") or []
    if not employee_id or not warehouse_id or not items_in:
        return JSONResponse({"ok": False, "error": "Xodim, ombor va kamida 1 mahsulot kerak"}, status_code=400)

    emp = db.query(Employee).filter(Employee.id == int(employee_id), Employee.is_active == True).first()
    if not emp:
        return JSONResponse({"ok": False, "error": "Xodim topilmadi"}, status_code=404)
    wh = db.query(Warehouse).filter(Warehouse.id == int(warehouse_id)).first()
    if not wh:
        return JSONResponse({"ok": False, "error": "Ombor topilmadi"}, status_code=404)
    if role == "sotuvchi":
        allowed = _get_pos_warehouses_for_user(db, current_user)
        if not any(w.id == wh.id for w in allowed):
            return JSONResponse({"ok": False, "error": "Bu omborga ruxsat yo'q"}, status_code=403)

    parsed_items: list = []
    for it in items_in:
        try:
            pid = int(it.get("product_id"))
            qty = float(it.get("quantity"))
            price = float(it.get("price"))
        except (TypeError, ValueError):
            continue
        if pid <= 0 or qty <= 0 or price < 0:
            continue
        product = db.query(Product).filter(Product.id == pid).first()
        if not product:
            return JSONResponse({"ok": False, "error": f"Mahsulot topilmadi: id={pid}"}, status_code=404)
        stock = db.query(Stock).filter(Stock.warehouse_id == wh.id, Stock.product_id == pid).with_for_update().first()
        avail = float(stock.quantity if stock else 0)
        if avail + 1e-6 < qty:
            return JSONResponse({
                "ok": False,
                "error": f"Qoldiq yetmaydi: {product.name} (omborda {avail:.2f}, kerak {qty:.2f})"
            }, status_code=400)
        parsed_items.append({"product": product, "qty": qty, "price": price})

    if not parsed_items:
        return JSONResponse({"ok": False, "error": "Yaroqli mahsulot topilmadi"}, status_code=400)

    last_order = db.query(Order).filter(Order.type == "sale").order_by(Order.id.desc()).first()
    new_number = f"S-{datetime.now().strftime('%Y%m%d')}-{(last_order.id + 1) if last_order else 1:04d}"
    order = Order(
        number=new_number,
        type="sale",
        partner_id=None,
        warehouse_id=wh.id,
        employee_id=emp.id,
        user_id=current_user.id,
        status="completed",
        payment_type="employee_advance",
        source="pos",
    )
    db.add(order)
    db.flush()

    total = 0.0
    item_notes = []
    for it in parsed_items:
        line_total = it["qty"] * it["price"]
        db.add(OrderItem(
            order_id=order.id,
            product_id=it["product"].id,
            quantity=it["qty"],
            price=it["price"],
            total=line_total,
        ))
        total += line_total
        item_notes.append(f"{it['product'].name} {it['qty']:g}×{it['price']:,.0f}")

    order.subtotal = total
    order.discount_percent = 0
    order.discount_amount = 0
    order.total = total
    order.paid = 0
    order.debt = 0

    for it in parsed_items:
        create_stock_movement(
            db=db,
            warehouse_id=wh.id,
            product_id=it["product"].id,
            quantity_change=-it["qty"],
            operation_type="sale",
            document_type="Sale",
            document_id=order.id,
            document_number=order.number,
            user_id=current_user.id,
            note=f"Xodim mahsulot xaridi: {emp.full_name}",
        )

    advance = EmployeeAdvance(
        employee_id=emp.id,
        cash_register_id=None,
        amount=total,
        advance_date=date_type.today(),
        note=f"POS Xodim mahsulot: {order.number} — {', '.join(item_notes)[:400]}",
        is_product=True,
        confirmed_at=datetime.now(),
    )
    db.add(advance)

    try:
        log_action(
            db, user=current_user, action="employee_product_sale",
            entity_type="order", entity_id=order.id, entity_number=order.number,
            details=f"Xodim mahsulot: {emp.full_name}, {total:,.0f} so'm, {len(parsed_items)} ta tovar",
            ip_address=request.client.host if request.client else "",
        )
    except Exception:
        pass

    db.commit()

    return JSONResponse({
        "ok": True,
        "order_number": order.number,
        "total": total,
        "employee_name": emp.full_name,
        "items_count": len(parsed_items),
    })


@router.get("/pos/my-operations")
async def sales_pos_my_operations(
    request: Request,
    date: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Sotuvchining tanlangan kundagi to'lovlari va harajatlari.

    Faqat current_user.id'ning yozuvlari ko'rsatiladi (boshqa sotuvchilar aralashmaydi).

    To'lovlar:  category='supplier_payment' (kontragentga to'lov)
    Harajatlar: category='expense'          (boshqa harajatlar)
    """
    from datetime import date as date_type, datetime as dt
    role = (current_user.role or "").strip()
    if role not in ("sotuvchi", "admin", "manager"):
        return JSONResponse({"ok": False, "error": "Ruxsat yo'q"}, status_code=403)

    target_date = date_type.today()
    if date:
        try:
            target_date = dt.strptime(date[:10], "%Y-%m-%d").date()
        except ValueError:
            return JSONResponse({"ok": False, "error": "Sana formati xato"}, status_code=400)

    base = (
        db.query(Payment)
        .filter(
            func.date(Payment.created_at) == target_date,
            Payment.user_id == current_user.id,
            Payment.type == "expense",
        )
    )

    payments_list = []
    expenses_list = []
    for p in base.order_by(Payment.created_at).all():
        partner_name = ""
        if p.partner_id:
            pn = db.query(Partner.name).filter(Partner.id == p.partner_id).first()
            partner_name = pn[0] if pn else ""
        cash_name = ""
        if p.cash_register_id:
            cn = db.query(CashRegister.name).filter(CashRegister.id == p.cash_register_id).first()
            cash_name = cn[0] if cn else ""
        rec = {
            "id": p.id,
            "number": p.number or "",
            "time": p.created_at.strftime("%H:%M") if p.created_at else "",
            "amount": float(p.amount or 0),
            "payment_type": (p.payment_type or "naqd"),
            "cash_name": cash_name,
            "partner_name": partner_name,
            "description": p.description or "",
        }
        cat = (p.category or "").strip()
        if cat == "supplier_payment":
            payments_list.append(rec)
        elif cat == "expense":
            expenses_list.append(rec)

    return JSONResponse({
        "ok": True,
        "date": target_date.strftime("%Y-%m-%d"),
        "date_display": target_date.strftime("%d.%m.%Y"),
        "payments": payments_list,
        "expenses": expenses_list,
        "totals": {
            "payments_count": len(payments_list),
            "payments_sum": sum(r["amount"] for r in payments_list),
            "expenses_count": len(expenses_list),
            "expenses_sum": sum(r["amount"] for r in expenses_list),
        },
    })


@router.get("/pos/z-report/open-days")
async def sales_pos_z_report_open_days(
    request: Request,
    days: int = 5,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Oxirgi N kun uchun (bugundan tashqari) sotuv vs Z holatini qaytaradi.

    Status qiymatlari:
      - "no_z":      DB'da sotuv bor, lekin Z fayl yaratilmagan
      - "incomplete": Z fayl bor, ammo uning closed_at'dan keyin yangi sotuvlar qo'shilgan
      - "ok":        Z fayl bor va orphan sotuv yo'q (banner'da ko'rsatilmaydi)
    """
    import os
    import json as _json
    from datetime import date as date_type, datetime as dt, timedelta

    role = (current_user.role or "").strip().lower()
    if role not in ("sotuvchi", "admin", "manager"):
        return JSONResponse({"ok": False, "error": "Ruxsat yo'q"}, status_code=403)

    # Manager uchun Z-hisobot shart emas — savdo joyida pul asosiy kassaga tushadi.
    # Banner faqat sotuvchilar (smena yopishi shart) va admin (monitoring) uchun.
    if role == "manager":
        return JSONResponse({"ok": True, "items": [], "reason": "manager_no_z"})

    days = max(1, min(int(days or 5), 14))

    pos_wh = None
    if role == "sotuvchi":
        pos_wh = _get_pos_warehouse_for_user(db, current_user)
        if not pos_wh:
            return JSONResponse({"ok": True, "items": [], "reason": "no_pos_warehouse"})
    wh_id = pos_wh.id if pos_wh else None

    today = date_type.today()
    items: list = []

    for offset in range(1, days + 1):
        d = today - timedelta(days=offset)

        q = db.query(Order).filter(
            func.date(Order.created_at) == d,
            Order.status.in_(("completed", "delivered")),
        )
        q = q.filter((Order.type == "sale") | (Order.type.is_(None)))
        if wh_id is not None:
            q = q.filter(Order.warehouse_id == wh_id)
        if role == "sotuvchi":
            q = q.filter(Order.user_id == current_user.id)
        rows = q.all()

        sales_count = sum(1 for o in rows if (o.type or "sale") == "sale")
        sales_total = sum(float(o.total or 0) for o in rows if (o.type or "sale") == "sale")
        if sales_count == 0:
            continue

        folder = os.path.join("data", "z_reports", d.strftime("%Y-%m-%d"))
        last_z = None
        if os.path.isdir(folder):
            for fname in os.listdir(folder):
                if not fname.endswith(".json"):
                    continue
                try:
                    with open(os.path.join(folder, fname), "r", encoding="utf-8") as f:
                        snap = _json.load(f)
                except (OSError, _json.JSONDecodeError):
                    continue
                if int(snap.get("user_id") or 0) != current_user.id:
                    continue
                if wh_id is not None and snap.get("warehouse_id") != wh_id:
                    continue
                ca = snap.get("closed_at") or ""
                if last_z is None or ca > (last_z.get("closed_at") or ""):
                    last_z = snap

        orphan_count = 0
        orphan_total = 0.0
        if last_z is None:
            status = "no_z"
        else:
            try:
                z_dt = datetime.fromisoformat(last_z.get("closed_at"))
            except (TypeError, ValueError):
                z_dt = None
            if z_dt is not None:
                for o in rows:
                    if (o.type or "sale") != "sale":
                        continue
                    if o.created_at and o.created_at > z_dt:
                        orphan_count += 1
                        orphan_total += float(o.total or 0)
            status = "incomplete" if orphan_count > 0 else "ok"

        if status == "ok":
            continue

        items.append({
            "date": d.strftime("%Y-%m-%d"),
            "date_display": d.strftime("%d.%m.%Y"),
            "status": status,
            "sales_count": sales_count,
            "sales_total": sales_total,
            "last_z": {
                "z_id": last_z.get("z_id"),
                "closed_at": last_z.get("closed_at"),
                "sales_total": float(last_z.get("sales_total") or 0),
                "sales_count": int(last_z.get("sales_count") or 0),
            } if last_z else None,
            "orphan_count": orphan_count,
            "orphan_total": orphan_total,
        })

    return JSONResponse({"ok": True, "items": items, "days_scanned": days})


@router.get("/pos/z-report/check")
async def sales_pos_z_report_check(
    request: Request,
    date: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Bugun (yoki tanlangan sanada) shu user shu omborda Z-hisobot mavjudligini tekshiradi.

    Frontend bu endpoint orqali preflight qiladi va dublikat oldini olish uchun
    foydalanuvchiga ogohlantirish chiqaradi.
    """
    import os
    import json as _json
    from datetime import date as date_type, datetime as dt
    role = (current_user.role or "").strip()
    if role not in ("sotuvchi", "admin", "manager"):
        return JSONResponse({"ok": False, "error": "Ruxsat yo'q"}, status_code=403)

    target_date = date_type.today()
    if date:
        try:
            target_date = dt.strptime(date[:10], "%Y-%m-%d").date()
        except ValueError:
            return JSONResponse({"ok": False, "error": "Sana formati xato"}, status_code=400)

    pos_wh = None
    if role == "sotuvchi":
        pos_wh = _get_pos_warehouse_for_user(db, current_user)
    wh_id = pos_wh.id if pos_wh else None

    folder = os.path.join("data", "z_reports", target_date.strftime("%Y-%m-%d"))
    existing: list = []
    if os.path.isdir(folder):
        try:
            for fname in os.listdir(folder):
                if not fname.endswith(".json"):
                    continue
                fpath = os.path.join(folder, fname)
                try:
                    with open(fpath, "r", encoding="utf-8") as f:
                        snap = _json.load(f)
                except (OSError, _json.JSONDecodeError):
                    continue
                if int(snap.get("user_id") or 0) != current_user.id:
                    continue
                if wh_id is not None and snap.get("warehouse_id") != wh_id:
                    continue
                existing.append({
                    "z_id": snap.get("z_id"),
                    "closed_at": snap.get("closed_at"),
                    "sales_total": float(snap.get("sales_total") or 0),
                    "sales_count": int(snap.get("sales_count") or 0),
                    "net_total": float(snap.get("net_total") or 0),
                })
        except OSError:
            pass

    existing.sort(key=lambda x: x.get("closed_at") or "", reverse=True)
    return JSONResponse({
        "ok": True,
        "exists": len(existing) > 0,
        "count": len(existing),
        "last": existing[0] if existing else None,
        "date": target_date.strftime("%Y-%m-%d"),
    })


def _build_pos_shift_snapshot(db, current_user, target_date, role, pos_wh):
    """POS smena snapshot ma'lumotini yasaydi (X va Z hisobot uchun umumiy chek).
    z_id / closed_at / doc_title chaqiruvchida to'ldiriladi."""
    from app.utils.z_cash_summary import compute_z_cash_summary, find_previous_z_remaining
    from datetime import datetime as _dt
    from sqlalchemy.orm import joinedload as _jl

    base_q = db.query(Order).filter(func.date(Order.created_at) == target_date)
    if role == "sotuvchi" and pos_wh:
        base_q = base_q.filter(Order.warehouse_id == pos_wh.id)
    completed = base_q.filter(Order.status.in_(("completed", "delivered"))).all()
    cancelled = base_q.filter(Order.status == "cancelled").all()
    sales = [o for o in completed if (o.type or "sale") == "sale"]
    returns_o = [o for o in completed if o.type == "return_sale"]
    sales_total = sum(float(o.total or 0) for o in sales)
    returns_total = sum(float(o.total or 0) for o in returns_o)

    by_type: dict = {}
    sale_order_ids = [o.id for o in sales]
    if sale_order_ids:
        try:
            pmt_q = db.query(Payment).filter(
                Payment.type == "income",
                or_(Payment.status == "confirmed", Payment.status.is_(None)),
                Payment.order_id.in_(sale_order_ids),
            ).all()
            cash_pt_map = {c.id: ((c.payment_type or "naqd").strip().lower()) for c in db.query(CashRegister).all()}
            sale_with_payments: set = set()
            for p in pmt_q:
                sale_with_payments.add(p.order_id)
                pt = cash_pt_map.get(p.cash_register_id, "naqd")
                if pt == "perechisleniye":
                    pt = "bank"
                if pt not in by_type:
                    by_type[pt] = {"count": 0, "sum": 0.0}
                by_type[pt]["count"] += 1
                by_type[pt]["sum"] += float(p.amount or 0)
            qarz_orders = [o for o in sales if o.id not in sale_with_payments]
            if qarz_orders:
                by_type["qarz"] = {"count": len(qarz_orders), "sum": sum(float(o.total or 0) for o in qarz_orders)}
        except Exception:
            for o in sales:
                pt = (o.payment_type or "naqd").lower()
                if pt not in by_type:
                    by_type[pt] = {"count": 0, "sum": 0.0}
                by_type[pt]["count"] += 1
                by_type[pt]["sum"] += float(o.total or 0)

    cash_snapshot: list = []
    try:
        from app.services.finance_service import cash_balance_formula
        user_full = db.query(User).options(_jl(User.cash_registers_list)).filter(User.id == current_user.id).first()
        user_cashes_z = list(getattr(user_full, "cash_registers_list", None) or []) if user_full else []
        wh_dept_id = getattr(pos_wh, "department_id", None) if pos_wh else None
        if role == "sotuvchi" and wh_dept_id:
            shop_cashes = [c for c in user_cashes_z if getattr(c, "department_id", None) == wh_dept_id]
            if not shop_cashes:
                shop_cashes = db.query(CashRegister).filter(
                    CashRegister.department_id == wh_dept_id,
                    CashRegister.is_active == True,
                ).all()
        else:
            shop_cashes = user_cashes_z
        for c in shop_cashes:
            bal, inc, exp = cash_balance_formula(db, c.id)
            cash_snapshot.append({"id": c.id, "name": c.name, "balance": float(bal or 0), "income": float(inc or 0), "expense": float(exp or 0)})
    except Exception:
        pass

    payments_made: list = []
    expenses_made: list = []
    try:
        op_rows = (
            db.query(Payment)
            .filter(
                func.date(Payment.created_at) == target_date,
                Payment.user_id == current_user.id,
                Payment.type == "expense",
                Payment.category.in_(("supplier_payment", "expense")),
            )
            .order_by(Payment.created_at)
            .all()
        )
        for p in op_rows:
            partner_name = ""
            if p.partner_id:
                pn = db.query(Partner.name).filter(Partner.id == p.partner_id).first()
                partner_name = pn[0] if pn else ""
            cash_name = ""
            if p.cash_register_id:
                cn = db.query(CashRegister.name).filter(CashRegister.id == p.cash_register_id).first()
                cash_name = cn[0] if cn else ""
            rec = {
                "number": p.number or "",
                "time": p.created_at.strftime("%H:%M") if p.created_at else "",
                "amount": float(p.amount or 0),
                "payment_type": (p.payment_type or "naqd"),
                "cash_name": cash_name,
                "partner_name": partner_name,
                "description": p.description or "",
            }
            if (p.category or "").strip() == "supplier_payment":
                payments_made.append(rec)
            else:
                expenses_made.append(rec)
    except Exception:
        pass

    cash_summary = compute_z_cash_summary(db, target_date, current_user.id, until_dt=_dt.now())
    prev_remaining, prev_zid = find_previous_z_remaining(
        user_id=current_user.id,
        warehouse_id=(pos_wh.id if pos_wh else None),
        before_closed_at=_dt.now().isoformat(),
    )
    cash_remaining = (
        prev_remaining
        + cash_summary["cash_sales_total"]
        - cash_summary["cash_expenses_total"]
        - cash_summary["cash_payments_out"]
        - cash_summary.get("cash_transfers_out", 0.0)  # 2026-05-26 fix
    )

    return {
        "date": target_date.strftime("%Y-%m-%d"),
        "user_id": current_user.id,
        "user": current_user.full_name or current_user.username,
        "role": role,
        "warehouse_id": pos_wh.id if pos_wh else None,
        "warehouse": pos_wh.name if pos_wh else "Barcha",
        "sales_count": len(sales),
        "sales_total": sales_total,
        "returns_count": len(returns_o),
        "returns_total": returns_total,
        "cancelled_count": len(cancelled),
        "cancelled_total": sum(float(o.total or 0) for o in cancelled),
        "net_total": sales_total - returns_total,
        "payment_breakdown": [{"type": k, "count": v["count"], "sum": v["sum"]} for k, v in by_type.items()],
        "cash_snapshot": cash_snapshot,
        "order_numbers": [o.number for o in completed],
        "payments_made": payments_made,
        "payments_made_sum": sum(r["amount"] for r in payments_made),
        "expenses_made": expenses_made,
        "expenses_made_sum": sum(r["amount"] for r in expenses_made),
        "cash_sales_pure": cash_summary["cash_sales_pure"],
        "cash_sales_split": cash_summary["cash_sales_split"],
        "cash_sales_total": cash_summary["cash_sales_total"],
        "cash_expenses_total": cash_summary["cash_expenses_total"],
        "cash_payments_out": cash_summary["cash_payments_out"],
        "cash_transfers_out": cash_summary.get("cash_transfers_out", 0.0),  # 2026-05-26 fix
        "previous_cash_remaining": prev_remaining,
        "previous_z_id": prev_zid,
        "cash_remaining": cash_remaining,
    }


@router.get("/pos/x-report/receipt", response_class=HTMLResponse)
async def sales_pos_x_report_receipt(
    request: Request,
    date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """X-hisobot 80mm chek (Z-chek bilan bir xil go'zal format)."""
    from datetime import date as date_type, datetime as dt
    role = (current_user.role or "").strip()
    if role not in ("sotuvchi", "admin", "manager"):
        return HTMLResponse("Ruxsat yo'q", status_code=403)
    target_date = date_type.today()
    if date:
        try:
            parsed = dt.strptime(date[:10], "%Y-%m-%d").date()
            if parsed <= date_type.today():
                target_date = parsed
        except ValueError:
            pass
    pos_wh = None
    if role == "sotuvchi":
        pos_wh = _get_pos_warehouse_for_user(db, current_user)
        if not pos_wh:
            return HTMLResponse("Sizga POS ombor biriktirilmagan", status_code=400)
    snap = _build_pos_shift_snapshot(db, current_user, target_date, role, pos_wh)
    snap["z_id"] = f"X-{target_date.strftime('%Y%m%d')}-U{current_user.id}-{dt.now().strftime('%H%M%S')}"
    snap["closed_at"] = dt.now().isoformat()
    snap["doc_title"] = "X-HISOBOT"
    return templates.TemplateResponse("reports/z_report_receipt.html", {"request": request, "snap": snap})


@router.post("/pos/z-report")
async def sales_pos_z_report(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Z-hisobot — smenani yopish: X-hisobot snapshot data/z_reports/ ga saqlanadi va audit log ga yoziladi."""
    import os
    import json as _json
    from datetime import date as date_type, datetime as dt
    role = (current_user.role or "").strip()
    if role not in ("sotuvchi", "admin", "manager"):
        return JSONResponse({"ok": False, "error": "Ruxsat yo'q"}, status_code=403)
    try:
        body = await request.json()
    except Exception:
        body = {}
    date_str = (body.get("date") or "").strip()
    target_date = date_type.today()
    if date_str:
        try:
            parsed = dt.strptime(date_str[:10], "%Y-%m-%d").date()
            if parsed > date_type.today():
                return JSONResponse({"ok": False, "error": "Kelajak sanasi bo'lishi mumkin emas"}, status_code=400)
            target_date = parsed
        except ValueError:
            return JSONResponse({"ok": False, "error": "Sana formati xato"}, status_code=400)

    pos_wh = None
    base_q = db.query(Order).filter(func.date(Order.created_at) == target_date)
    if role == "sotuvchi":
        pos_wh = _get_pos_warehouse_for_user(db, current_user)
        if not pos_wh:
            return JSONResponse({"ok": False, "error": "Sizga POS ombor biriktirilmagan"}, status_code=400)
        base_q = base_q.filter(Order.warehouse_id == pos_wh.id)

    completed = base_q.filter(Order.status.in_(("completed", "delivered"))).all()
    cancelled = base_q.filter(Order.status == "cancelled").all()
    sales = [o for o in completed if (o.type or "sale") == "sale"]
    returns_o = [o for o in completed if o.type == "return_sale"]
    sales_total = sum(float(o.total or 0) for o in sales)
    returns_total = sum(float(o.total or 0) for o in returns_o)

    by_type: dict = {}
    sale_order_ids = [o.id for o in sales]
    if sale_order_ids:
        try:
            pmt_q = db.query(Payment).filter(
                Payment.type == "income",
                or_(Payment.status == "confirmed", Payment.status.is_(None)),
                Payment.order_id.in_(sale_order_ids),
            ).all()
            cash_pt_map = {c.id: ((c.payment_type or "naqd").strip().lower()) for c in db.query(CashRegister).all()}
            sale_with_payments: set = set()
            for p in pmt_q:
                sale_with_payments.add(p.order_id)
                pt = cash_pt_map.get(p.cash_register_id, "naqd")
                if pt == "perechisleniye":
                    pt = "bank"
                if pt not in by_type:
                    by_type[pt] = {"count": 0, "sum": 0.0}
                by_type[pt]["count"] += 1
                by_type[pt]["sum"] += float(p.amount or 0)
            # To'lovsiz buyurtmalar → qarz
            qarz_orders = [o for o in sales if o.id not in sale_with_payments]
            if qarz_orders:
                by_type["qarz"] = {
                    "count": len(qarz_orders),
                    "sum": sum(float(o.total or 0) for o in qarz_orders),
                }
        except Exception:
            # Fallback: Order.payment_type dan o'qish
            for o in sales:
                pt = (o.payment_type or "naqd").lower()
                if pt not in by_type:
                    by_type[pt] = {"count": 0, "sum": 0.0}
                by_type[pt]["count"] += 1
                by_type[pt]["sum"] += float(o.total or 0)

    cash_snapshot: list = []
    try:
        from app.services.finance_service import cash_balance_formula
        from sqlalchemy.orm import joinedload as _jl
        user_full = db.query(User).options(_jl(User.cash_registers_list)).filter(User.id == current_user.id).first()
        user_cashes_z = list(getattr(user_full, "cash_registers_list", None) or []) if user_full else []
        wh_dept_id = getattr(pos_wh, "department_id", None) if pos_wh else None
        if role == "sotuvchi" and wh_dept_id:
            shop_cashes = [c for c in user_cashes_z if getattr(c, "department_id", None) == wh_dept_id]
            if not shop_cashes:
                shop_cashes = db.query(CashRegister).filter(
                    CashRegister.department_id == wh_dept_id,
                    CashRegister.is_active == True,
                ).all()
        else:
            shop_cashes = user_cashes_z
        for c in shop_cashes:
            bal, inc, exp = cash_balance_formula(db, c.id)
            cash_snapshot.append({"id": c.id, "name": c.name, "balance": float(bal or 0), "income": float(inc or 0), "expense": float(exp or 0)})
    except Exception:
        pass

    payments_made: list = []
    expenses_made: list = []
    try:
        op_rows = (
            db.query(Payment)
            .filter(
                func.date(Payment.created_at) == target_date,
                Payment.user_id == current_user.id,
                Payment.type == "expense",
                Payment.category.in_(("supplier_payment", "expense")),
            )
            .order_by(Payment.created_at)
            .all()
        )
        for p in op_rows:
            partner_name = ""
            if p.partner_id:
                pn = db.query(Partner.name).filter(Partner.id == p.partner_id).first()
                partner_name = pn[0] if pn else ""
            cash_name = ""
            if p.cash_register_id:
                cn = db.query(CashRegister.name).filter(CashRegister.id == p.cash_register_id).first()
                cash_name = cn[0] if cn else ""
            rec = {
                "number": p.number or "",
                "time": p.created_at.strftime("%H:%M") if p.created_at else "",
                "amount": float(p.amount or 0),
                "payment_type": (p.payment_type or "naqd"),
                "cash_name": cash_name,
                "partner_name": partner_name,
                "description": p.description or "",
            }
            if (p.category or "").strip() == "supplier_payment":
                payments_made.append(rec)
            else:
                expenses_made.append(rec)
    except Exception:
        pass

    # Naqd kassa hisoboti — savdo, harajat, to'lov + oldingi Z'dan boshlang'ich qoldiq
    from app.utils.z_cash_summary import compute_z_cash_summary, find_previous_z_remaining
    _close_dt = dt.now()
    _now_iso = _close_dt.isoformat()
    cash_summary = compute_z_cash_summary(db, target_date, current_user.id, until_dt=_close_dt)
    prev_remaining, prev_zid = find_previous_z_remaining(
        user_id=current_user.id,
        warehouse_id=(pos_wh.id if pos_wh else None),
        before_closed_at=_now_iso,
    )
    cash_remaining = (
        prev_remaining
        + cash_summary["cash_sales_total"]
        - cash_summary["cash_expenses_total"]
        - cash_summary["cash_payments_out"]
        - cash_summary.get("cash_transfers_out", 0.0)  # 2026-05-26 fix
    )

    snapshot = {
        "z_id": f"Z-{target_date.strftime('%Y%m%d')}-U{current_user.id}-{dt.now().strftime('%H%M%S')}",
        "date": target_date.strftime("%Y-%m-%d"),
        "closed_at": _now_iso,
        "user_id": current_user.id,
        "user": current_user.full_name or current_user.username,
        "role": role,
        "warehouse_id": pos_wh.id if pos_wh else None,
        "warehouse": pos_wh.name if pos_wh else "Barcha",
        "sales_count": len(sales),
        "sales_total": sales_total,
        "returns_count": len(returns_o),
        "returns_total": returns_total,
        "cancelled_count": len(cancelled),
        "cancelled_total": sum(float(o.total or 0) for o in cancelled),
        "net_total": sales_total - returns_total,
        "payment_breakdown": [{"type": k, "count": v["count"], "sum": v["sum"]} for k, v in by_type.items()],
        "cash_snapshot": cash_snapshot,
        "order_numbers": [o.number for o in completed],
        "payments_made": payments_made,
        "payments_made_sum": sum(r["amount"] for r in payments_made),
        "expenses_made": expenses_made,
        "expenses_made_sum": sum(r["amount"] for r in expenses_made),
        # Naqd kassa hisoboti
        "cash_sales_pure": cash_summary["cash_sales_pure"],
        "cash_sales_split": cash_summary["cash_sales_split"],
        "cash_sales_total": cash_summary["cash_sales_total"],
        "cash_expenses_total": cash_summary["cash_expenses_total"],
        "cash_payments_out": cash_summary["cash_payments_out"],
        "previous_cash_remaining": prev_remaining,
        "previous_z_id": prev_zid,
        "cash_remaining": cash_remaining,
    }

    out_path = None
    try:
        out_dir = os.path.join("data", "z_reports", target_date.strftime("%Y-%m-%d"))
        os.makedirs(out_dir, exist_ok=True)
        out_path = os.path.join(out_dir, f"{snapshot['z_id']}.json")
        with open(out_path, "w", encoding="utf-8") as f:
            _json.dump(snapshot, f, ensure_ascii=False, indent=2)
    except Exception as e:
        return JSONResponse({"ok": False, "error": f"Snapshot saqlashda xato: {e}"}, status_code=500)

    try:
        log_action(
            db, user=current_user, action="z_report",
            entity_type="pos_shift", entity_id=0, entity_number=snapshot["z_id"],
            details=f"Z-hisobot: {target_date}, Sotuv {sales_total:,.0f}, Qaytarish {returns_total:,.0f}, NET {sales_total - returns_total:,.0f}",
            ip_address=request.client.host if request.client else "",
        )
    except Exception:
        pass

    return JSONResponse({"ok": True, "snapshot_id": snapshot["z_id"], "path": out_path})


@router.post("/pos/draft/save")
async def sales_pos_draft_save(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Chekni saqlash — savatdagi tovarlarni vaqtinchalik saqlab qo'yish."""
    role = (current_user.role or "").strip()
    if role not in ("sotuvchi", "admin", "manager"):
        return JSONResponse({"ok": False, "error": "Ruxsat yo'q"}, status_code=403)
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"ok": False, "error": "JSON xato"}, status_code=400)
    items = body.get("items")
    if not items or not isinstance(items, list):
        return JSONResponse({"ok": False, "error": "Savat bo'sh. Kamida bitta mahsulot qo'shing."}, status_code=400)
    warehouse = _get_pos_warehouse_for_user(db, current_user)
    name = (body.get("name") or "").strip() or None
    items_json = json.dumps(items, ensure_ascii=False)
    draft = PosDraft(
        user_id=current_user.id,
        warehouse_id=warehouse.id if warehouse else None,
        name=name,
        items_json=items_json,
    )
    db.add(draft)
    db.commit()
    db.refresh(draft)
    return JSONResponse({"ok": True, "id": draft.id, "message": "Chek saqlandi."})


@router.get("/pos/drafts")
async def sales_pos_drafts_list(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Saqlangan cheklar ro'yxati."""
    role = (current_user.role or "").strip()
    if role not in ("sotuvchi", "admin", "manager"):
        return JSONResponse([], status_code=200)
    drafts = (
        db.query(PosDraft)
        .filter(PosDraft.user_id == current_user.id)
        .order_by(PosDraft.created_at.desc())
        .limit(50)
        .all()
    )
    out = []
    for d in drafts:
        try:
            items = json.loads(d.items_json or "[]")
        except Exception:
            items = []
        total = sum((float(x.get("price") or 0) * float(x.get("quantity") or 0)) for x in items)
        out.append({
            "id": d.id,
            "name": d.name or f"Chek #{d.id}",
            "created_at": d.created_at.strftime("%d.%m.%Y %H:%M") if d.created_at else "-",
            "total": round(total, 2),
            "item_count": len(items),
        })
    return out


@router.get("/pos/draft/{draft_id}")
async def sales_pos_draft_get(
    draft_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Bitta saqlangan chekni olish (savatga yuklash uchun)."""
    role = (current_user.role or "").strip()
    if role not in ("sotuvchi", "admin", "manager"):
        return JSONResponse({"ok": False, "error": "Ruxsat yo'q"}, status_code=403)
    draft = db.query(PosDraft).filter(PosDraft.id == draft_id, PosDraft.user_id == current_user.id).first()
    if not draft:
        return JSONResponse({"ok": False, "error": "Chek topilmadi"}, status_code=404)
    try:
        items = json.loads(draft.items_json or "[]")
    except Exception:
        items = []
    return JSONResponse({"ok": True, "items": items})


@router.post("/pos/complete")
async def sales_pos_complete(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """POS savatni sotuv qilish. Naqd mijoz → pul kassaga; boshqa kontragent → qarz."""
    ensure_orders_payment_due_date_column(db)
    role = (current_user.role or "").strip()
    if role not in ("sotuvchi", "admin", "manager"):
        return RedirectResponse(url="/?error=pos_access", status_code=303)
    form = await request.form()
    payment_type = (form.get("payment_type") or "").strip().lower()
    if payment_type not in ("naqd", "plastik", "click", "terminal", "split"):
        return RedirectResponse(url="/sales/pos?error=payment", status_code=303)
    warehouse = _get_pos_warehouse_for_user(db, current_user)
    wh_id_form = form.get("warehouse_id")
    if wh_id_form:
        try:
            wh_id = int(wh_id_form)
            allowed = _get_pos_warehouses_for_user(db, current_user)
            chosen = next((w for w in allowed if w.id == wh_id), None)
            if chosen:
                warehouse = chosen
        except (TypeError, ValueError):
            pass
    default_partner = _get_pos_partner(db)
    if not warehouse or not default_partner:
        return RedirectResponse(url="/sales/pos?error=config", status_code=303)
    partner_id_form = form.get("partner_id")
    partner = default_partner
    if partner_id_form and str(partner_id_form).strip().isdigit():
        try:
            pid = int(partner_id_form)
            p = db.query(Partner).filter(Partner.id == pid, Partner.is_active == True).first()
            if p:
                partner = p
        except (ValueError, TypeError):
            pass
    product_ids = [int(x) for x in form.getlist("product_id") if str(x).strip().isdigit()]
    quantities = []
    for q in form.getlist("quantity"):
        try:
            quantities.append(float(q))
        except (ValueError, TypeError):
            pass
    prices = []
    for p in form.getlist("price"):
        try:
            prices.append(float(p))
        except (ValueError, TypeError):
            pass
    if not product_ids or len(quantities) < len(product_ids):
        return RedirectResponse(url="/sales/pos?error=empty", status_code=303)
    price_type = _get_pos_price_type(db)
    last_order = db.query(Order).filter(Order.type == "sale").order_by(Order.id.desc()).first()
    new_number = f"S-{datetime.now().strftime('%Y%m%d')}-{(last_order.id + 1) if last_order else 1:04d}"
    order = Order(
        number=new_number,
        type="sale",
        partner_id=partner.id,
        warehouse_id=warehouse.id,
        price_type_id=price_type.id if price_type else None,
        user_id=current_user.id if current_user else None,
        status="draft",
        payment_type=payment_type,
    )
    db.add(order)
    db.commit()
    db.refresh(order)
    total_order = 0.0
    items_for_stock = []
    for i in range(min(len(product_ids), len(quantities))):
        pid, qty = product_ids[i], float(quantities[i])
        if not pid or qty <= 0:
            continue
        price = prices[i] if i < len(prices) and prices[i] >= 0 else None
        if price is None or price < 0:
            pp = db.query(ProductPrice).filter(ProductPrice.product_id == pid, ProductPrice.price_type_id == order.price_type_id).first()
            if pp:
                price = pp.sale_price or 0
            else:
                prod = db.query(Product).filter(Product.id == pid).first()
                price = (prod.sale_price or prod.purchase_price or 0) if prod else 0
        total_row = qty * price
        db.add(OrderItem(order_id=order.id, product_id=pid, quantity=qty, price=price, total=total_row))
        total_order += total_row
        items_for_stock.append((pid, qty))
    order.subtotal = total_order
    discount_percent = 0.0
    discount_amount = 0.0
    try:
        discount_percent = float(form.get("discount_percent") or 0)
    except (ValueError, TypeError):
        pass
    try:
        discount_amount = float(form.get("discount_amount") or 0)
    except (ValueError, TypeError):
        pass
    if discount_percent < 0 or discount_percent > 100:
        discount_percent = 0.0
    if discount_amount < 0 or discount_amount > total_order:
        discount_amount = 0.0
    discount_sum = (total_order * discount_percent / 100.0) + discount_amount
    if discount_sum > total_order:
        discount_sum = total_order
    order.discount_percent = discount_percent
    order.discount_amount = discount_amount
    order.total = total_order - discount_sum
    is_cash_client = (partner.id == default_partner.id)
    if is_cash_client:
        order.paid = order.total
        order.debt = 0
    else:
        order.paid = 0
        order.debt = order.total
        # D4 audit fix: kredit limit tekshiruvi qarz mijozlar uchun
        from app.services.partner_credit import check_credit_limit
        ok, err = check_credit_limit(partner, float(order.debt or 0))
        if not ok:
            db.rollback()
            from urllib.parse import quote
            return RedirectResponse(
                url="/sales/pos?error=" + quote(err),
                status_code=303,
            )
        due_str = (form.get("payment_due_date") or "").strip()
        if due_str:
            try:
                order.payment_due_date = datetime.strptime(due_str, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                order.payment_due_date = (datetime.now() + timedelta(days=7)).date()
        else:
            order.payment_due_date = (datetime.now() + timedelta(days=7)).date()
    # with_for_update() — bir vaqtda 2 so'rov bir xil zaxirani olishini oldini olish
    for pid, qty in items_for_stock:
        stock = db.query(Stock).filter(
            Stock.warehouse_id == order.warehouse_id,
            Stock.product_id == pid
        ).with_for_update().first()
        if not stock or (stock.quantity or 0) < qty:
            prod = db.query(Product).filter(Product.id == pid).first()
            name = prod.name if prod else f"#{pid}"
            mavjud = float(stock.quantity or 0) if stock else 0
            order.status = "cancelled"
            db.commit()
            detail = f"Yetarli yo'q: {name} (savatda: {qty}, omborda: {mavjud:.0f})"
            url = "/sales/pos?error=stock&detail=" + quote(detail)
            if warehouse and warehouse.id:
                url += "&warehouse_id=" + str(warehouse.id)
            return RedirectResponse(url=url, status_code=303)
    for pid, qty in items_for_stock:
        create_stock_movement(
            db=db,
            warehouse_id=order.warehouse_id,
            product_id=pid,
            quantity_change=-qty,
            operation_type="sale",
            document_type="Sale",
            document_id=order.id,
            document_number=order.number,
            user_id=current_user.id if current_user else None,
            note=f"Sotuv (POS {payment_type}): {order.number}",
            created_at=order.date,
        )
    order.status = "completed"
    db.commit()
    # F4 realtime: dashboard v2 ga push (silent fail)
    try:
        from app.services.realtime_bus import publish_event
        publish_event("sale_created", {
            "order_id": order.id,
            "number": order.number,
            "amount": float(order.total or 0),
        })
    except Exception:
        pass
    if is_cash_client:
        department_id = getattr(warehouse, "department_id", None) if warehouse else None
        if not department_id and current_user:
            department_id = getattr(current_user, "department_id", None)
        def _map_pay_type(pt: str) -> str:
            pt = (pt or "").strip().lower()
            if pt == "naqd":
                return "cash"
            if pt == "click":
                return "click"
            if pt == "terminal":
                return "terminal"
            return "card"

        total_to_pay = float(order.total or 0)
        if total_to_pay > 0:
            parts = None
            if payment_type == "split":
                raw = (form.get("payment_splits") or "").strip()
                try:
                    parts = json.loads(raw) if raw else None
                except Exception:
                    parts = None
                if not isinstance(parts, list) or not parts:
                    return RedirectResponse(url="/sales/pos?error=payment", status_code=303)
                cleaned = []
                sum_parts = 0.0
                allowed = ("naqd", "plastik", "click", "terminal")
                for p in parts:
                    if not isinstance(p, dict):
                        continue
                    t = (p.get("type") or "").strip().lower()
                    if t not in allowed:
                        continue
                    try:
                        amt = float(p.get("amount") or 0)
                    except (ValueError, TypeError):
                        amt = 0.0
                    if amt <= 0:
                        continue
                    cleaned.append({"type": t, "amount": amt})
                    sum_parts += amt
                if not cleaned or abs(sum_parts - total_to_pay) >= 0.01:
                    return RedirectResponse(url="/sales/pos?error=payment", status_code=303)
                parts = cleaned
            else:
                parts = [{"type": payment_type, "amount": total_to_pay}]

            today_str = datetime.now().strftime('%Y%m%d')
            today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            last_pay = db.query(Payment).filter(
                Payment.number.like(f"PAY-{today_str}-%"),
                Payment.created_at >= today_start,
            ).order_by(Payment.id.desc()).first()
            if last_pay and last_pay.number:
                try:
                    seq = int(last_pay.number.split("-")[-1]) + 1
                except (ValueError, IndexError):
                    seq = db.query(Payment).filter(Payment.created_at >= today_start).count() + 1
            else:
                seq = 1
            for part in parts:
                pt = part["type"]
                amt = float(part["amount"] or 0)
                cash_register = _get_pos_cash_register(db, pt, department_id, current_user=current_user)
                if not cash_register:
                    return RedirectResponse(url="/sales/pos?error=payment", status_code=303)
                pay_number = f"PAY-{today_str}-{seq:04d}"
                seq += 1
                db.add(Payment(
                    number=pay_number,
                    type="income",
                    cash_register_id=cash_register.id,
                    partner_id=order.partner_id,
                    order_id=order.id,
                    amount=amt,
                    payment_type=_map_pay_type(pt),
                    category="sale",
                    description=f"POS sotuv {order.number}",
                    user_id=current_user.id if current_user else None,
                ))
                if getattr(cash_register, "balance", None) is not None:
                    db.flush()
                    _sync_cash_balance(db, cash_register.id)
            db.commit()
    else:
        if order.partner_id:
            from app.services.partner_balance_service import recompute_partner_balance
            db.flush()
            recompute_partner_balance(db, order.partner_id, reason="sale_create",
                                      ref=order.number,
                                      actor=current_user.username if current_user else None)
        db.commit()
    log_action(db, user=current_user, action="create", entity_type="sale",
               entity_id=order.id, entity_number=order.number,
               details=f"Summa: {order.total:,.0f}, To'lov: {payment_type}, Partner: {partner.name}",
               ip_address=request.client.host if request.client else "")
    db.commit()
    check_low_stock_and_notify(db)
    try:
        from app.bot.services.audit_watchdog import audit_sale
        audit_sale(order.id)
    except Exception:
        pass
    return RedirectResponse(url="/sales/pos?success=1&number=" + order.number, status_code=303)


@router.get("/pos/receipt", response_class=HTMLResponse)
async def sales_pos_receipt(
    request: Request,
    number: str = "",
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """POS sotuv cheki — chop etish sahifasi."""
    if not number or not number.strip():
        return HTMLResponse("<html><body><p>Hujjat raqami ko'rsatilmagan.</p></body></html>", status_code=400)
    order = (
        db.query(Order)
        .options(
            joinedload(Order.items).joinedload(OrderItem.product),
            joinedload(Order.partner),
            joinedload(Order.user),
        )
        .filter(Order.number == number.strip(), Order.type == "sale")
        .first()
    )
    if not order:
        return HTMLResponse("<html><body><p>Hujjat topilmadi.</p></body></html>", status_code=404)
    receipt_barcode_b64 = None
    try:
        writer = ImageWriter()
        writer.set_options({
            "module_width": 0.35,
            "module_height": 14,
            "font_size": 10,
            "dpi": 600,
        })
        buf = io.BytesIO()
        code128 = barcode.get("code128", order.number, writer=writer)
        code128.write(buf)
        buf.seek(0)
        receipt_barcode_b64 = base64.b64encode(buf.getvalue()).decode("ascii")
    except Exception:
        pass
    return templates.TemplateResponse("sales/pos_receipt.html", {
        "request": request,
        "order": order,
        "receipt_barcode_b64": receipt_barcode_b64,
    })


# ---------- Savdodan qaytarish ----------
@router.get("/returns", response_class=HTMLResponse)
async def sales_returns_list(request: Request, db: Session = Depends(get_db), current_user: User = Depends(require_auth)):
    """Savdodan qaytarish — bajarilgan sotuvlar ro'yxati (faqat joriy foydalanuvchi)."""
    q = db.query(Order).filter(
        Order.type == "sale",
        Order.status.in_(("completed", "delivered"))
    )
    # Admin/manager barcha sotuvlarni ko'radi, sotuvchi faqat o'zinikini
    if current_user.role not in ("admin", "manager"):
        q = q.filter(Order.user_id == current_user.id)
    orders = q.options(
        joinedload(Order.partner),
        joinedload(Order.warehouse),
    ).order_by(Order.date.desc()).limit(QUERY_LIMIT_DEFAULT).all()
    success = request.query_params.get("success")
    number = request.query_params.get("number", "")
    warehouse_name = unquote(request.query_params.get("warehouse", "") or "")
    error = request.query_params.get("error")
    error_detail = unquote(request.query_params.get("detail", "") or "")
    rq = db.query(Order).filter(Order.type == "return_sale")
    if current_user.role not in ("admin", "manager"):
        rq = rq.filter(Order.user_id == current_user.id)
    return_docs = rq.options(
        joinedload(Order.partner),
        joinedload(Order.warehouse),
    ).order_by(Order.created_at.desc()).limit(100).all()
    return templates.TemplateResponse("sales/returns_list.html", {
        "request": request,
        "orders": orders,
        "return_docs": return_docs,
        "page_title": "Savdodan qaytarish",
        "current_user": current_user,
        "success": success,
        "number": number,
        "warehouse_name": warehouse_name,
        "error": error,
        "error_detail": error_detail,
    })


@router.get("/return/{order_id}", response_class=HTMLResponse)
async def sales_return_form(
    request: Request,
    order_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Savdodan qaytarish — tanlangan sotuv bo'yicha qaytarish miqdorlarini kiritish."""
    order = db.query(Order).filter(
        Order.id == order_id,
        Order.type == "sale",
        Order.status.in_(("completed", "delivered"))
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Sotuv topilmadi yoki bajarilmagan.")
    error = request.query_params.get("error")
    error_detail = unquote(request.query_params.get("detail", "") or "")
    return templates.TemplateResponse("sales/return_form.html", {
        "request": request,
        "order": order,
        "page_title": "Savdodan qaytarish",
        "current_user": current_user,
        "error": error,
        "error_detail": error_detail,
    })


@router.post("/return/create")
async def sales_return_create(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Savdodan qaytarishni rasmiylashtirish."""
    form = await request.form()
    order_id_raw = form.get("order_id")
    if not order_id_raw or not str(order_id_raw).strip().isdigit():
        return RedirectResponse(url="/sales/returns?error=empty&detail=" + quote("Sotuv tanlanmadi."), status_code=303)
    order_id = int(order_id_raw)
    sale = db.query(Order).filter(
        Order.id == order_id,
        Order.type == "sale",
        Order.status.in_(("completed", "delivered"))
    ).options(joinedload(Order.items)).first()
    if not sale:
        return RedirectResponse(url="/sales/returns?error=not_found&detail=" + quote("Sotuv topilmadi."), status_code=303)
    product_ids = [int(x) for x in form.getlist("product_id") if str(x).strip().isdigit()]
    quantities_raw = form.getlist("quantity_return")
    quantities = []
    for q in quantities_raw:
        try:
            quantities.append(float(q))
        except (ValueError, TypeError):
            quantities.append(0)
    if not product_ids or all(q <= 0 for q in quantities[:len(product_ids)]):
        return RedirectResponse(
            url="/sales/return/" + str(order_id) + "?error=empty&detail=" + quote("Kamida bitta mahsulot uchun qaytarish miqdorini kiriting."),
            status_code=303
        )
    sale_items_by_product = {item.product_id: item for item in sale.items}
    for i in range(min(len(product_ids), len(quantities))):
        pid, qty = product_ids[i], quantities[i]
        if qty <= 0:
            continue
        item = sale_items_by_product.get(pid)
        if not item:
            prod = db.query(Product).filter(Product.id == pid).first()
            name = prod.name if prod else "#" + str(pid)
            return RedirectResponse(
                url="/sales/return/" + str(order_id) + "?error=qty&detail=" + quote(f"'{name}' ushbu sotuvda yo'q."),
                status_code=303
            )
        sold_qty = item.quantity or 0
        if qty > sold_qty + 1e-6:
            name = (item.product.name if item.product else "") or ("#" + str(pid))
            return RedirectResponse(
                url="/sales/return/" + str(order_id) + "?error=qty&detail=" + quote(f"'{name}' uchun qaytarish miqdori sotilgan miqdordan oshmasin (sotilgan: {sold_qty:.3f}, kiritilgan: {qty:.3f})."),
                status_code=303
            )
    from datetime import date as date_type
    today_start = date_type.today()
    # Do'kon omborlaridan qaytarilsa — o'sha omborga, qolganlaridan — Vozvrat omboriga
    sale_wh = db.query(Warehouse).filter(Warehouse.id == sale.warehouse_id).first()
    if sale_wh and "do'kon" in (sale_wh.name or "").lower():
        return_warehouse_id = sale.warehouse_id
    else:
        vozvrat_wh = db.query(Warehouse).filter(Warehouse.name.ilike("%vozvrat%"), Warehouse.is_active == True).first()
        return_warehouse_id = vozvrat_wh.id if vozvrat_wh else sale.warehouse_id
    if not return_warehouse_id:
        return RedirectResponse(
            url="/sales/returns?error=no_warehouse&detail=" + quote("Ombor topilmadi. Avval ombor yarating."),
            status_code=303
        )
    count = db.query(Order).filter(
        Order.type == "return_sale",
        func.date(Order.created_at) == today_start
    ).count()
    new_number = f"R-{datetime.now().strftime('%Y%m%d')}-{count + 1:04d}"
    return_order = Order(
        number=new_number,
        type="return_sale",
        partner_id=sale.partner_id,
        warehouse_id=return_warehouse_id,
        price_type_id=sale.price_type_id,
        user_id=current_user.id if current_user else None,
        status="completed",
        payment_type=sale.payment_type,
        note=f"Savdodan qaytarish: {sale.number}",
    )
    db.add(return_order)
    db.commit()
    db.refresh(return_order)
    total_return = 0.0
    for i in range(min(len(product_ids), len(quantities))):
        pid, qty = product_ids[i], quantities[i]
        if not pid or qty <= 0:
            continue
        item = sale_items_by_product.get(pid)
        if not item:
            continue
        price = item.price or 0
        total_row = qty * price
        db.add(OrderItem(order_id=return_order.id, product_id=pid, quantity=qty, price=price, total=total_row))
        total_return += total_row
        create_stock_movement(
            db=db,
            warehouse_id=return_warehouse_id,
            product_id=pid,
            quantity_change=+qty,
            operation_type="return_sale",
            document_type="SaleReturn",
            document_id=return_order.id,
            document_number=return_order.number,
            user_id=current_user.id if current_user else None,
            note=f"Savdodan qaytarish: {sale.number} -> {return_order.number}",
            created_at=return_order.date,
        )
    return_order.subtotal = total_return
    return_order.total = total_return
    return_order.paid = total_return
    return_order.debt = 0
    db.flush()
    # === Refund: original sotuv naqd to'langan bo'lsa kassadan proporsional chiqim ===
    from app.services.refund_service import compute_return_refund
    returned_lines = [(product_ids[i], quantities[i]) for i in range(min(len(product_ids), len(quantities)))
                      if product_ids[i] and quantities[i] > 0]
    rinfo = compute_return_refund(db, sale, returned_lines)
    if rinfo["return_total"] > 0:
        return_order.total = rinfo["return_total"]
        return_order.subtotal = rinfo["return_total"]
        return_order.paid = rinfo["return_total"]
    has_child = db.query(Order.id).filter(Order.parent_order_id == return_order.id).first()
    if rinfo["refund_cash"] > 0 and rinfo["refund_cash_register_id"] and not has_child:
        _today = datetime.now().strftime('%Y%m%d')
        _last = db.query(Payment).filter(Payment.number.like(f"PAY-{_today}-%")).order_by(Payment.number.desc()).first()
        _seq = (int(_last.number.split("-")[-1]) + 1) if (_last and _last.number) else 1
        db.add(Payment(
            number=f"PAY-{_today}-{_seq:04d}",
            date=datetime.now(),
            type="expense",
            category="sale_return",
            payment_type="cash",
            status="confirmed",
            partner_id=sale.partner_id,
            order_id=return_order.id,
            cash_register_id=rinfo["refund_cash_register_id"],
            amount=rinfo["refund_cash"],
            description=f"Qaytarish refund: {return_order.number} ({sale.number})",
            user_id=current_user.id if current_user else None,
        ))
        db.flush()
        from app.services.finance_service import sync_cash_balance
        sync_cash_balance(db, rinfo["refund_cash_register_id"])
        if sale.partner_id:
            from app.services.partner_balance_service import recompute_partner_balance
            recompute_partner_balance(db, sale.partner_id, reason="sale_return_refund", ref=return_order.number)
    db.commit()
    wh_name = ""
    if return_warehouse_id:
        wh = db.query(Warehouse).filter(Warehouse.id == return_warehouse_id).first()
        wh_name = (wh.name or "").strip()
    params = "success=1&number=" + quote(return_order.number)
    if wh_name:
        params += "&warehouse=" + quote(wh_name)
    return RedirectResponse(url="/sales/returns?" + params, status_code=303)


@router.get("/return/document/{number}", response_class=HTMLResponse)
async def sales_return_document(
    request: Request,
    number: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Qaytarish hujjati (R-...) — ko'rish / chop etish."""
    doc = (
        db.query(Order)
        .options(
            joinedload(Order.items).joinedload(OrderItem.product),
            joinedload(Order.partner),
            joinedload(Order.warehouse),
            joinedload(Order.user),
        )
        .filter(Order.number == number.strip(), Order.type == "return_sale")
        .first()
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Qaytarish hujjati topilmadi.")
    return templates.TemplateResponse("sales/return_document.html", {
        "request": request,
        "doc": doc,
        "page_title": "Qaytarish " + doc.number,
        "current_user": current_user,
    })


@router.post("/return/revert/{return_order_id}")
async def sales_return_revert(
    return_order_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Qaytarish tasdiqini bekor qilish (faqat admin)."""
    doc = (
        db.query(Order)
        .options(joinedload(Order.items))
        .filter(Order.id == return_order_id, Order.type == "return_sale")
        .first()
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Qaytarish hujjati topilmadi.")
    if doc.status != "completed":
        return RedirectResponse(
            url="/sales/returns?error=revert&detail=" + quote("Faqat tasdiqlangan qaytarishning tasdiqini bekor qilish mumkin."),
            status_code=303
        )
    wh_id = doc.warehouse_id
    if not wh_id:
        return RedirectResponse(
            url="/sales/returns?error=revert&detail=" + quote("Hujjatda ombor ko'rsatilmagan."),
            status_code=303
        )
    for item in doc.items:
        create_stock_movement(
            db=db,
            warehouse_id=wh_id,
            product_id=item.product_id,
            quantity_change=-(item.quantity or 0),
            operation_type="return_sale_revert",
            document_type="SaleReturnRevert",
            document_id=doc.id,
            document_number=doc.number,
            user_id=current_user.id if current_user else None,
            note=f"Qaytarish tasdiqini bekor: {doc.number}",
            created_at=doc.date,
        )
    # Refund Payment'ni o'chirish (kassa naqdini qaytarish)
    refund_pays = db.query(Payment).filter(
        Payment.order_id == doc.id, Payment.category == "sale_return", Payment.type == "expense"
    ).all()
    _registers = set()
    for rp in refund_pays:
        if rp.cash_register_id:
            _registers.add(rp.cash_register_id)
        db.delete(rp)
    db.flush()
    from app.services.finance_service import sync_cash_balance
    for _cr in _registers:
        sync_cash_balance(db, _cr)
    if doc.partner_id:
        from app.services.partner_balance_service import recompute_partner_balance
        recompute_partner_balance(db, doc.partner_id, reason="sale_return_revert", ref=doc.number)
    doc.status = "cancelled"
    db.commit()
    return RedirectResponse(url="/sales/return/document/" + doc.number + "?reverted=1", status_code=303)


@router.post("/return/delete/{return_order_id}")
async def sales_return_delete(
    return_order_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Qaytarish hujjatini o'chirish (faqat admin)."""
    doc = db.query(Order).filter(Order.id == return_order_id, Order.type == "return_sale").first()
    if not doc:
        raise HTTPException(status_code=404, detail="Qaytarish hujjati topilmadi.")
    if doc.status != "cancelled":
        return RedirectResponse(
            url="/sales/returns?error=delete&detail=" + quote("Faqat tasdiqni bekor qilgandan keyin o'chirish mumkin. Avval tasdiqni bekor qiling."),
            status_code=303
        )
    number = doc.number
    for item in list(doc.items):
        db.delete(item)
    db.delete(doc)
    db.commit()
    return RedirectResponse(url="/sales/returns?deleted=1&number=" + quote(number), status_code=303)


@router.get("/return/edit/{return_order_id}", response_class=HTMLResponse)
async def sales_return_edit_form(
    request: Request,
    return_order_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Qaytarish hujjatini tahrirlash (faqat tasdiqni bekor qilingan hujjat)."""
    doc = (
        db.query(Order)
        .options(
            joinedload(Order.items).joinedload(OrderItem.product),
            joinedload(Order.partner),
            joinedload(Order.warehouse),
        )
        .filter(Order.id == return_order_id, Order.type == "return_sale")
        .first()
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Qaytarish hujjati topilmadi.")
    if doc.status != "cancelled":
        return RedirectResponse(
            url="/sales/return/document/" + doc.number + "?error=edit&detail=" + quote("Faqat tasdiqni bekor qilingan hujjatni tahrirlash mumkin."),
            status_code=303
        )
    return templates.TemplateResponse("sales/return_edit.html", {
        "request": request,
        "doc": doc,
        "page_title": "Qaytarishni tahrirlash " + doc.number,
        "current_user": current_user,
    })


@router.post("/return/update")
async def sales_return_update(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Qaytarish hujjati qatorlarini yangilash — faqat bekor qilingan hujjat."""
    form = await request.form()
    order_id_raw = form.get("order_id")
    if not order_id_raw or not str(order_id_raw).strip().isdigit():
        return RedirectResponse(url="/sales/returns?error=update", status_code=303)
    return_order_id = int(order_id_raw)
    doc = (
        db.query(Order)
        .options(joinedload(Order.items))
        .filter(Order.id == return_order_id, Order.type == "return_sale")
        .first()
    )
    if not doc or doc.status != "cancelled":
        return RedirectResponse(url="/sales/returns?error=update&detail=" + quote("Hujjat topilmadi yoki tahrirlash mumkin emas."), status_code=303)
    product_ids = [int(x) for x in form.getlist("product_id") if str(x).strip().isdigit()]
    quantities = []
    for q in form.getlist("quantity"):
        try:
            quantities.append(float(q))
        except (ValueError, TypeError):
            quantities.append(0)
    prices = []
    for p in form.getlist("price"):
        try:
            prices.append(float(p))
        except (ValueError, TypeError):
            prices.append(0)
    items_by_pid = {item.product_id: item for item in doc.items}
    total_return = 0.0
    for i in range(min(len(product_ids), len(quantities))):
        pid, qty = product_ids[i], quantities[i]
        if not pid or qty < 0:
            continue
        item = items_by_pid.get(pid)
        if not item:
            continue
        price = prices[i] if i < len(prices) and prices[i] >= 0 else (item.price or 0)
        item.quantity = qty
        item.price = price
        item.total = qty * price
        total_return += item.total
    doc.subtotal = total_return
    doc.total = total_return
    doc.paid = total_return
    doc.debt = 0
    db.commit()
    return RedirectResponse(url="/sales/return/document/" + doc.number + "?updated=1", status_code=303)


@router.post("/return/confirm/{return_order_id}")
async def sales_return_confirm(
    return_order_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Qaytarishni qayta tasdiqlash (faqat bekor qilingan hujjat): omborga qoldiq qo'shish."""
    doc = (
        db.query(Order)
        .options(joinedload(Order.items))
        .filter(Order.id == return_order_id, Order.type == "return_sale")
        .first()
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Qaytarish hujjati topilmadi.")
    if doc.status != "cancelled":
        return RedirectResponse(
            url="/sales/returns?error=confirm&detail=" + quote("Faqat bekor qilingan hujjatni qayta tasdiqlash mumkin."),
            status_code=303
        )
    wh_id = doc.warehouse_id
    if not wh_id:
        return RedirectResponse(url="/sales/returns?error=confirm&detail=" + quote("Hujjatda ombor ko'rsatilmagan."), status_code=303)
    # Atomik UPDATE WHERE — double-confirm xavfini oldini olish
    from sqlalchemy import text as _text
    claim = db.execute(
        _text("UPDATE orders SET status='completed' WHERE id=:id AND type='return_sale' AND status='cancelled'"),
        {"id": return_order_id}
    )
    if claim.rowcount == 0:
        return RedirectResponse(url="/sales/returns?error=confirm&detail=" + quote("Hujjat allaqachon tasdiqlangan."), status_code=303)
    for item in doc.items:
        create_stock_movement(
            db=db,
            warehouse_id=wh_id,
            product_id=item.product_id,
            quantity_change=+(item.quantity or 0),
            operation_type="return_sale",
            document_type="SaleReturn",
            document_id=doc.id,
            document_number=doc.number,
            user_id=current_user.id if current_user else None,
            note=f"Qaytarish qayta tasdiqlandi: {doc.number}",
            created_at=doc.date,
        )
    # Status allaqachon atomik UPDATE WHERE bilan o'zgartirildi
    db.commit()
    return RedirectResponse(url="/sales/return/document/" + doc.number + "?confirmed=1", status_code=303)


# ==================== POS: Ta'minotchiga to'lov ====================
@router.post("/pos/pay-supplier")
async def pos_pay_supplier(
    request: Request,
    partner_id: int = Form(...),
    amount: float = Form(...),
    payment_type: str = Form("naqd"),
    note: str = Form(""),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Sotuvchi tanlangan kassadan ta'minotchiga to'lov qiladi."""
    if amount <= 0:
        return RedirectResponse(url="/sales/pos?error=payment&detail=Summa+noto'g'ri", status_code=303)
    pt = (payment_type or "naqd").strip().lower()
    if pt not in ("naqd", "plastik", "click", "terminal"):
        pt = "naqd"
    partner = db.query(Partner).filter(Partner.id == partner_id).first()
    if not partner:
        return RedirectResponse(url="/sales/pos?error=payment&detail=Kontragent+topilmadi", status_code=303)
    # Sotuvchining tanlangan turdagi kassasini topish
    department_id = getattr(current_user, "department_id", None)
    cash_register = _get_pos_cash_register(db, pt, department_id, current_user=current_user)
    if not cash_register:
        return RedirectResponse(url="/sales/pos?error=payment&detail=Kassa+topilmadi", status_code=303)
    # To'lov yaratish (chiqim)
    today_str = datetime.now().strftime('%Y%m%d')
    last_pay = db.query(Payment).filter(Payment.number.like(f"PAY-{today_str}-%")).order_by(Payment.number.desc()).first()
    next_seq = int(last_pay.number.split("-")[-1]) + 1 if last_pay else 1
    pay_number = f"PAY-{today_str}-{next_seq:04d}"
    db.add(Payment(
        number=pay_number,
        type="expense",
        cash_register_id=cash_register.id,
        partner_id=partner_id,
        amount=amount,
        payment_type=pt,
        category="supplier_payment",
        description=note or f"Ta'minotchiga to'lov: {partner.name}",
        user_id=current_user.id,
    ))
    # Kassa balansini kamaytirish
    if getattr(cash_register, "balance", None) is not None:
        db.flush()
        _sync_cash_balance(db, cash_register.id)
    # Partner balansini qayta hisoblash (chiqim Payment yuqorida yaratildi)
    if partner_id:
        from app.services.partner_balance_service import recompute_partner_balance
        db.flush()
        recompute_partner_balance(db, partner_id, reason="supplier_payment",
                                  ref=pay_number,
                                  actor=current_user.username if current_user else None)
    db.commit()
    return RedirectResponse(url="/sales/pos?success=1&number=" + quote(f"To'lov: {partner.name} ga {amount:,.0f} so'm ({pt})"), status_code=303)


# ==================== POS: Harajat yozish ====================
@router.post("/pos/expense")
async def pos_expense(
    request: Request,
    expense_type: str = Form(...),
    amount: float = Form(...),
    payment_type: str = Form("naqd"),
    note: str = Form(""),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Sotuvchi tanlangan kassadan harajat yozadi (naqd/plastik/click/terminal)."""
    if amount <= 0:
        return RedirectResponse(url="/sales/pos?error=payment&detail=Summa+noto'g'ri", status_code=303)
    pt = (payment_type or "naqd").strip().lower()
    if pt not in ("naqd", "plastik", "click", "terminal"):
        pt = "naqd"
    # Harajat turi
    expense_type_name = "Boshqa"
    if expense_type != "other":
        et = db.query(ExpenseType).filter(ExpenseType.id == int(expense_type)).first()
        if et:
            expense_type_name = et.name
    # Sotuvchining tanlangan turdagi kassasini topish
    department_id = getattr(current_user, "department_id", None)
    cash_register = _get_pos_cash_register(db, pt, department_id, current_user=current_user)
    if not cash_register:
        return RedirectResponse(url="/sales/pos?error=payment&detail=Kassa+topilmadi", status_code=303)
    # To'lov yaratish (chiqim — harajat)
    today_str = datetime.now().strftime('%Y%m%d')
    last_pay = db.query(Payment).filter(Payment.number.like(f"PAY-{today_str}-%")).order_by(Payment.number.desc()).first()
    next_seq = int(last_pay.number.split("-")[-1]) + 1 if last_pay else 1
    pay_number = f"PAY-{today_str}-{next_seq:04d}"
    description = f"Harajat: {expense_type_name}"
    if note:
        description += f" — {note}"
    db.add(Payment(
        number=pay_number,
        type="expense",
        cash_register_id=cash_register.id,
        amount=amount,
        payment_type=pt,
        category="expense",
        description=description,
        user_id=current_user.id,
    ))
    # Kassa balansini kamaytirish
    if getattr(cash_register, "balance", None) is not None:
        db.flush()
        _sync_cash_balance(db, cash_register.id)
    db.commit()
    return RedirectResponse(url="/sales/pos?success=1&number=" + quote(f"Harajat: {expense_type_name} — {amount:,.0f} so'm ({pt})"), status_code=303)
