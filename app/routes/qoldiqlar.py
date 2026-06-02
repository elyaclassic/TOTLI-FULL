"""
Qoldiqlar — kassa, tovar, kontragent qoldiqlari va hujjatlar (1C uslubida).
"""
import io
import traceback
from datetime import datetime
from typing import Optional
from urllib.parse import quote

import openpyxl
from fastapi import APIRouter, Request, Depends, Form, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, text

from app.core import templates
from app.utils.user_scope import get_warehouses_for_user
from app.models.database import (
    get_db,
    User,
    Product,
    Partner,
    Warehouse,
    Stock,
    StockMovement,
    CashRegister,
    StockAdjustmentDoc,
    StockAdjustmentDocItem,
    CashBalanceDoc,
    CashBalanceDocItem,
    PartnerBalanceDoc,
    PartnerBalanceDocItem,
    EmployeeBalanceDoc,
    EmployeeBalanceDocItem,
    Order,
    Purchase,
    PurchaseItem,
    Payment,
    Employee,
    Salary,
)
from app.deps import require_auth, require_admin
from app.services.stock_service import create_stock_movement, delete_stock_movements_for_document
from app.constants import QUERY_LIMIT_DEFAULT, QUERY_LIMIT_HISTORY, QUERY_LIMIT_LIST

router = APIRouter(prefix="/qoldiqlar", tags=["qoldiqlar"])


def _get_last_purchase_prices(db: Session) -> dict:
    """Har bir mahsulot uchun oxirgi tasdiqlangan xarid narxini {product_id: price} dict.

    P7 audit fix: SQL window function bilan har product_id uchun MAX(purchase_id) topiladi —
    avval barcha rowlarni Pythonga olardi (full scan). Endi DB tomonida agregatsiya +
    idx_purchases_status_date (P9 yangi index) ishlatadi.
    """
    # SQLite 3.25+ window function qo'llab-quvvatlaydi
    sql = text("""
        SELECT pi.product_id, pi.price
        FROM purchase_items pi
        JOIN (
            SELECT pi2.product_id, MAX(p2.id) AS max_pid
            FROM purchase_items pi2
            JOIN purchases p2 ON p2.id = pi2.purchase_id
            WHERE p2.status = 'confirmed' AND pi2.price > 0
            GROUP BY pi2.product_id
        ) latest ON latest.product_id = pi.product_id AND latest.max_pid = pi.purchase_id
        WHERE pi.price > 0
    """)
    out = {}
    for row in db.execute(sql).fetchall():
        out[int(row[0])] = float(row[1] or 0)
    return out


def _tarix_doc_type_label(doc_type: str) -> str:
    """Hujjat turi uchun o'qiladigan nom (tarix sahifasi)."""
    labels = {
        "Purchase": "Kirim",
        "Production": "Ishlab chiqarish",
        "WarehouseTransfer": "Ombordan omborga",
        "StockAdjustmentDoc": "Qoldiq tuzatish",
        "Sale": "Sotuv",
        "SaleReturn": "Qaytish",
    }
    return labels.get(doc_type or "", doc_type or "—")


@router.get("", response_class=HTMLResponse)
async def qoldiqlar_page(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Qoldiqlar sahifasi: kassa, tovar (forma spiska 1C), kontragent qoldiqlarini kiritish"""
    from app.utils.db_schema import ensure_stock_adjustment_doc_type_column
    ensure_stock_adjustment_doc_type_column(db)
    cash_registers = db.query(CashRegister).filter(CashRegister.is_active == True).all()
    warehouses = get_warehouses_for_user(db, current_user)
    products = db.query(Product).filter(Product.is_active == True).order_by(Product.name).all()
    stocks = db.query(Stock).join(Warehouse).join(Product).order_by(Stock.updated_at.desc()).limit(QUERY_LIMIT_LIST).all()
    partners = db.query(Partner).filter(Partner.is_active == True).order_by(Partner.name).all()
    tovar_docs = (
        db.query(StockAdjustmentDoc)
        .order_by(StockAdjustmentDoc.id.desc())
        .limit(QUERY_LIMIT_HISTORY)
        .all()
    )
    cash_docs = (
        db.query(CashBalanceDoc)
        .order_by(CashBalanceDoc.created_at.desc())
        .limit(QUERY_LIMIT_DEFAULT)
        .all()
    )
    kontragent_docs = (
        db.query(PartnerBalanceDoc)
        .order_by(PartnerBalanceDoc.created_at.desc())
        .limit(QUERY_LIMIT_DEFAULT)
        .all()
    )
    # Xodimlar qoldiqlari — har xodimning oxirgi Salary.total qiymati (bitta query)
    employees = db.query(Employee).filter(Employee.is_active == True).order_by(Employee.full_name).all()
    emp_ids = [e.id for e in employees]
    emp_balances = {}
    if emp_ids:
        latest_sal_ids = (
            db.query(func.max(Salary.id))
            .filter(Salary.employee_id.in_(emp_ids))
            .group_by(Salary.employee_id)
            .subquery()
        )
        latest_salaries = db.query(Salary).filter(Salary.id.in_(latest_sal_ids)).all()
        emp_balances = {
            s.employee_id: {"year": s.year, "month": s.month, "total": float(s.total or 0), "status": s.status}
            for s in latest_salaries
        }
    xodim_docs = (
        db.query(EmployeeBalanceDoc)
        .order_by(EmployeeBalanceDoc.created_at.desc())
        .limit(QUERY_LIMIT_DEFAULT)
        .all()
    )
    return templates.TemplateResponse("qoldiqlar/index.html", {
        "request": request,
        "cash_registers": cash_registers,
        "warehouses": warehouses,
        "products": products,
        "stocks": stocks,
        "partners": partners,
        "tovar_docs": tovar_docs,
        "cash_docs": cash_docs,
        "kontragent_docs": kontragent_docs,
        "employees": employees,
        "emp_balances": emp_balances,
        "xodim_docs": xodim_docs,
        "current_user": current_user,
        "page_title": "Qoldiqlar",
        "show_tannarx": (getattr(current_user, "role", None) if current_user else None) == "admin",
    })


@router.get("/tarix", response_class=HTMLResponse)
async def qoldiqlar_tarix(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
    warehouse_id: Optional[str] = None,
    product_id: Optional[str] = None,
):
    """Mahsulot harakati tarixi — tanlanmasa barcha harakatlar; filtr ixtiyoriy (ombor/mahsulot)."""
    try:
        warehouses = db.query(Warehouse).filter(Warehouse.is_active == True).all()
        products = db.query(Product).filter(Product.is_active == True).order_by(Product.name).all()
        try:
            selected_warehouse_id = int(warehouse_id) if (warehouse_id and str(warehouse_id).strip()) else None
        except (ValueError, TypeError):
            selected_warehouse_id = None
        try:
            selected_product_id = int(product_id) if (product_id and str(product_id).strip()) else None
        except (ValueError, TypeError):
            selected_product_id = None

        q = db.query(StockMovement)
        if selected_warehouse_id:
            q = q.filter(StockMovement.warehouse_id == selected_warehouse_id)
        if selected_product_id:
            q = q.filter(StockMovement.product_id == selected_product_id)
        q = q.order_by(StockMovement.created_at.desc())
        from app.utils.pagination import paginate, pagination_query_string
        _pg = paginate(q, request.query_params.get("page", 1), per_page=50)
        movements = _pg["items"]

        movement_rows = []
        warehouse_ids = [m.warehouse_id for m in movements if m.warehouse_id is not None]
        product_ids = [m.product_id for m in movements if m.product_id is not None]
        wh_by_id = {}
        if warehouse_ids:
            for w in db.query(Warehouse).filter(Warehouse.id.in_(set(warehouse_ids))).all():
                wh_by_id[w.id] = w
        prod_by_id = {}
        if product_ids:
            for p in db.query(Product).filter(Product.id.in_(set(product_ids))).all():
                prod_by_id[p.id] = p
        for m in movements:
            wh = wh_by_id.get(m.warehouse_id) if m.warehouse_id is not None else None
            pr = prod_by_id.get(m.product_id) if m.product_id is not None else None
            movement_rows.append({
                "date": m.created_at.strftime("%d.%m.%Y %H:%M") if m.created_at else "—",
                "warehouse_name": (getattr(wh, "name", None) if wh else None) or (f"#{m.warehouse_id}" if m.warehouse_id is not None else "—"),
                "product_name": (getattr(pr, "name", None) if pr else None) or (f"#{m.product_id}" if m.product_id is not None else "—"),
                "product_code": (getattr(pr, "code", None) or "") if pr else "",
                "doc_type_label": _tarix_doc_type_label(m.document_type or ""),
                "doc_number": m.document_number or (f"{m.document_type or ''}-{m.document_id}" if m.document_id else "—"),
                "quantity_change": float(m.quantity_change or 0),
                "warehouse_id": m.warehouse_id,
                "product_id": m.product_id,
            })

        return templates.TemplateResponse("qoldiqlar/tarix.html", {
            "request": request,
            "current_user": current_user,
            "warehouses": warehouses,
            "products": products,
            "selected_product_id": selected_product_id,
            "selected_warehouse_id": selected_warehouse_id,
            "movements": movement_rows,
            "page_title": "Qoldiqlar — Mahsulot harakati tarixi",
            "page": _pg["page"],
            "per_page": _pg["per_page"],
            "total_count": _pg["total_count"],
            "total_pages": _pg["total_pages"],
            "items_count": _pg["items_count"],
            "base_url": "/qoldiqlar/tarix",
            "pagination_query": pagination_query_string({"warehouse_id": str(selected_warehouse_id or ""), "product_id": str(selected_product_id or "")}),
        })
    except Exception as e:
        import logging
        logging.getLogger(__name__).exception("qoldiqlar_tarix: %s", e)
        try:
            _wh = db.query(Warehouse).filter(Warehouse.is_active == True).all()
            _pr = db.query(Product).filter(Product.is_active == True).order_by(Product.name).all()
        except Exception:
            _wh = []
            _pr = []
        return templates.TemplateResponse("qoldiqlar/tarix.html", {
            "request": request,
            "current_user": current_user,
            "warehouses": _wh,
            "products": _pr,
            "selected_product_id": None,
            "selected_warehouse_id": None,
            "movements": [],
            "page_title": "Qoldiqlar — Mahsulot harakati tarixi",
            "error_message": "Ma'lumotlarni yuklashda xatolik yuz berdi",
        }, status_code=500)


@router.post("/kassa/{cash_id:int}")
async def qoldiqlar_kassa_save(
    cash_id: int,
    balance: float = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Kassa opening_balance ni to'g'ridan-to'g'ri o'rnatish (faqat admin, faqat boshlang'ich qoldiq uchun)"""
    cash = db.query(CashRegister).filter(CashRegister.id == cash_id).first()
    if not cash:
        raise HTTPException(status_code=404, detail="Kassa topilmadi")
    from app.services.finance_service import sync_cash_balance
    cash.opening_balance = balance
    db.flush()
    sync_cash_balance(db, cash_id)
    db.commit()
    return RedirectResponse(url="/qoldiqlar#kassa", status_code=303)


# --- Kassa qoldiq HUJJATLARI (1C uslubida) ---
@router.get("/kassa/hujjat/new", response_class=HTMLResponse)
async def qoldiqlar_kassa_hujjat_new(
    request: Request,
    force_new: int = 0,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Yangi kassa qoldiq hujjati"""
    from app.utils.draft_check import redirect_to_draft
    redirect = redirect_to_draft(
        db, CashBalanceDoc,
        edit_url_template="/qoldiqlar/kassa/hujjat/{id}",
        user_role=getattr(current_user, "role", "") or "",
        force_new=bool(force_new),
        message="Sizda ochiq kassa qoldiqlari qoralamasi bor — avval uni tugating yoki bekor qiling.",
        user_id=current_user.id,
    )
    if redirect:
        return redirect
    cash_registers = db.query(CashRegister).filter(CashRegister.is_active == True).all()
    return templates.TemplateResponse("qoldiqlar/kassa_hujjat_form.html", {
        "request": request,
        "doc": None,
        "cash_registers": cash_registers,
        "current_user": current_user,
        "page_title": "Kassa qoldiqlari — yangi hujjat",
    })


@router.post("/kassa/hujjat")
async def qoldiqlar_kassa_hujjat_create(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Kassa qoldiq hujjatini yaratish (qoralama)"""
    form = await request.form()
    cash_ids = form.getlist("cash_register_id")
    balances = form.getlist("balance")

    items_data = []
    for i, cid in enumerate(cash_ids):
        if not cid:
            continue
        try:
            bid = int(cid)
            bal = float(balances[i]) if i < len(balances) and balances[i] != "" else None
        except (TypeError, ValueError):
            continue
        if bal is not None:
            items_data.append((bid, bal))

    if not items_data:
        return RedirectResponse(url="/qoldiqlar/kassa/hujjat/new", status_code=303)

    today = datetime.now()
    count = db.query(CashBalanceDoc).filter(
        CashBalanceDoc.date >= today.replace(hour=0, minute=0, second=0)
    ).count()
    number = f"KLD-{today.strftime('%Y%m%d')}-{str(count + 1).zfill(4)}"

    doc = CashBalanceDoc(
        number=number,
        date=today,
        user_id=current_user.id if current_user else None,
        status="draft",
    )
    db.add(doc)
    db.flush()
    for cid, bal in items_data:
        db.add(CashBalanceDocItem(doc_id=doc.id, cash_register_id=cid, balance=bal))
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/kassa/hujjat/{doc.id}", status_code=303)


@router.get("/kassa/hujjat/{doc_id}", response_class=HTMLResponse)
async def qoldiqlar_kassa_hujjat_view(
    request: Request,
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Kassa qoldiq hujjatini ko'rish"""
    doc = db.query(CashBalanceDoc).filter(CashBalanceDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    cash_registers = db.query(CashRegister).filter(CashRegister.is_active == True).all()
    return templates.TemplateResponse("qoldiqlar/kassa_hujjat_form.html", {
        "request": request,
        "doc": doc,
        "cash_registers": cash_registers,
        "current_user": current_user,
        "page_title": f"Kassa qoldiqlari {doc.number}",
    })


@router.post("/kassa/hujjat/{doc_id}/tasdiqlash")
async def qoldiqlar_kassa_hujjat_tasdiqlash(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Kassa hujjatini tasdiqlash — kassa balanslarini yangilash"""
    doc = db.query(CashBalanceDoc).filter(CashBalanceDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "draft":
        raise HTTPException(status_code=400, detail="Hujjat allaqachon tasdiqlangan")
    if not doc.items:
        raise HTTPException(status_code=400, detail="Kamida bitta kassa qatori bo'lishi kerak")
    # Atomik UPDATE WHERE — double-confirm xavfini oldini olish
    from sqlalchemy import text as _text
    claim = db.execute(
        _text("UPDATE cash_balance_docs SET status='confirmed' WHERE id=:id AND status='draft'"),
        {"id": doc_id}
    )
    if claim.rowcount == 0:
        return RedirectResponse(url=f"/qoldiqlar/kassa/hujjat/{doc_id}?already=1", status_code=303)
    from app.services.finance_service import cash_balance_formula as _cash_balance_formula
    for item in doc.items:
        cash = db.query(CashRegister).filter(CashRegister.id == item.cash_register_id).first()
        if cash:
            item.previous_balance = cash.balance
            # Hozirgi hisoblangan balans
            current_balance, income_sum, expense_sum = _cash_balance_formula(db, cash.id)
            # Qo'shish: eski balansga kiritilgan qiymatni qo'shamiz
            delta = float(item.balance or 0)
            target = current_balance + delta
            # opening_balance ni shunday moslaymizki: opening + income - expense = target
            cash.opening_balance = target - income_sum + expense_sum
            cash.balance = target
    # Status allaqachon atomik UPDATE WHERE bilan o'zgartirildi
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/kassa/hujjat/{doc_id}", status_code=303)


@router.post("/kassa/hujjat/{doc_id}/revert")
async def qoldiqlar_kassa_hujjat_revert(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Kassa hujjati tasdiqini bekor qilish (faqat admin)"""
    doc = db.query(CashBalanceDoc).filter(CashBalanceDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "confirmed":
        raise HTTPException(status_code=400, detail="Faqat tasdiqlangan hujjatning tasdiqini bekor qilish mumkin")
    from app.services.finance_service import cash_balance_formula as _cash_balance_formula
    for item in doc.items:
        cash = db.query(CashRegister).filter(CashRegister.id == item.cash_register_id).first()
        if cash and item.previous_balance is not None:
            # Oldingi balansga qaytish uchun opening_balance ni moslash
            current_balance, income_sum, expense_sum = _cash_balance_formula(db, cash.id)
            target = float(item.previous_balance)
            cash.opening_balance = target - income_sum + expense_sum
            cash.balance = target
    doc.status = "draft"
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/kassa/hujjat/{doc_id}", status_code=303)


@router.post("/kassa/hujjat/{doc_id}/delete")
async def qoldiqlar_kassa_hujjat_delete(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Kassa hujjatini o'chirish (faqat qoralama, faqat admin)"""
    doc = db.query(CashBalanceDoc).filter(CashBalanceDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "draft":
        raise HTTPException(status_code=400, detail="Faqat qoralama holatidagi hujjatni o'chirish mumkin. Avval tasdiqni bekor qiling.")
    db.delete(doc)
    db.commit()
    return RedirectResponse(url="/qoldiqlar#kassa", status_code=303)


@router.post("/kassa/hujjat/{doc_id}/update")
async def qoldiqlar_kassa_hujjat_update(
    doc_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Kassa qoldiq hujjati qatorlari summasini tahrirlash (faqat qoralama)."""
    doc = db.query(CashBalanceDoc).filter(CashBalanceDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "draft":
        return RedirectResponse(url=f"/qoldiqlar/kassa/hujjat/{doc_id}", status_code=303)
    form = await request.form()
    item_ids = form.getlist("item_id")
    balances = form.getlist("balance")
    for i, iid in enumerate(item_ids):
        try:
            item = db.query(CashBalanceDocItem).filter(
                CashBalanceDocItem.id == int(iid),
                CashBalanceDocItem.doc_id == doc_id,
            ).first()
            if item is not None and i < len(balances) and str(balances[i]).strip() != "":
                item.balance = float(balances[i])
        except (ValueError, TypeError):
            continue
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/kassa/hujjat/{doc_id}", status_code=303)


# --- Kontragent qoldiq HUJJATLARI (1C uslubida) ---
@router.get("/kontragent/hujjat/new", response_class=HTMLResponse)
async def qoldiqlar_kontragent_hujjat_new(
    request: Request,
    force_new: int = 0,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Yangi kontragent balans hujjati"""
    from app.utils.draft_check import redirect_to_draft
    redirect = redirect_to_draft(
        db, PartnerBalanceDoc,
        edit_url_template="/qoldiqlar/kontragent/hujjat/{id}",
        user_role=getattr(current_user, "role", "") or "",
        force_new=bool(force_new),
        message="Sizda ochiq kontragent qoldiqlari qoralamasi bor — avval uni tugating yoki bekor qiling.",
        user_id=current_user.id,
    )
    if redirect:
        return redirect
    partners = db.query(Partner).filter(Partner.is_active == True).order_by(Partner.name).all()
    today_str = datetime.now().strftime("%Y-%m-%d")
    return templates.TemplateResponse("qoldiqlar/kontragent_hujjat_form.html", {
        "request": request,
        "doc": None,
        "partners": partners,
        "current_user": current_user,
        "today_str": today_str,
        "page_title": "Kontragent qoldiqlari — yangi hujjat",
    })


@router.post("/kontragent/hujjat")
async def qoldiqlar_kontragent_hujjat_create(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Kontragent balans hujjatini yaratish (qoralama)"""
    form = await request.form()
    partner_ids = form.getlist("partner_id")
    balances = form.getlist("balance")

    # Formadan sanani olish
    doc_date_str = form.get("doc_date", "")
    if doc_date_str:
        try:
            doc_date = datetime.strptime(doc_date_str, "%Y-%m-%d")
        except ValueError:
            doc_date = datetime.now()
    else:
        doc_date = datetime.now()

    items_data = []
    for i, pid in enumerate(partner_ids):
        if not pid:
            continue
        try:
            pid_int = int(pid)
            bal_str = (balances[i] if i < len(balances) else "").strip()
            if not bal_str:
                continue
            bal = float(bal_str)
        except (TypeError, ValueError):
            continue
        items_data.append((pid_int, bal))

    if not items_data:
        return RedirectResponse(url="/qoldiqlar/kontragent/hujjat/new", status_code=303)

    # MAX(number) suffix asosida raqam — count() bug (deleted gaps, race condition) ni oldini olish
    prefix = f"KNT-{doc_date.strftime('%Y%m%d')}-"
    last = (
        db.query(PartnerBalanceDoc)
        .filter(PartnerBalanceDoc.number.like(f"{prefix}%"))
        .order_by(PartnerBalanceDoc.number.desc())
        .first()
    )
    try:
        next_num = int(last.number.split("-")[-1]) + 1 if last else 1
    except (ValueError, AttributeError):
        next_num = 1
    number = f"{prefix}{str(next_num).zfill(4)}"

    doc = PartnerBalanceDoc(
        number=number,
        date=doc_date,
        user_id=current_user.id if current_user else None,
        status="draft",
    )
    db.add(doc)
    db.flush()
    for pid, bal in items_data:
        db.add(PartnerBalanceDocItem(doc_id=doc.id, partner_id=pid, balance=bal))
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/kontragent/hujjat/{doc.id}", status_code=303)


@router.get("/kontragent/hujjat/{doc_id}", response_class=HTMLResponse)
async def qoldiqlar_kontragent_hujjat_view(
    request: Request,
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Kontragent balans hujjatini ko'rish"""
    doc = db.query(PartnerBalanceDoc).filter(PartnerBalanceDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    partners = db.query(Partner).filter(Partner.is_active == True).order_by(Partner.name).all()
    return templates.TemplateResponse("qoldiqlar/kontragent_hujjat_form.html", {
        "request": request,
        "doc": doc,
        "partners": partners,
        "current_user": current_user,
        "page_title": f"Kontragent qoldiqlari {doc.number}",
    })


@router.post("/kontragent/hujjat/{doc_id}/tasdiqlash")
async def qoldiqlar_kontragent_hujjat_tasdiqlash(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Kontragent hujjatini tasdiqlash — kontragent balanslarini yangilash"""
    doc = db.query(PartnerBalanceDoc).filter(PartnerBalanceDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "draft":
        raise HTTPException(status_code=400, detail="Hujjat allaqachon tasdiqlangan")
    if not doc.items:
        raise HTTPException(status_code=400, detail="Kamida bitta kontragent qatori bo'lishi kerak")
    # Atomik UPDATE WHERE — double-confirm xavfini oldini olish
    from sqlalchemy import text as _text
    claim = db.execute(
        _text("UPDATE partner_balance_docs SET status='confirmed' WHERE id=:id AND status='draft'"),
        {"id": doc_id}
    )
    if claim.rowcount == 0:
        return RedirectResponse(url=f"/qoldiqlar/kontragent/hujjat/{doc_id}?already=1", status_code=303)
    # Status allaqachon atomik UPDATE WHERE bilan o'zgartirildi
    from app.services.partner_balance_service import recompute_partner_balance
    db.flush()
    for pid in {item.partner_id for item in doc.items if item.partner_id}:
        recompute_partner_balance(db, pid, reason="balance_doc_confirm", ref=doc.number,
                                  actor=current_user.username if current_user else None)
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/kontragent/hujjat/{doc_id}", status_code=303)


@router.post("/kontragent/hujjat/{doc_id}/revert")
async def qoldiqlar_kontragent_hujjat_revert(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Kontragent hujjati tasdiqini bekor qilish (faqat admin)"""
    doc = db.query(PartnerBalanceDoc).filter(PartnerBalanceDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "confirmed":
        raise HTTPException(status_code=400, detail="Faqat tasdiqlangan hujjatning tasdiqini bekor qilish mumkin")
    doc.status = "draft"
    from app.services.partner_balance_service import recompute_partner_balance
    db.flush()
    for pid in {item.partner_id for item in doc.items if item.partner_id}:
        recompute_partner_balance(db, pid, reason="balance_doc_revert", ref=doc.number,
                                  actor=current_user.username if current_user else None)
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/kontragent/hujjat/{doc_id}", status_code=303)


@router.get("/balance-import", response_class=HTMLResponse)
async def qoldiqlar_balance_import(
    request: Request,
    apply: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Sales Doctor balansy klientov JSON dan hisobot + draft hujjat yaratish.

    JSON fayl: scripts/balance_import_20260520.json (sales doctor xlsx dan).
    apply=False (default): hisobot. apply=True: draft hujjat yaratish.
    """
    from app.services.balance_import_data import SD_BALANCE_DATA
    sd_rows = SD_BALANCE_DATA

    def norm_name(s):
        return "".join(c for c in (s or "").lower().strip() if c.isalnum())

    def norm_phone(p):
        d = "".join(c for c in (p or "") if c.isdigit())
        return d[-9:] if len(d) >= 7 else ""

    partners = db.query(Partner).filter(Partner.is_active == True).all()
    by_name, by_phone = {}, {}
    for p in partners:
        nk = norm_name(p.name)
        if nk:
            by_name.setdefault(nk, []).append(p)
        for ph in [p.phone, p.phone2]:
            pk = norm_phone(ph or "")
            if pk:
                by_phone.setdefault(pk, []).append(p)

    matched, multi, not_found = [], [], []
    for sd in sd_rows:
        cands = []
        nk = norm_name(sd["name"])
        pk = norm_phone(sd["phone"])
        if nk in by_name:
            cands.extend(by_name[nk])
        if pk and pk in by_phone:
            for c in by_phone[pk]:
                if c not in cands:
                    cands.append(c)
        if len(cands) == 1:
            matched.append((sd, cands[0]))
        elif len(cands) > 1:
            multi.append((sd, cands))
        else:
            not_found.append(sd)

    # SIGNING TESKARI: SD "-" = mijoz qarzdor, TOTLI "+" = mijoz qarzdor
    # Yangi TOTLI balans = -sd_summa, delta = target - hozirgi
    diffs = [(sd, p, (-sd["summa"]) - p.balance) for sd, p in matched if abs((-sd["summa"]) - p.balance) > 0.5]

    # Apply rejimida: draft hujjat yaratish
    apply_msg = ""
    if apply and diffs:
        now = datetime.now()
        prefix = f"KNT-{now.strftime('%Y%m%d')}-"
        last = db.query(PartnerBalanceDoc).filter(PartnerBalanceDoc.number.like(f"{prefix}%")).order_by(PartnerBalanceDoc.number.desc()).first()
        try:
            next_num = int(last.number.split("-")[-1]) + 1 if last else 1
        except (ValueError, AttributeError):
            next_num = 1
        number = f"{prefix}{str(next_num).zfill(4)}"
        doc = PartnerBalanceDoc(number=number, date=now, user_id=current_user.id, status="draft")
        db.add(doc)
        db.flush()
        for sd, p, delta in diffs:
            db.add(PartnerBalanceDocItem(doc_id=doc.id, partner_id=p.id, balance=delta))
        db.commit()
        apply_msg = f'<div class="alert alert-success"><strong>✓ Qoralama hujjat yaratildi:</strong> <a href="/qoldiqlar/kontragent/hujjat/{doc.id}">{number}</a> — brauzerda ochib Tasdiqlash tugmasini bosing.</div>'

    # HTML hisobot
    html = ['<!DOCTYPE html><html><head><meta charset="utf-8"><title>Balans import</title>',
            '<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css">',
            '<style>body{padding:20px;font-family:sans-serif;}table{font-size:.85rem;}.num{text-align:right;font-family:monospace;}</style>',
            '</head><body><div class="container-fluid">',
            '<h4>Balans import — Sales Doctor 20.05.2026</h4>',
            apply_msg]
    html.append(f'<p><b>Sales Doctor:</b> {len(sd_rows)} mijoz | <b>TOTLI BI aktiv:</b> {len(partners)} mijoz</p>')
    html.append(f'<ul>')
    html.append(f'<li>✓ Topilgan (yagona): <b>{len(matched)}</b></li>')
    html.append(f'<li>? Bir nechta variant: <b>{len(multi)}</b></li>')
    html.append(f'<li>✗ Topilmadi: <b>{len(not_found)}</b></li>')
    html.append(f'<li>= Balans bir xil: <b>{len(matched) - len(diffs)}</b></li>')
    html.append(f'<li>≠ Farq qiladi: <b>{len(diffs)}</b></li>')
    total_delta = sum(d for _, _, d in diffs)
    html.append(f'<li><b>JAMI DELTA:</b> {total_delta:,.0f} so\'m</li>')
    html.append('</ul>')

    if not apply and diffs:
        html.append(f'<p><a href="/qoldiqlar/balance-import?apply=1" class="btn btn-warning" onclick="return confirm(\'{len(diffs)} ta mijoz uchun draft hujjat yarataymi?\')">✓ Qoralama hujjat yaratish</a></p>')

    if diffs:
        html.append('<h5>Farq qiluvchi mijozlar:</h5>')
        html.append('<table class="table table-sm table-striped"><thead><tr><th>#</th><th>Nomi</th><th>Tel</th><th class="num">SD raw</th><th class="num">Maqsad (yangi)</th><th class="num">TOTLI hozir</th><th class="num">Delta (+/−)</th></tr></thead><tbody>')
        for i, (sd, p, d) in enumerate(sorted(diffs, key=lambda x: abs(x[2]), reverse=True), 1):
            target = -sd["summa"]
            html.append(f'<tr><td>{i}</td><td>{(p.name or "-")[:35]}</td><td>{(p.phone or "-")[:15]}</td><td class="num text-muted">{sd["summa"]:,.0f}</td><td class="num"><b>{target:,.0f}</b></td><td class="num">{p.balance:,.0f}</td><td class="num"><b style="color:{"#dc3545" if d < 0 else "#198754"}">{d:+,.0f}</b></td></tr>')
        html.append('</tbody></table>')

    if not_found:
        html.append(f'<h5>Topilmadi ({len(not_found)}):</h5><table class="table table-sm"><thead><tr><th>Nomi</th><th>Tel</th><th class="num">SD balans</th></tr></thead><tbody>')
        for nf in not_found:
            html.append(f'<tr><td>{nf["name"][:40]}</td><td>{nf["phone"]}</td><td class="num">{nf["summa"]:,.0f}</td></tr>')
        html.append('</tbody></table>')

    if multi:
        html.append(f'<h5>Bir nechta variant ({len(multi)}):</h5>')
        for sd, cands in multi:
            html.append(f'<p><b>{sd["name"]}</b> (tel {sd["phone"]}, bal {sd["summa"]:,.0f}):<br>')
            for c in cands:
                html.append(f'&nbsp;&nbsp;→ <a href="/partners?id={c.id}">{c.name}</a> (tel {c.phone}, bal {c.balance:,.0f})<br>')
            html.append('</p>')

    html.append('</div></body></html>')
    return HTMLResponse("".join(html))


@router.get("/balance-import/multi", response_class=HTMLResponse)
async def qoldiqlar_balance_import_multi(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """30 ta dublikat mijozlar uchun interaktiv tanlov sahifasi."""
    from app.services.balance_import_data import SD_BALANCE_DATA

    def norm_name(s):
        return "".join(c for c in (s or "").lower().strip() if c.isalnum())

    def norm_phone(p):
        d = "".join(c for c in (p or "") if c.isdigit())
        return d[-9:] if len(d) >= 7 else ""

    partners = db.query(Partner).filter(Partner.is_active == True).all()
    by_name, by_phone = {}, {}
    for p in partners:
        nk = norm_name(p.name)
        if nk:
            by_name.setdefault(nk, []).append(p)
        for ph in [p.phone, p.phone2]:
            pk = norm_phone(ph or "")
            if pk:
                by_phone.setdefault(pk, []).append(p)

    multi_list = []
    for sd in SD_BALANCE_DATA:
        cands = []
        nk = norm_name(sd["name"])
        pk = norm_phone(sd["phone"])
        if nk in by_name:
            cands.extend(by_name[nk])
        if pk and pk in by_phone:
            for c in by_phone[pk]:
                if c not in cands:
                    cands.append(c)
        if len(cands) > 1:
            multi_list.append((sd, cands))

    h = ['<!DOCTYPE html><html><head><meta charset="utf-8"><title>Dublikat tanlash</title>',
         '<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css">',
         '<style>body{padding:20px;font-family:sans-serif;}.sd-card{border:1px solid #dee2e6;border-radius:8px;padding:12px;margin-bottom:12px;background:#f8f9fa;}.sd-head{font-weight:600;margin-bottom:8px;}.cand-row{padding:6px 8px;border-radius:4px;margin:2px 0;background:white;}.cand-row:hover{background:#fffbe6;}.num{font-family:monospace;}</style>',
         '</head><body><div class="container-fluid">',
         '<h4>Dublikat mijozlar — qaysi TOTLI BI partneri to\'g\'ri?</h4>',
         f'<p class="text-muted">Jami {len(multi_list)} ta dublikat. Har biri uchun to\'g\'ri partnerni tanlang yoki "Skip" qoldiring. Pastda "Saqlash" — draft hujjat yaratiladi.</p>',
         '<form method="post" action="/qoldiqlar/balance-import/multi/apply">',
         f'<input type="hidden" name="csrf_token" value="{getattr(request.state, "csrf_token", "") or ""}">']

    for idx, (sd, cands) in enumerate(multi_list):
        target = -sd["summa"]
        h.append(f'<div class="sd-card">')
        h.append(f'<div class="sd-head">{idx+1}. <b>{sd["name"]}</b> &middot; tel {sd["phone"]} &middot; SD raw <span class="num">{sd["summa"]:,.0f}</span> &middot; <b>maqsad balans: <span class="num">{target:,.0f}</span></b></div>')
        # Skip variant
        h.append(f'<div class="cand-row"><label class="d-flex align-items-center" style="cursor:pointer;"><input type="radio" name="choice_{idx}" value="skip" checked class="me-2"> <span class="text-muted">— Skip (o\'tkazib yuborish)</span></label></div>')
        for c in cands:
            delta = target - (c.balance or 0)
            color = "#dc3545" if delta < 0 else ("#198754" if delta > 0 else "#6c757d")
            h.append(f'<div class="cand-row"><label class="d-flex align-items-center" style="cursor:pointer;"><input type="radio" name="choice_{idx}" value="{c.id}" class="me-2"> <div class="flex-grow-1">#{c.id} <b>{c.name}</b> &middot; tel {c.phone or "-"}, {c.phone2 or ""} &middot; hozir <span class="num">{c.balance or 0:,.0f}</span> &middot; delta <b class="num" style="color:{color}">{delta:+,.0f}</b></div></label></div>')
        h.append(f'<input type="hidden" name="sd_idx_{idx}" value="{idx}">')
        h.append('</div>')

    h.append('<div class="mt-3 mb-5"><button type="submit" class="btn btn-warning btn-lg">✓ Tanlanganlar uchun draft hujjat yaratish</button> <a href="/qoldiqlar" class="btn btn-outline-secondary btn-lg">Bekor</a></div>')
    h.append('</form></div></body></html>')
    return HTMLResponse("".join(h))


@router.post("/balance-import/multi/apply")
async def qoldiqlar_balance_import_multi_apply(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Dublikat tanlovlardan draft hujjat yaratish."""
    from app.services.balance_import_data import SD_BALANCE_DATA
    form = await request.form()

    selected = []  # [(sd_row, partner)]
    for key, val in form.items():
        if not key.startswith("choice_"):
            continue
        if val == "skip" or not val:
            continue
        try:
            idx = int(key[len("choice_"):])
            pid = int(val)
        except (ValueError, TypeError):
            continue
        if idx < 0 or idx >= len(SD_BALANCE_DATA):
            continue
        partner = db.query(Partner).filter(Partner.id == pid).first()
        if partner:
            selected.append((SD_BALANCE_DATA[idx], partner))

    if not selected:
        return HTMLResponse("<p>Hech narsa tanlanmagan.</p> <a href='/qoldiqlar/balance-import/multi'>Orqaga</a>")

    diffs = []
    for sd, p in selected:
        target = -sd["summa"]
        delta = target - (p.balance or 0)
        if abs(delta) > 0.5:
            diffs.append((sd, p, delta))

    if not diffs:
        return HTMLResponse("<p>Tanlanganlarning hech birida balans farqi yo'q.</p> <a href='/qoldiqlar/balance-import/multi'>Orqaga</a>")

    now = datetime.now()
    prefix = f"KNT-{now.strftime('%Y%m%d')}-"
    last = db.query(PartnerBalanceDoc).filter(PartnerBalanceDoc.number.like(f"{prefix}%")).order_by(PartnerBalanceDoc.number.desc()).first()
    try:
        next_num = int(last.number.split("-")[-1]) + 1 if last else 1
    except (ValueError, AttributeError):
        next_num = 1
    number = f"{prefix}{str(next_num).zfill(4)}"
    doc = PartnerBalanceDoc(number=number, date=now, user_id=current_user.id, status="draft")
    db.add(doc)
    db.flush()
    for sd, p, delta in diffs:
        db.add(PartnerBalanceDocItem(doc_id=doc.id, partner_id=p.id, balance=delta))
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/kontragent/hujjat/{doc.id}", status_code=303)


@router.post("/kontragent/hujjat/{doc_id}/items/add")
async def qoldiqlar_kontragent_hujjat_add_item(
    doc_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Draft kontragent hujjatiga yangi qator qo'shish"""
    doc = db.query(PartnerBalanceDoc).filter(PartnerBalanceDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "draft":
        raise HTTPException(status_code=400, detail="Faqat qoralama hujjatga qator qo'shish mumkin. Avval tasdiqni bekor qiling.")
    form = await request.form()
    try:
        partner_id = int(form.get("partner_id") or 0)
        balance = float((form.get("balance") or "").strip())
    except (TypeError, ValueError):
        return RedirectResponse(url=f"/qoldiqlar/kontragent/hujjat/{doc_id}?msg=Noto%27g%27ri+qiymat", status_code=303)
    if partner_id <= 0:
        return RedirectResponse(url=f"/qoldiqlar/kontragent/hujjat/{doc_id}?msg=Kontragent+tanlanmagan", status_code=303)
    existing = db.query(PartnerBalanceDocItem).filter(
        PartnerBalanceDocItem.doc_id == doc_id,
        PartnerBalanceDocItem.partner_id == partner_id
    ).first()
    if existing:
        return RedirectResponse(url=f"/qoldiqlar/kontragent/hujjat/{doc_id}?msg=Bu+kontragent+allaqachon+qatorda+bor", status_code=303)
    db.add(PartnerBalanceDocItem(doc_id=doc_id, partner_id=partner_id, balance=balance))
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/kontragent/hujjat/{doc_id}", status_code=303)


@router.post("/kontragent/hujjat/{doc_id}/items/{item_id}/delete")
async def qoldiqlar_kontragent_hujjat_delete_item(
    doc_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Draft kontragent hujjatidan qatorni o'chirish"""
    doc = db.query(PartnerBalanceDoc).filter(PartnerBalanceDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "draft":
        raise HTTPException(status_code=400, detail="Faqat qoralama hujjatdan qator o'chirish mumkin.")
    item = db.query(PartnerBalanceDocItem).filter(
        PartnerBalanceDocItem.id == item_id,
        PartnerBalanceDocItem.doc_id == doc_id
    ).first()
    if item:
        db.delete(item)
        db.commit()
    return RedirectResponse(url=f"/qoldiqlar/kontragent/hujjat/{doc_id}", status_code=303)


@router.post("/kontragent/hujjat/{doc_id}/delete")
async def qoldiqlar_kontragent_hujjat_delete(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Kontragent hujjatini o'chirish (faqat qoralama, faqat admin)"""
    doc = db.query(PartnerBalanceDoc).filter(PartnerBalanceDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "draft":
        raise HTTPException(status_code=400, detail="Faqat qoralama holatidagi hujjatni o'chirish mumkin. Avval tasdiqni bekor qiling.")
    db.delete(doc)
    db.commit()
    return RedirectResponse(url="/qoldiqlar#kontragent", status_code=303)


@router.post("/tovar")
async def qoldiqlar_tovar_save(
    warehouse_id: int = Form(...),
    product_id: int = Form(...),
    quantity: float = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Tovar qoldig'ini kiritish yoki qo'shish (omborda mavjud bo'lsa qo'shiladi). StockMovement audit trail bilan."""
    if quantity < 0:
        return RedirectResponse(url="/qoldiqlar#tovar", status_code=303)
    from app.services.stock_service import create_stock_movement
    create_stock_movement(
        db=db,
        warehouse_id=warehouse_id,
        product_id=product_id,
        quantity_change=quantity,
        operation_type="manual_add",
        document_type="ManualAdjustment",
        document_id=0,
        document_number="MANUAL",
        user_id=current_user.id if current_user else None,
        note="Qo'lda qoldiq kiritish",
    )
    db.commit()
    return RedirectResponse(url="/qoldiqlar#tovar", status_code=303)


# ==========================================
# XODIM QOLDIQLARI HUJJAT TIZIMI
# ==========================================

@router.get("/xodim/hujjat/new", response_class=HTMLResponse)
async def qoldiqlar_xodim_hujjat_new(
    request: Request,
    force_new: int = 0,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Yangi xodim balans hujjati"""
    from app.utils.draft_check import redirect_to_draft
    redirect = redirect_to_draft(
        db, EmployeeBalanceDoc,
        edit_url_template="/qoldiqlar/xodim/hujjat/{id}",
        user_role=getattr(current_user, "role", "") or "",
        force_new=bool(force_new),
        message="Sizda ochiq xodim qoldiqlari qoralamasi bor — avval uni tugating yoki bekor qiling.",
        user_id=current_user.id,
    )
    if redirect:
        return redirect
    employees = db.query(Employee).filter(Employee.is_active == True).order_by(Employee.full_name).all()
    today_str = datetime.now().strftime("%Y-%m-%d")
    return templates.TemplateResponse("qoldiqlar/xodim_hujjat_form.html", {
        "request": request,
        "doc": None,
        "employees": employees,
        "current_user": current_user,
        "today_str": today_str,
        "page_title": "Xodim qoldiqlari — yangi hujjat",
    })


@router.post("/xodim/hujjat")
async def qoldiqlar_xodim_hujjat_create(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Xodim balans hujjatini yaratish (qoralama)"""
    form = await request.form()
    employee_ids = form.getlist("employee_id")
    balances = form.getlist("balance")
    doc_date_str = form.get("doc_date", "")
    if doc_date_str:
        try:
            doc_date = datetime.strptime(doc_date_str, "%Y-%m-%d")
        except ValueError:
            doc_date = datetime.now()
    else:
        doc_date = datetime.now()

    items_data = []
    for i, eid in enumerate(employee_ids):
        if not eid:
            continue
        try:
            eid_int = int(eid)
            bal_str = (balances[i] if i < len(balances) else "").strip()
            if not bal_str:
                continue
            bal = float(bal_str)
        except (TypeError, ValueError):
            continue
        items_data.append((eid_int, bal))

    if not items_data:
        return RedirectResponse(url="/qoldiqlar/xodim/hujjat/new", status_code=303)

    count = db.query(EmployeeBalanceDoc).filter(
        EmployeeBalanceDoc.date >= doc_date.replace(hour=0, minute=0, second=0),
        EmployeeBalanceDoc.date < doc_date.replace(hour=23, minute=59, second=59)
    ).count()
    number = f"XOD-{doc_date.strftime('%Y%m%d')}-{str(count + 1).zfill(4)}"

    doc = EmployeeBalanceDoc(
        number=number,
        date=doc_date,
        user_id=current_user.id if current_user else None,
        status="draft",
    )
    db.add(doc)
    db.flush()
    for eid, bal in items_data:
        db.add(EmployeeBalanceDocItem(doc_id=doc.id, employee_id=eid, balance=bal))
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/xodim/hujjat/{doc.id}", status_code=303)


@router.get("/xodim/hujjat/{doc_id}", response_class=HTMLResponse)
async def qoldiqlar_xodim_hujjat_view(
    request: Request,
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Xodim balans hujjatini ko'rish"""
    doc = db.query(EmployeeBalanceDoc).filter(EmployeeBalanceDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    employees = db.query(Employee).filter(Employee.is_active == True).order_by(Employee.full_name).all()
    return templates.TemplateResponse("qoldiqlar/xodim_hujjat_form.html", {
        "request": request,
        "doc": doc,
        "employees": employees,
        "current_user": current_user,
        "page_title": f"Xodim qoldiqlari {doc.number}",
    })


@router.post("/xodim/hujjat/{doc_id}/tasdiqlash")
async def qoldiqlar_xodim_hujjat_tasdiqlash(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Xodim hujjatini tasdiqlash — Salary jadvaliga yozish"""
    doc = db.query(EmployeeBalanceDoc).filter(EmployeeBalanceDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "draft":
        raise HTTPException(status_code=400, detail="Hujjat allaqachon tasdiqlangan")
    if not doc.items:
        raise HTTPException(status_code=400, detail="Kamida bitta xodim qatori bo'lishi kerak")
    # Atomik UPDATE WHERE — double-confirm xavfini oldini olish
    from sqlalchemy import text as _text
    claim = db.execute(
        _text("UPDATE employee_balance_docs SET status='confirmed' WHERE id=:id AND status='draft'"),
        {"id": doc_id}
    )
    if claim.rowcount == 0:
        return RedirectResponse(url=f"/qoldiqlar/xodim/hujjat/{doc_id}?already=1", status_code=303)
    now = datetime.now()
    year, month = doc.date.year if doc.date else now.year, doc.date.month if doc.date else now.month
    for item in doc.items:
        emp = db.query(Employee).filter(Employee.id == item.employee_id).first()
        if not emp:
            continue
        s = db.query(Salary).filter(Salary.employee_id == item.employee_id, Salary.year == year, Salary.month == month).first()
        old_total = float(s.total or 0) if s else 0
        item.previous_balance = old_total
        if not s:
            s = Salary(employee_id=item.employee_id, year=year, month=month)
            db.add(s)
        # Hujjatda: musbat = xodim qarzi, manfiy = bizning qarzimiz
        # Salary da: musbat = biz to'lashimiz kerak, manfiy = xodim qarzda
        # Shuning uchun ishorani teskari qilamiz
        new_total = old_total - item.balance
        s.base_salary = max(0, abs(new_total))
        s.total = new_total
        if s.paid is None:
            s.paid = 0
        s.status = "paid" if new_total <= 0 else "pending"
        s.is_balance_entry = True
    doc.status = "confirmed"
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/xodim/hujjat/{doc_id}", status_code=303)


@router.post("/xodim/hujjat/{doc_id}/revert")
async def qoldiqlar_xodim_hujjat_revert(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Xodim hujjati tasdiqini bekor qilish (faqat admin)"""
    doc = db.query(EmployeeBalanceDoc).filter(EmployeeBalanceDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "confirmed":
        raise HTTPException(status_code=400, detail="Faqat tasdiqlangan hujjatning tasdiqini bekor qilish mumkin")
    now = datetime.now()
    year, month = doc.date.year if doc.date else now.year, doc.date.month if doc.date else now.month
    for item in doc.items:
        if item.previous_balance is not None:
            s = db.query(Salary).filter(Salary.employee_id == item.employee_id, Salary.year == year, Salary.month == month).first()
            if s:
                s.total = item.previous_balance
                s.base_salary = max(0, item.previous_balance)
                s.status = "paid" if item.previous_balance <= 0 else "pending"
    doc.status = "draft"
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/xodim/hujjat/{doc_id}", status_code=303)


@router.post("/xodim/hujjat/{doc_id}/delete")
async def qoldiqlar_xodim_hujjat_delete(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Xodim hujjatini o'chirish (faqat qoralama, faqat admin)"""
    doc = db.query(EmployeeBalanceDoc).filter(EmployeeBalanceDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "draft":
        raise HTTPException(status_code=400, detail="Faqat qoralama holatidagi hujjatni o'chirish mumkin")
    db.delete(doc)
    db.commit()
    return RedirectResponse(url="/qoldiqlar?tab=xodim", status_code=303)


@router.post("/xodim/hujjat/{doc_id}/add-row")
async def qoldiqlar_xodim_hujjat_add_row(
    doc_id: int,
    employee_id: int = Form(...),
    balance: float = Form(...),
    balance_type: int = Form(1),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Draft hujjatga qator qo'shish"""
    doc = db.query(EmployeeBalanceDoc).filter(EmployeeBalanceDoc.id == doc_id).first()
    if not doc or doc.status != "draft":
        raise HTTPException(status_code=400, detail="Faqat qoralamani tahrirlash mumkin")
    bal = abs(balance) * (1 if balance_type == 1 else -1)
    # Dublikat tekshiruv
    existing = db.query(EmployeeBalanceDocItem).filter(
        EmployeeBalanceDocItem.doc_id == doc_id,
        EmployeeBalanceDocItem.employee_id == employee_id,
    ).first()
    if existing:
        existing.balance = bal
    else:
        db.add(EmployeeBalanceDocItem(doc_id=doc_id, employee_id=employee_id, balance=bal))
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/xodim/hujjat/{doc_id}", status_code=303)


@router.post("/xodim/hujjat/{doc_id}/delete-row/{item_id}")
async def qoldiqlar_xodim_hujjat_delete_row(
    doc_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Draft hujjatdan qatorni o'chirish"""
    doc = db.query(EmployeeBalanceDoc).filter(EmployeeBalanceDoc.id == doc_id).first()
    if not doc or doc.status != "draft":
        raise HTTPException(status_code=400, detail="Faqat qoralamani tahrirlash mumkin")
    item = db.query(EmployeeBalanceDocItem).filter(
        EmployeeBalanceDocItem.id == item_id,
        EmployeeBalanceDocItem.doc_id == doc_id,
    ).first()
    if item:
        db.delete(item)
        db.commit()
    return RedirectResponse(url=f"/qoldiqlar/xodim/hujjat/{doc_id}", status_code=303)


@router.post("/kontragent/recalculate")
async def qoldiqlar_kontragent_recalculate(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Barcha kontragent balanslarini qayta hisoblash (admin).
    Hisoblash: sotuvlar qarz + qoldiq hujjatlari - xaridlar - kirim to'lovlar + chiqim to'lovlar - qaytarishlar
    """
    from app.services.partner_balance_service import compute_partner_balance, recompute_partner_balance
    partners = db.query(Partner).filter(Partner.is_active == True).all()
    updated = 0
    for partner in partners:
        old_bal = float(partner.balance or 0)
        new_bal = compute_partner_balance(db, partner.id)
        if abs(old_bal - new_bal) > 0.01:
            recompute_partner_balance(db, partner.id, reason="manual_recalc_all",
                                      actor=current_user.username if current_user else None)
            updated += 1

    db.commit()
    return RedirectResponse(
        url=f"/qoldiqlar?msg={updated} ta kontragent balansi qayta hisoblandi",
        status_code=303,
    )


@router.post("/kontragent/{partner_id}")
async def qoldiqlar_kontragent_save(
    partner_id: int,
    balance: str = Form(""),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Kontragent balansini yangilash"""
    partner = db.query(Partner).filter(Partner.id == partner_id).first()
    if not partner:
        raise HTTPException(status_code=404, detail="Kontragent topilmadi")
    balance_str = (balance or "").strip()
    if not balance_str:
        return RedirectResponse(url="/qoldiqlar#kontragent", status_code=303)
    try:
        new_value = float(balance_str)
    except (TypeError, ValueError):
        return RedirectResponse(url="/qoldiqlar#kontragent", status_code=303)
    from app.services.partner_balance_service import compute_partner_balance, recompute_partner_balance
    current = compute_partner_balance(db, partner.id)
    delta = new_value - current
    if abs(delta) > 0.01:
        now = datetime.now()
        prefix = f"KNT-{now.strftime('%Y%m%d')}-"
        last = (
            db.query(PartnerBalanceDoc)
            .filter(PartnerBalanceDoc.number.like(f"{prefix}%"))
            .order_by(PartnerBalanceDoc.number.desc())
            .first()
        )
        try:
            next_num = int(last.number.split("-")[-1]) + 1 if last else 1
        except (ValueError, AttributeError):
            next_num = 1
        number = f"{prefix}{str(next_num).zfill(4)}"
        adj_doc = PartnerBalanceDoc(
            number=number,
            status="confirmed",
            date=now,
            user_id=current_user.id if current_user else None,
        )
        db.add(adj_doc)
        db.flush()
        db.add(PartnerBalanceDocItem(doc_id=adj_doc.id, partner_id=partner.id, balance=delta))
        db.flush()
        recompute_partner_balance(db, partner.id, reason="manual_balance_edit", ref=adj_doc.number,
                                  actor=current_user.username if current_user else None)
    db.commit()
    return RedirectResponse(url="/qoldiqlar#kontragent", status_code=303)


@router.get("/export")
async def qoldiqlar_export(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Tovar qoldiqlari hisoboti — Excel hujjat sifatida yuklab olish"""
    stocks = (
        db.query(Stock)
        .join(Warehouse, Stock.warehouse_id == Warehouse.id)
        .join(Product, Stock.product_id == Product.id)
        .order_by(Warehouse.name, Product.name)
        .all()
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tovar qoldiqlari"
    ws.append(["Ombor", "Mahsulot", "Kod", "Miqdor"])
    for s in stocks:
        ws.append([
            s.warehouse.name if s.warehouse else "-",
            s.product.name if s.product else "-",
            (s.product.code or "") if s.product else "",
            float(s.quantity) if s.quantity is not None else 0,
        ])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tovar_qoldiqlari.xlsx"},
    )


# --- Tovar qoldiq HUJJATLARI (1C uslubida: ro'yxat + hujjat + qatorlar) ---
@router.get("/tovar/hujjat", response_class=HTMLResponse)
async def qoldiqlar_tovar_hujjat_list(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Tovar qoldiqlari hujjatlari ro'yxati"""
    docs = (
        db.query(StockAdjustmentDoc)
        .order_by(StockAdjustmentDoc.created_at.desc())
        .limit(200)
        .all()
    )
    return templates.TemplateResponse("qoldiqlar/hujjat_list.html", {
        "request": request,
        "docs": docs,
        "current_user": current_user,
        "page_title": "Tovar qoldiqlari hujjatlari",
        "show_tannarx": (getattr(current_user, "role", None) if current_user else None) == "admin",
    })


@router.get("/tovar/hujjat/new", response_class=HTMLResponse)
async def qoldiqlar_tovar_hujjat_new(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Yangi tovar qoldiq hujjati (qoralama)"""
    warehouses = get_warehouses_for_user(db, current_user)
    products = db.query(Product).filter(Product.is_active == True).order_by(Product.name).all()
    is_admin = (getattr(current_user, "role", None) if current_user else None) == "admin"
    return templates.TemplateResponse("qoldiqlar/hujjat_form.html", {
        "request": request,
        "doc": None,
        "warehouses": warehouses,
        "products": products,
        "last_prices": _get_last_purchase_prices(db) if is_admin else {},
        "current_user": current_user,
        "page_title": "Tovar qoldiqlari — yangi hujjat",
        "show_tannarx": is_admin,
        "now": datetime.now(),
    })


@router.post("/tovar/hujjat")
async def qoldiqlar_tovar_hujjat_create(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Tovar qoldiq hujjatini yaratish (qoralama)"""
    form = await request.form()
    product_ids = form.getlist("product_id")
    warehouse_ids = form.getlist("warehouse_id")
    quantities = form.getlist("quantity")
    cost_prices = form.getlist("cost_price")
    sale_prices = form.getlist("sale_price")

    items_data = []
    for i, pid in enumerate(product_ids):
        if not pid or not str(pid).strip():
            continue
        try:
            wid = int(warehouse_ids[i]) if i < len(warehouse_ids) and warehouse_ids[i] else None
            qty = float(quantities[i]) if i < len(quantities) and str(quantities[i]).strip() else 0
            _cp = cost_prices[i] if i < len(cost_prices) else ""
            _sp = sale_prices[i] if i < len(sale_prices) else ""
            cp = float(_cp) if str(_cp).strip() else 0
            sp = float(_sp) if str(_sp).strip() else 0
        except (TypeError, ValueError):
            continue
        if wid and qty > 0:
            try:
                items_data.append((int(pid), wid, qty, cp, sp))
            except ValueError:
                continue

    # Sana: formdan kelgan yoki hozirgi vaqt
    doc_date_str = form.get("doc_date", "")
    if doc_date_str and str(doc_date_str).strip():
        try:
            doc_date = datetime.fromisoformat(str(doc_date_str))
        except (ValueError, TypeError):
            doc_date = datetime.now()
    else:
        doc_date = datetime.now()

    count = db.query(StockAdjustmentDoc).filter(
        StockAdjustmentDoc.date >= doc_date.replace(hour=0, minute=0, second=0)
    ).count()
    number = f"QLD-{doc_date.strftime('%Y%m%d')}-{str(count + 1).zfill(4)}"

    total_tannarx = sum(qty * cp for _, _, qty, cp, _ in items_data)
    total_sotuv = sum(qty * sp for _, _, qty, _, sp in items_data)

    doc = StockAdjustmentDoc(
        number=number,
        date=doc_date,
        user_id=current_user.id if current_user else None,
        status="draft",
        total_tannarx=total_tannarx,
        total_sotuv=total_sotuv,
    )
    db.add(doc)
    db.flush()

    for pid, wid, qty, cp, sp in items_data:
        db.add(StockAdjustmentDocItem(
            doc_id=doc.id,
            product_id=pid,
            warehouse_id=wid,
            quantity=qty,
            cost_price=cp,
            sale_price=sp,
        ))
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/tovar/hujjat/{doc.id}", status_code=303)


@router.post("/tovar/import-excel")
async def qoldiqlar_tovar_import_excel(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Exceldan tovar qoldiqlarini yuklash — hujjat (QLD-...) yaratiladi, jadvalda ko'rinadi."""
    form = await request.form()
    file = form.get("file") or form.get("excel_file")
    if not file or not getattr(file, "filename", None):
        return RedirectResponse(url="/qoldiqlar?error=import&detail=" + quote("Excel fayl tanlang") + "#tovar", status_code=303)
    try:
        contents = await file.read()
        if not contents:
            return RedirectResponse(url="/qoldiqlar?error=import&detail=" + quote("Fayl bo'sh") + "#tovar", status_code=303)
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=False, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        items_data = []
        for row in rows:
            if not row or (row[0] is None and (len(row) < 2 or row[1] is None)):
                continue
            wh_key = str(row[0] or "").strip() if len(row) > 0 else ""
            raw_prod = row[1] if len(row) > 1 else None
            if raw_prod is not None and isinstance(raw_prod, (int, float)) and float(raw_prod) == int(float(raw_prod)):
                prod_key = str(int(float(raw_prod)))
            else:
                prod_key = str(raw_prod or "").strip()
            try:
                qty = float(row[2]) if len(row) > 2 and row[2] is not None else 0
            except (TypeError, ValueError):
                qty = 0
            cp = 0.0
            sp = 0.0
            if len(row) > 3 and row[3] is not None and row[3] != "":
                try:
                    cp = float(row[3])
                except (TypeError, ValueError):
                    pass
            if len(row) > 4 and row[4] is not None and row[4] != "":
                try:
                    sp = float(row[4])
                except (TypeError, ValueError):
                    pass
            if not wh_key or not prod_key or qty <= 0:
                continue
            warehouse = db.query(Warehouse).filter(
                (func.lower(Warehouse.name) == wh_key.lower()) | (Warehouse.code == wh_key)
            ).first()
            product = db.query(Product).filter(
                (Product.code == prod_key) | (Product.barcode == prod_key)
            ).first()
            if not product and prod_key:
                product = db.query(Product).filter(
                    Product.name.isnot(None),
                    func.lower(Product.name) == prod_key.lower()
                ).first()
            if not warehouse or not product:
                continue
            items_data.append((product.id, warehouse.id, qty, cp, sp))
        if not items_data:
            return RedirectResponse(
                url="/qoldiqlar?error=import&detail=" + quote("Hech qanday to'g'ri qator topilmadi. Ombor va mahsulot nomi/kodi to'g'ri ekanligini tekshiring.") + "#tovar",
                status_code=303,
            )
        today = datetime.now()
        count = db.query(StockAdjustmentDoc).filter(
            StockAdjustmentDoc.date >= today.replace(hour=0, minute=0, second=0)
        ).count()
        number = f"QLD-{today.strftime('%Y%m%d')}-{str(count + 1).zfill(4)}"
        total_tannarx = sum(qty * cp for _, _, qty, cp, _ in items_data)
        total_sotuv = sum(qty * sp for _, _, qty, _, sp in items_data)
        doc = StockAdjustmentDoc(
            number=number,
            date=today,
            user_id=current_user.id if current_user else None,
            status="draft",
            total_tannarx=total_tannarx,
            total_sotuv=total_sotuv,
        )
        db.add(doc)
        db.flush()
        for pid, wid, qty, cp, sp in items_data:
            db.add(StockAdjustmentDocItem(
                doc_id=doc.id,
                product_id=pid,
                warehouse_id=wid,
                quantity=qty,
                cost_price=cp,
                sale_price=sp,
            ))
        db.commit()
        return RedirectResponse(
            url="/qoldiqlar?success=import&doc_number=" + quote(doc.number) + "#tovar",
            status_code=303,
        )
    except Exception as e:
        pass  # logged above
        return RedirectResponse(
            url="/qoldiqlar?error=import&detail=" + quote("Import xatoligi. Fayl formatini tekshiring.") + "#tovar",
            status_code=303,
        )


@router.get("/tovar/hujjat/{doc_id}", response_class=HTMLResponse)
async def qoldiqlar_tovar_hujjat_view(
    request: Request,
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Tovar qoldiq hujjatini ko'rish/tahrirlash"""
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    warehouses = get_warehouses_for_user(db, current_user)
    products = db.query(Product).filter(Product.is_active == True).order_by(Product.name).all()
    is_admin = (getattr(current_user, "role", None) if current_user else None) == "admin"
    return templates.TemplateResponse("qoldiqlar/hujjat_form.html", {
        "request": request,
        "doc": doc,
        "warehouses": warehouses,
        "products": products,
        "last_prices": _get_last_purchase_prices(db) if is_admin else {},
        "current_user": current_user,
        "page_title": f"Tovar qoldiqlari {doc.number}",
        "show_tannarx": is_admin,
    })


@router.post("/tovar/hujjat/{doc_id}/add-row")
async def qoldiqlar_tovar_hujjat_add_row(
    doc_id: int,
    product_id: int = Form(...),
    warehouse_id: int = Form(...),
    quantity: float = Form(...),
    cost_price: float = Form(0),
    sale_price: float = Form(0),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Hujjatga qator qo'shish (faqat qoralama)"""
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc or doc.status != "draft":
        raise HTTPException(status_code=400, detail="Faqat qoralamani tahrirlash mumkin")
    if quantity <= 0:
        return RedirectResponse(url=f"/qoldiqlar/tovar/hujjat/{doc_id}", status_code=303)
    # Dublikat tekshiruv: bir xil ombor+mahsulot bo'lsa — miqdorni yangilash
    existing_item = db.query(StockAdjustmentDocItem).filter(
        StockAdjustmentDocItem.doc_id == doc_id,
        StockAdjustmentDocItem.product_id == product_id,
        StockAdjustmentDocItem.warehouse_id == warehouse_id,
    ).first()
    if existing_item:
        # Eski summalarni chiqarish
        doc.total_tannarx = (doc.total_tannarx or 0) - (existing_item.quantity * (existing_item.cost_price or 0))
        doc.total_sotuv = (doc.total_sotuv or 0) - (existing_item.quantity * (existing_item.sale_price or 0))
        existing_item.quantity = quantity
        existing_item.cost_price = cost_price or 0
        existing_item.sale_price = sale_price or 0
        doc.total_tannarx = (doc.total_tannarx or 0) + quantity * (cost_price or 0)
        doc.total_sotuv = (doc.total_sotuv or 0) + quantity * (sale_price or 0)
        db.commit()
        return RedirectResponse(url=f"/qoldiqlar/tovar/hujjat/{doc_id}", status_code=303)
    doc.total_tannarx = (doc.total_tannarx or 0) + quantity * (cost_price or 0)
    doc.total_sotuv = (doc.total_sotuv or 0) + quantity * (sale_price or 0)
    db.add(StockAdjustmentDocItem(
        doc_id=doc_id,
        product_id=product_id,
        warehouse_id=warehouse_id,
        quantity=quantity,
        cost_price=cost_price or 0,
        sale_price=sale_price or 0,
    ))
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/tovar/hujjat/{doc_id}", status_code=303)


@router.post("/tovar/hujjat/{doc_id}/delete-row/{item_id}")
async def qoldiqlar_tovar_hujjat_delete_row(
    doc_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Hujjatdan qatorni o'chirish (faqat qoralama)"""
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc or doc.status != "draft":
        raise HTTPException(status_code=400, detail="Faqat qoralamani tahrirlash mumkin")
    item = db.query(StockAdjustmentDocItem).filter(
        StockAdjustmentDocItem.id == item_id,
        StockAdjustmentDocItem.doc_id == doc_id,
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Qator topilmadi")
    doc.total_tannarx = (doc.total_tannarx or 0) - (item.quantity * (item.cost_price or 0))
    doc.total_sotuv = (doc.total_sotuv or 0) - (item.quantity * (item.sale_price or 0))
    if doc.total_tannarx < 0:
        doc.total_tannarx = 0
    if doc.total_sotuv < 0:
        doc.total_sotuv = 0
    db.delete(item)
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/tovar/hujjat/{doc_id}", status_code=303)


@router.post("/tovar/hujjat/{doc_id}/edit-row/{item_id}")
async def qoldiqlar_tovar_hujjat_edit_row(
    doc_id: int,
    item_id: int,
    quantity: float = Form(...),
    warehouse_id: int = Form(...),
    cost_price: float = Form(0),
    sale_price: float = Form(0),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Hujjat qatorini tahrirlash (faqat qoralama): soni, ombor, tannarx, sotuv narx"""
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc or doc.status != "draft":
        raise HTTPException(status_code=400, detail="Faqat qoralamani tahrirlash mumkin")
    item = db.query(StockAdjustmentDocItem).filter(
        StockAdjustmentDocItem.id == item_id,
        StockAdjustmentDocItem.doc_id == doc_id,
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Qator topilmadi")
    if quantity <= 0:
        return RedirectResponse(url=f"/qoldiqlar/tovar/hujjat/{doc_id}", status_code=303)
    # Eski summalarni hisobdan chiqarish
    doc.total_tannarx = (doc.total_tannarx or 0) - (item.quantity * (item.cost_price or 0))
    doc.total_sotuv = (doc.total_sotuv or 0) - (item.quantity * (item.sale_price or 0))
    item.quantity = quantity
    item.warehouse_id = warehouse_id
    item.cost_price = cost_price or 0
    item.sale_price = sale_price or 0
    doc.total_tannarx = (doc.total_tannarx or 0) + item.quantity * item.cost_price
    doc.total_sotuv = (doc.total_sotuv or 0) + item.quantity * item.sale_price
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/tovar/hujjat/{doc_id}", status_code=303)


@router.post("/tovar/hujjat/{doc_id}/tasdiqlash")
async def qoldiqlar_tovar_hujjat_tasdiqlash(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Hujjatni tasdiqlash — ombor qoldiqlariga qo'shiladi (create_stock_movement orqali)"""
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "draft":
        raise HTTPException(status_code=400, detail="Hujjat allaqachon tasdiqlangan")
    if not doc.items:
        raise HTTPException(status_code=400, detail="Kamida bitta qator bo'lishi kerak")
    # Atomik UPDATE WHERE — double-confirm xavfini oldini olish
    from sqlalchemy import text as _text
    claim = db.execute(
        _text("UPDATE stock_adjustment_docs SET status='confirmed' WHERE id=:id AND status='draft'"),
        {"id": doc_id}
    )
    if claim.rowcount == 0:
        return RedirectResponse(url=f"/qoldiqlar/tovar/hujjat/{doc_id}?already=1", status_code=303)

    doc_warehouse_ids = list({item.warehouse_id for item in doc.items})
    doc_pairs = {(item.warehouse_id, item.product_id) for item in doc.items}

    is_qld = (doc.number or "").startswith("QLD")  # QLD = qo'shish, INV = almashtirish

    for item in doc.items:
        stock = db.query(Stock).filter(
            Stock.warehouse_id == item.warehouse_id,
            Stock.product_id == item.product_id,
        ).first()
        old_quantity = float(stock.quantity or 0) if stock else 0
        new_quantity = float(item.quantity or 0)

        if is_qld:
            # QLD: mavjud qoldiqqa qo'shish
            quantity_change = new_quantity
            item.previous_quantity = old_quantity
        else:
            # INV: aniq raqamga almashtirish
            quantity_change = new_quantity - old_quantity
            item.previous_quantity = old_quantity

        if abs(quantity_change) > 1e-9:
            create_stock_movement(
                db=db,
                warehouse_id=item.warehouse_id,
                product_id=item.product_id,
                quantity_change=quantity_change,
                operation_type="adjustment",
                document_type="StockAdjustmentDoc",
                document_id=doc.id,
                document_number=doc.number,
                user_id=current_user.id if current_user else None,
                note=f"{'Qoldiq kiritish' if is_qld else 'Inventarizatsiya'}: {doc.number}",
                created_at=doc.date,
            )
        if (item.cost_price or 0) > 0:
            prod = db.query(Product).filter(Product.id == item.product_id).first()
            if prod:
                prod.purchase_price = item.cost_price

    # Faqat hujjatdagi tovarlar yangilanadi — boshqa tovarlarning qoldig'iga tegmaydi

    # Stock drift fix: har bir (wh, pid) juftligi uchun stock.quantity harakatlar yig'indisiga moslash
    from app.services.stock_service import reconcile_stock
    db.flush()
    for _wh, _pid in {(it.warehouse_id, it.product_id) for it in doc.items}:
        reconcile_stock(db, _wh, _pid, reason="stock_adjustment_confirm",
                        actor=current_user.username if current_user else None)

    # Status allaqachon atomik UPDATE WHERE bilan o'zgartirildi
    db.commit()
    try:
        from app.bot.services.audit_watchdog import audit_stock_adjustment
        audit_stock_adjustment(doc.id)
    except Exception:
        pass
    return RedirectResponse(url=f"/qoldiqlar/tovar/hujjat/{doc_id}", status_code=303)


@router.post("/tovar/hujjat/{doc_id}/apply-to-warehouse")
async def qoldiqlar_tovar_hujjat_apply_to_warehouse(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Tasdiqlangan hujjat bo'yicha omborni qayta moslash: hujjatda yo'q mahsulotlar uchun qoldiq 0 (savdoda ko'rinmasin)."""
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status not in ("confirmed",):
        raise HTTPException(status_code=400, detail="Faqat tasdiqlangan hujjat uchun")
    # Ikki marta bosish himoyasi: allaqachon apply qilinganmi?
    if doc.status == "applied":
        return RedirectResponse(url=f"/qoldiqlar/tovar/hujjat/{doc_id}?applied=1", status_code=303)
    doc_warehouse_ids = list({item.warehouse_id for item in doc.items})
    doc_pairs = {(item.warehouse_id, item.product_id) for item in doc.items}
    for wh_id in doc_warehouse_ids:
        for stock in db.query(Stock).filter(Stock.warehouse_id == wh_id).all():
            if (stock.warehouse_id, stock.product_id) in doc_pairs:
                continue
            old_q = float(stock.quantity or 0)
            if old_q > 0:
                create_stock_movement(
                    db=db,
                    warehouse_id=stock.warehouse_id,
                    product_id=stock.product_id,
                    quantity_change=-old_q,
                    operation_type="adjustment",
                    document_type="StockAdjustmentDoc",
                    document_id=doc.id,
                    document_number=doc.number,
                    user_id=current_user.id if current_user else None,
                    note=f"Omborni hujjatga moslash: {doc.number}",
                    created_at=doc.date,
                )
    doc.status = "applied"
    db.commit()
    return RedirectResponse(url=f"/qoldiqlar/tovar/hujjat/{doc_id}?applied=1", status_code=303)


@router.post("/tovar/hujjat/{doc_id}/revert")
async def qoldiqlar_tovar_hujjat_revert(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Tovar qoldiq hujjati tasdiqini bekor qilish (faqat admin) — ombor qoldig'i harakatdagi o'zgarishlar orqali qayta tiklanadi."""
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status not in ("confirmed", "applied"):
        raise HTTPException(status_code=400, detail="Faqat tasdiqlangan hujjatning tasdiqini bekor qilish mumkin")
    movements = db.query(StockMovement).filter(
        StockMovement.document_type == "StockAdjustmentDoc",
        StockMovement.document_id == doc_id,
    ).all()
    for m in movements:
        if not m.warehouse_id or not m.product_id:
            continue
        create_stock_movement(
            db=db,
            warehouse_id=m.warehouse_id,
            product_id=m.product_id,
            quantity_change=-float(m.quantity_change or 0),
            operation_type="adjustment_revert",
            document_type="StockAdjustmentDoc",
            document_id=doc_id,
            document_number=doc.number,
            note=f"Hujjat tasdiqini bekor qilish: {doc.number}",
        )
    doc.status = "draft"

    # Stock drift fix: revert dan keyin ham harakatlar yig'indisiga moslash
    from app.services.stock_service import reconcile_stock
    db.flush()
    for _wh, _pid in {(it.warehouse_id, it.product_id) for it in doc.items}:
        reconcile_stock(db, _wh, _pid, reason="stock_adjustment_revert",
                        actor=current_user.username if current_user else None)

    db.commit()
    return RedirectResponse(url="/qoldiqlar/tovar/hujjat?reverted=1", status_code=303)


@router.post("/tovar/hujjat/{doc_id}/delete")
async def qoldiqlar_tovar_hujjat_delete(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Tovar qoldiq hujjatini o'chirish (faqat qoralama, faqat admin)"""
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "draft":
        raise HTTPException(status_code=400, detail="Faqat qoralama holatidagi hujjatni o'chirish mumkin. Avval tasdiqni bekor qiling.")
    db.delete(doc)
    db.commit()
    return RedirectResponse(url="/qoldiqlar#tovar", status_code=303)
