"""
Ombor — qoldiqlar, eksport/import, ombordan omborga o'tkazish.
"""
import io
import traceback
from datetime import datetime
from urllib.parse import quote, unquote

import openpyxl
from fastapi import APIRouter, Request, Depends, Form, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, JSONResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_, and_, func, text
from typing import Optional

from app.core import templates
from app.models.database import (
    get_db,
    User,
    Warehouse,
    Stock,
    Product,
    Purchase,
    PurchaseItem,
    Production,
    Recipe,
    StockAdjustmentDoc,
    StockAdjustmentDocItem,
    WarehouseTransfer,
    WarehouseTransferItem,
    Unit,
    ProductPrice,
    StockMovement,
)
from app.services.stock_service import create_stock_movement, delete_stock_movements_for_document
from app.deps import require_auth, require_admin
from app.utils.user_scope import get_warehouses_for_user

router = APIRouter(prefix="/warehouse", tags=["warehouse"])
inventory_router = APIRouter(prefix="/inventory", tags=["inventory"])


def _product_kg_per_unit(name: str) -> float:
    """Mahsulot nomidan 1 dona ning kg og'irligini aniqlash."""
    import re
    n = (name or "").lower()
    m_gr = re.search(r'(\d+)\s*gr', n)
    if m_gr:
        return int(m_gr.group(1)) / 1000.0
    m_g = re.search(r'(\d+)\s*g(?:\b|\))', n)
    if m_g:
        return int(m_g.group(1)) / 1000.0
    m_kg = re.search(r'([\d.]+)\s*kg', n)
    if m_kg:
        return float(m_kg.group(1))
    return 1.0


def _warehouses_for_user(db: Session, user: User):
    """Foydalanuvchi uchun ko'rinadigan omborlar: sozlamada belgilangan yoki admin/raxbar uchun barcha."""
    return get_warehouses_for_user(db, user)


@router.get("", response_class=HTMLResponse)
async def warehouse_list(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    warehouses = _warehouses_for_user(db, current_user)
    wh_ids = [w.id for w in warehouses]
    stocks_q = db.query(Stock).join(Product).join(Warehouse).filter(Stock.quantity > 0)
    if wh_ids:
        stocks_q = stocks_q.filter(Stock.warehouse_id.in_(wh_ids))
    stocks = stocks_q.all()
    stock_sources = {}
    for s in stocks:
        items = []
        purchases = (
            db.query(Purchase)
            .join(PurchaseItem, Purchase.id == PurchaseItem.purchase_id)
            .filter(
                Purchase.warehouse_id == s.warehouse_id,
                PurchaseItem.product_id == s.product_id,
                Purchase.status == "confirmed",
            )
            .order_by(Purchase.date.desc())
            .limit(3)
            .all()
        )
        for p in purchases:
            items.append((p.number, f"/purchases/edit/{p.id}", p.date.strftime("%d.%m.%Y") if p.date else ""))
        out_wh_id = s.warehouse_id
        # Production query - max_stage muammosini oldini olish
        # Bazada max_stage ustuni bo'lmasligi mumkin, shuning uchun faqat mavjud ustunlarni ishlatamiz
        try:
            # Faqat kerakli ustunlarni tanlab, max_stage ni o'z ichiga olmaydigan so'rov
            # SQLAlchemy modelda max_stage bor, lekin bazada yo'q bo'lishi mumkin
            result = db.execute(
                text("""
                    SELECT p.id, p.number, p.date
                    FROM productions p
                    INNER JOIN recipes r ON p.recipe_id = r.id
                    WHERE p.status = 'completed'
                      AND r.product_id = :product_id
                      AND (
                          p.output_warehouse_id = :warehouse_id
                          OR (p.output_warehouse_id IS NULL AND p.warehouse_id = :warehouse_id)
                      )
                    ORDER BY p.date DESC
                    LIMIT 3
                """),
                {"product_id": s.product_id, "warehouse_id": out_wh_id}
            )
            productions = result.fetchall()
            for pr in productions:
                try:
                    if pr.date and hasattr(pr.date, 'strftime'):
                        pr_date = pr.date.strftime("%d.%m.%Y")
                    elif pr.date:
                        from datetime import datetime as _dt
                        pr_date = _dt.fromisoformat(str(pr.date)[:19]).strftime("%d.%m.%Y")
                    else:
                        pr_date = ""
                except Exception:
                    pr_date = str(pr.date)[:10] if pr.date else ""
                items.append((pr.number, "/production/orders", pr_date))
        except Exception as prod_error:
            # Database da max_stage yoki boshqa ustunlar yo'q bo'lishi mumkin - e'tiborsiz qoldiramiz
            print(f"Production query error (warehouse {s.warehouse_id}, product {s.product_id}): {prod_error}")
            import traceback
            pass  # logged above
        adj_docs = (
            db.query(StockAdjustmentDoc)
            .join(StockAdjustmentDocItem, StockAdjustmentDoc.id == StockAdjustmentDocItem.doc_id)
            .filter(
                StockAdjustmentDoc.status == "confirmed",
                StockAdjustmentDocItem.warehouse_id == s.warehouse_id,
                StockAdjustmentDocItem.product_id == s.product_id,
            )
            .order_by(StockAdjustmentDoc.date.desc())
            .limit(3)
            .distinct()
            .all()
        )
        for doc in adj_docs:
            items.append((doc.number, f"/qoldiqlar/tovar/hujjat/{doc.id}", doc.date.strftime("%d.%m.%Y") if doc.date else ""))
        items.sort(key=lambda x: x[2] or "", reverse=True)
        stock_sources[s.id] = items[:8]
    # Hujjatlar ro'yxati — mahsulotlar hujjat ichida ko'riladi
    qoldiq_docs = (
        db.query(StockAdjustmentDoc)
        .order_by(StockAdjustmentDoc.date.desc(), StockAdjustmentDoc.id.desc())
        .limit(100)
        .all()
    )
    purchase_docs = (
        db.query(Purchase)
        .filter(Purchase.status.in_(["confirmed", "draft"]))
        .order_by(Purchase.date.desc())
        .limit(80)
        .all()
    )
    return templates.TemplateResponse("warehouse/list.html", {
        "request": request,
        "warehouses": warehouses,
        "stocks": stocks,
        "stock_sources": stock_sources,
        "qoldiq_docs": qoldiq_docs,
        "purchase_docs": purchase_docs,
        "current_user": current_user,
        "page_title": "Ombor qoldiqlari",
        "show_tannarx": (getattr(current_user, "role", None) if current_user else None) == "admin",
    })


@router.post("/stock/{stock_id}/zero")
async def warehouse_stock_zero(
    stock_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    stock = db.query(Stock).filter(Stock.id == stock_id).first()
    if not stock:
        raise HTTPException(status_code=404, detail="Qoldiq topilmadi")
    old_qty = float(stock.quantity or 0)
    if old_qty > 0:
        from app.services.stock_service import create_stock_movement
        create_stock_movement(
            db=db,
            warehouse_id=stock.warehouse_id,
            product_id=stock.product_id,
            quantity_change=-old_qty,
            operation_type="manual_zero",
            document_type="ManualAdjustment",
            document_id=stock.id,
            document_number="ZERO",
            user_id=current_user.id if current_user else None,
            note="Qoldiq qo'lda 0 ga tushirildi",
        )
    db.commit()
    return RedirectResponse(url="/warehouse", status_code=303)


@router.get("/export")
async def warehouse_export(db: Session = Depends(get_db), current_user: User = Depends(require_auth)):
    stocks = db.query(Stock).join(Product).join(Warehouse).filter(Stock.quantity > 0).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Qoldiqlar"
    ws.append(["Ombor nomi", "Ombor kodi", "Mahsulot kodi", "Mahsulot nomi", "Qoldiq", "Tannarx (so'm)", "Summa (so'm)"])
    for s in stocks:
        pr, wh = s.product, s.warehouse
        tannarx = (pr.purchase_price or 0) if pr else 0
        summa = s.quantity * tannarx
        ws.append([
            wh.name if wh else "",
            wh.code if wh else "",
            pr.code if pr else "",
            pr.name if pr else "",
            s.quantity,
            tannarx,
            summa,
        ])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ombor_qoldiqlari.xlsx"},
    )


@router.get("/template")
async def warehouse_template(db: Session = Depends(get_db), current_user: User = Depends(require_auth)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Andoza"
    ws.append(["Ombor nomi (yoki kodi)", "Mahsulot nomi (yoki kodi)", "Qoldiq", "Tannarx (so'm)", "Sotuv narxi (so'm)"])
    ws.append(["Xom ashyo ombori", "Yong'oq", 30, "", ""])
    ws.append(["Xom ashyo ombori", "Bodom", 100, "", ""])
    for col in range(1, 6):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=qoldiqlar_andoza.xlsx"},
    )


@router.post("/import")
async def warehouse_import(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    form = await request.form()
    file = form.get("file") or form.get("excel_file")
    if not file or not getattr(file, "filename", None):
        return RedirectResponse(url="/warehouse?error=import&detail=" + quote("Excel fayl tanlang"), status_code=303)
    try:
        contents = await file.read()
        if not contents:
            return RedirectResponse(url="/warehouse?error=import&detail=" + quote("Fayl bo'sh"), status_code=303)
        if len(contents) > 5 * 1024 * 1024:
            return RedirectResponse(url="/warehouse?error=import&detail=" + quote("Fayl hajmi 5MB dan oshmasligi kerak"), status_code=303)
        if contents[:2] != b"PK":
            return RedirectResponse(url="/warehouse?error=import&detail=" + quote("Fayl .xlsx formati bo'lishi kerak"), status_code=303)
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=False, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        updated, skipped = 0, 0
        missing_products, missing_warehouses = [], []
        for row in rows:
            if not row or (row[0] is None and row[1] is None):
                continue
            wh_key = str(row[0] or "").strip() if len(row) > 0 else ""
            raw_prod = row[1] if len(row) > 1 else None
            prod_key = str(int(float(raw_prod))) if raw_prod is not None and isinstance(raw_prod, (int, float)) and float(raw_prod) == int(float(raw_prod)) else str(raw_prod or "").strip()
            try:
                qty = float(row[2]) if len(row) > 2 and row[2] is not None else 0
                qty = max(0, qty)
            except (TypeError, ValueError):
                qty = 0
            tannarx = None
            sotuv_narxi = None
            if len(row) > 3 and row[3] is not None and row[3] != "":
                try:
                    tannarx = float(row[3])
                except (TypeError, ValueError):
                    pass
            if len(row) > 4 and row[4] is not None and row[4] != "":
                try:
                    sotuv_narxi = float(row[4])
                except (TypeError, ValueError):
                    pass
            if not wh_key or not prod_key:
                skipped += 1
                continue
            warehouse = db.query(Warehouse).filter(
                (func.lower(Warehouse.name) == wh_key.lower()) | (Warehouse.code == wh_key)
            ).first()
            # Kod/shtrixkod va nom — katta-kichik harf farqisiz
            product = db.query(Product).filter(
                or_(
                    and_(Product.code.isnot(None), Product.code != "", func.lower(Product.code) == prod_key.lower()),
                    and_(Product.barcode.isnot(None), Product.barcode != "", func.lower(Product.barcode) == prod_key.lower()),
                )
            ).first()
            if not product and prod_key:
                product = db.query(Product).filter(
                    Product.name.isnot(None),
                    func.lower(func.trim(Product.name)) == prod_key.strip().lower(),
                ).first()
            if not warehouse:
                if wh_key and wh_key not in missing_warehouses:
                    missing_warehouses.append(wh_key)
                skipped += 1
                continue
            if not product:
                if prod_key and prod_key not in missing_products:
                    missing_products.append(prod_key)
                skipped += 1
                continue
            stock = db.query(Stock).filter(
                Stock.warehouse_id == warehouse.id,
                Stock.product_id == product.id,
            ).first()
            if stock:
                stock.quantity = qty
            else:
                db.add(Stock(warehouse_id=warehouse.id, product_id=product.id, quantity=qty))
            if tannarx is not None:
                product.purchase_price = tannarx
            if sotuv_narxi is not None:
                product.sale_price = sotuv_narxi
            updated += 1
        db.commit()
        detail = f"Yuklandi: {updated} ta"
        if skipped:
            detail += f", o'tkazib yuborildi: {skipped} ta"
            if missing_products:
                sample = ", ".join(missing_products[:5])
                if len(missing_products) > 5:
                    sample += f" va yana {len(missing_products) - 5} ta"
                detail += f". Mahsulot topilmadi: {sample}"
            if missing_warehouses:
                sample = ", ".join(missing_warehouses[:3])
                if len(missing_warehouses) > 3:
                    sample += f" va yana {len(missing_warehouses) - 3} ta"
                detail += f". Ombor topilmadi: {sample}"
        return RedirectResponse(url="/warehouse?success=import&detail=" + quote(detail), status_code=303)
    except Exception as e:
        pass  # logged above
        return RedirectResponse(url="/warehouse?error=import&detail=" + quote("Import xatoligi. Fayl formatini tekshiring."), status_code=303)


@router.get("/transfers", response_class=HTMLResponse)
async def warehouse_transfers_list(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    transfers = db.query(WarehouseTransfer).order_by(WarehouseTransfer.date.desc()).limit(200).all()
    role = (getattr(current_user, "role", None) or "").strip().lower()
    if role in ("manager", "menejer", "sotuvchi"):
        wh_ids = [w.id for w in _warehouses_for_user(db, current_user)]
        if wh_ids:
            transfers = [t for t in transfers if (t.from_warehouse_id in wh_ids or t.to_warehouse_id in wh_ids)]
    error = request.query_params.get("error")
    return templates.TemplateResponse("warehouse/transfers_list.html", {
        "request": request,
        "current_user": current_user,
        "transfers": transfers,
        "page_title": "Ombordan omborga o'tkazish",
        "error_message": unquote(error) if error else None,
    })


@router.get("/transfers/new", response_class=HTMLResponse)
async def warehouse_transfer_new(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    warehouses = _warehouses_for_user(db, current_user)
    products = db.query(Product).filter(Product.is_active == True).order_by(Product.name).all()
    stocks = db.query(Stock).filter(Stock.quantity > 0).all()
    stock_by_warehouse_product = {}
    for s in stocks:
        wid, pid = str(s.warehouse_id), str(s.product_id)
        if wid not in stock_by_warehouse_product:
            stock_by_warehouse_product[wid] = {}
        stock_by_warehouse_product[wid][pid] = s.quantity
    products_list = [{"id": p.id, "name": (p.name or ""), "code": (p.code or "")} for p in products]
    return templates.TemplateResponse("warehouse/transfer_form.html", {
        "request": request,
        "current_user": current_user,
        "transfer": None,
        "warehouses": warehouses,
        "products": products,
        "products_list": products_list,
        "stock_by_warehouse_product": stock_by_warehouse_product,
        "now": datetime.now(),
        "page_title": "Ombordan omborga o'tkazish (yaratish)",
        "show_tannarx": (getattr(current_user, "role", None) if current_user else None) == "admin",
    })


@router.get("/transfers/{transfer_id}", response_class=HTMLResponse)
async def warehouse_transfer_edit(
    request: Request,
    transfer_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    transfer = db.query(WarehouseTransfer).filter(WarehouseTransfer.id == transfer_id).first()
    if not transfer:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    warehouses = _warehouses_for_user(db, current_user)
    products = db.query(Product).filter(Product.is_active == True).order_by(Product.name).all()
    stocks = db.query(Stock).filter(Stock.quantity > 0).all()
    stock_by_warehouse_product = {}
    for s in stocks:
        wid, pid = str(s.warehouse_id), str(s.product_id)
        if wid not in stock_by_warehouse_product:
            stock_by_warehouse_product[wid] = {}
        stock_by_warehouse_product[wid][pid] = s.quantity
    products_list = [{"id": p.id, "name": (p.name or ""), "code": (p.code or "")} for p in products]
    return templates.TemplateResponse("warehouse/transfer_form.html", {
        "request": request,
        "current_user": current_user,
        "transfer": transfer,
        "warehouses": warehouses,
        "products": products,
        "products_list": products_list,
        "stock_by_warehouse_product": stock_by_warehouse_product,
        "now": transfer.date or datetime.now(),
        "page_title": f"O'tkazish {transfer.number}",
        "show_tannarx": (getattr(current_user, "role", None) if current_user else None) == "admin",
    })


@router.post("/transfers/create")
async def warehouse_transfer_create(
    request: Request,
    from_warehouse_id: int = Form(...),
    to_warehouse_id: int = Form(...),
    note: str = Form(""),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    if from_warehouse_id == to_warehouse_id:
        return RedirectResponse(url="/warehouse/transfers/new?error=" + quote("Qayerdan va qayerga bir xil bo'lmasin."), status_code=303)
    form = await request.form()
    today = datetime.now()
    count = db.query(WarehouseTransfer).filter(
        WarehouseTransfer.date >= today.replace(hour=0, minute=0, second=0)
    ).count()
    number = f"OT-{today.strftime('%Y%m%d')}-{str(count + 1).zfill(4)}"
    transfer = WarehouseTransfer(
        number=number,
        from_warehouse_id=from_warehouse_id,
        to_warehouse_id=to_warehouse_id,
        status="draft",
        user_id=current_user.id,
        note=note or None,
    )
    db.add(transfer)
    db.commit()
    db.refresh(transfer)
    for key, value in form.items():
        if key.startswith("product_id_") and value:
            try:
                pid = int(value)
                qkey = "quantity_" + key.replace("product_id_", "")
                qty = float(form.get(qkey, "0").replace(",", "."))
                if pid and qty > 0:
                    db.add(WarehouseTransferItem(transfer_id=transfer.id, product_id=pid, quantity=qty))
            except (ValueError, TypeError):
                pass
    db.commit()
    return RedirectResponse(url=f"/warehouse/transfers/{transfer.id}", status_code=303)


@router.post("/transfers/{transfer_id}/save")
async def warehouse_transfer_save(
    request: Request,
    transfer_id: int,
    from_warehouse_id: int = Form(...),
    to_warehouse_id: int = Form(...),
    note: str = Form(""),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    transfer = db.query(WarehouseTransfer).filter(WarehouseTransfer.id == transfer_id).first()
    if not transfer or transfer.status != "draft":
        raise HTTPException(status_code=404, detail="Hujjat topilmadi yoki tahrirlab bo'lmaydi")
    if from_warehouse_id == to_warehouse_id:
        return RedirectResponse(url=f"/warehouse/transfers/{transfer_id}?error=" + quote("Qayerdan va qayerga bir xil bo'lmasin."), status_code=303)
    transfer.from_warehouse_id = from_warehouse_id
    transfer.to_warehouse_id = to_warehouse_id
    transfer.note = note or None
    form = await request.form()
    db.query(WarehouseTransferItem).filter(WarehouseTransferItem.transfer_id == transfer_id).delete()
    for key, value in form.items():
        if key.startswith("product_id_") and value:
            try:
                pid = int(value)
                qkey = "quantity_" + key.replace("product_id_", "")
                qty = float(form.get(qkey, "0").replace(",", "."))
                if pid and qty > 0:
                    db.add(WarehouseTransferItem(transfer_id=transfer_id, product_id=pid, quantity=qty))
            except (ValueError, TypeError):
                pass
    db.commit()
    return RedirectResponse(url="/warehouse/transfers?saved=1", status_code=303)


@router.post("/transfers/{transfer_id}/confirm")
async def warehouse_transfer_confirm(
    transfer_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    transfer = db.query(WarehouseTransfer).filter(WarehouseTransfer.id == transfer_id).first()
    if not transfer:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if transfer.status == "confirmed":
        return RedirectResponse(url=f"/warehouse/transfers/{transfer_id}?error=" + quote("Hujjat allaqachon tasdiqlangan."), status_code=303)
    items = db.query(WarehouseTransferItem).filter(WarehouseTransferItem.transfer_id == transfer_id).all()
    if not items:
        return RedirectResponse(url=f"/warehouse/transfers/{transfer_id}?error=" + quote("Kamida bitta mahsulot qo'shing."), status_code=303)
    for item in items:
        src = db.query(Stock).filter(
            Stock.warehouse_id == transfer.from_warehouse_id,
            Stock.product_id == item.product_id,
        ).first()
        need = float(item.quantity or 0)
        have = float(src.quantity or 0) if src else 0
        if not src or (have + 1e-6 < need):
            prod = db.query(Product).filter(Product.id == item.product_id).first()
            name = prod.name if prod else f"#{item.product_id}"
            avail_display = "0" if abs(have) < 1e-6 else ("%.6f" % have).rstrip("0").rstrip(".")
            return RedirectResponse(
                url=f"/warehouse/transfers/{transfer_id}?error=" + quote(f"Qayerdan omborda «{name}» yetarli emas (kerak: {item.quantity}, mavjud: {avail_display})"),
                status_code=303,
            )
    from app.services.stock_service import create_stock_movement
    for item in items:
        # Stock.quantity ni faqat create_stock_movement o'zgartiradi (ikki marta o'zgarmaslik uchun)
        create_stock_movement(
            db=db,
            warehouse_id=transfer.from_warehouse_id,
            product_id=item.product_id,
            quantity_change=-item.quantity,
            operation_type="transfer_out",
            document_type="WarehouseTransfer",
            document_id=transfer.id,
            document_number=transfer.number,
            user_id=current_user.id if current_user else None,
            note=f"O'tkazma chiqim: {transfer.number}",
            created_at=transfer.date,
        )
        create_stock_movement(
            db=db,
            warehouse_id=transfer.to_warehouse_id,
            product_id=item.product_id,
            quantity_change=+item.quantity,
            operation_type="transfer_in",
            document_type="WarehouseTransfer",
            document_id=transfer.id,
            document_number=transfer.number,
            user_id=current_user.id if current_user else None,
            note=f"O'tkazma kirim: {transfer.number}",
            created_at=transfer.date,
        )
    transfer.status = "confirmed"
    db.commit()
    return RedirectResponse(url=f"/warehouse/transfers/{transfer_id}?confirmed=1", status_code=303)


@router.post("/transfers/{transfer_id}/revert")
async def warehouse_transfer_revert(
    transfer_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    transfer = db.query(WarehouseTransfer).filter(WarehouseTransfer.id == transfer_id).first()
    if not transfer:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if transfer.status != "confirmed":
        return RedirectResponse(url="/warehouse/transfers?error=" + quote("Faqat tasdiqlangan hujjatning tasdiqini bekor qilish mumkin."), status_code=303)
    from app.services.stock_service import create_stock_movement, delete_stock_movements_for_document
    items = db.query(WarehouseTransferItem).filter(WarehouseTransferItem.transfer_id == transfer_id).all()
    for item in items:
        # Teskari harakatlar: dest dan chiqim, src ga kirim
        create_stock_movement(
            db=db,
            warehouse_id=transfer.to_warehouse_id,
            product_id=item.product_id,
            quantity_change=-item.quantity,
            operation_type="transfer_out",
            document_type="WarehouseTransfer",
            document_id=transfer.id,
            document_number=f"{transfer.number}-REVERT",
            user_id=current_user.id if current_user else None,
            note=f"O'tkazma bekor: {transfer.number}",
            created_at=transfer.date,
        )
        create_stock_movement(
            db=db,
            warehouse_id=transfer.from_warehouse_id,
            product_id=item.product_id,
            quantity_change=+item.quantity,
            operation_type="transfer_in",
            document_type="WarehouseTransfer",
            document_id=transfer.id,
            document_number=f"{transfer.number}-REVERT",
            user_id=current_user.id if current_user else None,
            note=f"O'tkazma bekor: {transfer.number}",
            created_at=transfer.date,
        )
    transfer.status = "draft"
    db.commit()
    return RedirectResponse(url="/warehouse/transfers?reverted=1", status_code=303)


@router.post("/transfers/{transfer_id}/delete")
async def warehouse_transfer_delete(
    transfer_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    transfer = db.query(WarehouseTransfer).filter(WarehouseTransfer.id == transfer_id).first()
    if not transfer:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if transfer.status == "confirmed":
        return RedirectResponse(
            url="/warehouse/transfers?error=" + quote("Tasdiqlangan hujjatni to'g'ridan-to'g'ri o'chirib bo'lmaydi. Avval tasdiqni bekor qiling."),
            status_code=303,
        )
    db.delete(transfer)
    db.commit()
    return RedirectResponse(url="/warehouse/transfers?deleted=1", status_code=303)


@router.get("/movement", response_class=HTMLResponse)
async def warehouse_movement(request: Request, current_user: User = Depends(require_auth)):
    return RedirectResponse(url="/warehouse/transfers", status_code=302)


@router.post("/transfer")
async def warehouse_transfer(
    request: Request,
    from_warehouse_id: int = Form(...),
    to_warehouse_id: int = Form(...),
    product_id: int = Form(...),
    quantity: float = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    if from_warehouse_id == to_warehouse_id:
        return RedirectResponse(url="/warehouse/movement?error=1&detail=" + quote("Qayerdan va qayerga ombor bir xil bo'lmasin."), status_code=303)
    if quantity <= 0:
        return RedirectResponse(url="/warehouse/movement?error=1&detail=" + quote("Miqdor 0 dan katta bo'lishi kerak."), status_code=303)
    source = db.query(Stock).filter(
        Stock.warehouse_id == from_warehouse_id,
        Stock.product_id == product_id,
    ).first()
    need_q = float(quantity or 0)
    have_q = float(source.quantity or 0) if source else 0
    if not source or (have_q + 1e-6 < need_q):
        product = db.query(Product).filter(Product.id == product_id).first()
        name = product.name if product else f"#{product_id}"
        avail_display = "0" if abs(have_q) < 1e-6 else ("%.6f" % have_q).rstrip("0").rstrip(".")
        return RedirectResponse(
            url="/warehouse/movement?error=1&detail=" + quote(f"Qayerdan omborda «{name}» yetarli emas (kerak: {quantity}, mavjud: {avail_display})"),
            status_code=303,
        )
    source.quantity -= quantity
    if source.quantity <= 0:
        source.quantity = 0
    dest = db.query(Stock).filter(
        Stock.warehouse_id == to_warehouse_id,
        Stock.product_id == product_id,
    ).first()
    if dest:
        dest.quantity += quantity
    else:
        db.add(Stock(warehouse_id=to_warehouse_id, product_id=product_id, quantity=quantity))
    db.commit()
    return RedirectResponse(url="/warehouse/movement?success=1", status_code=303)


# ==========================================
# INVENTORY ROUTES (inventory_router)
# ==========================================

def _ensure_inventory_columns(db: Session) -> None:
    """SQLite: stock_adjustment_docs.warehouse_id va stock_adjustment_doc_items.previous_quantity ustunlari yo'q bo'lsa qo'shadi."""
    try:
        r = db.execute(text("PRAGMA table_info(stock_adjustment_docs)"))
        cols_doc = [row[1] for row in r.fetchall()]
        if "warehouse_id" not in cols_doc:
            db.execute(text("ALTER TABLE stock_adjustment_docs ADD COLUMN warehouse_id INTEGER REFERENCES warehouses(id)"))
            db.commit()
        r = db.execute(text("PRAGMA table_info(stock_adjustment_doc_items)"))
        cols_item = [row[1] for row in r.fetchall()]
        if "previous_quantity" not in cols_item:
            db.execute(text("ALTER TABLE stock_adjustment_doc_items ADD COLUMN previous_quantity REAL"))
            db.commit()
    except Exception:
        db.rollback()


def _parse_doc_date(s: str):
    """Sana matnini parse qiladi: YYYY-MM-DDTHH:MM, dd.mm.yyyy HH:MM, dd.mm.yyyy va boshqa formatlar."""
    import re
    if not s or not str(s).strip():
        return None
    s = str(s).strip()
    try:
        return datetime.fromisoformat(s.replace("Z", "")[:19])
    except (ValueError, TypeError):
        pass
    m = re.match(r"(\d{1,2})[./](\d{1,2})[./](\d{4})\s+(\d{1,2}):(\d{2})", s)
    if m:
        try:
            return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)), int(m.group(4)), int(m.group(5)))
        except (ValueError, TypeError):
            pass
    m = re.match(r"(\d{1,2})[./](\d{1,2})[./](\d{4})", s)
    if m:
        try:
            return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except (ValueError, TypeError):
            pass
    return None


def _parse_quantity(value) -> float:
    if value is None or str(value).strip() == "":
        return 0.0
    try:
        return float(str(value).strip().replace(",", "."))
    except (TypeError, ValueError):
        return 0.0


def _next_inventory_number(db: Session, date_str: str) -> str:
    prefix = f"INV-{date_str}-"
    rows = db.query(StockAdjustmentDoc.number).filter(
        StockAdjustmentDoc.number.like(f"{prefix}%")
    ).all()
    max_suffix = 0
    for (num,) in rows:
        if num and num.startswith(prefix):
            try:
                suf = int(num[len(prefix):].strip())
                if suf > max_suffix:
                    max_suffix = suf
            except (ValueError, TypeError):
                pass
    return f"{prefix}{str(max_suffix + 1).zfill(4)}"


@inventory_router.get("", response_class=HTMLResponse)
async def inventory_list_page(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    migration_warning = None
    try:
        docs = (
            db.query(StockAdjustmentDoc)
            .filter(StockAdjustmentDoc.warehouse_id.isnot(None))
            .order_by(StockAdjustmentDoc.created_at.desc())
            .limit(200)
            .all()
        )
    except Exception:
        try:
            _ensure_inventory_columns(db)
            docs = (
                db.query(StockAdjustmentDoc)
                .filter(StockAdjustmentDoc.warehouse_id.isnot(None))
                .order_by(StockAdjustmentDoc.created_at.desc())
                .limit(200)
                .all()
            )
        except Exception:
            docs = []
            migration_warning = "Inventarizatsiya uchun bazada warehouse_id ustuni kerak. Loyiha ildizida: alembic upgrade head"
    message = request.query_params.get("message", "").strip()
    return templates.TemplateResponse("inventory/list.html", {
        "request": request,
        "docs": docs,
        "current_user": current_user,
        "page_title": "Inventarizatsiya",
        "migration_warning": migration_warning,
        "message": message,
    })


@inventory_router.get("/new", response_class=HTMLResponse)
async def inventory_new_page(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    warehouses = db.query(Warehouse).filter(Warehouse.is_active == True).order_by(Warehouse.name).all()
    return templates.TemplateResponse("inventory/new.html", {
        "request": request,
        "warehouses": warehouses,
        "current_user": current_user,
        "page_title": "Inventarizatsiya — yangi hujjat",
    })


@inventory_router.post("/create-draft")
async def inventory_create_draft(
    warehouse_id: int = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    wh = db.query(Warehouse).filter(Warehouse.id == warehouse_id).first()
    if not wh:
        return RedirectResponse(url="/inventory/new?message=Ombor topilmadi.", status_code=303)
    today = datetime.now()
    doc = StockAdjustmentDoc(
        number="INV-PENDING",
        date=today,
        warehouse_id=warehouse_id,
        user_id=current_user.id,
        status="draft",
        total_tannarx=0,
        total_sotuv=0,
    )
    db.add(doc)
    db.flush()
    doc.number = f"INV-PENDING-{doc.id}"
    db.commit()
    db.refresh(doc)
    return RedirectResponse(url=f"/inventory/{doc.id}/edit", status_code=303)


@inventory_router.get("/{doc_id}/edit", response_class=HTMLResponse)
async def inventory_edit_page(
    doc_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "draft":
        return RedirectResponse(url=f"/inventory/{doc_id}", status_code=303)
    if not doc.warehouse_id:
        return RedirectResponse(url="/inventory", status_code=303)
    warehouse = doc.warehouse
    wh_id = doc.warehouse_id
    stocks_by_product = {}
    for s in db.query(Stock).filter(Stock.warehouse_id == wh_id).all():
        pid = s.product_id
        stocks_by_product[pid] = stocks_by_product.get(pid, 0) + float(s.quantity or 0)
    by_product = {}
    for item in doc.items:
        pid = item.product_id
        if pid not in by_product:
            prod = item.product
            by_product[pid] = {
                "item_id": item.id,
                "product_id": pid,
                "product_name": (prod.name or "") if prod else "",
                "product_code": (prod.code or "") if prod else "",
                "current_quantity": stocks_by_product.get(pid, 0),
                "actual_quantity": float(item.quantity or 0),
                "cost_price": float(item.cost_price or 0),
                "sale_price": float(item.sale_price or 0),
            }
    product_ids_in_doc = set(by_product.keys())
    if product_ids_in_doc:
        products_to_add = db.query(Product).filter(Product.is_active == True).filter(~Product.id.in_(product_ids_in_doc)).order_by(Product.name).all()
    else:
        products_to_add = db.query(Product).filter(Product.is_active == True).order_by(Product.name).all()
    products_data = sorted(by_product.values(), key=lambda x: (x["product_name"].lower(), x["product_id"]))
    show_tannarx = getattr(current_user, "role", None) == "admin"
    inv_date = doc.date or datetime.now()
    doc_date_value = inv_date.strftime("%Y-%m-%dT%H:%M") if inv_date else ""
    doc_date_display = inv_date.strftime("%d.%m.%Y %H:%M") if inv_date else ""
    return templates.TemplateResponse("inventory/edit.html", {
        "request": request,
        "doc": doc,
        "warehouse": warehouse,
        "products_data": products_data,
        "products_to_add": products_to_add,
        "show_tannarx": show_tannarx,
        "current_user": current_user,
        "doc_date_value": doc_date_value,
        "doc_date_display": doc_date_display,
        "page_title": "Inventarizatsiya — tahrirlash",
    })


@inventory_router.post("/{doc_id}/load")
async def inventory_load_warehouse(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/inventory", status_code=303)
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc or doc.status != "draft" or not doc.warehouse_id:
        return RedirectResponse(url="/inventory", status_code=303)
    wh_id = doc.warehouse_id
    stocks = (
        db.query(Stock, Product, Unit)
        .join(Product, Stock.product_id == Product.id)
        .outerjoin(Unit, Product.unit_id == Unit.id)
        .filter(Stock.warehouse_id == wh_id)
        .filter(Product.is_active == True)
        .all()
    )
    by_product = {}
    for stock, product, unit in stocks:
        pid = product.id
        qty = float(stock.quantity or 0)
        if pid not in by_product:
            by_product[pid] = qty
        else:
            by_product[pid] += qty
    existing_ids = {item.product_id for item in doc.items}
    for pid, qty in by_product.items():
        if qty is None or float(qty or 0) == 0:
            continue
        if pid in existing_ids:
            continue
        product = db.query(Product).filter(Product.id == pid).first()
        if not product:
            continue
        cost = float(product.purchase_price or 0)
        sale = float(product.sale_price or 0)
        if (product.sale_price or 0) <= 0:
            pp = db.query(ProductPrice).filter(ProductPrice.product_id == pid).first()
            if pp:
                sale = float(pp.sale_price or 0)
        db.add(StockAdjustmentDocItem(
            doc_id=doc_id,
            product_id=pid,
            warehouse_id=wh_id,
            quantity=qty,
            cost_price=cost,
            sale_price=sale,
        ))
        existing_ids.add(pid)
    db.commit()
    return RedirectResponse(url=f"/inventory/{doc_id}/edit?message=Qoldiq tovarlar yuklandi.", status_code=303)


@inventory_router.post("/{doc_id}/add-product")
async def inventory_add_product(
    doc_id: int,
    product_id: int = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc or doc.status != "draft" or not doc.warehouse_id:
        return RedirectResponse(url="/inventory", status_code=303)
    if any(item.product_id == product_id for item in doc.items):
        return RedirectResponse(url=f"/inventory/{doc_id}/edit?message=Ushbu tovar allaqachon jadvalda.", status_code=303)
    product = db.query(Product).filter(Product.id == product_id, Product.is_active == True).first()
    if not product:
        return RedirectResponse(url=f"/inventory/{doc_id}/edit?message=Tovar topilmadi.", status_code=303)
    cost = float(product.purchase_price or 0)
    sale = float(product.sale_price or 0)
    if (product.sale_price or 0) <= 0:
        pp = db.query(ProductPrice).filter(ProductPrice.product_id == product_id).first()
        if pp:
            sale = float(pp.sale_price or 0)
    db.add(StockAdjustmentDocItem(
        doc_id=doc_id,
        product_id=product_id,
        warehouse_id=doc.warehouse_id,
        quantity=0,
        cost_price=cost,
        sale_price=sale,
    ))
    db.commit()
    return RedirectResponse(url=f"/inventory/{doc_id}/edit?message=Tovar qo'shildi.", status_code=303)


@inventory_router.post("/{doc_id}/remove-zero-balance")
async def inventory_remove_zero_balance(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/inventory", status_code=303)
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc or doc.status != "draft" or not doc.warehouse_id:
        return RedirectResponse(url="/inventory", status_code=303)
    wh_id = doc.warehouse_id
    removed = 0
    for item in list(doc.items):
        total = sum(
            float(s.quantity or 0)
            for s in db.query(Stock).filter(
                Stock.warehouse_id == wh_id,
                Stock.product_id == item.product_id,
            ).all()
        )
        if total <= 0:
            db.delete(item)
            removed += 1
    db.commit()
    msg = f"0 qoldiqli {removed} ta qator olib tashlandi." if removed else "0 qoldiqli qatorlar yo'q."
    return RedirectResponse(url=f"/inventory/{doc_id}/edit?message=" + quote(msg), status_code=303)


@inventory_router.post("/{doc_id}/remove-item")
async def inventory_remove_item(
    request: Request,
    doc_id: int,
    item_id: int = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Inventarizatsiya jadvalidan bitta qatorni o'chirish."""
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc or doc.status != "draft":
        return JSONResponse({"success": False, "error": "Hujjat topilmadi yoki tasdiqlanган"})
    item = db.query(StockAdjustmentDocItem).filter(
        StockAdjustmentDocItem.id == item_id,
        StockAdjustmentDocItem.doc_id == doc_id,
    ).first()
    if not item:
        return JSONResponse({"success": False, "error": "Qator topilmadi"})
    db.delete(item)
    db.commit()
    return JSONResponse({"success": True})


@inventory_router.post("/{doc_id}/save")
async def inventory_save_draft(
    doc_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/inventory", status_code=303)
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc or doc.status != "draft":
        return RedirectResponse(url="/inventory", status_code=303)
    try:
        form = await request.form()
        doc_date_str = form.get("doc_date")
        if doc_date_str:
            parsed = _parse_doc_date(doc_date_str)
            if parsed:
                doc.date = parsed
        if doc.number and doc.number.startswith("INV-PENDING") and doc.date:
            date_str = doc.date.strftime("%Y%m%d")
            doc.number = _next_inventory_number(db, date_str)
        item_ids = form.getlist("item_id")
        quantities = form.getlist("actual_quantity")
        total_tannarx = 0.0
        total_sotuv = 0.0
        for i, iid in enumerate(item_ids):
            if not iid:
                continue
            try:
                item_id = int(iid)
            except (TypeError, ValueError):
                continue
            item = db.query(StockAdjustmentDocItem).filter(
                StockAdjustmentDocItem.id == item_id,
                StockAdjustmentDocItem.doc_id == doc_id,
            ).first()
            if not item:
                continue
            raw = quantities[i] if i < len(quantities) else None
            if raw is not None and str(raw).strip() != "":
                qty = _parse_quantity(raw)
            else:
                qty = float(item.quantity or 0)
            item.quantity = qty
            total_tannarx += qty * float(item.cost_price or 0)
            total_sotuv += qty * float(item.sale_price or 0)
        doc.total_tannarx = total_tannarx
        doc.total_sotuv = total_sotuv
        try:
            db.commit()
        except Exception as commit_err:
            db.rollback()
            if "UNIQUE" in str(commit_err) and "number" in str(commit_err).lower() and doc.date:
                date_str = doc.date.strftime("%Y%m%d")
                doc.number = _next_inventory_number(db, date_str)
                db.commit()
            else:
                raise commit_err
        return RedirectResponse(url=f"/inventory/{doc_id}/edit?message=Saqlandi.", status_code=303)
    except Exception as e:
        db.rollback()
        import logging
        logging.getLogger(__name__).exception("inventory save error doc_id=%s: %s", doc_id, e)
        return RedirectResponse(
            url=f"/inventory/{doc_id}/edit?error=" + quote("Saqlashda xatolik yuz berdi"),
            status_code=303,
        )


@inventory_router.get("/{doc_id}", response_class=HTMLResponse)
async def inventory_view_page(
    doc_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status == "draft":
        return RedirectResponse(url=f"/inventory/{doc_id}/edit", status_code=303)
    if not doc.warehouse_id:
        return RedirectResponse(url="/inventory", status_code=303)
    warehouse = doc.warehouse
    rows = []
    for item in doc.items:
        prod = item.product
        rows.append({
            "product_name": (prod.name or "") if prod else "",
            "product_code": (prod.code or "") if prod else "",
            "quantity": float(item.quantity or 0),
            "cost_price": float(item.cost_price or 0),
            "sale_price": float(item.sale_price or 0),
        })
    show_tannarx = getattr(current_user, "role", None) == "admin"
    return templates.TemplateResponse("inventory/view.html", {
        "request": request,
        "doc": doc,
        "warehouse": warehouse,
        "rows": rows,
        "show_tannarx": show_tannarx,
        "current_user": current_user,
        "page_title": "Inventarizatsiya — " + (doc.number or ""),
    })


@inventory_router.post("/{doc_id}/confirm")
async def inventory_confirm(
    doc_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc or doc.status != "draft" or not doc.warehouse_id:
        return RedirectResponse(url="/inventory", status_code=303)
    if not doc.items:
        return RedirectResponse(url=f"/inventory/{doc_id}/edit?message=Jadval bo'sh. Avval qoldiq tovarlarni yuklang yoki tovar qo'shing.", status_code=303)
    form = await request.form()
    doc_date_str = form.get("doc_date")
    if doc_date_str:
        parsed = _parse_doc_date(doc_date_str)
        if parsed:
            doc.date = parsed
    if doc.number and doc.number.startswith("INV-PENDING") and doc.date:
        date_str = doc.date.strftime("%Y%m%d")
        doc.number = _next_inventory_number(db, date_str)
    item_ids = form.getlist("item_id")
    quantities = form.getlist("actual_quantity")
    for i, iid in enumerate(item_ids):
        if not iid:
            continue
        try:
            item_id = int(iid)
        except (TypeError, ValueError):
            continue
        item = db.query(StockAdjustmentDocItem).filter(
            StockAdjustmentDocItem.id == item_id,
            StockAdjustmentDocItem.doc_id == doc_id,
        ).first()
        if not item:
            continue
        raw = quantities[i] if i < len(quantities) else None
        if raw is not None and str(raw).strip() != "":
            qty = _parse_quantity(raw)
        else:
            qty = float(item.quantity or 0)
        item.quantity = qty
    total_tannarx = sum(float(it.quantity or 0) * float(it.cost_price or 0) for it in doc.items)
    total_sotuv = sum(float(it.quantity or 0) * float(it.sale_price or 0) for it in doc.items)
    doc.total_tannarx = total_tannarx
    doc.total_sotuv = total_sotuv
    db.commit()
    db.refresh(doc)
    pairs = set((item.warehouse_id, item.product_id) for item in doc.items)
    for wh_id, prod_id in pairs:
        rows = db.query(Stock).filter(Stock.warehouse_id == wh_id, Stock.product_id == prod_id).all()
        if len(rows) > 1:
            total = sum(float(r.quantity or 0) for r in rows)
            keep = rows[0]
            keep.quantity = total
            if hasattr(keep, "updated_at"):
                keep.updated_at = datetime.now()
            for r in rows[1:]:
                db.delete(r)
    db.commit()
    # Items ni oldindan list ga olish (commit dan keyin lazy-load muammosi uchun)
    items_snapshot = [
        {
            "item": item,
            "warehouse_id": item.warehouse_id,
            "product_id": item.product_id,
            "quantity": float(item.quantity or 0),
        }
        for item in doc.items
    ]
    for snap in items_snapshot:
        # Hujjat sanasidagi qoldiqni aniqlash (shu sanadan oldingi oxirgi movement)
        doc_date = doc.date or datetime.now()
        last_mv = (
            db.query(StockMovement)
            .filter(
                StockMovement.warehouse_id == snap["warehouse_id"],
                StockMovement.product_id == snap["product_id"],
                StockMovement.created_at <= doc_date,
            )
            .order_by(StockMovement.id.desc())
            .first()
        )
        if last_mv:
            old_qty = float(last_mv.quantity_after or 0)
        else:
            # Movement yo'q — Stock jadvalidan olish
            stocks = db.query(Stock).filter(
                Stock.warehouse_id == snap["warehouse_id"],
                Stock.product_id == snap["product_id"],
            ).all()
            old_qty = sum(float(s.quantity or 0) for s in stocks)
        new_qty = snap["quantity"]
        if hasattr(snap["item"], "previous_quantity"):
            snap["item"].previous_quantity = old_qty
        quantity_change = new_qty - old_qty
        if abs(quantity_change) > 1e-9:
            create_stock_movement(
                db=db,
                warehouse_id=snap["warehouse_id"],
                product_id=snap["product_id"],
                quantity_change=quantity_change,
                operation_type="adjustment",
                document_type="StockAdjustmentDoc",
                document_id=doc.id,
                document_number=doc.number,
                user_id=current_user.id,
                note=f"Inventarizatsiya: {doc.number}",
                created_at=doc.date,
            )
            # Stock.quantity = new_qty + (hujjat sanasidan keyingi harakatlar)
            stock_row = db.query(Stock).filter(
                Stock.warehouse_id == snap["warehouse_id"],
                Stock.product_id == snap["product_id"],
            ).first()
            if stock_row:
                # Hujjat sanasidan keyingi harakatlar yig'indisi
                from sqlalchemy import func as sqla_func
                after_changes = db.query(
                    sqla_func.coalesce(sqla_func.sum(StockMovement.quantity_change), 0)
                ).filter(
                    StockMovement.warehouse_id == snap["warehouse_id"],
                    StockMovement.product_id == snap["product_id"],
                    StockMovement.created_at > doc_date,
                    StockMovement.operation_type != "adjustment",
                ).scalar() or 0
                stock_row.quantity = new_qty + float(after_changes)
    doc.status = "confirmed"
    db.commit()
    return RedirectResponse(url=f"/inventory/{doc_id}?message=Tasdiqlandi.", status_code=303)


@inventory_router.post("/{doc_id}/revoke")
async def inventory_revoke(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
    return_to: Optional[str] = Form(None),
):
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "confirmed":
        if return_to == "list":
            return RedirectResponse(url="/qoldiqlar/tovar/hujjat", status_code=303)
        return RedirectResponse(url=f"/inventory/{doc_id}", status_code=303)
    for item in doc.items:
        stock = db.query(Stock).filter(
            Stock.warehouse_id == item.warehouse_id,
            Stock.product_id == item.product_id,
        ).first()
        prev = getattr(item, "previous_quantity", None)
        if stock is not None and prev is not None:
            stock.quantity = prev
            if hasattr(stock, "updated_at"):
                stock.updated_at = datetime.now()
    delete_stock_movements_for_document(db, "StockAdjustmentDoc", doc_id)
    doc.status = "draft"
    db.commit()
    if return_to == "list":
        return RedirectResponse(url="/qoldiqlar/tovar/hujjat?reverted=1", status_code=303)
    return RedirectResponse(url=f"/inventory/{doc_id}/edit?message=Tasdiqlash bekor qilindi.", status_code=303)


@inventory_router.post("/{doc_id}/delete")
async def inventory_delete(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    if doc.status != "draft":
        return RedirectResponse(url="/inventory?message=Faqat qoralama hujjat o'chiriladi.", status_code=303)
    for item in list(doc.items):
        db.delete(item)
    db.delete(doc)
    db.commit()
    return RedirectResponse(url="/inventory?message=Hujjat o'chirildi.", status_code=303)


@inventory_router.get("/{doc_id}/print", response_class=HTMLResponse)
async def inventory_print_page(
    doc_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    doc = db.query(StockAdjustmentDoc).filter(StockAdjustmentDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Hujjat topilmadi")
    warehouse = doc.warehouse
    rows = []
    for item in doc.items:
        prod = item.product
        rows.append({
            "product_name": (prod.name or "") if prod else "",
            "product_code": (prod.code or "") if prod else "",
            "quantity": float(item.quantity or 0),
            "cost_price": float(item.cost_price or 0),
            "sale_price": float(item.sale_price or 0),
        })
    show_tannarx = getattr(current_user, "role", None) == "admin"
    return templates.TemplateResponse("inventory/print.html", {
        "request": request,
        "doc": doc,
        "warehouse": warehouse,
        "rows": rows,
        "show_tannarx": show_tannarx,
        "current_user": current_user,
    })


# ============================================================
# Otxod ishlab chiqarish va Utilizatsiya
# ============================================================

@router.get("/otxod/{warehouse_id}", response_class=HTMLResponse)
async def warehouse_otxod_form(
    warehouse_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Vozvrat omboridagi mahsulotlarni tanlab Otxod Holva ishlab chiqarish."""
    warehouse = db.query(Warehouse).filter(Warehouse.id == warehouse_id).first()
    if not warehouse:
        raise HTTPException(status_code=404, detail="Ombor topilmadi")
    stocks = (
        db.query(Stock)
        .filter(Stock.warehouse_id == warehouse_id, Stock.quantity > 0)
        .options(joinedload(Stock.product).joinedload(Product.unit))
        .all()
    )
    items = []
    for s in stocks:
        if not s.product:
            continue
        unit_name = s.product.unit.name if s.product.unit else "kg"
        unit_code = (s.product.unit.code if s.product.unit else "kg").lower()
        kg_per = _product_kg_per_unit(s.product.name) if unit_code == "dona" else 1.0
        items.append({
            "stock_id": s.id,
            "product_id": s.product.id,
            "product_name": s.product.name,
            "product_code": s.product.code or "",
            "quantity": float(s.quantity or 0),
            "unit": unit_name,
            "unit_code": unit_code,
            "kg_per": kg_per,
        })
    return templates.TemplateResponse("warehouse/otxod_form.html", {
        "request": request,
        "current_user": current_user,
        "warehouse": warehouse,
        "items": items,
        "mode": "otxod",
        "page_title": "Otxod ishlab chiqarish",
    })


@router.post("/otxod/{warehouse_id}/confirm")
async def warehouse_otxod_confirm(
    warehouse_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Tanlangan mahsulotlarni chiqarib, Otxod Holva sifatida kirim qilish."""
    form = await request.form()
    product_ids = form.getlist("product_id")
    quantities = form.getlist("qty")

    if not product_ids:
        return RedirectResponse(
            url=f"/warehouse/otxod/{warehouse_id}?error=" + quote("Kamida bitta mahsulot tanlang."),
            status_code=303,
        )

    # Hujjat raqami
    today_str = datetime.now().strftime("%Y%m%d")
    count = db.query(StockAdjustmentDoc).filter(
        StockAdjustmentDoc.number.like(f"OTX-{today_str}%")
    ).count()
    doc_number = f"OTX-{today_str}-{count + 1:04d}"

    # Hujjat yaratish
    doc = StockAdjustmentDoc(
        number=doc_number,
        date=datetime.now(),
        warehouse_id=warehouse_id,
        user_id=current_user.id if current_user else None,
        status="confirmed",
    )
    db.add(doc)
    db.flush()

    total_kg = 0.0
    for i in range(len(product_ids)):
        try:
            pid = int(product_ids[i])
            qty = float(quantities[i])
        except (ValueError, TypeError, IndexError):
            continue
        if qty <= 0:
            continue

        product = db.query(Product).options(joinedload(Product.unit)).filter(Product.id == pid).first()
        if not product:
            continue

        unit_code = (product.unit.code if product.unit else "kg").lower()
        kg_per = _product_kg_per_unit(product.name) if unit_code == "dona" else 1.0
        kg = qty * kg_per
        total_kg += kg

        # Vozvrat omboridan chiqim
        create_stock_movement(
            db=db,
            warehouse_id=warehouse_id,
            product_id=pid,
            quantity_change=-qty,
            operation_type="otxod_chiqim",
            document_type="StockAdjustmentDoc",
            document_id=doc.id,
            document_number=doc_number,
            user_id=current_user.id if current_user else None,
            note=f"Otxod ishlab chiqarish: {product.name} {qty:.2f} → {kg:.2f} kg",
            created_at=doc.date,
        )

        # Hujjat itemlari
        db.add(StockAdjustmentDocItem(
            doc_id=doc.id,
            product_id=pid,
            warehouse_id=warehouse_id,
            quantity=-qty,
            previous_quantity=0,
            cost_price=product.purchase_price or 0,
            sale_price=product.sale_price or 0,
        ))

    if total_kg <= 0:
        db.rollback()
        return RedirectResponse(
            url=f"/warehouse/otxod/{warehouse_id}?error=" + quote("Miqdor kiritilmadi."),
            status_code=303,
        )

    # Otxod Holva mahsulotini topish (P291, id=305)
    otxod_product = db.query(Product).filter(Product.code == "P291").first()
    if not otxod_product:
        otxod_product = db.query(Product).filter(Product.name.ilike("%otxod holva%")).first()
    if not otxod_product:
        db.rollback()
        return RedirectResponse(
            url=f"/warehouse/otxod/{warehouse_id}?error=" + quote("'Otxod Holva' mahsuloti topilmadi. Avval yarating."),
            status_code=303,
        )

    # Yarim tayyor Maxsulot aralash omboriga (id=6) Otxod Holva kirim
    target_warehouse_id = 6
    target_wh = db.query(Warehouse).filter(Warehouse.id == target_warehouse_id).first()
    if not target_wh:
        target_wh = db.query(Warehouse).filter(Warehouse.name.ilike("%yarim tayyor%aralash%")).first()
        target_warehouse_id = target_wh.id if target_wh else warehouse_id

    create_stock_movement(
        db=db,
        warehouse_id=target_warehouse_id,
        product_id=otxod_product.id,
        quantity_change=+total_kg,
        operation_type="otxod_kirim",
        document_type="StockAdjustmentDoc",
        document_id=doc.id,
        document_number=doc_number,
        user_id=current_user.id if current_user else None,
        note=f"Otxod Holva kirim: {total_kg:.2f} kg ({doc_number})",
        created_at=doc.date,
    )

    db.commit()
    return RedirectResponse(
        url=f"/production?success=otxod&detail=" + quote(f"{doc_number}: {total_kg:.2f} kg Otxod Holva yaratildi"),
        status_code=303,
    )


@router.get("/utilizatsiya/{warehouse_id}", response_class=HTMLResponse)
async def warehouse_utilizatsiya_form(
    warehouse_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Omboridagi mahsulotlarni tanlab utilizatsiya qilish."""
    warehouse = db.query(Warehouse).filter(Warehouse.id == warehouse_id).first()
    if not warehouse:
        raise HTTPException(status_code=404, detail="Ombor topilmadi")
    stocks = (
        db.query(Stock)
        .filter(Stock.warehouse_id == warehouse_id, Stock.quantity > 0)
        .options(joinedload(Stock.product).joinedload(Product.unit))
        .all()
    )
    items = []
    for s in stocks:
        if not s.product:
            continue
        unit_name = s.product.unit.name if s.product.unit else "kg"
        unit_code = (s.product.unit.code if s.product.unit else "kg").lower()
        kg_per = _product_kg_per_unit(s.product.name) if unit_code == "dona" else 1.0
        items.append({
            "stock_id": s.id,
            "product_id": s.product.id,
            "product_name": s.product.name,
            "product_code": s.product.code or "",
            "quantity": float(s.quantity or 0),
            "unit": unit_name,
            "unit_code": unit_code,
            "kg_per": kg_per,
        })
    return templates.TemplateResponse("warehouse/otxod_form.html", {
        "request": request,
        "current_user": current_user,
        "warehouse": warehouse,
        "items": items,
        "mode": "utilizatsiya",
        "page_title": "Utilizatsiya",
    })


@router.post("/utilizatsiya/{warehouse_id}/confirm")
async def warehouse_utilizatsiya_confirm(
    warehouse_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Tanlangan mahsulotlarni utilizatsiya qilib ombordan o'chirish."""
    form = await request.form()
    product_ids = form.getlist("product_id")
    quantities = form.getlist("qty")

    if not product_ids:
        return RedirectResponse(
            url=f"/warehouse/utilizatsiya/{warehouse_id}?error=" + quote("Kamida bitta mahsulot tanlang."),
            status_code=303,
        )

    today_str = datetime.now().strftime("%Y%m%d")
    count = db.query(StockAdjustmentDoc).filter(
        StockAdjustmentDoc.number.like(f"UTL-{today_str}%")
    ).count()
    doc_number = f"UTL-{today_str}-{count + 1:04d}"

    doc = StockAdjustmentDoc(
        number=doc_number,
        date=datetime.now(),
        warehouse_id=warehouse_id,
        user_id=current_user.id if current_user else None,
        status="confirmed",
    )
    db.add(doc)
    db.flush()

    removed_count = 0
    for i in range(len(product_ids)):
        try:
            pid = int(product_ids[i])
            qty = float(quantities[i])
        except (ValueError, TypeError, IndexError):
            continue
        if qty <= 0:
            continue

        product = db.query(Product).filter(Product.id == pid).first()
        if not product:
            continue

        create_stock_movement(
            db=db,
            warehouse_id=warehouse_id,
            product_id=pid,
            quantity_change=-qty,
            operation_type="utilizatsiya",
            document_type="StockAdjustmentDoc",
            document_id=doc.id,
            document_number=doc_number,
            user_id=current_user.id if current_user else None,
            note=f"Utilizatsiya: {product.name} {qty:.2f}",
            created_at=doc.date,
        )

        db.add(StockAdjustmentDocItem(
            doc_id=doc.id,
            product_id=pid,
            warehouse_id=warehouse_id,
            quantity=-qty,
            previous_quantity=0,
            cost_price=product.purchase_price or 0,
            sale_price=product.sale_price or 0,
        ))
        removed_count += 1

    if removed_count == 0:
        db.rollback()
        return RedirectResponse(
            url=f"/warehouse/utilizatsiya/{warehouse_id}?error=" + quote("Miqdor kiritilmadi."),
            status_code=303,
        )

    db.commit()
    return RedirectResponse(
        url=f"/production?success=utilizatsiya&detail=" + quote(f"{doc_number}: {removed_count} ta mahsulot utilizatsiya qilindi"),
        status_code=303,
    )
