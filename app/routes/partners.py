"""
Kontragentlar (partners) — ro'yxat, qo'shish, tahrir, o'chirish, export/import.
"""
import io
import json
import os
from fastapi import APIRouter, Request, Depends, Form, HTTPException, File, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session
import openpyxl

from app.core import templates
from app.models.database import get_db, User, Partner, Order, Purchase, Agent, PriceType
from app.deps import require_auth, require_admin

router = APIRouter(prefix="/partners", tags=["partners"])

_ROUTE_EXPORT_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "route_partner_ids.json")


def _load_route_partner_ids():
    try:
        with open(_ROUTE_EXPORT_PATH, "r", encoding="utf-8") as f:
            payload = json.load(f)
        return set(payload.get("matched_partner_ids") or []), payload.get("generated_at"), payload.get("export_total")
    except (FileNotFoundError, json.JSONDecodeError):
        return set(), None, None


@router.get("", response_class=HTMLResponse)
async def partners_list(
    request: Request,
    type: str = "all",
    sort: str = None,
    order: str = "asc",
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    from app.utils.sort_helpers import parse_sort, apply_sort
    query = db.query(Partner).filter(Partner.is_active == True)
    if type != "all":
        query = query.filter(Partner.type == type)
    # Saralash (whitelist orqali, SQL injection himoyasi)
    sort_allowed = {
        "name": Partner.name,
        "type": Partner.type,
        "phone": Partner.phone,
        "address": Partner.address,
        "balance": Partner.balance,
        "id": Partner.id,
    }
    sort_col, sort_dir = parse_sort(sort, order, sort_allowed, default_col=Partner.id, default_dir="asc")
    query = apply_sort(query, sort_col, sort_dir)
    partners = query.all()
    try:
        from app.config.maps_config import YANDEX_MAPS_API_KEY
        yandex_apikey = YANDEX_MAPS_API_KEY or ""
    except Exception:
        yandex_apikey = ""
    agents = db.query(Agent).filter(Agent.is_active == True).order_by(Agent.full_name).all()
    price_types = db.query(PriceType).filter(PriceType.is_active == True).order_by(PriceType.id).all()

    route_ids, route_generated_at, route_export_total = _load_route_partner_ids()
    not_in_route_count = sum(1 for p in partners if p.id not in route_ids) if route_ids else 0

    return templates.TemplateResponse("partners/list.html", {
        "request": request,
        "partners": partners,
        "agents": agents,
        "price_types": price_types,
        "current_type": type,
        "current_sort": sort or "",
        "current_order": sort_dir,
        "current_user": current_user,
        "page_title": "Kontragentlar",
        "yandex_maps_apikey": yandex_apikey,
        "route_partner_ids": route_ids,
        "route_generated_at": route_generated_at,
        "route_export_total": route_export_total,
        "not_in_route_count": not_in_route_count,
    })


@router.post("/add")
async def partner_add(
    request: Request,
    name: str = Form(...),
    type: str = Form(...),
    phone: str = Form(""),
    address: str = Form(""),
    credit_limit: float = Form(0),
    discount_percent: float = Form(0),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    if credit_limit < 0:
        raise HTTPException(status_code=400, detail="Kredit limit manfiy bo'lishi mumkin emas")
    if discount_percent < 0 or discount_percent > 100:
        raise HTTPException(status_code=400, detail="Chegirma 0-100 oralig'ida bo'lishi kerak")
    existing_by_name = db.query(Partner).filter(Partner.name == name).first()
    if existing_by_name:
        raise HTTPException(status_code=400, detail=f"'{name}' nomli kontragent allaqachon mavjud!")
    if phone and phone.strip():
        existing_by_phone = db.query(Partner).filter(Partner.phone == phone).first()
        if existing_by_phone:
            raise HTTPException(status_code=400, detail=f"'{phone}' telefon raqamli kontragent allaqachon mavjud!")
    partner = Partner(
        name=name,
        code=None,
        type=type,
        phone=phone,
        address=address,
        credit_limit=credit_limit,
        discount_percent=discount_percent,
    )
    db.add(partner)
    db.commit()
    return RedirectResponse(url="/partners", status_code=303)


@router.get("/detail/{partner_id}")
async def partner_detail_json(
    partner_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Kontragent tafsilotlari (JSON) — edit modal uchun."""
    p = db.query(Partner).filter(Partner.id == partner_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Kontragent topilmadi")
    return {
        "id": p.id,
        "name": p.name or "",
        "type": p.type or "customer",
        "phone": p.phone or "",
        "phone2": p.phone2 or "",
        "address": p.address or "",
        "legal_name": p.legal_name or "",
        "contact_person": p.contact_person or "",
        "landmark": p.landmark or "",
        "notes": p.notes or "",
        "visit_day": p.visit_day,
        "category": p.category or "",
        "region": p.region or "",
        "credit_limit": float(p.credit_limit or 0),
        "discount_percent": float(p.discount_percent or 0),
        "price_type_id": p.price_type_id,
        "agent_id": p.agent_id,
        "latitude": p.latitude,
        "longitude": p.longitude,
        "customer_type": p.customer_type or "",
        "sales_channel": p.sales_channel or "",
        "product_categories": p.product_categories or "",
        "inn": p.inn or "",
        "account": p.account or "",
        "bank": p.bank or "",
        "mfo": p.mfo or "",
        "oked": p.oked or "",
        "pinfl": p.pinfl or "",
        "contract_number": p.contract_number or "",
        "contract_date": str(p.contract_date) if p.contract_date else "",
        "extra_agent_ids": [pa.agent_id for pa in (p.partner_agents or []) if pa.agent_id],
        "agents": _build_partner_agents_payload(p),
    }


def _build_partner_agents_payload(p) -> list[dict]:
    """Agent ro'yxati: asosiy (Partner.agent_id) + qo'shimcha (PartnerAgent).
    Eski Partner.agent_id partner_agents'da bo'lmasa, birinchi sifatida qo'shiladi."""
    result = []
    pa_list = list(p.partner_agents or [])
    pa_agent_ids = {pa.agent_id for pa in pa_list if pa.agent_id}
    if p.agent_id and p.agent_id not in pa_agent_ids:
        result.append({
            "agent_id": p.agent_id,
            "visit_type": "weekly",
            "visit_days": str(p.visit_day) if p.visit_day is not None else "",
            "position": 1,
        })
    for pa in sorted(pa_list, key=lambda x: (x.position or 99, x.id)):
        result.append({
            "agent_id": pa.agent_id,
            "visit_type": pa.visit_type or "weekly",
            "visit_days": pa.visit_days or "",
            "position": pa.position or 2,
        })
    return result


@router.post("/edit/{partner_id}")
async def partner_edit(
    request: Request,
    partner_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    form = await request.form()
    name = form.get("name", "").strip()
    type_ = form.get("type", "customer").strip()
    phone = form.get("phone", "").strip()
    address = form.get("address", "").strip()
    try:
        credit_limit = float(form.get("credit_limit", 0) or 0)
    except (ValueError, TypeError):
        credit_limit = 0
    try:
        discount_percent = float(form.get("discount_percent", 0) or 0)
    except (ValueError, TypeError):
        discount_percent = 0
    agent_id_raw = form.get("agent_id", "").strip()

    if not name:
        raise HTTPException(status_code=400, detail="Ism kiritilmadi")
    if credit_limit < 0:
        raise HTTPException(status_code=400, detail="Kredit limit manfiy bo'lishi mumkin emas")
    if discount_percent < 0 or discount_percent > 100:
        raise HTTPException(status_code=400, detail="Chegirma 0-100 oralig'ida bo'lishi kerak")
    partner = db.query(Partner).filter(Partner.id == partner_id).first()
    if not partner:
        raise HTTPException(status_code=404, detail="Kontragent topilmadi")
    existing_by_name = db.query(Partner).filter(Partner.name == name, Partner.id != partner_id).first()
    if existing_by_name:
        raise HTTPException(status_code=400, detail=f"'{name}' nomli kontragent allaqachon mavjud!")
    if phone:
        existing_by_phone = db.query(Partner).filter(Partner.phone == phone, Partner.id != partner_id).first()
        if existing_by_phone:
            raise HTTPException(status_code=400, detail=f"'{phone}' telefon raqamli kontragent allaqachon mavjud!")
    # Asosiy maydonlar
    partner.name = name
    partner.type = type_
    partner.phone = phone
    partner.address = address
    partner.credit_limit = credit_limit
    partner.discount_percent = discount_percent
    price_type_id_raw = form.get("price_type_id", "").strip()
    if price_type_id_raw and price_type_id_raw.isdigit():
        partner.price_type_id = int(price_type_id_raw)
    else:
        partner.price_type_id = None
    try:
        partner.agent_id = int(agent_id_raw) if agent_id_raw and agent_id_raw.isdigit() else None
    except (ValueError, TypeError):
        partner.agent_id = None
    # Qo'shimcha maydonlar
    legal_name = form.get("legal_name", "").strip()
    contact_person = form.get("contact_person", "").strip()
    phone2 = form.get("phone2", "").strip()
    landmark = form.get("landmark", "").strip()
    notes = form.get("notes", "").strip()
    visit_day_raw = form.get("visit_day", "").strip()
    category = form.get("category", "").strip()
    region = form.get("region", "").strip()
    customer_type = form.get("customer_type", "").strip()
    sales_channel = form.get("sales_channel", "").strip()
    product_categories = form.get("product_categories", "").strip()
    inn = form.get("inn", "").strip()
    account = form.get("account", "").strip()
    bank = form.get("bank", "").strip()
    mfo = form.get("mfo", "").strip()
    oked = form.get("oked", "").strip()
    pinfl = form.get("pinfl", "").strip()
    contract_number = form.get("contract_number", "").strip()
    contract_date_raw = form.get("contract_date", "").strip()
    latitude_raw = form.get("latitude", "").strip()
    longitude_raw = form.get("longitude", "").strip()

    partner.legal_name = legal_name or partner.legal_name
    partner.contact_person = contact_person or partner.contact_person
    partner.phone2 = phone2 or partner.phone2
    partner.landmark = landmark or partner.landmark
    partner.notes = notes if notes else partner.notes
    partner.category = category or partner.category
    partner.region = region or partner.region
    partner.customer_type = customer_type or partner.customer_type
    partner.sales_channel = sales_channel or partner.sales_channel
    partner.product_categories = product_categories or partner.product_categories
    partner.inn = inn or partner.inn
    partner.account = account or partner.account
    partner.bank = bank or partner.bank
    partner.mfo = mfo or partner.mfo
    partner.oked = oked or partner.oked
    partner.pinfl = pinfl or partner.pinfl
    partner.contract_number = contract_number or partner.contract_number
    # Tashrif kuni
    if visit_day_raw != "":
        try:
            partner.visit_day = int(visit_day_raw)
        except (ValueError, TypeError):
            partner.visit_day = None
    # GPS
    if latitude_raw:
        try:
            partner.latitude = float(latitude_raw)
        except (ValueError, TypeError):
            pass
    if longitude_raw:
        try:
            partner.longitude = float(longitude_raw)
        except (ValueError, TypeError):
            pass
    # Shartnoma sanasi
    if contract_date_raw:
        from datetime import datetime as dt
        try:
            partner.contract_date = dt.strptime(contract_date_raw, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            pass
    # Agentlar to'liq ro'yxati (per-agent visit_type + visit_days)
    # Form maydonlari: agent_ids[], visit_types[], visit_days_csv[]
    # Birinchi agent (position=1) — asosiy (Partner.agent_id legacy), keyingilari qo'shimcha
    from app.models.database import PartnerAgent
    agent_ids_raw = form.getlist("agent_ids")
    visit_types_raw = form.getlist("visit_types")
    visit_days_raw = form.getlist("visit_days_csv")

    parsed_rows = []
    for i, aid_raw in enumerate(agent_ids_raw):
        try:
            aid = int(aid_raw)
            if aid <= 0:
                continue
        except (ValueError, TypeError):
            continue
        if any(r[0] == aid for r in parsed_rows):
            continue  # duplicate skip
        vtype = (visit_types_raw[i] if i < len(visit_types_raw) else "weekly") or "weekly"
        vdays = (visit_days_raw[i] if i < len(visit_days_raw) else "") or ""
        parsed_rows.append((aid, vtype, vdays, i + 1))

    # Eski partner_agents ni butunlay o'chirib qaytadan yozish
    db.query(PartnerAgent).filter(PartnerAgent.partner_id == partner_id).delete()
    # Asosiy agent — birinchi qatordan (legacy Partner.agent_id va visit_day uchun)
    if parsed_rows:
        first_aid, first_vtype, first_vdays, _ = parsed_rows[0]
        partner.agent_id = first_aid
        first_day_list = [d for d in first_vdays.split(",") if d.strip()]
        try:
            partner.visit_day = int(first_day_list[0]) if first_day_list else None
        except (ValueError, TypeError):
            partner.visit_day = None
        # Qo'shimcha agentlar (position 2+) — partner_agents jadvalga
        for aid, vtype, vdays, pos in parsed_rows[1:]:
            db.add(PartnerAgent(
                partner_id=partner_id,
                agent_id=aid,
                visit_type=vtype,
                visit_days=vdays,
                position=pos,
            ))
    else:
        partner.agent_id = None
        partner.visit_day = None
    db.commit()
    return RedirectResponse(url="/partners", status_code=303)


@router.get("/duplicates", response_class=HTMLResponse)
async def partner_duplicates(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Dublikat partnerlarni topib guruhlash (nom yoki telefon bo'yicha)."""
    def norm_name(s):
        return "".join(c for c in (s or "").lower().strip() if c.isalnum())

    def norm_phone(p):
        d = "".join(c for c in (p or "") if c.isdigit())
        return d[-9:] if len(d) >= 7 else ""

    partners = db.query(Partner).filter(Partner.is_active == True).order_by(Partner.id).all()

    # Union-Find guruhlash: nom (norm) va telefon (norm)
    parent = {}
    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x
    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    for p in partners:
        parent[p.id] = p.id

    key_to_pid = {}
    for p in partners:
        keys = []
        nk = norm_name(p.name)
        if nk and len(nk) >= 3:
            keys.append(("n", nk))
        for ph in (p.phone, p.phone2):
            pk = norm_phone(ph or "")
            if pk:
                keys.append(("p", pk))
        for k in keys:
            if k in key_to_pid:
                union(p.id, key_to_pid[k])
            else:
                key_to_pid[k] = p.id

    # Guruhlarni yig'ish
    from collections import defaultdict
    groups_map = defaultdict(list)
    for p in partners:
        groups_map[find(p.id)].append(p)

    # Faqat 2+ a'zo bo'lgan guruhlar
    duplicate_groups = [v for v in groups_map.values() if len(v) > 1]

    # Filtr: faqat Sales Doctor mijozlari bilan bog'liq guruhlar (?source=sd)
    sd_only = request.query_params.get("source") == "sd"
    if sd_only:
        try:
            from app.services.balance_import_data import SD_BALANCE_DATA
        except ImportError:
            SD_BALANCE_DATA = []
        sd_keys = set()
        for sd in SD_BALANCE_DATA:
            nk = norm_name(sd["name"])
            if nk and len(nk) >= 3:
                sd_keys.add(("n", nk))
            pk = norm_phone(sd["phone"])
            if pk:
                sd_keys.add(("p", pk))
        def group_matches_sd(g):
            for pp in g:
                nk = norm_name(pp.name)
                if ("n", nk) in sd_keys:
                    return True
                for ph in (pp.phone, pp.phone2):
                    pk = norm_phone(ph or "")
                    if pk and ("p", pk) in sd_keys:
                        return True
            return False
        duplicate_groups = [g for g in duplicate_groups if group_matches_sd(g)]

    # Eng katta absolyut balansli partnerni birinchi qo'yish (ehtimol primary)
    for g in duplicate_groups:
        g.sort(key=lambda x: -abs(x.balance or 0))
    duplicate_groups.sort(key=lambda g: -sum(abs(p.balance or 0) for p in g))

    h = ['<!DOCTYPE html><html><head><meta charset="utf-8"><title>Dublikat partnerlar</title>',
         '<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css">',
         '<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css">',
         '<style>body{padding:20px;font-family:sans-serif;}.dup-group{border:1px solid #dee2e6;border-radius:8px;padding:12px;margin-bottom:12px;background:#f8f9fa;}.cand-row{padding:6px 8px;border-radius:4px;margin:2px 0;background:white;display:flex;align-items:center;gap:8px;}.num{font-family:monospace;}.muted{color:#6c757d;}</style>',
         '</head><body><div class="container-fluid">',
         '<div class="mb-3"><a href="/partners" class="btn btn-outline-secondary btn-sm"><i class="bi bi-arrow-left"></i> Kontragentlarga qaytish</a></div>',
         '<h4>Dublikat partnerlar</h4>',
         '<div class="mb-3">',
         f'<a href="/partners/duplicates" class="btn btn-sm {"btn-primary" if not sd_only else "btn-outline-primary"}">Barchasi</a> ',
         f'<a href="/partners/duplicates?source=sd" class="btn btn-sm {"btn-primary" if sd_only else "btn-outline-primary"}">Faqat Sales Doctor mijozlari</a>',
         '</div>',
         f'<p class="text-muted">Topildi: <b>{len(duplicate_groups)}</b> ta guruh, jami <b>{sum(len(g) for g in duplicate_groups)}</b> partner.</p>',
         '<div class="alert alert-info small mb-3">',
         '<b>Belgilar:</b> <span class="badge bg-light text-dark">◉ radio</span> — primary (asosiy partner, qabul qiladi). ',
         '<span class="badge bg-light text-dark">☑ checkbox</span> — merge qilinsinmi? Default barchasi belgilangan. ',
         '<b>Saqlash uchun</b> — qaysi partnerni alohida qoldirmoqchi bo\'lsangiz checkbox\'ni o\'chiring (masalan boshqa filial bo\'lsa).',
         '</div>',
         '<p class="alert alert-warning small"><b>Ehtiyot:</b> merge qaytarib bo\'lmaydi. Har guruhni diqqat bilan tekshiring.</p>']

    csrf_token = getattr(request.state, "csrf_token", "") or ""
    for gi, group in enumerate(duplicate_groups):
        total_bal = sum(p.balance or 0 for p in group)
        h.append('<div class="dup-group">')
        h.append(f'<div><b>Guruh {gi+1}</b> &middot; {len(group)} partner &middot; jami balans <span class="num">{total_bal:,.0f}</span></div>')
        h.append(f'<form method="post" action="/partners/merge" class="mt-2" onsubmit="return confirm(\'Guruh {gi+1}: tanlangan primary\\\'ga merge qilamizmi? Qaytarib bo\\\'lmaydi.\');">')
        h.append(f'<input type="hidden" name="csrf_token" value="{csrf_token}">')
        for i, p in enumerate(group):
            checked = "checked" if i == 0 else ""
            phones = [p.phone, p.phone2]
            ph_str = ", ".join(x for x in phones if x) or "—"
            bal_color = "#dc3545" if (p.balance or 0) > 0 else ("#198754" if (p.balance or 0) < 0 else "#6c757d")
            h.append(f'<div class="cand-row">')
            h.append(f'<input type="radio" name="primary_id" value="{p.id}" {checked} title="Primary (asosiy)">')
            h.append(f'<input type="checkbox" name="member_id" value="{p.id}" checked title="Merge qilinsinmi?" class="merge-cb">')
            h.append(f'<div class="flex-grow-1"><b>#{p.id} {p.name}</b> &middot; <span class="muted">tel</span> {ph_str} &middot; <span class="muted">manzil</span> {(p.address or "-")[:25]} &middot; <span class="muted">balans</span> <b class="num" style="color:{bal_color}">{p.balance or 0:,.0f}</b> &middot; <span class="muted">tip</span> {p.type or "-"}</div>')
            h.append('</div>')
        h.append('<button type="submit" class="btn btn-warning btn-sm mt-2">⚐ Merge qilish (tanlangan primary)</button>')
        h.append('</form></div>')

    if not duplicate_groups:
        h.append('<div class="alert alert-success">Dublikat topilmadi!</div>')

    h.append('</div></body></html>')
    return HTMLResponse("".join(h))


@router.post("/merge")
async def partner_merge(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Tanlangan primary partnerga qolganlarni merge qilish."""
    from sqlalchemy import inspect, text
    from app.models.database import PartnerAgent

    form = await request.form()
    try:
        primary_id = int(form.get("primary_id") or 0)
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="primary_id noto'g'ri")
    if primary_id <= 0:
        raise HTTPException(status_code=400, detail="primary tanlanmagan")

    member_ids_raw = form.getlist("member_id")
    other_ids = []
    for x in member_ids_raw:
        try:
            mid = int(x)
            if mid != primary_id and mid > 0:
                other_ids.append(mid)
        except (ValueError, TypeError):
            continue
    if not other_ids:
        return RedirectResponse(url="/partners/duplicates?msg=hech+nima+merge+qilinmadi", status_code=303)

    primary = db.query(Partner).filter(Partner.id == primary_id).first()
    others = db.query(Partner).filter(Partner.id.in_(other_ids)).all()
    if not primary or len(others) != len(other_ids):
        raise HTTPException(status_code=404, detail="Partner topilmadi")

    # 2) partner_agents — UniqueConstraint(partner_id, agent_id) — alohida ishlatish
    existing_agents = {pa.agent_id for pa in db.query(PartnerAgent).filter(PartnerAgent.partner_id == primary_id).all()}
    for o_id in other_ids:
        for pa in db.query(PartnerAgent).filter(PartnerAgent.partner_id == o_id).all():
            if pa.agent_id in existing_agents:
                db.delete(pa)
            else:
                pa.partner_id = primary_id
                existing_agents.add(pa.agent_id)

    # 3) Boshqa barcha jadvallar — generic UPDATE foreign key
    db.flush()
    inspector = inspect(db.bind)
    other_ids_csv = ",".join(str(x) for x in other_ids)
    for table_name in inspector.get_table_names():
        if table_name in ("partners", "partner_agents"):
            continue
        try:
            fks = inspector.get_foreign_keys(table_name)
        except Exception:
            continue
        for fk in fks:
            if fk.get("referred_table") == "partners":
                col = fk["constrained_columns"][0]
                db.execute(text(
                    f'UPDATE {table_name} SET {col} = {primary_id} WHERE {col} IN ({other_ids_csv})'
                ))

    # 4) Eski partnerlar — soft-delete
    for o in others:
        o.is_active = False
        if not (o.name or "").startswith("[→#"):
            o.name = f"[→#{primary_id}] {o.name}"

    # Balansni qayta hisoblash — barcha hujjatlar primary'ga ko'chgach
    from app.services.partner_balance_service import recompute_partner_balance
    db.flush()
    recompute_partner_balance(db, primary_id, reason="partner_merge", ref=str(primary_id),
                              actor=current_user.username if current_user else None)
    db.commit()
    return RedirectResponse(url=f"/partners/duplicates?merged=1&primary={primary_id}", status_code=303)


@router.post("/bulk-assign-agent")
async def partner_bulk_assign_agent(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Tanlangan kontragentlarni ommaviy tahrirlash (agent, tashrif kuni, kategoriya, hudud)"""
    form = await request.form()
    partner_ids = form.getlist("partner_ids")
    partner_ids = [int(x) for x in partner_ids if str(x).isdigit()]
    if not partner_ids:
        return RedirectResponse(url="/partners", status_code=303)

    updates = {}
    agent_id = form.get("agent_id", "").strip()
    visit_day = form.get("visit_day", "").strip()
    category = form.get("category", "").strip()
    region = form.get("region", "").strip()

    if agent_id:
        if agent_id == "__clear__":
            updates[Partner.agent_id] = None
        else:
            agent = db.query(Agent).filter(Agent.id == int(agent_id)).first()
            if agent:
                updates[Partner.agent_id] = agent.id
    if visit_day:
        updates[Partner.visit_day] = None if visit_day == "__clear__" else visit_day
    if category:
        updates[Partner.category] = None if category == "__clear__" else category
    if region:
        updates[Partner.region] = None if region == "__clear__" else region

    if updates:
        db.query(Partner).filter(Partner.id.in_(partner_ids)).update(updates, synchronize_session="fetch")
        db.commit()
    return RedirectResponse(url="/partners", status_code=303)


@router.post("/delete/{partner_id}")
async def partner_delete(
    partner_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    partner = db.query(Partner).filter(Partner.id == partner_id).first()
    if not partner:
        raise HTTPException(status_code=404, detail="Kontragent topilmadi")
    has_orders = db.query(Order).filter(Order.partner_id == partner_id).first()
    has_purchases = db.query(Purchase).filter(Purchase.partner_id == partner_id).first()
    if has_orders or has_purchases:
        raise HTTPException(
            status_code=400,
            detail="Bu kontragent bilan bog'liq buyurtmalar yoki kirimlar mavjud. O'chirish mumkin emas.",
        )
    db.delete(partner)
    db.commit()
    return RedirectResponse(url="/partners", status_code=303)


@router.get("/export")
async def export_partners(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    partners = db.query(Partner).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Partners"
    ws.append(["ID", "Kod", "Nomi", "Turi", "Telefon", "Manzil", "Kredit Limit", "Chegirma %"])
    for p in partners:
        ws.append([p.id, p.code, p.name, p.type, p.phone, p.address, p.credit_limit, p.discount_percent])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=kontragentlar.xlsx"},
    )


@router.get("/template")
async def template_partners():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append([
        "Nomi", "Turi", "Telefon", "Manzil", "Kredit Limit", "Chegirma %",
        "Agent", "Tashrif kuni", "Latitude", "Longitude", "Izoh"
    ])
    ws.append(["Mijoz MCHJ", "customer", "+998901234567", "Toshkent", 1000000, 0, "Akbarjon", 1, 41.311081, 69.240562, "Demo yozuv"])
    ws.append(["Yetkazib Beruvchi", "supplier", "+998909876543", "Samarqand", 0, 0, "", "", "", "", ""])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=kontragent_andoza.xlsx"},
    )


@router.post("/import")
async def import_partners(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    def _parse_visit_day(value):
        if value in (None, ""):
            return None
        if isinstance(value, (int, float)):
            iv = int(value)
            return iv if iv in (0, 1, 2, 3, 4, 5, 6) else None
        raw = str(value).strip().lower()
        day_map = {
            "0": 0, "yak": 0, "yakshanba": 0, "воскресенье": 0,
            "1": 1, "dush": 1, "dushanba": 1, "понедельник": 1,
            "2": 2, "sesh": 2, "seshanba": 2, "вторник": 2,
            "3": 3, "chor": 3, "chorshanba": 3, "среда": 3,
            "4": 4, "pay": 4, "payshanba": 4, "четверг": 4,
            "5": 5, "juma": 5, "пятница": 5,
            "6": 6, "shan": 6, "shanba": 6, "суббота": 6,
        }
        return day_map.get(raw)

    def _parse_float(value):
        if value in (None, ""):
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Fayl hajmi 5MB dan oshmasligi kerak")
    if contents[:2] != b"PK":
        raise HTTPException(status_code=400, detail="Fayl .xlsx formati bo'lishi kerak")
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        if not row[0]:
            continue
        values = list(row) + [None] * max(0, 11 - len(row))
        name, type_, phone, address, credit_limit, discount_percent, agent_name, visit_day, latitude, longitude, notes = values[:11]
        if name is None:
            continue
        partner = db.query(Partner).filter(Partner.name == name).first()
        agent_id = None
        if agent_name not in (None, ""):
            agent = (
                db.query(Agent)
                .filter(Agent.full_name.ilike(str(agent_name).strip()))
                .first()
            )
            if agent:
                agent_id = agent.id
        if not partner:
            count = db.query(Partner).count()
            code = f"P{count + 1:04d}"
            partner = Partner(
                code=code,
                name=name,
                type=type_ or "customer",
                phone=phone or "",
                address=address or "",
                credit_limit=credit_limit or 0,
                discount_percent=discount_percent or 0,
                agent_id=agent_id,
                visit_day=_parse_visit_day(visit_day),
                latitude=_parse_float(latitude),
                longitude=_parse_float(longitude),
                notes=str(notes).strip() if notes not in (None, "") else "",
            )
            db.add(partner)
        else:
            if phone is not None:
                partner.phone = phone
            if address is not None:
                partner.address = address
            if credit_limit is not None:
                partner.credit_limit = credit_limit
            if discount_percent is not None:
                partner.discount_percent = discount_percent
            if agent_id is not None:
                partner.agent_id = agent_id
            parsed_visit_day = _parse_visit_day(visit_day)
            if parsed_visit_day is not None:
                partner.visit_day = parsed_visit_day
            parsed_lat = _parse_float(latitude)
            parsed_lng = _parse_float(longitude)
            if parsed_lat is not None:
                partner.latitude = parsed_lat
            if parsed_lng is not None:
                partner.longitude = parsed_lng
            if notes not in (None, ""):
                partner.notes = str(notes).strip()
        db.commit()
    return RedirectResponse(url="/partners", status_code=303)
