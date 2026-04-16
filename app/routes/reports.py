"""
Hisobotlar — savdo, qoldiq, qarzdorlik va Excel export.
"""
import io
import json
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Request, Depends, File, Form, UploadFile, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, or_, and_
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.core import templates
from app.models.database import get_db, Order, OrderItem, Stock, StockMovement, Product, Partner, Warehouse, User, Production, Recipe, StockAdjustmentDoc, StockAdjustmentDocItem, Employee, Purchase, PurchaseItem, WarehouseTransfer, Payment, ProductPrice, ExpenseDoc, ExpenseDocItem, ExpenseType, Salary, PartnerBalanceDoc, PartnerBalanceDocItem, AuditLog
from app.deps import get_current_user, require_auth, require_admin
from app.utils.user_scope import get_warehouses_for_user
from app.utils.rate_limit import check_api_rate_limit

router = APIRouter(prefix="/reports", tags=["reports"])


def _check_export_rate_limit(request: Request):
    """Export endpointlar uchun rate limit (daqiqada max 60)."""
    if check_api_rate_limit(request):
        raise HTTPException(status_code=429, detail="Juda ko'p so'rov. Biroz kuting.")


def get_allowed_report_types(user: User) -> list:
    """Foydalanuvchiga ruxsat berilgan hisobot turlarini qaytaradi."""
    if not user:
        return []
    # Admin uchun barcha hisobotlar
    if user.role == "admin":
        return ["sales", "stock", "debts", "production", "employees", "profit", "partner_reconciliation"]
    # allowed_sections bo'sh yoki None bo'lsa, hech narsa ko'rsatilmaydi
    if not user.allowed_sections:
        return []
    try:
        sections = json.loads(user.allowed_sections) if isinstance(user.allowed_sections, str) else user.allowed_sections
        if not isinstance(sections, list):
            return []
        # allowed_sections ichida "reports_sales", "reports_stock" kabi formatda bo'lishi mumkin
        report_types = []
        for s in sections:
            if isinstance(s, str) and s.startswith("reports_"):
                report_type = s.replace("reports_", "")
                if report_type in ["sales", "stock", "debts", "production", "employees", "profit", "partner_reconciliation"]:
                    report_types.append(report_type)
        return report_types
    except (json.JSONDecodeError, TypeError, AttributeError):
        return []


@router.get("", response_class=HTMLResponse)
async def reports_index(request: Request, current_user: User = Depends(require_auth)):
    """Hisobotlar bosh sahifasi"""
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    allowed_types = get_allowed_report_types(current_user)
    return templates.TemplateResponse("reports/index.html", {
        "request": request,
        "page_title": "Hisobotlar",
        "current_user": current_user,
        "allowed_report_types": allowed_types,
    })


@router.get("/form", response_class=HTMLResponse)
async def reports_form(request: Request, db: Session = Depends(get_db), current_user: User = Depends(require_auth)):
    """Hisobotlar formasi — hisobot turi va filtrlarni tanlash"""
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    today = datetime.now()
    start_date = today.replace(day=1).strftime("%Y-%m-%d")
    end_date = today.strftime("%Y-%m-%d")
    warehouses = get_warehouses_for_user(db, current_user)
    allowed_types = get_allowed_report_types(current_user)
    return templates.TemplateResponse("reports/form.html", {
        "request": request,
        "page_title": "Hisobotlar formasi",
        "current_user": current_user,
        "start_date": start_date,
        "end_date": end_date,
        "warehouses": warehouses,
        "allowed_report_types": allowed_types,
    })


@router.get("/sales", response_class=HTMLResponse)
async def report_sales(
    request: Request,
    start_date: str = None,
    end_date: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    if not start_date:
        start_date = datetime.now().replace(day=1).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now().strftime("%Y-%m-%d")
    orders = db.query(Order).filter(
        Order.type == "sale",
        Order.date >= start_date,
        Order.date <= end_date + " 23:59:59",
    ).all()
    total = sum(o.total or 0 for o in orders)
    return templates.TemplateResponse("reports/sales.html", {
        "request": request,
        "orders": orders,
        "total": total,
        "start_date": start_date,
        "end_date": end_date,
        "page_title": "Savdo hisoboti",
        "current_user": current_user,
    })


@router.get("/sales/export")
async def report_sales_export(
    request: Request,
    start_date: str = None,
    end_date: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    _check_export_rate_limit(request)
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    if not start_date:
        start_date = datetime.now().replace(day=1).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now().strftime("%Y-%m-%d")
    orders = db.query(Order).filter(
        Order.type == "sale",
        Order.date >= start_date,
        Order.date <= end_date + " 23:59:59",
    ).order_by(Order.date.desc()).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Savdo"
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A1"] = "Savdo hisoboti"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Davr: {start_date} — {end_date}"
    ws.append(["№", "Sana", "Buyurtma", "Mijoz", "Jami", "Holat"])
    for c in range(1, 7):
        ws.cell(row=4, column=c).fill = header_fill
        ws.cell(row=4, column=c).font = Font(bold=True, color="FFFFFF")
    for i, o in enumerate(orders, 1):
        ws.append([
            i,
            o.date.strftime("%d.%m.%Y %H:%M") if o.date else "",
            o.number or "",
            o.partner.name if o.partner else "",
            float(o.total or 0),
            o.status or "",
        ])
    total = sum(o.total or 0 for o in orders)
    ws.append(["", "", "", "JAMI:", total, ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=savdo_{start_date}_{end_date}.xlsx"},
    )


@router.get("/stock", response_class=HTMLResponse)
async def report_stock(
    request: Request,
    warehouse_id: str = None,
    report_date: str = None,
    low: int = None,
    merged: int = None,
    cleared: int = None,
    recalculated: int = None,
    cleanup: int = None,
    msg: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Qoldiq hisoboti — ombor bo'yicha qoldiqlar. report_date berilsa shu sanadagi qoldiq (harakatlar bo'yicha)."""
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    warehouses = get_warehouses_for_user(db, current_user)
    wh_ids = [w.id for w in warehouses]
    wh_id = None
    if warehouse_id is not None and str(warehouse_id).strip() != "":
        try:
            wid = int(warehouse_id)
            if not wh_ids or wid in wh_ids:
                wh_id = wid
        except (ValueError, TypeError):
            pass
    if report_date and str(report_date).strip():
        values = _stock_report_as_of_date(db, report_date.strip()[:10], wh_id)
    else:
        values = _stock_report_filtered(db, wh_id)
    stocks = [{"warehouse": v["warehouse"], "product": v["product"], "quantity": v["quantity"]} for v in values]
    if low:
        filtered = []
        for v in stocks:
            try:
                qty = float(v.get("quantity") or 0)
                p = v.get("product")
                min_s = float(getattr(p, "min_stock", 0) or 0) if p else 0
                if min_s > 0 and qty < min_s:
                    filtered.append(v)
                elif min_s <= 0 and 0 < qty < 10:
                    filtered.append(v)
            except Exception:
                continue
        stocks = filtered
    total_sum = 0.0
    for v in stocks:
        qty = float(v.get("quantity") or 0)
        price = float(getattr(v.get("product"), "purchase_price", None) or 0)
        total_sum += qty * price
    today_str = datetime.now().strftime("%Y-%m-%d")
    return templates.TemplateResponse("reports/stock.html", {
        "request": request,
        "stocks": stocks,
        "total_sum": total_sum,
        "warehouses": warehouses,
        "selected_warehouse_id": wh_id,
        "report_date": (report_date or "").strip()[:10] or None,
        "today": today_str,
        "merged": merged,
        "cleared": cleared,
        "recalculated": recalculated,
        "cleanup": cleanup,
        "msg": msg,
        "page_title": "Qoldiq hisoboti",
        "current_user": current_user,
        "show_tannarx": (getattr(current_user, "role", None) if current_user else None) == "admin",
        "only_low": 1 if low else 0,
    })


def _document_type_label(doc_type: str) -> str:
    """Hujjat turi uchun o'qiladigan nom"""
    labels = {
        "Purchase": "Kirim (sotib olish)",
        "Production": "Ishlab chiqarish",
        "WarehouseTransfer": "Ombordan omborga",
        "StockAdjustmentDoc": "Qoldiq tuzatish",
        "Sale": "Sotuv",
        "SaleReturn": "Qaytish",
    }
    return labels.get(doc_type, doc_type or "—")


def _document_url(doc_type: str, doc_id: int) -> str:
    """Hujjat turi va ID bo'yicha ko'rish havolasi"""
    if doc_type == "Purchase":
        return f"/purchases/edit/{doc_id}"
    if doc_type == "Production":
        return f"/production/{doc_id}/materials"
    if doc_type == "WarehouseTransfer":
        return f"/warehouse/transfers/{doc_id}"
    if doc_type == "StockAdjustmentDoc":
        return f"/qoldiqlar/tovar/hujjat/{doc_id}"
    if doc_type == "Sale":
        return f"/sales/edit/{doc_id}"
    return "#"


@router.get("/stock/no-history", response_class=HTMLResponse)
async def report_stock_no_history(
    request: Request,
    warehouse_id: int = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Faqat qoldigi bor, lekin harakat tarixi yo'q mahsulotlar — tanlangan ombor bo'yicha (qoldiq qayerdandur paydo bo'lib qolganlar)."""
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    warehouses = db.query(Warehouse).filter(Warehouse.is_active == True).order_by(Warehouse.name).all()
    rows = []
    warehouse = None
    if warehouse_id:
        warehouse = db.query(Warehouse).filter(Warehouse.id == warehouse_id).first()
        if warehouse:
            # Ombor bo'yicha barcha Stock (qoldiq > 0), product_id bo'yicha yig'indisi
            from collections import defaultdict
            by_product = defaultdict(float)
            for s in db.query(Stock).filter(Stock.warehouse_id == warehouse_id).all():
                q = float(s.quantity or 0)
                if q > 0:
                    by_product[s.product_id] += q
            product_ids = list(by_product.keys())
            if product_ids:
                products_by_id = {p.id: p for p in db.query(Product).filter(Product.id.in_(product_ids)).all()}
                # Harakat tarixi yo'q (warehouse_id, product_id) larni qoldiramiz
                for pid in product_ids:
                    has_movement = (
                        db.query(StockMovement)
                        .filter(
                            StockMovement.warehouse_id == warehouse_id,
                            StockMovement.product_id == pid,
                        )
                        .limit(1)
                        .first()
                    )
                    if not has_movement and pid in products_by_id:
                        rows.append({
                            "product": products_by_id[pid],
                            "product_id": pid,
                            "quantity": by_product[pid],
                        })
                rows.sort(key=lambda r: ((r["product"].name or "").lower(), r["product_id"]))
    return templates.TemplateResponse("reports/stock_no_history.html", {
        "request": request,
        "warehouses": warehouses,
        "warehouse": warehouse,
        "selected_warehouse_id": warehouse_id,
        "rows": rows,
        "page_title": "Tarixsiz qoldiqlar",
        "current_user": current_user,
    })


@router.post("/stock/no-history/create-inventory", response_class=RedirectResponse)
async def report_stock_no_history_create_inventory(
    request: Request,
    warehouse_id: int = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Tarixsiz qoldiqlar ro'yxatidagi mahsulotlar uchun inventarizatsiya hujjati yaratib, tahrirga yo'naltiradi. Haqiqiy qoldiqni 0 qilsangiz — qoldiq tozalanadi, tasdiqlanganda harakat tarixi yoziladi."""
    from urllib.parse import quote
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    if not warehouse_id:
        return RedirectResponse(url="/reports/stock/no-history?error=" + quote("Ombor tanlang."), status_code=303)
    warehouse = db.query(Warehouse).filter(Warehouse.id == warehouse_id).first()
    if not warehouse:
        return RedirectResponse(url="/reports/stock/no-history?error=" + quote("Ombor topilmadi."), status_code=303)
    # Tarixsiz qoldiqlar ro'yxati (qoldiq > 0, StockMovement yo'q)
    from collections import defaultdict
    by_product = defaultdict(float)
    for s in db.query(Stock).filter(Stock.warehouse_id == warehouse_id).all():
        q = float(s.quantity or 0)
        if q > 0:
            by_product[s.product_id] += q
    product_ids = [pid for pid in by_product.keys()]
    no_history_ids = []
    for pid in product_ids:
        has_movement = (
            db.query(StockMovement)
            .filter(StockMovement.warehouse_id == warehouse_id, StockMovement.product_id == pid)
            .limit(1)
            .first()
        )
        if not has_movement:
            no_history_ids.append(pid)
    if not no_history_ids:
        return RedirectResponse(
            url="/reports/stock/no-history?warehouse_id=" + str(warehouse_id) + "&message=" + quote("Bu omborda tarixsiz qoldiq yo'q."),
            status_code=303,
        )
    # Qoralama inventarizatsiya hujjati
    today = datetime.now()
    doc = StockAdjustmentDoc(
        number="INV-PENDING",
        date=today,
        warehouse_id=warehouse_id,
        user_id=current_user.id,
        status="draft",
        total_tannarx=0,
        total_sotuv=0,
    )
    db.add(doc)
    db.flush()
    doc.number = f"INV-PENDING-{doc.id}"
    # Har bir tarixsiz mahsulot uchun qator (haqiqiy = hisobiy — keyin 0 qilish mumkin)
    products = {p.id: p for p in db.query(Product).filter(Product.id.in_(no_history_ids)).all()}
    for pid in no_history_ids:
        if pid not in products:
            continue
        product = products[pid]
        qty = by_product[pid]
        cost = float(product.purchase_price or 0)
        sale = float(product.sale_price or 0)
        if (product.sale_price or 0) <= 0:
            pp = db.query(ProductPrice).filter(ProductPrice.product_id == pid).first()
            if pp:
                sale = float(getattr(pp, "sale_price", 0) or 0)
        db.add(StockAdjustmentDocItem(
            doc_id=doc.id,
            product_id=pid,
            warehouse_id=warehouse_id,
            quantity=qty,
            cost_price=cost,
            sale_price=sale,
        ))
    db.commit()
    msg = quote(f"Inventarizatsiya yaratildi ({len(no_history_ids)} ta mahsulot). Haqiqiy qoldiqni 0 qilsangiz — qoldiq tozalanadi.")
    return RedirectResponse(url=f"/inventory/{doc.id}/edit?message={msg}", status_code=303)


def _load_movement_doc_filters(db: Session, movements: list) -> tuple:
    """Stock source: tasdiqlangan adjustment va completed production ID'larini yuklab qaytaradi.
    Returns: (confirmed_adj_ids: set, completed_production_ids: set, adj_doc_dates: dict)"""
    adj_ids = [m.document_id for m in movements if (m.document_type or "") == "StockAdjustmentDoc" and m.document_id]
    adj_doc_dates = {}
    confirmed_adj_ids = set()
    if adj_ids:
        for doc in db.query(StockAdjustmentDoc).filter(
            StockAdjustmentDoc.id.in_(adj_ids),
            StockAdjustmentDoc.status == "confirmed",
        ).all():
            adj_doc_dates[doc.id] = doc.date
            confirmed_adj_ids.add(doc.id)
    prod_ids = [m.document_id for m in movements if (m.document_type or "") == "Production" and m.document_id]
    completed_production_ids = set()
    if prod_ids:
        for p in db.query(Production).filter(Production.id.in_(prod_ids)).all():
            if getattr(p, "status", None) == "completed":
                completed_production_ids.add(p.id)
    return confirmed_adj_ids, completed_production_ids, adj_doc_dates


def _apply_document_dates(db: Session, rows: list) -> None:
    """Purchase/Sale/Transfer uchun hujjat sanasini row['date'] ga o'rnatadi (mutates rows)."""
    purchase_ids = [r["document_id"] for r in rows if (r.get("document_type") or "") == "Purchase" and r.get("document_id")]
    sale_ids = [r["document_id"] for r in rows if (r.get("document_type") or "") in ("Sale", "SaleReturn", "SaleReturnRevert") and r.get("document_id")]
    transfer_ids = [r["document_id"] for r in rows if (r.get("document_type") or "") == "WarehouseTransfer" and r.get("document_id")]
    purchases_by_id = {p.id: p for p in db.query(Purchase).filter(Purchase.id.in_(purchase_ids)).all()} if purchase_ids else {}
    orders_by_id = {o.id: o for o in db.query(Order).filter(Order.id.in_(sale_ids)).all()} if sale_ids else {}
    transfers_by_id = {t.id: t for t in db.query(WarehouseTransfer).filter(WarehouseTransfer.id.in_(transfer_ids)).all()} if transfer_ids else {}
    for r in rows:
        doc_type = r.get("document_type") or ""
        doc_id = r.get("document_id")
        if doc_type == "Purchase" and doc_id and doc_id in purchases_by_id and purchases_by_id[doc_id].date:
            r["date"] = purchases_by_id[doc_id].date.strftime("%d.%m.%Y %H:%M")
        elif doc_type in ("Sale", "SaleReturn", "SaleReturnRevert") and doc_id and doc_id in orders_by_id and orders_by_id[doc_id].date:
            r["date"] = orders_by_id[doc_id].date.strftime("%d.%m.%Y %H:%M")
        elif doc_type == "WarehouseTransfer" and doc_id and doc_id in transfers_by_id and transfers_by_id[doc_id].date:
            r["date"] = transfers_by_id[doc_id].date.strftime("%d.%m.%Y %H:%M")


def _check_production_quantity_mismatch(db: Session, rows: list) -> None:
    """Production qatorlari uchun hujjatdagi miqdor vs harakatdagi miqdorni solishtirish (mutates rows).
    Production sanasini ham ustun qilib qo'yadi."""
    from sqlalchemy.orm import joinedload as _jl
    from app.utils.production_order import production_output_quantity_for_stock
    prod_ids = [r["document_id"] for r in rows if (r.get("document_type") or "") == "Production" and r.get("document_id")]
    if not prod_ids:
        return
    productions_by_id = {
        p.id: p for p in db.query(Production).options(_jl(Production.recipe)).filter(Production.id.in_(prod_ids)).all()
    }
    for r in rows:
        if (r.get("document_type") or "") != "Production" or not r.get("document_id"):
            continue
        prod = productions_by_id.get(r["document_id"])
        if not prod:
            continue
        if prod.date:
            r["date"] = prod.date.strftime("%d.%m.%Y %H:%M")
        if not prod.recipe:
            continue
        expected = production_output_quantity_for_stock(db, prod, prod.recipe)
        change = r.get("quantity_change") or 0
        if abs(expected - change) > 0.001:
            r["quantity_mismatch"] = True
            r["document_quantity"] = expected


@router.get("/stock/source", response_class=HTMLResponse)
async def report_stock_source(
    request: Request,
    warehouse_id: int = None,
    product_id: int = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Berilgan ombor + mahsulot uchun qoldiq manbai — barcha harakatlar (qaysi hujjatdan kirgan/chiqqan)."""
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    if not warehouse_id or not product_id:
        return RedirectResponse(url="/reports/stock", status_code=303)
    warehouse = db.query(Warehouse).filter(Warehouse.id == warehouse_id).first()
    product = db.query(Product).filter(Product.id == product_id).first()
    if not warehouse or not product:
        return RedirectResponse(url="/reports/stock", status_code=303)
    movements = (
        db.query(StockMovement)
        .filter(
            StockMovement.warehouse_id == warehouse_id,
            StockMovement.product_id == product_id,
        )
        .order_by(StockMovement.created_at.asc())
        .all()
    )
    confirmed_adj_ids, completed_production_ids, doc_dates = _load_movement_doc_filters(db, movements)
    # Bir xil hujjat (document_type, document_id) uchun bitta qator — dublikat harakatlar birlashtiriladi
    rows = []
    seen_doc = set()  # (document_type, document_id)
    transfers_by_id = {}  # keyed by document_id for wrong_warehouse check
    for m in movements:
        # Qoldiq tuzatish: faqat tasdiqlangan hujjat ko'rinsin
        if (m.document_type or "") == "StockAdjustmentDoc" and m.document_id not in confirmed_adj_ids:
            continue
        # Ishlab chiqarish: tasdiqni bekor qilingan yoki qoralama — harakatni ko'rsatma (qoldiq allaqachon tuzatilgan)
        if (m.document_type or "") == "Production" and m.document_id and m.document_id not in completed_production_ids:
            continue
        key = (m.document_type or "", m.document_id)
        if key in seen_doc:
            continue
        seen_doc.add(key)
        if (m.document_type or "") == "StockAdjustmentDoc" and m.document_id and m.document_id in doc_dates and doc_dates[m.document_id]:
            display_date = doc_dates[m.document_id].strftime("%d.%m.%Y %H:%M")
        else:
            display_date = m.created_at.strftime("%d.%m.%Y %H:%M") if m.created_at else "—"
        row = {
            "date": display_date,
            "document_type": m.document_type or "",
            "document_type_label": _document_type_label(m.document_type or ""),
            "document_number": m.document_number or f"{m.document_type}-{m.document_id}",
            "document_id": m.document_id,
            "document_url": _document_url(m.document_type or "", m.document_id),
            "quantity_change": float(m.quantity_change or 0),
            "quantity_after": float(m.quantity_after or 0),
            "movement_id": getattr(m, "id", None),
            "_sort_at": m.created_at if m.created_at else datetime.min,
        }
        # Ombordan omborga: hujjatda bu ombor ko'rsatilmagan bo'lsa ogohlantirish
        if (m.document_type or "") == "WarehouseTransfer" and m.document_id:
            if m.document_id not in transfers_by_id:
                t = db.query(WarehouseTransfer).filter(WarehouseTransfer.id == m.document_id).first()
                transfers_by_id[m.document_id] = t
            t = transfers_by_id.get(m.document_id)
            if t:
                if (m.quantity_change or 0) > 0:  # transfer_in — hujjatda "omborga" bu ombor bo'lishi kerak
                    if t.to_warehouse_id != warehouse_id:
                        row["wrong_warehouse"] = True
                else:  # transfer_out — hujjatda "ombordan" bu ombor bo'lishi kerak
                    if t.from_warehouse_id != warehouse_id:
                        row["wrong_warehouse"] = True
        rows.append(row)
    # Hujjat sanalari va Production miqdor mos kelishini tekshirish (helperlar)
    _apply_document_dates(db, rows)
    _check_production_quantity_mismatch(db, rows)
    # Tartib: haqiqiy vaqt bo'yicha (created_at), matn sanasi emas — oxirgi qator = hozirgi qoldiq
    rows.sort(key=lambda r: (r.get("_sort_at") or datetime.min, r.get("document_id") or 0))
    # Qoldiq (harakatdan keyin) — ketma-ket yig'indi (sana tartibida)
    if rows:
        balance = 0.0
        for r in rows:
            balance += r["quantity_change"]
            r["quantity_after"] = round(balance, 6)
    for r in rows:
        r.pop("_sort_at", None)
    current_stock = db.query(Stock).filter(
        Stock.warehouse_id == warehouse_id,
        Stock.product_id == product_id,
    ).all()
    current_qty = sum(float(s.quantity or 0) for s in current_stock)
    # Dona mahsulotlar uchun ko'rsatishda butun son (216, 6), boshqalar uchun 3 xona kasr
    _unit_str = ((getattr(product, "unit", None) and (product.unit.name or "") or "") + " " + (getattr(product, "unit", None) and (product.unit.code or "") or "")).lower()
    is_dona = product and "dona" in _unit_str
    msg = request.query_params.get("msg") or ""
    error = request.query_params.get("error") or ""
    removed = request.query_params.get("removed")
    return templates.TemplateResponse("reports/stock_source.html", {
        "request": request,
        "warehouse": warehouse,
        "product": product,
        "movements": rows,
        "current_qty": current_qty,
        "is_dona": is_dona,
        "page_title": "Qoldiq manbai",
        "current_user": current_user,
        "msg": msg,
        "error": error,
        "removed": removed,
    })


@router.post("/stock/source/remove-movement")
async def report_stock_source_remove_movement(
    request: Request,
    movement_id: int = Form(...),
    warehouse_id: int = Form(...),
    product_id: int = Form(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Bitta harakatni olib tashlash (masalan hujjatda bu ombor ko'rsatilmagan noto'g'ri yozuv). Qoldiq qaytariladi."""
    from urllib.parse import quote
    m = db.query(StockMovement).filter(StockMovement.id == movement_id).first()
    if not m or m.warehouse_id != warehouse_id or m.product_id != product_id:
        return RedirectResponse(
            url=f"/reports/stock/source?warehouse_id={warehouse_id}&product_id={product_id}&error=" + quote("Harakat topilmadi yoki bu ombor+mahsulotga tegishli emas."),
            status_code=303,
        )
    add_back = -(float(m.quantity_change or 0))
    stocks = db.query(Stock).filter(Stock.warehouse_id == m.warehouse_id, Stock.product_id == m.product_id).all()
    if stocks:
        stocks[0].quantity = (float(stocks[0].quantity or 0) + add_back)
        if hasattr(stocks[0], "updated_at"):
            stocks[0].updated_at = datetime.now()
    elif add_back > 0:
        db.add(Stock(warehouse_id=m.warehouse_id, product_id=m.product_id, quantity=add_back))
    db.delete(m)
    db.commit()
    return RedirectResponse(
        url=f"/reports/stock/source?warehouse_id={warehouse_id}&product_id={product_id}&removed=1&msg=" + quote("Noto'g'ri harakat olib tashlandi, qoldiq qaytarildi."),
        status_code=303,
    )


@router.post("/stock/cleanup-orphan-movements")
async def report_stock_cleanup_orphan_movements(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """O'chirilgan barcha hujjat turlariga tegishli harakatlarni tozalash: ombordan omborga, ishlab chiqarish, kirim, qoldiq tuzatish. Qoldiq harakat ta'siriga qarab qaytariladi."""
    from urllib.parse import quote
    doc_types_tables = [
        ("WarehouseTransfer", WarehouseTransfer),
        ("Production", Production),
        ("Purchase", Purchase),
        ("StockAdjustmentDoc", StockAdjustmentDoc),
    ]
    existing_ids_by_type = {}
    for doc_type, model in doc_types_tables:
        try:
            existing_ids_by_type[doc_type] = {r[0] for r in db.query(model.id).all()}
        except Exception:
            existing_ids_by_type[doc_type] = set()
    all_movements = (
        db.query(StockMovement)
        .filter(
            StockMovement.document_type.in_([d[0] for d in doc_types_tables]),
            StockMovement.document_id.isnot(None),
        )
        .all()
    )
    orphans = [
        m for m in all_movements
        if (m.document_type or "") in existing_ids_by_type and m.document_id not in existing_ids_by_type.get(m.document_type, set())
    ]
    reverted = 0
    for m in orphans:
        wh_id, prod_id = m.warehouse_id, m.product_id
        if wh_id is None or prod_id is None:
            db.delete(m)
            reverted += 1
            continue
        add_back = -(float(m.quantity_change or 0))
        stocks = db.query(Stock).filter(Stock.warehouse_id == wh_id, Stock.product_id == prod_id).all()
        if stocks:
            stocks[0].quantity = (float(stocks[0].quantity or 0) + add_back)
            if hasattr(stocks[0], "updated_at"):
                stocks[0].updated_at = datetime.now()
        elif add_back > 0:
            db.add(Stock(warehouse_id=wh_id, product_id=prod_id, quantity=add_back))
        db.delete(m)
        reverted += 1
    db.commit()
    msg = quote(f"O'chirilgan hujjatlar harakatlari: {reverted} ta tozalandi (ombordan omborga, ishlab chiqarish, kirim, qoldiq tuzatish). Qoldiq qaytarildi.")
    return RedirectResponse(url=f"/reports/stock?cleanup_orphan=1&msg={msg}", status_code=303)


@router.post("/stock/cleanup-orphan-sale-movements")
async def report_stock_cleanup_orphan_sale_movements(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """O'chirilgan sotuv/qaytish hujjatlariga tegishli StockMovement larni tozalash va qoldiqni qaytarish (hujjat yo'q, qoldiq qolib ketgan holatlar)."""
    from urllib.parse import quote
    existing_order_ids = {r[0] for r in db.query(Order.id).all()}
    sale_doc_types = ("Sale", "SaleReturn", "SaleReturnRevert")
    orphans = (
        db.query(StockMovement)
        .filter(
            StockMovement.document_type.in_(sale_doc_types),
            StockMovement.document_id.isnot(None),
        )
        .all()
    )
    orphans = [m for m in orphans if m.document_id not in existing_order_ids]
    reverted = 0
    for m in orphans:
        wh_id, prod_id = m.warehouse_id, m.product_id
        if wh_id is None or prod_id is None:
            db.delete(m)
            reverted += 1
            continue
        # Harakat ta'sirini bekor qilish: qoldiqga (-quantity_change) qo'shamiz
        add_back = -(float(m.quantity_change or 0))
        stocks = db.query(Stock).filter(Stock.warehouse_id == wh_id, Stock.product_id == prod_id).all()
        if stocks:
            stocks[0].quantity = (float(stocks[0].quantity or 0) + add_back)
            if hasattr(stocks[0], "updated_at"):
                stocks[0].updated_at = datetime.now()
        elif add_back > 0:
            db.add(Stock(warehouse_id=wh_id, product_id=prod_id, quantity=add_back))
        db.delete(m)
        reverted += 1
    db.commit()
    msg = quote(f"O'chirilgan sotuv/qaytish harakatlari: {reverted} ta tozalandi, qoldiq qaytarildi.")
    return RedirectResponse(url=f"/reports/stock?cleanup_sale=1&msg={msg}", status_code=303)


@router.post("/stock/merge-duplicates")
async def report_stock_merge_duplicates(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Bir xil (ombor, mahsulot) uchun bitta Stock qatori qoldiradi, qolganlarini o'chirib yig'indini bitta qatorga yozadi (faqat admin)."""
    from collections import defaultdict
    all_stocks = db.query(Stock).all()
    by_key = defaultdict(list)
    for s in all_stocks:
        by_key[(s.warehouse_id, s.product_id)].append(s)
    merged = 0
    for key, group in by_key.items():
        if len(group) <= 1:
            continue
        total = sum(float(s.quantity or 0) for s in group)
        keep = group[0]
        keep.quantity = total
        if hasattr(keep, "updated_at"):
            keep.updated_at = datetime.now()
        for s in group[1:]:
            db.delete(s)
            merged += 1
    db.commit()
    return RedirectResponse(
        url=f"/reports/stock?merged={merged}",
        status_code=303,
    )


@router.post("/stock/recalculate-from-movements")
async def report_stock_recalculate_from_movements(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """O'CHIRILGAN: Bu endpoint xavfli edi — tarixiy nomuvofiq movementlar sababli
    qoldiqlarni 0 ga tushirib ma'lumotni yo'qotgan (2026-04-04 incident).
    Endi ishlamaydi."""
    raise HTTPException(
        status_code=410,
        detail="Bu funksiya o'chirildi. Tarixiy harakatlardan qoldiq qayta hisoblash xavfli edi. Inventarizatsiya hujjati orqali tuzating.",
    )
    # Eski kod (o'chirilgan):
    _disabled = """
    from collections import defaultdict
    # O'chirilgan ishlab chiqarish hujjatlariga tegishli "orfan" harakatlarni o'chirish (qoldiq to'g'ri tushadi)
    existing_production_ids = {r[0] for r in db.query(Production.id).all()}
    orphan_production_movements = db.query(StockMovement).filter(
        StockMovement.document_type == "Production",
        StockMovement.document_id.isnot(None),
    ).all()
    deleted_orphans = 0
    for m in orphan_production_movements:
        if m.document_id not in existing_production_ids:
            db.delete(m)
            deleted_orphans += 1
    if deleted_orphans:
        db.flush()
    # Tasdiqlangan qoldiq tuzatish hujjatlarini aniqlash (boshqa hujjat turlari hammasi hisobga olinadi)
    adj_ids = db.query(StockMovement.document_id).filter(
        StockMovement.document_type == "StockAdjustmentDoc",
        StockMovement.document_id.isnot(None),
    ).distinct().all()
    adj_ids = [r[0] for r in adj_ids if r[0]]
    confirmed_adj_ids = set()
    if adj_ids:
        for doc in db.query(StockAdjustmentDoc).filter(
            StockAdjustmentDoc.id.in_(adj_ids),
            StockAdjustmentDoc.status == "confirmed",
        ).all():
            confirmed_adj_ids.add(doc.id)
    # Har bir (warehouse_id, product_id) uchun harakatlarni created_at bo'yicha olamiz
    movements = (
        db.query(StockMovement)
        .filter(
            StockMovement.warehouse_id.isnot(None),
            StockMovement.product_id.isnot(None),
        )
        .order_by(StockMovement.created_at.asc())
        .all()
    )
    # (warehouse_id, product_id) -> har bir movement uchun dedup (bitta hujjatda bir nechta mahsulot bo'lishi mumkin)
    by_key = defaultdict(list)
    for m in movements:
        if (m.document_type or "") == "StockAdjustmentDoc" and m.document_id not in confirmed_adj_ids:
            continue
        key = (m.warehouse_id, m.product_id)
        by_key[key].append(m)
    totals = {}
    for key, lst in by_key.items():
        seen_doc = set()
        s = 0.0
        for m in lst:
            # Dedup: hujjat + mahsulot + ombor (bitta Sale/Purchase da ko'p mahsulot bo'ladi)
            doc_key = (m.document_type or "", m.document_id, m.product_id, m.warehouse_id)
            if doc_key in seen_doc:
                continue
            seen_doc.add(doc_key)
            s += float(m.quantity_change or 0)
        totals[key] = s
    updated = 0
    created = 0
    for (wh_id, prod_id), qty in totals.items():
        qty = max(0.0, qty)
        stock = db.query(Stock).filter(
            Stock.warehouse_id == wh_id,
            Stock.product_id == prod_id,
        ).first()
        if stock:
            stock.quantity = qty
            if hasattr(stock, "updated_at"):
                stock.updated_at = datetime.now()
            updated += 1
        else:
            db.add(Stock(warehouse_id=wh_id, product_id=prod_id, quantity=qty))
            created += 1
    db.commit()
    from urllib.parse import quote
    msg_parts = [f"Stock qoldiqlari harakatlar tarixidan qayta hisoblandi: {updated} yangilandi, {created} yangi qator."]
    if deleted_orphans:
        msg_parts.append(f" O'chirilgan ishlab chiqarish hujjatlariga tegishli {deleted_orphans} ta harakat olib tashlandi.")
    msg = quote("".join(msg_parts))
    return RedirectResponse(url=f"/reports/stock?recalculated=1&msg={msg}", status_code=303)
    """  # noqa — eski kod string ichida (ishlamaydi)


@router.post("/stock/clear")
async def report_stock_clear(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """O'CHIRILGAN: Stock jadvalini to'liq tozalash xavfli edi — endi ishlamaydi."""
    raise HTTPException(
        status_code=410,
        detail="Bu funksiya o'chirildi. Stock jadvalini tozalash xavfli edi.",
    )


@router.post("/stock/cleanup-orphans")
async def report_stock_cleanup_orphans(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """O'chirilgan ma'lumotlarni tozalash: Product yoki Warehouse mavjud bo'lmagan Stock, StockMovement, StockAdjustmentDocItem yozuvlarini o'chirish (faqat admin)."""
    from urllib.parse import quote
    valid_product_ids = {r[0] for r in db.query(Product.id).all()}
    valid_warehouse_ids = {r[0] for r in db.query(Warehouse.id).all()}
    valid_doc_ids = {r[0] for r in db.query(StockAdjustmentDoc.id).all()}

    # Stock: product_id yoki warehouse_id mavjud emas
    deleted_stock = 0
    for s in db.query(Stock).all():
        if (s.product_id not in valid_product_ids) or (s.warehouse_id not in valid_warehouse_ids):
            db.delete(s)
            deleted_stock += 1

    # StockMovement: product_id yoki warehouse_id mavjud emas
    deleted_movements = 0
    for m in db.query(StockMovement).all():
        if (m.product_id not in valid_product_ids) or (m.warehouse_id not in valid_warehouse_ids):
            db.delete(m)
            deleted_movements += 1

    # StockAdjustmentDocItem: product_id, warehouse_id yoki doc_id mavjud emas
    deleted_items = 0
    for it in db.query(StockAdjustmentDocItem).all():
        if (
            (it.product_id is not None and it.product_id not in valid_product_ids)
            or (it.warehouse_id is not None and it.warehouse_id not in valid_warehouse_ids)
            or (it.doc_id is not None and it.doc_id not in valid_doc_ids)
        ):
            db.delete(it)
            deleted_items += 1

    db.commit()
    msg = f"Tozalandi: Stock {deleted_stock}, StockMovement {deleted_movements}, Qoldiq hujjat qatorlari {deleted_items}."
    return RedirectResponse(url=f"/reports/stock?cleanup=1&msg=" + quote(msg), status_code=303)


def _stock_report_as_of_date(db: Session, report_date, wh_id: int = None):
    """Berilgan sanagacha bo'lgan harakatlar bo'yicha qoldiqni hisoblaydi. report_date — date yoki YYYY-MM-DD string."""
    if isinstance(report_date, str):
        try:
            report_date = datetime.strptime(report_date.strip()[:10], "%Y-%m-%d").date()
        except (ValueError, TypeError):
            report_date = datetime.now().date()
    if not isinstance(report_date, datetime):
        report_date = datetime.combine(report_date, datetime.max.time())
    else:
        report_date = report_date.replace(hour=23, minute=59, second=59, microsecond=999999)
    q = (
        db.query(StockMovement)
        .filter(StockMovement.created_at <= report_date)
        .order_by(StockMovement.warehouse_id, StockMovement.product_id, StockMovement.created_at.desc())
    )
    if wh_id:
        q = q.filter(StockMovement.warehouse_id == wh_id)
    rows = q.all()
    last_by_key = {}
    for m in rows:
        key = (m.warehouse_id, m.product_id)
        if key not in last_by_key:
            last_by_key[key] = float(m.quantity_after or 0)
    if not last_by_key:
        return []
    wh_ids = list({k[0] for k in last_by_key})
    prod_ids = list({k[1] for k in last_by_key})
    warehouses = {w.id: w for w in db.query(Warehouse).filter(Warehouse.id.in_(wh_ids)).all()}
    products = {p.id: p for p in db.query(Product).filter(Product.id.in_(prod_ids)).all()}
    result = []
    for (wid, pid), qty in last_by_key.items():
        if qty == 0:
            continue
        wh = warehouses.get(wid)
        prod = products.get(pid)
        if wh and prod:
            result.append({"warehouse": wh, "product": prod, "quantity": qty})
    return sorted(result, key=lambda x: ((x["warehouse"].name or "").lower(), (x["product"].name or "").lower()))


def _stock_report_filtered(db: Session, wh_id: int = None):
    """Stock jadvalidan qoldiq > 0 bo'lgan barcha qatorlarni qaytaradi."""
    q = (
        db.query(Stock)
        .join(Product, Stock.product_id == Product.id)
        .join(Warehouse, Stock.warehouse_id == Warehouse.id)
        .order_by(Warehouse.name, Product.name)
    )
    if wh_id:
        q = q.filter(Stock.warehouse_id == wh_id)
    rows = q.all()
    aggregated = {}
    for s in rows:
        key = (s.warehouse_id, s.product_id)
        if key not in aggregated:
            aggregated[key] = {"warehouse": s.warehouse, "product": s.product, "quantity": 0}
        aggregated[key]["quantity"] += float(s.quantity or 0)
    aggregated = {k: v for k, v in aggregated.items() if float(v.get("quantity") or 0) != 0}
    return sorted(aggregated.values(), key=lambda x: ((x["warehouse"].name or "").lower(), (x["product"].name or "").lower()))


@router.get("/stock/export")
async def report_stock_export(
    request: Request,
    warehouse_id: str = None,
    report_date: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    _check_export_rate_limit(request)
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    wh_id = None
    if warehouse_id is not None and str(warehouse_id).strip() != "":
        try:
            wh_id = int(warehouse_id)
        except (ValueError, TypeError):
            wh_id = None
    if report_date and str(report_date).strip():
        values = _stock_report_as_of_date(db, report_date.strip()[:10], wh_id)
    else:
        values = _stock_report_filtered(db, wh_id)
    stocks = [{"warehouse": v["warehouse"], "product": v["product"], "quantity": v["quantity"]} for v in values]
    wb = Workbook()
    ws = wb.active
    ws.title = "Qoldiq"
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A1"] = "Qoldiq hisoboti" + (" — " + report_date[:10] if report_date and report_date.strip() else "")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = report_date[:10] + " sana bo'yicha" if report_date and report_date.strip() else datetime.now().strftime("%d.%m.%Y %H:%M")
    ws.append(["Ombor", "Mahsulot", "Kod", "Qoldiq", "Minimal", "Narx", "Summa"])
    for c in range(1, 8):
        ws.cell(row=4, column=c).fill = header_fill
        ws.cell(row=4, column=c).font = Font(bold=True, color="FFFFFF")
    for s in stocks:
        p = s.product
        wh = s.warehouse
        min_s = getattr(p, "min_stock", 0) or 0
        price = getattr(p, "purchase_price", 0) or 0
        ws.append([
            wh.name if wh else "",
            p.name if p else "",
            (p.barcode or p.code or "") if p else "",
            float(s.quantity or 0),
            float(min_s),
            float(price),
            float((s.quantity or 0) * price),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn_date = (report_date or "").strip()[:10].replace("-", "") if report_date else datetime.now().strftime("%Y%m%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=qoldiq_{fn_date}.xlsx"},
    )


@router.get("/stock/andoza")
async def report_stock_andoza(current_user: User = Depends(require_auth)):
    """Qoldiqlar uchun Excel andoza (Tannarx va Sotuv narxi ixtiyoriy)."""
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    wb = Workbook()
    ws = wb.active
    ws.title = "Andoza"
    ws.append(["Ombor nomi (yoki kodi)", "Mahsulot nomi (yoki kodi)", "Qoldiq", "Tannarx (so'm)", "Sotuv narxi (so'm)"])
    ws.append(["Xom ashyo ombori", "Yong'oq", 30, "", ""])
    ws.append(["Xom ashyo ombori", "Bodom", 100, "", ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=qoldiqlar_andoza.xlsx"},
    )


@router.post("/stock/import")
async def report_stock_import(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Exceldan qoldiqlarni yuklash — hujjat qoralama holatida yaratiladi. doc_date bo'lsa shu sana ishlatiladi."""
    from urllib.parse import quote
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    form = await request.form()
    file = form.get("file") or form.get("excel_file")
    if not file or not getattr(file, "filename", None):
        return RedirectResponse(url="/reports/stock?error=" + quote("Excel fayl tanlang"), status_code=303)
    try:
        contents = await file.read() if hasattr(file, "read") else (getattr(file, "file", None) and file.file.read() or b"")
    except Exception:
        contents = b""
    if not contents:
        return RedirectResponse(url="/reports/stock?error=" + quote("Fayl bo'sh"), status_code=303)
    doc_date_str = (form.get("doc_date") or "").strip()
    try:
        if doc_date_str:
            doc_date = datetime.strptime(doc_date_str, "%Y-%m-%d").replace(hour=23, minute=59, second=0, microsecond=0)
        else:
            doc_date = datetime.now()
    except ValueError:
        doc_date = datetime.now()
    wb = load_workbook(io.BytesIO(contents))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    items_data = []  # (product_id, warehouse_id, qty, tannarx, sotuv_narxi)
    for row in rows:
        if not row or (row[0] is None or row[0] == "") or (row[1] is None or row[1] == ""):
            continue
        wh_key = str(row[0] or "").strip()
        raw_prod = row[1]
        if raw_prod is not None and isinstance(raw_prod, (int, float)) and float(raw_prod) == int(float(raw_prod)):
            product_key = str(int(float(raw_prod)))
        else:
            product_key = str(raw_prod or "").strip()
        try:
            qty = float(row[2]) if row[2] is not None and row[2] != "" else 0
        except (TypeError, ValueError):
            qty = 0
        tannarx = 0.0
        sotuv_narxi = 0.0
        if len(row) > 3 and row[3] is not None and row[3] != "":
            try:
                tannarx = float(row[3])
            except (TypeError, ValueError):
                pass
        if len(row) > 4 and row[4] is not None and row[4] != "":
            try:
                sotuv_narxi = float(row[4])
            except (TypeError, ValueError):
                pass
        wh = db.query(Warehouse).filter(
            (func.lower(Warehouse.name) == wh_key.lower()) | (Warehouse.code == wh_key)
        ).first()
        product = db.query(Product).filter(
            (Product.code == product_key) | (Product.barcode == product_key)
        ).first()
        if not product and product_key:
            product = db.query(Product).filter(
                Product.name.isnot(None),
                func.lower(Product.name) == product_key.lower()
            ).first()
        if not wh or not product:
            continue
        if tannarx > 0:
            product.purchase_price = tannarx
        if sotuv_narxi > 0:
            product.sale_price = sotuv_narxi
        items_data.append((product.id, wh.id, qty, tannarx, sotuv_narxi))
    if not items_data:
        return RedirectResponse(
            url="/reports/stock?error=" + quote("Hech qanday to'g'ri qator topilmadi"),
            status_code=303,
        )
    doc_date_start = doc_date.replace(hour=0, minute=0, second=0, microsecond=0)
    doc_date_end = doc_date_start + timedelta(days=1)
    count = db.query(StockAdjustmentDoc).filter(
        StockAdjustmentDoc.date >= doc_date_start,
        StockAdjustmentDoc.date < doc_date_end,
    ).count()
    number = f"QLD-{doc_date.strftime('%Y%m%d')}-{str(count + 1).zfill(4)}"
    total_tannarx = sum(qty * cp for _, _, qty, cp, _ in items_data)
    total_sotuv = sum(qty * sp for _, _, qty, _, sp in items_data)
    doc = StockAdjustmentDoc(
        number=number,
        date=doc_date,
        user_id=current_user.id if current_user else None,
        status="draft",
        total_tannarx=total_tannarx,
        total_sotuv=total_sotuv,
    )
    db.add(doc)
    db.flush()
    for pid, wid, qty, cp, sp in items_data:
        db.add(StockAdjustmentDocItem(
            doc_id=doc.id,
            product_id=pid,
            warehouse_id=wid,
            quantity=qty,
            cost_price=cp,
            sale_price=sp,
        ))
    db.commit()
    return RedirectResponse(
        url=f"/qoldiqlar/tovar/hujjat/{doc.id}?from=import&msg=" + quote("Hujjat qoralama. Qoldiq hisobotida ko'rinishi uchun «Tasdiqlash» bosing."),
        status_code=303,
    )


@router.get("/production", response_class=HTMLResponse)
async def report_production(
    request: Request,
    start_date: str = None,
    end_date: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Ishlab chiqarish hisoboti"""
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    if not start_date:
        start_date = datetime.now().replace(day=1).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now().strftime("%Y-%m-%d")
    q = (
        db.query(Production)
        .filter(
            Production.date >= start_date,
            Production.date <= end_date + " 23:59:59",
        )
        .order_by(Production.date.desc())
    )
    productions = q.all()
    total_qty = sum(p.quantity for p in productions if p.status == "completed")
    return templates.TemplateResponse("reports/production.html", {
        "request": request,
        "productions": productions,
        "total_qty": total_qty,
        "start_date": start_date,
        "end_date": end_date,
        "page_title": "Ishlab chiqarish hisoboti",
        "current_user": current_user,
    })


@router.get("/employees", response_class=HTMLResponse)
async def report_employees(request: Request, db: Session = Depends(get_db), current_user: User = Depends(require_auth)):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    allowed = get_allowed_report_types(current_user)
    if "employees" not in allowed and current_user.role != "admin":
        return RedirectResponse(url="/reports", status_code=303)
    employees = db.query(Employee).order_by(Employee.full_name).all()
    return templates.TemplateResponse("reports/employees.html", {
        "request": request,
        "employees": employees,
        "page_title": "Xodimlar hisoboti",
        "current_user": current_user,
    })


@router.get("/debts", response_class=HTMLResponse)
async def report_debts(
    request: Request,
    overdue_days: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    # Barcha kontragentlar — balans != 0 (savdo, xarid, qoldiq hujjati — barchasi)
    q = db.query(Partner).filter(Partner.is_active == True, Partner.balance != 0)
    if overdue_days and int(overdue_days) > 0:
        cutoff = datetime.now() - timedelta(days=int(overdue_days))
        overdue_partner_ids = (
            db.query(Order.partner_id)
            .filter(
                Order.type == "sale",
                Order.debt > 0,
                Order.created_at < cutoff,
                Order.partner_id.isnot(None),
            )
            .distinct()
            .all()
        )
        ids = [r[0] for r in overdue_partner_ids if r and r[0]]
        if ids:
            q = q.filter(Partner.id.in_(ids))
        else:
            q = q.filter(Partner.id == -1)
    debtors = q.order_by(Partner.name).all()
    # Mijozlar qarzi (balance > 0)
    total_debt = sum(p.balance for p in debtors if p.balance > 0)
    # Ta'minotchilarga qarzimiz (balance < 0)
    total_credit = sum(abs(p.balance) for p in debtors if p.balance < 0)
    return templates.TemplateResponse("reports/debts.html", {
        "request": request,
        "debtors": debtors,
        "total_debt": total_debt,
        "total_credit": total_credit,
        "page_title": "Qarzdorlik hisoboti",
        "current_user": current_user,
    })


@router.get("/debts/export")
async def report_debts_export(request: Request, db: Session = Depends(get_db), current_user: User = Depends(require_auth)):
    _check_export_rate_limit(request)
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    debtors = db.query(Partner).filter(Partner.is_active == True, Partner.balance != 0).order_by(Partner.name).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Qarzdorlik"
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A1"] = "Qarzdorlik hisoboti"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws.append(["Kod", "Mijoz", "Telefon", "Balans (qarz +)", "Kredit limiti"])
    for c in range(1, 6):
        ws.cell(row=4, column=c).fill = header_fill
        ws.cell(row=4, column=c).font = Font(bold=True, color="FFFFFF")
    for p in debtors:
        ws.append([
            p.code or "",
            p.name or "",
            p.phone or "",
            float(p.balance or 0),
            float(p.credit_limit or 0),
        ])
    total = sum(p.balance for p in debtors if (p.balance or 0) > 0)
    ws.append(["", "", "JAMI QARZ:", total, ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=qarzdorlik_{datetime.now().strftime('%Y%m%d')}.xlsx"},
    )


def _build_partner_movements(db: Session, partner_id: int, date_from: datetime, date_to: datetime, period_only: bool):
    """
    Kontragent uchun harakatlar ro'yxati (1C uslubida).
    period_only=True: faqat [date_from, date_to] oralig'idagi qatorlar.
    Qaytadi: (rows, opening_debit, opening_credit) yoki period_only=True bo'lsa (rows, 0, 0).
    Balans: Debit = kontragent bizga qarzdor (sotuv), Credit = to'lov/xarid/qaytarish.
    """
    partner = db.query(Partner).filter(Partner.id == partner_id).first()
    if not partner:
        return [], 0.0, 0.0
    date_from_start = date_from.replace(hour=0, minute=0, second=0, microsecond=0)
    date_to_end = date_to.replace(hour=23, minute=59, second=59, microsecond=999999)
    rows = []

    # Sotuvlar (debit) va qaytarishlar (credit) — bekor qilinganlar bundan mustasno
    q_orders = db.query(Order).filter(
        Order.partner_id == partner_id,
        Order.type.in_(["sale", "return_sale"]),
        Order.status.notin_(["cancelled", "draft"]),
    )
    if period_only:
        q_orders = q_orders.filter(Order.date >= date_from_start, Order.date <= date_to_end)
    for o in q_orders.order_by(Order.date):
        debit = float(o.total or 0) if o.type == "sale" else 0.0
        credit = float(o.total or 0) if o.type == "return_sale" else 0.0
        doc_label = f"{'Sotuv' if o.type == 'sale' else 'Qaytarish'} {o.number or ''} {o.date.strftime('%d.%m.%Y %H:%M') if o.date else ''}".strip()
        rows.append({
            "date": o.date,
            "doc_type": "Sotuv" if o.type == "sale" else "Qaytarish",
            "doc_number": o.number or "",
            "doc_label": doc_label,
            "doc_url": f"/sales/edit/{o.id}",
            "debit": debit,
            "credit": credit,
        })

    # To'lovlar: income = credit (ular bizga to'ladı), expense = debit (biz ularga to'ladık) — faqat tasdiqlangan
    q_payments = db.query(Payment).filter(Payment.partner_id == partner_id)
    if hasattr(Payment, "status"):
        q_payments = q_payments.filter(or_(Payment.status == "confirmed", Payment.status.is_(None)))
    if period_only:
        q_payments = q_payments.filter(Payment.date >= date_from_start, Payment.date <= date_to_end)
    for p in q_payments.order_by(Payment.date):
        if p.type == "income":
            doc_label = f"To'lov (kirim) {p.number or ''} {p.date.strftime('%d.%m.%Y %H:%M') if p.date else ''}".strip()
            rows.append({
                "date": p.date,
                "doc_type": "To'lov (kirim)",
                "doc_number": p.number or "",
                "doc_label": doc_label,
                "doc_url": f"/finance/payment/{p.id}/edit",
                "debit": 0.0,
                "credit": float(p.amount or 0),
            })
        else:
            doc_label = f"To'lov (chiqim) {p.number or ''} {p.date.strftime('%d.%m.%Y %H:%M') if p.date else ''}".strip()
            rows.append({
                "date": p.date,
                "doc_type": "To'lov (chiqim)",
                "doc_number": p.number or "",
                "doc_label": doc_label,
                "doc_url": f"/finance/payment/{p.id}/edit",
                "debit": float(p.amount or 0),
                "credit": 0.0,
            })

    # Xaridlar (biz yetkazuvchiga qarzdormiz — credit) — faqat tasdiqlangan
    q_purchases = db.query(Purchase).filter(
        Purchase.partner_id == partner_id,
        Purchase.status == "confirmed",
    )
    if period_only:
        q_purchases = q_purchases.filter(Purchase.date >= date_from_start, Purchase.date <= date_to_end)
    for p in q_purchases.order_by(Purchase.date):
        total_val = float((p.total or 0) + (p.total_expenses or 0))
        doc_label = f"Tovarlar kirimi (xarid) {p.number or ''} {p.date.strftime('%d.%m.%Y %H:%M') if p.date else ''}".strip()
        rows.append({
            "date": p.date,
            "doc_type": "Xarid",
            "doc_number": p.number or "",
            "doc_label": doc_label,
            "doc_url": f"/purchases/edit/{p.id}",
            "debit": 0.0,
            "credit": total_val,
        })

    # Kontragent qoldiq hujjatlari (tasdiqlangan) — boshlang'ich qoldiq kiritish
    q_balance_items = (
        db.query(PartnerBalanceDocItem, PartnerBalanceDoc)
        .join(PartnerBalanceDoc, PartnerBalanceDocItem.doc_id == PartnerBalanceDoc.id)
        .filter(
            PartnerBalanceDocItem.partner_id == partner_id,
            PartnerBalanceDoc.status == "confirmed",
        )
    )
    if period_only:
        q_balance_items = q_balance_items.filter(
            PartnerBalanceDoc.date >= date_from_start,
            PartnerBalanceDoc.date <= date_to_end,
        )
    for item, doc in q_balance_items.order_by(PartnerBalanceDoc.date):
        # Kiritilgan qoldiq summasi (to'liq) — musbat = uning qarzi (debit), manfiy = bizning qarzimiz (credit)
        bal = float(item.balance or 0)
        debit = bal if bal > 0 else 0.0
        credit = -bal if bal < 0 else 0.0
        doc_label = f"Qoldiq kiritish {doc.number or ''} {doc.date.strftime('%d.%m.%Y %H:%M') if doc.date else ''}".strip()
        rows.append({
            "date": doc.date,
            "doc_type": "Qoldiq kiritish",
            "doc_number": doc.number or "",
            "doc_label": doc_label,
            "doc_url": f"/qoldiqlar/kontragent/hujjat/{doc.id}",
            "debit": debit,
            "credit": credit,
        })

    rows.sort(key=lambda r: r["date"])
    opening_debit = 0.0
    opening_credit = 0.0
    if not period_only:
        # Opening = barcha harakatlar perioddan oldin
        for r in rows:
            if r["date"] < date_from_start:
                opening_debit += r["debit"]
                opening_credit += r["credit"]
        rows = [r for r in rows if date_from_start <= r["date"] <= date_to_end]
    return rows, opening_debit, opening_credit


@router.get("/partner-reconciliation", response_class=HTMLResponse)
async def report_partner_reconciliation(
    request: Request,
    partner_id: Optional[str] = None,
    date_from: str = None,
    date_to: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Kontragentlar hisob-kitobini solishtirish hisoboti (1C uslubida)."""
    # partner_id bo'sh string yoki noto'g'ri kelishi mumkin
    try:
        partner_id = int(partner_id) if partner_id and str(partner_id).strip() else None
    except (ValueError, TypeError):
        partner_id = None
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    allowed = get_allowed_report_types(current_user)
    if "partner_reconciliation" not in allowed:
        return RedirectResponse(url="/reports", status_code=303)
    partners = db.query(Partner).filter(Partner.is_active == True).order_by(Partner.name).all()
    today = datetime.now()
    if not date_from:
        date_from = (today.replace(day=1)).strftime("%Y-%m-%d")
    if not date_to:
        date_to = today.strftime("%Y-%m-%d")
    try:
        dt_from = datetime.strptime(date_from, "%Y-%m-%d")
        dt_to = datetime.strptime(date_to, "%Y-%m-%d")
    except (ValueError, TypeError):
        dt_from = today.replace(day=1)
        dt_to = today
    rows = []
    opening_debit = opening_credit = 0.0
    total_debit = total_credit = 0.0
    partner_obj = None
    products_purchased = []
    products_sold = []
    # Umumiy hisobot (barcha kontragentlar)
    summary_rows = []
    if not partner_id:
        # Barcha kontragentlar bo'yicha umumiy
        all_partners = db.query(Partner).filter(Partner.is_active == True, Partner.balance != 0).order_by(Partner.name).all()
        grand_total_debit = 0.0
        grand_total_credit = 0.0
        for p in all_partners:
            p_rows, p_od, p_oc = _build_partner_movements(db, p.id, dt_from, dt_to, period_only=False)
            p_total_debit = sum(r["debit"] for r in p_rows)
            p_total_credit = sum(r["credit"] for r in p_rows)
            p_opening = p_od - p_oc
            p_closing = p_opening + p_total_debit - p_total_credit
            grand_total_debit += p_total_debit
            grand_total_credit += p_total_credit
            summary_rows.append({
                "partner": p,
                "opening_balance": p_opening,
                "total_debit": p_total_debit,
                "total_credit": p_total_credit,
                "closing_balance": p_closing,
            })
    if partner_id:
        partner_obj = db.query(Partner).filter(Partner.id == partner_id).first()
        if partner_obj:
            rows, opening_debit, opening_credit = _build_partner_movements(db, partner_id, dt_from, dt_to, period_only=False)
            total_debit = sum(r["debit"] for r in rows)
            total_credit = sum(r["credit"] for r in rows)
            date_from_start = dt_from.replace(hour=0, minute=0, second=0, microsecond=0)
            date_to_end = dt_to.replace(hour=23, minute=59, second=59, microsecond=999999)
            purchases_in_period = (
                db.query(PurchaseItem, Product)
                .join(Purchase, PurchaseItem.purchase_id == Purchase.id)
                .join(Product, PurchaseItem.product_id == Product.id)
                .filter(
                    Purchase.partner_id == partner_id,
                    Purchase.status == "confirmed",
                    Purchase.date >= date_from_start,
                    Purchase.date <= date_to_end,
                )
            ).all()
            by_product_purchase = {}
            for pi, prod in purchases_in_period:
                key = prod.id
                if key not in by_product_purchase:
                    by_product_purchase[key] = {"product_name": prod.name or "", "product_code": prod.code or "", "quantity": 0.0, "total": 0.0}
                by_product_purchase[key]["quantity"] += float(pi.quantity or 0)
                by_product_purchase[key]["total"] += float(pi.total or 0)
            products_purchased = sorted(by_product_purchase.values(), key=lambda x: -x["total"])
            orders_in_period = (
                db.query(OrderItem, Product)
                .join(Order, OrderItem.order_id == Order.id)
                .join(Product, OrderItem.product_id == Product.id)
                .filter(
                    Order.partner_id == partner_id,
                    Order.type == "sale",
                    Order.status.notin_(["cancelled", "draft"]),
                    Order.date >= date_from_start,
                    Order.date <= date_to_end,
                )
            ).all()
            by_product_sale = {}
            for oi, prod in orders_in_period:
                key = prod.id
                if key not in by_product_sale:
                    by_product_sale[key] = {"product_name": prod.name or "", "product_code": prod.code or "", "quantity": 0.0, "total": 0.0}
                by_product_sale[key]["quantity"] += float(oi.quantity or 0)
                by_product_sale[key]["total"] += float(oi.total or 0)
            products_sold = sorted(by_product_sale.values(), key=lambda x: -x["total"])
    opening_balance = opening_debit - opening_credit
    closing_balance = opening_balance + total_debit - total_credit
    return templates.TemplateResponse("reports/partner_reconciliation.html", {
        "request": request,
        "partners": partners,
        "partner_id": partner_id,
        "partner": partner_obj,
        "date_from": date_from,
        "date_to": date_to,
        "rows": rows,
        "opening_balance": opening_balance,
        "closing_balance": closing_balance,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "opening_debit": opening_debit,
        "opening_credit": opening_credit,
        "products_purchased": products_purchased,
        "products_sold": products_sold,
        "summary_rows": summary_rows,
        "page_title": "Kontragentlar hisob-kitobini solishtirish",
        "current_user": current_user,
    })


def _partner_recon_parse_dates(date_from: Optional[str], date_to: Optional[str]) -> tuple:
    """Returns: (dt_from, dt_to, iso_from, iso_to)."""
    today = datetime.now()
    if not date_from:
        date_from = today.replace(day=1).strftime("%Y-%m-%d")
    if not date_to:
        date_to = today.strftime("%Y-%m-%d")
    try:
        dt_from = datetime.strptime(date_from, "%Y-%m-%d")
        dt_to = datetime.strptime(date_to, "%Y-%m-%d")
    except (ValueError, TypeError):
        dt_from = today.replace(day=1)
        dt_to = today
    return dt_from, dt_to, date_from, date_to


def _partner_product_analytics(db: Session, partner_id: int, dt_from: datetime, dt_to: datetime) -> tuple:
    """Kontragent bilan davrda xarid qilingan/sotilgan mahsulotlar bo'yicha yig'ma.
    Returns: (products_purchased: list, products_sold: list)."""
    date_from_start = dt_from.replace(hour=0, minute=0, second=0, microsecond=0)
    date_to_end = dt_to.replace(hour=23, minute=59, second=59, microsecond=999999)

    def _aggregate(pairs):
        out = {}
        for item, prod in pairs:
            key = prod.id
            if key not in out:
                out[key] = {"product_name": prod.name or "", "product_code": prod.code or "", "quantity": 0.0, "total": 0.0}
            out[key]["quantity"] += float(item.quantity or 0)
            out[key]["total"] += float(item.total or 0)
        return sorted(out.values(), key=lambda x: -x["total"])

    purchase_pairs = (
        db.query(PurchaseItem, Product)
        .join(Purchase, PurchaseItem.purchase_id == Purchase.id)
        .join(Product, PurchaseItem.product_id == Product.id)
        .filter(Purchase.partner_id == partner_id,
                Purchase.date >= date_from_start,
                Purchase.date <= date_to_end)
    ).all()
    sale_pairs = (
        db.query(OrderItem, Product)
        .join(Order, OrderItem.order_id == Order.id)
        .join(Product, OrderItem.product_id == Product.id)
        .filter(Order.partner_id == partner_id,
                Order.type == "sale",
                Order.date >= date_from_start,
                Order.date <= date_to_end)
    ).all()
    return _aggregate(purchase_pairs), _aggregate(sale_pairs)


@router.get("/partner-reconciliation/export")
async def report_partner_reconciliation_export(
    request: Request,
    partner_id: int = None,
    date_from: str = None,
    date_to: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Kontragent solishtirish hisobotini Excelga eksport."""
    _check_export_rate_limit(request)
    if not current_user:
        return RedirectResponse(url="/reports", status_code=303)
    if not partner_id:
        return RedirectResponse(url="/reports/partner-reconciliation?error=partner", status_code=303)
    partner = db.query(Partner).filter(Partner.id == partner_id).first()
    if not partner:
        return RedirectResponse(url="/reports/partner-reconciliation", status_code=303)
    dt_from, dt_to, date_from, date_to = _partner_recon_parse_dates(date_from, date_to)
    rows, opening_debit, opening_credit = _build_partner_movements(db, partner_id, dt_from, dt_to, period_only=False)
    total_debit = sum(r["debit"] for r in rows)
    total_credit = sum(r["credit"] for r in rows)
    opening_balance = opening_debit - opening_credit
    closing_balance = opening_balance + total_debit - total_credit
    products_purchased, products_sold = _partner_product_analytics(db, partner_id, dt_from, dt_to)

    from openpyxl.styles import Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Solishtirish"

    # Stillar
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="017449", end_color="017449", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    balance_fill = PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid")
    total_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    num_fmt = '#,##0'
    p_name = partner.name or "Kontragent"

    # Sarlavha
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "HISOB-KITOBLARNI SOLISHTIRISH"
    c.font = Font(bold=True, size=16, color="017449")
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:E2")
    c = ws["A2"]
    c.value = f"TOTLI HOLVA  va  {p_name}"
    c.font = Font(bold=True, size=12)
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:E3")
    c = ws["A3"]
    c.value = f"Davr: {date_from}  —  {date_to}"
    c.font = Font(size=11, color="555555")
    c.alignment = Alignment(horizontal="center")

    # Jadval sarlavhalari — 2 qatorli merged
    # 5-qator: Hujjatlar | TOTLI HOLVA (merged B5:C5) | Kontragent (merged D5:E5)
    # 6-qator: (bo'sh) | DT Haqdor | KT Qarzdor | DT Haqdor | KT Qarzdor
    ws.merge_cells("A5:A6")
    c = ws.cell(row=5, column=1, value="HUJJATLAR")
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=6, column=1).fill = header_fill

    ws.merge_cells("B5:C5")
    c = ws.cell(row=5, column=2, value="TOTLI HOLVA")
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal="center")
    ws.cell(row=5, column=3).fill = header_fill

    ws.merge_cells("D5:E5")
    c = ws.cell(row=5, column=4, value=p_name.upper())
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal="center")
    ws.cell(row=5, column=5).fill = header_fill

    sub_headers = {2: "DT Haqdor (so'm)", 3: "KT Qarzdor (so'm)", 4: "DT Haqdor (so'm)", 5: "KT Qarzdor (so'm)"}
    for ci, h in sub_headers.items():
        c = ws.cell(row=6, column=ci, value=h)
        c.fill = subheader_fill
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

    def _set_balance_row(rn, label, ob):
        ws.cell(row=rn, column=1, value=label).font = Font(bold=True)
        for ci in range(1, 6):
            ws.cell(row=rn, column=ci).fill = balance_fill
            ws.cell(row=rn, column=ci).border = thin_border
        vals = [
            ob if ob > 0 else None,
            -ob if ob < 0 else None,
            -ob if ob < 0 else None,
            ob if ob > 0 else None,
        ]
        for ci, v in enumerate(vals, 2):
            cell = ws.cell(row=rn, column=ci)
            if v:
                cell.value = v
                cell.number_format = num_fmt
            else:
                cell.value = "—"
            cell.alignment = Alignment(horizontal="right")
            cell.font = Font(bold=True)

    row_num = 7
    # Davr boshiga qoldiq
    _set_balance_row(row_num, f"Davr boshiga qoldiq ({date_from})", opening_balance)
    row_num += 1

    # Hujjatlar
    for r in rows:
        ws.cell(row=row_num, column=1, value=r.get("doc_label") or f"{r.get('doc_type', '')} {r.get('doc_number', '')}")
        vals = [
            (r["debit"], r["credit"]),   # TOTLI HOLVA: DT, KT
            (r["credit"], r["debit"]),   # Kontragent: DT, KT
        ]
        col = 2
        for dt_val, kt_val in vals:
            for v in [dt_val, kt_val]:
                cell = ws.cell(row=row_num, column=col)
                if v:
                    cell.value = v
                    cell.number_format = num_fmt
                else:
                    cell.value = "—"
                cell.alignment = Alignment(horizontal="right")
                cell.border = thin_border
                col += 1
        ws.cell(row=row_num, column=1).border = thin_border
        row_num += 1

    # Jami davr
    ws.cell(row=row_num, column=1, value="Jami davr:").font = Font(bold=True)
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="right")
    for ci in range(1, 6):
        ws.cell(row=row_num, column=ci).fill = total_fill
        ws.cell(row=row_num, column=ci).border = thin_border
        ws.cell(row=row_num, column=ci).font = Font(bold=True)
    for ci, v in [(2, total_debit), (3, total_credit), (4, total_credit), (5, total_debit)]:
        cell = ws.cell(row=row_num, column=ci, value=v)
        cell.number_format = num_fmt
        cell.alignment = Alignment(horizontal="right")
    row_num += 1

    # Davr oxiriga qoldiq
    _set_balance_row(row_num, f"Davr oxiriga qoldiq ({date_to})", closing_balance)
    row_num += 2

    # Xulosa
    ws.cell(row=row_num, column=1, value="Bizning foydamizga (kontragent qarzdor):").font = Font(bold=True)
    c = ws.cell(row=row_num, column=2, value=closing_balance if closing_balance > 0 else 0)
    c.number_format = num_fmt
    c.font = Font(bold=True, color="CC0000")
    row_num += 1
    ws.cell(row=row_num, column=1, value="Kontragent foydasiga (biz qarzdormiz):").font = Font(bold=True)
    c = ws.cell(row=row_num, column=2, value=-closing_balance if closing_balance < 0 else 0)
    c.number_format = num_fmt
    c.font = Font(bold=True, color="2E7D32")

    ws.column_dimensions["A"].width = 52
    for col_letter in ["B", "C", "D", "E"]:
        ws.column_dimensions[col_letter].width = 22
    ws.print_area = f"A1:E{row_num}"

    def _build_product_sheet(ws_p, title, color, items, label_empty):
        ws_p.merge_cells("A1:D1")
        c = ws_p["A1"]
        c.value = title
        c.font = Font(bold=True, size=14, color=color)
        c.alignment = Alignment(horizontal="center")
        ws_p.merge_cells("A2:D2")
        c = ws_p["A2"]
        c.value = f"Kontragent: {p_name}  |  Davr: {date_from} — {date_to}"
        c.font = Font(size=11, color="555555")
        c.alignment = Alignment(horizontal="center")
        p_headers = ["Mahsulot", "Kod", "Miqdor", "Summa (so'm)"]
        for ci, h in enumerate(p_headers, 1):
            cell = ws_p.cell(row=4, column=ci, value=h)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        total_sum = 0
        for i, p in enumerate(items, 5):
            ws_p.cell(row=i, column=1, value=p["product_name"]).border = thin_border
            ws_p.cell(row=i, column=2, value=p["product_code"]).border = thin_border
            c = ws_p.cell(row=i, column=3, value=p["quantity"])
            c.number_format = '#,##0'
            c.border = thin_border
            c.alignment = Alignment(horizontal="right")
            c = ws_p.cell(row=i, column=4, value=p["total"])
            c.number_format = num_fmt
            c.border = thin_border
            c.alignment = Alignment(horizontal="right")
            total_sum += p["total"]
        if items:
            tr = 5 + len(items)
            ws_p.cell(row=tr, column=3, value="JAMI:").font = Font(bold=True)
            ws_p.cell(row=tr, column=3).alignment = Alignment(horizontal="right")
            c = ws_p.cell(row=tr, column=4, value=total_sum)
            c.number_format = num_fmt
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="right")
            for ci in range(1, 5):
                ws_p.cell(row=tr, column=ci).fill = total_fill
                ws_p.cell(row=tr, column=ci).border = thin_border
        else:
            ws_p.cell(row=5, column=1, value=label_empty).font = Font(color="999999")
        ws_p.column_dimensions["A"].width = 40
        ws_p.column_dimensions["B"].width = 15
        ws_p.column_dimensions["C"].width = 12
        ws_p.column_dimensions["D"].width = 20

    # Varaq: Kontragentdan xarid qilingan mahsulotlar
    ws_purchase = wb.create_sheet("Xarid qilingan", 1)
    _build_product_sheet(ws_purchase, "Kontragentdan xarid qilingan mahsulotlar", "2E7D32", products_purchased, "Davrda xarid qilinmagan.")

    # Varaq: Kontragentga sotilgan mahsulotlar
    ws_sale = wb.create_sheet("Sotilgan", 2)
    _build_product_sheet(ws_sale, "Kontragentga sotilgan mahsulotlar", "1565C0", products_sold, "Davrda sotuv bo'lmagan.")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"kontragent_solishtirish_{partner.id}_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ==========================================
# FOYDA HISOBOTI
# ==========================================

def _parse_profit_date_range(date_from: Optional[str], date_to: Optional[str]) -> tuple:
    """date_from/date_to parsing + default: oy boshi - bugun. Returns: (dt_from, dt_to, iso_from, iso_to)."""
    today = datetime.now()
    if not date_from:
        date_from = today.replace(day=1).strftime("%Y-%m-%d")
    if not date_to:
        date_to = today.strftime("%Y-%m-%d")
    try:
        dt_from = datetime.strptime(date_from, "%Y-%m-%d")
        dt_to = datetime.strptime(date_to, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    except (ValueError, TypeError):
        dt_from = today.replace(day=1)
        dt_to = today.replace(hour=23, minute=59, second=59)
    return dt_from, dt_to, date_from, date_to


def _compute_sales_and_cogs(db: Session, dt_from: datetime, dt_to: datetime) -> tuple:
    """Returns: (sale_orders, revenue, cogs, sale_items)."""
    sale_orders = (
        db.query(Order)
        .filter(Order.type == "sale", Order.status != "cancelled",
                Order.date >= dt_from, Order.date <= dt_to)
        .all()
    )
    revenue = sum(float(o.total or 0) for o in sale_orders)
    sale_order_ids = [o.id for o in sale_orders]
    sale_items = []
    cogs = 0.0
    if sale_order_ids:
        sale_items = (
            db.query(OrderItem, Product)
            .join(Product, OrderItem.product_id == Product.id)
            .filter(OrderItem.order_id.in_(sale_order_ids))
            .all()
        )
        for oi, prod in sale_items:
            cogs += float(prod.purchase_price or 0) * float(oi.quantity or 0)
    return sale_orders, revenue, cogs, sale_items


def _compute_operating_expenses(db: Session, dt_from: datetime, dt_to: datetime) -> tuple:
    """Returns: (total_expenses, expense_list) — ExpenseDoc confirmed hujjatlari turiga qarab guruhlangan."""
    expense_docs = (
        db.query(ExpenseDoc)
        .filter(ExpenseDoc.status == "confirmed",
                ExpenseDoc.date >= dt_from, ExpenseDoc.date <= dt_to)
        .all()
    )
    expense_by_type: dict = {}
    if expense_docs:
        expense_doc_ids = [e.id for e in expense_docs]
        items = db.query(ExpenseDocItem).filter(ExpenseDocItem.expense_doc_id.in_(expense_doc_ids)).all()
        for ei in items:
            et = ei.expense_type
            type_name = et.name if et else "Boshqa"
            category = et.category if et else "Boshqa"
            if type_name not in expense_by_type:
                expense_by_type[type_name] = {"name": type_name, "category": category, "amount": 0.0}
            expense_by_type[type_name]["amount"] += float(ei.amount or 0)
    total_expenses = sum(v["amount"] for v in expense_by_type.values())
    expense_list = sorted(expense_by_type.values(), key=lambda x: -x["amount"])
    return total_expenses, expense_list


def _compute_salary_total(db: Session, dt_from: datetime, dt_to: datetime) -> float:
    """Davr oylari uchun jami Salary.total."""
    months = set()
    d = dt_from.replace(day=1)
    while d <= dt_to:
        months.add((d.year, d.month))
        if d.month == 12:
            d = d.replace(year=d.year + 1, month=1)
        else:
            d = d.replace(month=d.month + 1)
    total = 0.0
    for y, m in months:
        sals = db.query(Salary).filter(Salary.year == y, Salary.month == m).all()
        total += sum(float(s.total or 0) for s in sals)
    return total


def _compute_daily_trend(sale_orders: list, sale_items: list) -> tuple:
    """Kunlik revenue + cogs → (labels, revenues, profits)."""
    daily_data: dict = {}
    for o in sale_orders:
        if not o.date:
            continue
        key = o.date.strftime("%Y-%m-%d")
        daily_data.setdefault(key, {"revenue": 0.0, "cogs": 0.0})["revenue"] += float(o.total or 0)
    order_by_id = {o.id: o for o in sale_orders}
    for oi, prod in sale_items:
        o = order_by_id.get(oi.order_id)
        if not o or not o.date:
            continue
        key = o.date.strftime("%Y-%m-%d")
        if key in daily_data:
            daily_data[key]["cogs"] += float(prod.purchase_price or 0) * float(oi.quantity or 0)
    labels = sorted(daily_data.keys())
    revenues = [daily_data[k]["revenue"] for k in labels]
    profits = [daily_data[k]["revenue"] - daily_data[k]["cogs"] for k in labels]
    return labels, revenues, profits


@router.get("/profit", response_class=HTMLResponse)
async def report_profit(
    request: Request,
    date_from: str = None,
    date_to: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    allowed = get_allowed_report_types(current_user)
    if "profit" not in allowed:
        return RedirectResponse(url="/reports", status_code=303)

    dt_from, dt_to, date_from, date_to = _parse_profit_date_range(date_from, date_to)

    # Sotuv + COGS + qaytarishlar
    sale_orders, revenue, cogs, sale_items = _compute_sales_and_cogs(db, dt_from, dt_to)
    sale_count = len(sale_orders)
    gross_profit = revenue - cogs
    return_orders = (
        db.query(Order)
        .filter(Order.type == "return_sale", Order.status != "cancelled",
                Order.date >= dt_from, Order.date <= dt_to)
        .all()
    )
    returns_total = sum(float(o.total or 0) for o in return_orders)

    # Xarid + operatsion xarajatlar + ish haqi
    purchases = (
        db.query(Purchase)
        .filter(Purchase.status == "confirmed", Purchase.date >= dt_from, Purchase.date <= dt_to)
        .all()
    )
    purchase_total = sum(float(p.total or 0) for p in purchases)
    purchase_expenses = sum(float(p.total_expenses or 0) for p in purchases)
    total_expenses, expense_list = _compute_operating_expenses(db, dt_from, dt_to)
    salary_total = _compute_salary_total(db, dt_from, dt_to)

    # To'lovlar
    _payment_sum = lambda ptype: float(db.query(func.coalesce(func.sum(Payment.amount), 0)).filter(
        Payment.type == ptype, Payment.status != "cancelled",
        Payment.date >= dt_from, Payment.date <= dt_to,
    ).scalar() or 0)
    payments_income = _payment_sum("income")
    payments_expense = _payment_sum("expense")

    # Yakuniy hisob + trend
    net_revenue = revenue - returns_total
    operating_expenses = total_expenses + salary_total
    net_profit = gross_profit - returns_total - operating_expenses
    daily_labels, daily_revenue, daily_profit = _compute_daily_trend(sale_orders, sale_items)

    return templates.TemplateResponse("reports/profit.html", {
        "request": request,
        "page_title": "Foyda hisoboti",
        "current_user": current_user,
        "date_from": date_from,
        "date_to": date_to,
        # Asosiy ko'rsatkichlar
        "revenue": revenue,
        "net_revenue": net_revenue,
        "cogs": cogs,
        "gross_profit": gross_profit,
        "returns_total": returns_total,
        "sale_count": sale_count,
        # Xarajatlar
        "purchase_total": purchase_total,
        "purchase_expenses": purchase_expenses,
        "total_expenses": total_expenses,
        "salary_total": salary_total,
        "operating_expenses": operating_expenses,
        "expense_list": expense_list,
        # Natija
        "net_profit": net_profit,
        # To'lovlar
        "payments_income": float(payments_income),
        "payments_expense": float(payments_expense),
        # Grafik
        "daily_labels": daily_labels,
        "daily_revenue": daily_revenue,
        "daily_profit": daily_profit,
        # Margin
        "gross_margin": round((gross_profit / revenue * 100) if revenue else 0, 1),
        "net_margin": round((net_profit / revenue * 100) if revenue else 0, 1),
    })


# ==================== SOTILGAN MAHSULOTLAR HISOBOTI ====================

@router.get("/sold-products", response_class=HTMLResponse)
async def sold_products_report(
    request: Request,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    warehouse_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_auth),
):
    """Sotilgan mahsulotlar — sana va ombor bo'yicha."""
    now = datetime.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = now.replace(hour=23, minute=59, second=59, microsecond=0)
    try:
        d_from = datetime.strptime(date_from, "%Y-%m-%dT%H:%M") if date_from and "T" in date_from else (datetime.strptime(date_from, "%Y-%m-%d") if date_from else today_start)
    except (ValueError, TypeError):
        d_from = today_start
    try:
        d_to = datetime.strptime(date_to, "%Y-%m-%dT%H:%M") if date_to and "T" in date_to else (datetime.strptime(date_to, "%Y-%m-%d").replace(hour=23, minute=59, second=59) if date_to else today_end)
    except (ValueError, TypeError):
        d_to = today_end

    warehouses = db.query(Warehouse).order_by(Warehouse.name).all()

    # Sotilgan mahsulotlar (OrderItem + Order)
    q = (
        db.query(
            Product.id,
            Product.name,
            func.sum(OrderItem.quantity).label("total_qty"),
            func.sum(OrderItem.total).label("total_sum"),
            func.count(func.distinct(Order.id)).label("order_count"),
        )
        .join(OrderItem, OrderItem.product_id == Product.id)
        .join(Order, Order.id == OrderItem.order_id)
        .filter(
            Order.type == "sale",
            Order.status == "completed",
            Order.created_at >= d_from,
            Order.created_at <= d_to,
        )
    )
    if warehouse_id:
        q = q.filter(Order.warehouse_id == warehouse_id)

    rows = q.group_by(Product.id, Product.name).order_by(func.sum(OrderItem.total).desc()).all()

    items = []
    grand_qty = 0
    grand_sum = 0
    for r in rows:
        qty = float(r.total_qty or 0)
        total = float(r.total_sum or 0)
        grand_qty += qty
        grand_sum += total
        items.append({
            "product_id": r.id,
            "product_name": r.name,
            "quantity": qty,
            "total": total,
            "order_count": r.order_count,
            "avg_price": round(total / qty, 0) if qty else 0,
        })

    return templates.TemplateResponse("reports/sold_products.html", {
        "request": request,
        "current_user": current_user,
        "page_title": "Sotilgan mahsulotlar",
        "items": items,
        "warehouses": warehouses,
        "date_from": d_from.strftime("%Y-%m-%d"),
        "date_to": d_to.strftime("%Y-%m-%d"),
        "selected_warehouse_id": warehouse_id,
        "grand_qty": grand_qty,
        "grand_sum": grand_sum,
    })
