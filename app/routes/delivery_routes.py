"""
Yetkazib berish — haydovchilar, yetkazishlar, xarita, supervayzer.
"""
from datetime import datetime
from typing import Optional
from fastapi import APIRouter, Request, Depends, Form, HTTPException
from sqlalchemy.orm import Session

from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from app.core import templates
from app.models.database import (
    get_db,
    Driver,
    DriverLocation,
    Delivery,
    Agent,
    AgentLocation,
    Visit,
    Partner,
    PartnerLocation,
    Order,
    OrderItem,
    User,
    AgentPayment,
    Payment,
    CashRegister,
    Product,
    Stock,
    Warehouse,
    EmployeeAdvance,
)
from app.deps import require_admin, require_admin_or_manager
from app.services.stock_service import create_stock_movement, delete_stock_movements_for_document
from app.constants import QUERY_LIMIT_DEFAULT, QUERY_LIMIT_HISTORY
from urllib.parse import quote
from sqlalchemy.orm import joinedload
from sqlalchemy import func
from app.logging_config import get_logger

logger = get_logger("delivery_routes")
router = APIRouter(tags=["delivery"])


def _resync_active_cash(db):
    """Inkassatsiya (agent/driver) oqimi confirmed kassa-Payment yaratganda/o'zgartirganda/
    o'chirganda kassa STORED balansini formula bilan qayta sinxronlash.

    Bu oqimlar avval sync_cash_balance umuman chaqirmasdi -> har inkassatsiyada
    CashRegister.balance formula'dan ortda qolib drift to'planardi (Asosiy kassa).
    Past chastotali supervisor amallari -> barcha faol kassani sync qilish xavfsiz."""
    from app.services.finance_service import sync_cash_balance
    db.flush()
    for _cr in db.query(CashRegister).filter(CashRegister.is_active == True).all():
        sync_cash_balance(db, _cr.id)


@router.get("/delivery", response_class=HTMLResponse)
async def delivery_list(request: Request, db: Session = Depends(get_db), current_user: User = Depends(require_admin_or_manager)):
    drivers = db.query(Driver).all()
    today = datetime.now().date()
    driver_ids = [d.id for d in drivers]
    # Batch loadlar (3N -> 3 ta umumiy query)
    last_loc_map: dict = {}
    today_count_map: dict = {}
    pending_count_map: dict = {}
    if driver_ids:
        # Oxirgi DriverLocation per driver — MAX(id) GROUP BY + load
        max_loc_ids = [
            r.max_id for r in
            db.query(DriverLocation.driver_id, func.max(DriverLocation.id).label("max_id"))
            .filter(DriverLocation.driver_id.in_(driver_ids))
            .group_by(DriverLocation.driver_id).all()
            if r.max_id
        ]
        if max_loc_ids:
            for loc in db.query(DriverLocation).filter(DriverLocation.id.in_(max_loc_ids)).all():
                last_loc_map[loc.driver_id] = loc
        # Bugungi yetkazishlar soni
        for r in (
            db.query(Delivery.driver_id, func.count(Delivery.id).label("cnt"))
            .filter(Delivery.driver_id.in_(driver_ids), Delivery.created_at >= today)
            .group_by(Delivery.driver_id).all()
        ):
            today_count_map[r.driver_id] = r.cnt
        # Pending yetkazishlar soni
        for r in (
            db.query(Delivery.driver_id, func.count(Delivery.id).label("cnt"))
            .filter(Delivery.driver_id.in_(driver_ids), Delivery.status == "pending")
            .group_by(Delivery.driver_id).all()
        ):
            pending_count_map[r.driver_id] = r.cnt
    for driver in drivers:
        driver.last_location = last_loc_map.get(driver.id)
        driver.today_deliveries = today_count_map.get(driver.id, 0)
        driver.pending_deliveries = pending_count_map.get(driver.id, 0)
    deliveries = db.query(Delivery).order_by(Delivery.created_at.desc()).limit(50).all()
    return templates.TemplateResponse("delivery/list.html", {
        "request": request,
        "current_user": current_user,
        "drivers": drivers,
        "deliveries": deliveries,
        "page_title": "Yetkazib berish",
    })


@router.post("/drivers/add")
async def driver_add(
    request: Request,
    full_name: str = Form(...),
    phone: str = Form(""),
    vehicle_number: str = Form(""),
    vehicle_type: str = Form(""),
    telegram_id: str = Form(""),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    last_driver = db.query(Driver).order_by(Driver.id.desc()).first()
    code = f"DR{str((last_driver.id if last_driver else 0) + 1).zfill(3)}"
    driver = Driver(
        code=code,
        full_name=full_name,
        phone=phone,
        vehicle_number=vehicle_number,
        vehicle_type=vehicle_type,
        telegram_id=telegram_id,
        is_active=True,
    )
    db.add(driver)
    db.commit()
    return RedirectResponse(url="/delivery", status_code=303)


@router.get("/delivery/{driver_id}", response_class=HTMLResponse)
async def driver_detail(
    request: Request,
    driver_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_manager),
):
    driver = db.query(Driver).filter(Driver.id == driver_id).first()
    if not driver:
        raise HTTPException(status_code=404, detail="Haydovchi topilmadi")
    locations = (
        db.query(DriverLocation)
        .filter(DriverLocation.driver_id == driver_id)
        .order_by(DriverLocation.recorded_at.desc())
        .limit(100)
        .all()
    )
    deliveries = (
        db.query(Delivery)
        .filter(Delivery.driver_id == driver_id)
        .order_by(Delivery.created_at.desc())
        .limit(30)
        .all()
    )
    return templates.TemplateResponse("delivery/detail.html", {
        "request": request,
        "current_user": current_user,
        "driver": driver,
        "locations": locations,
        "deliveries": deliveries,
        "page_title": f"Haydovchi: {driver.full_name}",
        "yandex_maps_apikey": _get_yandex_apikey(),
    })


@router.get("/map", response_class=HTMLResponse)
async def map_view(request: Request, db: Session = Depends(get_db), current_user: User = Depends(require_admin_or_manager)):
    agents = db.query(Agent).filter(Agent.is_active == True).all()
    agent_markers = []
    for agent in agents:
        last_loc = (
            db.query(AgentLocation)
            .filter(AgentLocation.agent_id == agent.id)
            .order_by(AgentLocation.recorded_at.desc())
            .first()
        )
        if last_loc:
            agent_markers.append({
                "id": agent.id,
                "name": agent.full_name,
                "type": "agent",
                "lat": last_loc.latitude,
                "lng": last_loc.longitude,
                "time": last_loc.recorded_at.strftime("%H:%M"),
            })
    drivers = db.query(Driver).filter(Driver.is_active == True).all()
    driver_markers = []
    for driver in drivers:
        last_loc = (
            db.query(DriverLocation)
            .filter(DriverLocation.driver_id == driver.id)
            .order_by(DriverLocation.recorded_at.desc())
            .first()
        )
        if last_loc:
            driver_markers.append({
                "id": driver.id,
                "name": driver.full_name,
                "type": "driver",
                "lat": last_loc.latitude,
                "lng": last_loc.longitude,
                "time": last_loc.recorded_at.strftime("%H:%M"),
                "vehicle": driver.vehicle_number,
            })
    # Agent dict (id -> name)
    agent_dict = {a.id: a.full_name for a in agents}
    # Partner mijozlar + agent statistika
    all_partners = db.query(Partner).filter(Partner.is_active == True).all()
    # Agent bo'yicha mijozlar soni (jami va xaritadagi)
    agent_partner_total = {}
    for p in all_partners:
        aid = p.agent_id or 0
        agent_partner_total[aid] = agent_partner_total.get(aid, 0) + 1

    partner_markers = []
    agent_partner_on_map = {}
    seen_ids = set()
    # 1) Partner.latitude/longitude dan
    for p in all_partners:
        if p.latitude and p.longitude:
            aid = p.agent_id or 0
            agent_partner_on_map[aid] = agent_partner_on_map.get(aid, 0) + 1
            seen_ids.add(p.id)
            partner_markers.append({
                "id": p.id, "name": p.name, "type": "partner",
                "lat": p.latitude, "lng": p.longitude,
                "address": p.address or "",
                "agent_id": aid,
                "agent_name": agent_dict.get(aid, "Tayinlanmagan"),
            })
    # 2) PartnerLocation dan (Partner.lat bo'sh bo'lganlar)
    partner_locations = db.query(PartnerLocation).all()
    for loc in partner_locations:
        if loc.partner_id in seen_ids or not loc.latitude or not loc.longitude:
            continue
        partner = db.query(Partner).filter(Partner.id == loc.partner_id).first()
        if partner:
            aid = partner.agent_id or 0
            agent_partner_on_map[aid] = agent_partner_on_map.get(aid, 0) + 1
            seen_ids.add(partner.id)
            partner_markers.append({
                "id": loc.partner_id, "name": partner.name, "type": "partner",
                "lat": loc.latitude, "lng": loc.longitude,
                "address": loc.address or partner.address or "",
                "agent_id": aid,
                "agent_name": agent_dict.get(aid, "Tayinlanmagan"),
            })
    # Agent ro'yxati (chap panel uchun)
    agent_list = []
    for a in agents:
        on_map = agent_partner_on_map.get(a.id, 0)
        total = agent_partner_total.get(a.id, 0)
        if total > 0:
            agent_list.append({"id": a.id, "name": a.full_name, "on_map": on_map, "total": total})
    # Tayinlanmagan
    unassigned_on_map = agent_partner_on_map.get(0, 0)
    unassigned_total = agent_partner_total.get(0, 0)
    try:
        from app.config.maps_config import MAP_PROVIDER
        map_provider = MAP_PROVIDER
    except Exception:
        map_provider = "yandex"
    try:
        from app.config.maps_config import YANDEX_MAPS_API_KEY
        yandex_apikey = YANDEX_MAPS_API_KEY or ""
    except Exception:
        yandex_apikey = ""
    return templates.TemplateResponse("map/index.html", {
        "request": request,
        "current_user": current_user,
        "agents": agents,
        "drivers": drivers,
        "partner_locations": partner_locations,
        "agent_markers": agent_markers,
        "driver_markers": driver_markers,
        "partner_markers": partner_markers,
        "region_markers": [],
        "map_provider": map_provider,
        "yandex_maps_apikey": yandex_apikey,
        "agent_list": agent_list,
        "unassigned_on_map": unassigned_on_map,
        "unassigned_total": unassigned_total,
        "page_title": "Xarita",
    })


@router.get("/supervisor", response_class=HTMLResponse)
async def supervisor_dashboard(request: Request, db: Session = Depends(get_db), current_user: User = Depends(require_admin_or_manager)):
    today = datetime.now().date()
    total_agents = db.query(Agent).filter(Agent.is_active == True).count()
    active_agents = 0
    for agent in db.query(Agent).filter(Agent.is_active == True).all():
        last_loc = (
            db.query(AgentLocation)
            .filter(AgentLocation.agent_id == agent.id, AgentLocation.recorded_at >= today)
            .first()
        )
        if last_loc:
            active_agents += 1
    today_visits = db.query(Visit).filter(Visit.visit_date >= today).count()
    today_orders = db.query(Order).filter(Order.type == "sale", Order.date >= today).all()
    today_sales_sum = sum(o.total for o in today_orders)
    total_drivers = db.query(Driver).filter(Driver.is_active == True).count()
    pending_deliveries = db.query(Delivery).filter(Delivery.status == "pending").count()
    today_delivered = (
        db.query(Delivery)
        .filter(Delivery.status == "delivered", Delivery.delivered_at >= today)
        .count()
    )
    agent_stats = []
    for agent in db.query(Agent).filter(Agent.is_active == True).all():
        visits = db.query(Visit).filter(Visit.agent_id == agent.id, Visit.visit_date >= today).count()
        last_loc = (
            db.query(AgentLocation)
            .filter(AgentLocation.agent_id == agent.id)
            .order_by(AgentLocation.recorded_at.desc())
            .first()
        )
        agent_stats.append({
            "agent": agent,
            "visits": visits,
            "last_seen": last_loc.recorded_at if last_loc else None,
            "is_online": last_loc and (datetime.now() - last_loc.recorded_at).seconds < 600 if last_loc else False,
        })
    agent_stats.sort(key=lambda x: x["visits"], reverse=True)
    # O'tgan hafta vizitlar
    from datetime import timedelta
    week_ago = today - timedelta(days=7)
    visits_week = db.query(Visit).filter(Visit.visit_date >= week_ago).count()
    # Yo'ldagi yetkazishlar
    in_transit = db.query(Delivery).filter(Delivery.status.in_(["pending", "in_transit"])).count()
    stats = {
        "total_agents": total_agents,
        "active_agents": active_agents,
        "today_visits": today_visits,
        "visits_today": today_visits,
        "visits_week": visits_week,
        "today_orders": len(today_orders),
        "today_sales": today_sales_sum,
        "total_drivers": total_drivers,
        "pending_deliveries": pending_deliveries,
        "today_delivered": today_delivered,
        "delivered_today": today_delivered,
        "in_transit": in_transit,
    }
    agents = db.query(Agent).filter(Agent.is_active == True).all()
    drivers = db.query(Driver).filter(Driver.is_active == True).all()
    # Agent va driver markerlar (mini xarita uchun)
    agent_markers = []
    for agent in agents:
        last_loc = (
            db.query(AgentLocation)
            .filter(AgentLocation.agent_id == agent.id)
            .order_by(AgentLocation.recorded_at.desc())
            .first()
        )
        if last_loc:
            agent.last_location = last_loc
            agent.is_online = (datetime.now() - last_loc.recorded_at).total_seconds() < 600
            if last_loc.latitude is not None and last_loc.longitude is not None:
                agent_markers.append({
                    "id": agent.id,
                    "name": agent.full_name,
                    "type": "agent",
                    "lat": float(last_loc.latitude),
                    "lng": float(last_loc.longitude),
                    "time": last_loc.recorded_at.strftime("%H:%M"),
                })
        else:
            agent.is_online = False
    driver_markers = []
    for driver in drivers:
        last_loc = (
            db.query(DriverLocation)
            .filter(DriverLocation.driver_id == driver.id)
            .order_by(DriverLocation.recorded_at.desc())
            .first()
        )
        if last_loc:
            driver.last_location = last_loc
            if last_loc.latitude is not None and last_loc.longitude is not None:
                driver_markers.append({
                    "id": driver.id,
                    "name": driver.full_name,
                    "type": "driver",
                    "lat": float(last_loc.latitude),
                    "lng": float(last_loc.longitude),
                    "time": last_loc.recorded_at.strftime("%H:%M"),
                    "vehicle": driver.vehicle_number or "",
                })
    recent_visits = db.query(Visit).filter(Visit.visit_date >= today).order_by(Visit.visit_date.desc()).limit(20).all()
    recent_deliveries = db.query(Delivery).order_by(Delivery.created_at.desc()).limit(10).all()
    # Agent buyurtmalari (bugungi)
    agent_orders = (
        db.query(Order)
        .filter(Order.source == "agent", Order.date >= today)
        .order_by(Order.created_at.desc())
        .limit(20)
        .all()
    )

    # Oylik savdo rejasi widget (joriy oy)
    from app.models.database import SalesPlan
    month_start = today.replace(day=1)
    current_period = today.strftime("%Y-%m")
    sales_plan = db.query(SalesPlan).filter(SalesPlan.period == current_period).first()
    plan_amount = float(sales_plan.amount) if sales_plan else 0.0
    sales_plan_rows = []
    for agent in agents:
        sold = db.query(func.coalesce(func.sum(Order.total), 0)).filter(
            Order.agent_id == agent.id,
            Order.source == "agent",
            Order.type == "sale",  # return_sale (obmen qaytarish) rejaga kirmasin
            Order.status.in_(("confirmed", "completed", "waiting_production", "delivered")),
            Order.date >= month_start,
        ).scalar() or 0.0
        sold = float(sold)
        pct = round((sold / plan_amount * 100), 1) if plan_amount > 0 else 0.0
        sales_plan_rows.append({"agent": agent, "sold": sold, "percent": pct})
    sales_plan_rows.sort(key=lambda x: x["sold"], reverse=True)

    return templates.TemplateResponse("supervisor/dashboard.html", {
        "request": request,
        "current_user": current_user,
        "stats": stats,
        "agents": agents,
        "drivers": drivers,
        "agent_markers": agent_markers,
        "driver_markers": driver_markers,
        "agent_stats": agent_stats[:10],
        "recent_visits": recent_visits,
        "recent_deliveries": recent_deliveries,
        "agent_orders": agent_orders,
        "agent_payments": [p for p in _get_pending_agent_payments(db)],
        "sales_plan_amount": plan_amount,
        "sales_plan_period": current_period,
        "sales_plan_rows": sales_plan_rows,
        "yandex_maps_apikey": _get_yandex_apikey(),
        "page_title": "Supervayzer",
        "now": datetime.now(),
    })


@router.get("/supervisor/agent-orders", response_class=HTMLResponse)
async def supervisor_agent_orders(
    request: Request,
    status: str = "all",
    agent_id: int = 0,
    date_from: str = None,
    date_to: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_manager),
):
    """Barcha agentlar buyurtmalari bitta board'da — bulk tasdiqlash + yo'lga chiqarish.

    Per-agent (/agents/{id}) sahifasidagi bulk UX (checkbox + bulk confirm/dispatch/revert/
    Excel + sana darchasi + qidiruv) shu yerda BARCHA agentlar uchun birlashgan. Bulk
    endpointlar order ID bilan ishlaydi (agent-agnostik): /supervisor/agent-orders/confirm,
    /sales/{id}/dispatch, /sales/{id}/revert. Status tab + agent filtri qo'shimcha.
    """
    from datetime import datetime as _dt
    from app.models.database import Delivery

    today = _dt.now().replace(hour=0, minute=0, second=0, microsecond=0)
    try:
        d_from = _dt.strptime(date_from, "%Y-%m-%d") if date_from else today
    except (ValueError, TypeError):
        d_from = today
    try:
        d_to = (_dt.strptime(date_to, "%Y-%m-%d") if date_to else today).replace(hour=23, minute=59, second=59)
    except (ValueError, TypeError):
        d_to = today.replace(hour=23, minute=59, second=59)
    sel_agent_id = agent_id if (agent_id and agent_id > 0) else None

    q = db.query(Order).filter(
        Order.source == "agent",
        Order.date >= d_from,
        Order.date <= d_to,
        Order.parent_order_id.is_(None),  # obmen child (sale) yashirin, parent ko'rsatiladi
    ).options(joinedload(Order.partner), joinedload(Order.items))
    if status and status != "all":
        q = q.filter(Order.status == status)
    if sel_agent_id:
        q = q.filter(Order.agent_id == sel_agent_id)
    orders = q.order_by(Order.id.desc()).all()

    agent_names = {a.id: a.full_name for a in db.query(Agent).all()}

    # Obmen parent -> child sale
    exchange_children = {}
    if orders:
        parent_ids = [o.id for o in orders if o.type == "return_sale"]
        if parent_ids:
            for ch in db.query(Order).filter(Order.parent_order_id.in_(parent_ids)).all():
                exchange_children[ch.parent_order_id] = ch

    # Har order uchun haydovchi + yetkazilgan vaqt (obmen child orqali ham)
    order_drivers, order_delivered_at = {}, {}
    if orders:
        order_ids = [o.id for o in orders]
        child_to_parent = {}
        for parent_id, ch in exchange_children.items():
            order_ids.append(ch.id)
            child_to_parent[ch.id] = parent_id
        deliveries = db.query(Delivery).filter(Delivery.order_id.in_(order_ids)).all()
        driver_map = {d.id: d.full_name for d in db.query(Driver).filter(
            Driver.id.in_([dl.driver_id for dl in deliveries if dl.driver_id])).all()}
        for dl in deliveries:
            if dl.driver_id and dl.driver_id in driver_map:
                order_drivers.setdefault(dl.order_id, driver_map[dl.driver_id])
                _pid = child_to_parent.get(dl.order_id)
                if _pid:
                    order_drivers.setdefault(_pid, driver_map[dl.driver_id])
            if dl.delivered_at:
                order_delivered_at.setdefault(dl.order_id, dl.delivered_at)
                _pid = child_to_parent.get(dl.order_id)
                if _pid:
                    order_delivered_at.setdefault(_pid, dl.delivered_at)
    # Topilma 2A: waiting_production buyurtma uchun qaysi Production (qaysi mahsulot) ko'rsatilsin
    from app.models.database import Production as _Production, Stock as _Stock
    production_info = {}
    missing_items = {}  # {order_id: [{product, need, have, missing}]}
    _wp_orders = [o for o in orders if o.status == "waiting_production"]
    _wp_ids = [o.id for o in _wp_orders]
    if _wp_ids:
        for p in db.query(_Production).options(joinedload(_Production.recipe)).filter(
            _Production.order_id.in_(_wp_ids)
        ).all():
            rn = getattr(getattr(p, "recipe", None), "name", None) or "—"
            production_info.setdefault(p.order_id, []).append(
                f"{p.number} · {rn} · {float(p.quantity or 0):g} · {p.status}"
            )
        # Yetishmayotgan mahsulotlarni hisoblash (har order uchun)
        for o in _wp_orders:
            missing = []
            for it in (o.items or []):
                if not it.product_id or not (it.quantity or 0) > 0:
                    continue
                wh_id = it.warehouse_id or o.warehouse_id
                if not wh_id:
                    continue
                stock = db.query(_Stock).filter(
                    _Stock.warehouse_id == wh_id, _Stock.product_id == it.product_id
                ).first()
                have = float(stock.quantity or 0) if stock else 0.0
                need = float(it.quantity or 0)
                gap = need - have
                if gap > 0.01:
                    pname = it.product.name if it.product else f"#{it.product_id}"
                    missing.append({"product": pname, "need": need, "have": have, "missing": gap})
            if missing:
                missing_items[o.id] = missing
    drivers = db.query(Driver).filter(Driver.is_active == True).order_by(Driver.full_name).all()
    agents = db.query(Agent).filter(Agent.is_active == True).order_by(Agent.full_name).all()
    draft_count = db.query(func.count(Order.id)).filter(Order.source == "agent", Order.status == "draft").scalar() or 0
    waiting_count = db.query(func.count(Order.id)).filter(Order.source == "agent", Order.status == "waiting_production").scalar() or 0

    total_count = sum(1 for o in orders if o.status != "cancelled")
    total_sum = sum(float(o.total or 0) for o in orders if o.status != "cancelled")
    is_today_only = (d_from.date() == today.date() and d_to.date() == today.date())
    if is_today_only:
        range_label = "Bugun"
    elif d_from.date() == d_to.date():
        range_label = d_from.strftime("%d.%m.%Y")
    else:
        range_label = f"{d_from.strftime('%d.%m')}–{d_to.strftime('%d.%m')}"

    return templates.TemplateResponse("supervisor/agent_orders.html", {
        "request": request,
        "current_user": current_user,
        "orders": orders,
        "drivers": drivers,
        "active_drivers": drivers,
        "agents": agents,
        "agent_names": agent_names,
        "sel_agent_id": sel_agent_id,
        "exchange_children": exchange_children,
        "order_drivers": order_drivers,
        "order_delivered_at": order_delivered_at,
        "current_status": status,
        "draft_count": draft_count,
        "waiting_count": waiting_count,
        "production_info": production_info,
        "missing_items": missing_items,
        "today_iso": _dt.now().date().isoformat(),
        "today_orders_count": total_count,
        "today_orders_total": total_sum,
        "range_label": range_label,
        "date_from": d_from.strftime("%Y-%m-%d"),
        "date_to": d_to.strftime("%Y-%m-%d"),
        "page_title": "Agent buyurtmalari",
    })


@router.get("/supervisor/agent-orders/partners")
async def supervisor_agent_order_partners(
    agent_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_manager),
):
    """Tanlangan agentning mijozlari (board'dagi 'Yangi buyurtma' modal uchun)."""
    partners = (
        db.query(Partner)
        .filter(Partner.agent_id == agent_id, Partner.is_active == True)
        .order_by(Partner.name)
        .all()
    )
    return {"partners": [{"id": p.id, "name": p.name, "phone": p.phone or ""} for p in partners]}


@router.get("/supervisor/agent-orders/products")
async def supervisor_agent_order_products(
    partner_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_manager),
):
    """Agent mahsulotlari + tanlangan mijoz narx turi bo'yicha narxlar."""
    from app.routes.api_agent_ops import _resolve_price_type_id, _product_price_for_type
    partner = db.query(Partner).filter(Partner.id == partner_id).first()
    if not partner:
        return {"products": [], "error": "Mijoz topilmadi"}
    pt = _resolve_price_type_id(partner)
    prods = (
        db.query(Product)
        .filter(Product.is_active == True, Product.is_for_agent == True)
        .order_by(Product.name)
        .all()
    )
    out = [
        {
            "id": p.id,
            "name": p.name,
            "price": float(_product_price_for_type(db, p, pt) or 0),
            "unit": (p.unit.name if getattr(p, "unit", None) else ""),
        }
        for p in prods
    ]
    return {"products": out, "discount_percent": float(partner.discount_percent or 0)}


@router.post("/supervisor/agent-orders/create")
async def supervisor_create_agent_order(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_manager),
):
    """Admin/menejer agent nomidan qo'lda buyurtma yaratadi (draft). Board'dan tasdiqlanadi.

    JSON: {agent_id, partner_id, items:[{product_id, qty}], note?}.
    Agent ilovasidagi create bilan bir xil narx/chegirma mantig'i, lekin token emas —
    agent_id formadan. MERGE qilinmaydi (har doim yangi mustaqil draft sale).
    """
    from app.routes.api_agent_ops import _resolve_price_type_id, _product_price_for_type
    try:
        body = await request.json()
        agent_id = int(body.get("agent_id") or 0)
        partner_id = int(body.get("partner_id") or 0)
        items = body.get("items", [])
        note = (body.get("note") or "").strip()[:200]

        agent = db.query(Agent).filter(Agent.id == agent_id).first()
        if not agent:
            return {"success": False, "error": "Agent tanlanmagan"}
        partner = db.query(Partner).filter(Partner.id == partner_id, Partner.agent_id == agent_id).first()
        if not partner:
            return {"success": False, "error": "Mijoz topilmadi (agentga biriktirilmagan)"}
        if not items:
            return {"success": False, "error": "Mahsulot tanlang"}

        warehouse = db.query(Warehouse).filter(Warehouse.name.ilike("%tayyor mahsulot%"), Warehouse.is_active == True).first()
        if not warehouse:
            warehouse = db.query(Warehouse).filter(Warehouse.name.ilike("%tayyor%"), Warehouse.is_active == True).first()
        if not warehouse:
            warehouse = db.query(Warehouse).filter(Warehouse.is_active == True).first()
        if not warehouse:
            return {"success": False, "error": "Ombor topilmadi"}

        today = datetime.now()
        prefix = f"AGT-{today.strftime('%Y%m%d')}"
        last = db.query(Order).filter(Order.number.like(f"{prefix}%")).order_by(Order.id.desc()).first()
        try:
            seq = int(last.number.split("-")[-1]) + 1 if last and last.number else 1
        except (ValueError, IndexError, AttributeError):
            seq = 1
        order_number = f"{prefix}-{seq:03d}"

        partner_discount = float(partner.discount_percent or 0)
        price_type_id = _resolve_price_type_id(partner)
        subtotal = 0.0
        order_items = []
        for it in items:
            try:
                pid = int(it.get("product_id") or 0)
                qty = float(it.get("qty", it.get("quantity", 0)) or 0)
            except (ValueError, TypeError):
                continue
            if pid <= 0 or qty <= 0:
                continue
            prod = db.query(Product).filter(Product.id == pid, Product.is_active == True, Product.is_for_agent == True).first()
            if not prod:
                continue
            price = _product_price_for_type(db, prod, price_type_id)
            line = qty * price
            subtotal += line
            order_items.append(OrderItem(
                product_id=prod.id, quantity=qty, price=price,
                discount_percent=partner_discount,
                total=line * (1 - partner_discount / 100),
            ))
        if not order_items:
            return {"success": False, "error": "Yaroqli mahsulot yo'q"}

        discount_amount = subtotal * partner_discount / 100
        total = subtotal - discount_amount
        order = Order(
            number=order_number, date=today, type="sale",
            partner_id=partner.id, warehouse_id=warehouse.id,
            agent_id=agent.id, source="agent", price_type_id=price_type_id,
            subtotal=subtotal, discount_percent=partner_discount, discount_amount=discount_amount,
            total=total, paid=0, debt=total, status="draft", payment_type="naqd",
            note=(note + f" [Qo'lda: {getattr(current_user, 'username', '')}]").strip(),
        )
        db.add(order)
        db.flush()
        for oi in order_items:
            oi.order_id = order.id
            db.add(oi)
        db.commit()
        logger.info(f"Supervisor manual agent order: {order_number} agent={agent.code} partner={partner.id} total={total}")
        return {"success": True, "order_id": order.id, "order_number": order_number, "total": total}
    except Exception as e:
        db.rollback()
        logger.error(f"supervisor_create_agent_order: {e}")
        return {"success": False, "error": "Server xatosi"}


@router.post("/supervisor/agent-orders/exchange-create")
async def supervisor_create_agent_exchange(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_manager),
):
    """Admin/menejer agent nomidan ALMASHTIRISH (obmen) yaratadi — draft.

    JSON: {agent_id, partner_id, return_items:[{product_id,qty}], new_items:[{product_id,qty}], note?}.
    Tarixsiz obmen (parent buyurtma kerak emas), agent ilovasidagi standalone-exchange bilan
    bir xil: return_sale (Vozvrat ombori) + child sale (Tayyor mahsulot). Narxlar server tomonda
    mijoz narx turi bo'yicha. Balans farqi = yangi − qaytgan (mijoz faqat farqni qarz oladi).
    """
    from app.routes.api_agent_ops import _resolve_price_type_id, _product_price_for_type, VOZVRAT_WAREHOUSE_ID
    try:
        body = await request.json()
        agent_id = int(body.get("agent_id") or 0)
        partner_id = int(body.get("partner_id") or 0)
        return_items = body.get("return_items", [])
        new_items = body.get("new_items", [])
        note = (body.get("note") or "").strip()[:200]

        agent = db.query(Agent).filter(Agent.id == agent_id).first()
        if not agent:
            return {"success": False, "error": "Agent tanlanmagan"}
        partner = db.query(Partner).filter(Partner.id == partner_id, Partner.agent_id == agent_id).first()
        if not partner:
            return {"success": False, "error": "Mijoz topilmadi (agentga biriktirilmagan)"}
        if not return_items or not new_items:
            return {"success": False, "error": "Qaytariladigan va yangi mahsulot kerak"}

        new_warehouse = db.query(Warehouse).filter(Warehouse.name.ilike("%tayyor mahsulot%"), Warehouse.is_active == True).first()
        if not new_warehouse:
            new_warehouse = db.query(Warehouse).filter(Warehouse.is_active == True).first()
        if not new_warehouse:
            return {"success": False, "error": "Ombor topilmadi"}

        price_type_id = _resolve_price_type_id(partner)
        partner_discount = float(partner.discount_percent or 0)

        def _lines(raw, with_discount):
            oi, subtotal = [], 0.0
            for it in raw:
                try:
                    pid = int(it.get("product_id") or 0)
                    qty = float(it.get("qty", it.get("quantity", 0)) or 0)
                except (ValueError, TypeError):
                    continue
                if pid <= 0 or qty <= 0:
                    continue
                prod = db.query(Product).filter(Product.id == pid, Product.is_active == True, Product.is_for_agent == True).first()
                if not prod:
                    continue
                price = _product_price_for_type(db, prod, price_type_id)
                line = qty * price
                subtotal += line
                disc = partner_discount if with_discount else 0
                oi.append(OrderItem(product_id=pid, quantity=qty, price=price, discount_percent=disc, total=line * (1 - disc / 100)))
            return oi, subtotal

        ret_oi, ret_subtotal = _lines(return_items, with_discount=False)
        new_oi, new_subtotal = _lines(new_items, with_discount=True)
        if not ret_oi or not new_oi:
            return {"success": False, "error": "Yaroqli mahsulotlar yo'q"}
        new_total = new_subtotal - (new_subtotal * partner_discount / 100)

        today = datetime.now()
        prefix = f"AGT-{today.strftime('%Y%m%d')}"
        last = db.query(Order).filter(Order.number.like(f"{prefix}%")).order_by(Order.id.desc()).first()
        try:
            seq = int(last.number.split("-")[-1]) + 1 if last and last.number else 1
        except (ValueError, IndexError, AttributeError):
            seq = 1

        ret_order = Order(
            number=f"{prefix}-{seq:03d}", date=today, type="return_sale",
            partner_id=partner_id, warehouse_id=VOZVRAT_WAREHOUSE_ID,
            agent_id=agent.id, source="agent", price_type_id=price_type_id,
            subtotal=ret_subtotal, discount_percent=0, discount_amount=0,
            total=ret_subtotal, paid=0, debt=0, status="draft", payment_type="naqd",
            note=f"OBMEN qaytarish (qo'lda: {getattr(current_user, 'username', '')}): {note}. Agent: {agent.code}",
        )
        db.add(ret_order)
        db.flush()
        for oi in ret_oi:
            oi.order_id = ret_order.id
            db.add(oi)

        obmen_debt = max(0.0, new_total - ret_subtotal)
        new_order = Order(
            number=f"{prefix}-{seq + 1:03d}", date=today, type="sale",
            partner_id=partner_id, warehouse_id=new_warehouse.id,
            agent_id=agent.id, source="agent", price_type_id=price_type_id,
            subtotal=new_subtotal, discount_percent=partner_discount,
            discount_amount=new_subtotal * partner_discount / 100,
            total=new_total, paid=0, debt=obmen_debt, status="draft", payment_type="naqd",
            note=f"OBMEN chiqarish (qo'lda): return={ret_order.number}, farq={obmen_debt:.0f}. Agent: {agent.code}",
            parent_order_id=ret_order.id,
        )
        db.add(new_order)
        db.flush()
        for oi in new_oi:
            oi.order_id = new_order.id
            db.add(oi)

        db.commit()
        logger.info(f"Supervisor manual exchange: {ret_order.number}/{new_order.number} agent={agent.code} diff={new_total - ret_subtotal}")
        return {
            "success": True,
            "return_order_number": ret_order.number, "new_order_number": new_order.number,
            "balance_diff": new_total - ret_subtotal,
        }
    except Exception as e:
        db.rollback()
        logger.error(f"supervisor_create_agent_exchange: {e}")
        return {"success": False, "error": "Server xatosi"}


@router.post("/supervisor/agent-orders/confirm/{order_id}")
async def supervisor_confirm_agent_order(
    request: Request,
    order_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_manager),
):
    """Agent buyurtmasini tasdiqlash — faqat status o'zgaradi.

    Yangi flow (POS bilan moslashtirilgan, 2026-05-12):
      1) confirm  → status='confirmed' (stock/balance/Delivery TEGMAYDI)
      2) dispatch (/sales/{id}/dispatch) → stock chiqarish + Delivery yaratish
      3) driver "Yetkazdim" → partner balance + delivered

    return_sale (obmen qaytarish) ham endi oddiy sotuv kabi dispatch → driver
    oqimidan o'tadi. Qaytgan tovar jismonan haydovchi "Yetkazdim" bosganda keladi
    (apply_return_stock_addition shu yerda emas, api_driver_ops.py da chaqiriladi).
    """
    order = db.query(Order).filter(Order.id == order_id, Order.source == "agent").first()
    if not order:
        return RedirectResponse(url="/supervisor/agent-orders?error=not_found", status_code=303)
    if order.status not in ("draft",):
        return RedirectResponse(url="/supervisor/agent-orders?error=already_confirmed", status_code=303)

    # D4 audit fix: agent buyurtma tasdiqlashda kredit limit tekshiruvi (soft guard).
    # Balance dispatch/delivered bosqichida yoziladi, lekin foydalanuvchini erta ogohlantirish uchun
    # qoldirilgan.
    if order.type != "return_sale" and order.partner_id and float(order.debt or 0) > 0:
        from app.services.partner_credit import check_credit_limit
        partner = db.query(Partner).filter(Partner.id == order.partner_id).first()
        ok, err = check_credit_limit(partner, float(order.debt or 0))
        if not ok:
            return RedirectResponse(
                url="/supervisor/agent-orders?error=" + quote(err),
                status_code=303,
            )

    # Atomik UPDATE WHERE — double-confirm xavfini oldini olish.
    from sqlalchemy import text as _text
    claim = db.execute(
        _text("UPDATE orders SET status='confirmed' WHERE id=:id AND source='agent' AND status='draft'"),
        {"id": order_id},
    )
    if claim.rowcount == 0:
        return RedirectResponse(url="/supervisor/agent-orders?error=already_confirmed", status_code=303)
    db.refresh(order)

    # Oddiy sotuv VA obmen (return_sale): faqat status va user_id.
    # Stock/balance/Delivery — dispatch bosqichida; obmen qaytgan tovar kirimi
    # haydovchi "Yetkazdim" bosganda (api_driver_ops.py) amalga oshiriladi.
    order.user_id = current_user.id
    # Obmen child (yangi sotuv) ham birga tasdiqlanadi — dispatch UI tugmasi
    # ex_child.status=='confirmed' shartiga bog'liq (agents/detail.html).
    # b881e3c'da bu blok early-return va apply_return_stock_addition bilan birga
    # noto'g'ri o'chirilgan edi (regressiya, 2 orphan: AGT-20260518-005,
    # AGT-20260519-011); endi tiklandi.
    if order.type == "return_sale":
        db.execute(
            _text("UPDATE orders SET status='confirmed', user_id=:uid "
                  "WHERE parent_order_id=:pid AND type='sale' AND status='draft'"),
            {"uid": current_user.id, "pid": order.id},
        )
    db.commit()
    try:
        from app.bot.services.audit_watchdog import audit_agent_order_confirm
        audit_agent_order_confirm(order.id)
    except Exception:
        pass
    return RedirectResponse(url="/supervisor/agent-orders", status_code=303)


@router.post("/supervisor/agent-orders/reject/{order_id}")
async def supervisor_reject_agent_order(
    request: Request,
    order_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_manager),
):
    """Agent buyurtmasini bekor qilish."""
    # Race condition oldini olish
    order = db.query(Order).filter(Order.id == order_id, Order.source == "agent").with_for_update().first()
    if not order:
        return RedirectResponse(url="/supervisor/agent-orders?error=not_found", status_code=303)
    # Bug 1 — paid > 0 bo'lgan orderni darhol bekor qilish taqiqlanadi (refund kerak)
    if (order.paid or 0) > 0:
        referer = request.headers.get("referer", "/supervisor/agent-orders")
        sep = "&" if "?" in referer else "?"
        return RedirectResponse(
            url=f"{referer}{sep}error=paid_block&detail=" + quote(
                f"Bu buyurtmaga {order.paid:,.0f} so'm to'lov qabul qilingan. Avval refund qiling, keyin bekor qilish mumkin."
            ),
            status_code=303,
        )
    # Agar tasdiqlangan bo'lsa — stock qaytarish + yetkazishni ham bekor qilish
    if order.status == "confirmed":
        for it in order.items:
            if not it.product_id or not (it.quantity or 0) > 0:
                continue
            wh_id = it.warehouse_id if it.warehouse_id else order.warehouse_id
            if not wh_id:
                continue
            create_stock_movement(
                db=db,
                warehouse_id=wh_id,
                product_id=it.product_id,
                quantity_change=+float(it.quantity or 0),
                operation_type="sale_revert",
                document_type="Sale",
                document_id=order.id,
                document_number=order.number,
                user_id=current_user.id if current_user else None,
                note=f"Agent buyurtma bekor (reject): {order.number}",
                created_at=datetime.now(),
            )
        from app.models.database import Delivery as DeliveryModel
        # Bug 6 — failed delivery ham cancelled bo'lsin
        for delivery in db.query(DeliveryModel).filter(
            DeliveryModel.order_id == order.id,
            DeliveryModel.status.in_(["pending", "in_progress", "failed"]),
        ).all():
            delivery.status = "cancelled"
    order.status = "cancelled"
    # Site 1: delivery_revert — recompute partner balance after order cancelled
    if order.partner_id:
        from app.services.partner_balance_service import recompute_partner_balance, recompute_partner_order_debts
        db.flush()
        recompute_partner_balance(
            db, order.partner_id,
            reason="delivery_revert",
            ref=order.number,
            actor=current_user.username if current_user else None,
        )
    db.commit()
    referer = request.headers.get("referer", "/supervisor/agent-orders")
    return RedirectResponse(url=referer, status_code=303)


@router.post("/supervisor/agent-orders/delete/{order_id}")
async def supervisor_delete_agent_order(
    request: Request,
    order_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_manager),
):
    """Agent buyurtmasini o'chirish (faqat bekor qilingan yoki draft).
    Obmen juftligi (return_sale + parent_order_id'li child sale) avtomatik birga o'chiriladi.
    """
    from app.models.database import OrderItem as OI, Delivery as DeliveryModel
    order = db.query(Order).filter(Order.id == order_id, Order.source == "agent").first()
    if not order:
        return RedirectResponse(url="/supervisor/agent-orders?error=not_found", status_code=303)
    if order.status == "confirmed":
        return RedirectResponse(
            url="/supervisor/agent-orders?error=confirmed_delete&detail=" + quote("Tasdiqlangan buyurtmani to'g'ridan-to'g'ri o'chirib bo'lmaydi. Avval bekor qiling."),
            status_code=303,
        )

    # Obmen juftligini topish (FK constraint uchun ikkalasini birga o'chirish kerak)
    to_delete = [order]
    if order.type == "return_sale":
        # Parent ret_order: child sale topib qo'shish
        child = db.query(Order).filter(Order.parent_order_id == order.id, Order.type == "sale").first()
        if child:
            if child.status not in ("draft", "cancelled"):
                return RedirectResponse(
                    url="/supervisor/agent-orders?error=obmen_pair&detail=" + quote(
                        f"Bu obmen juftligi: child {child.number} status='{child.status}'. Avval ikkalasini bekor qiling."
                    ),
                    status_code=303,
                )
            to_delete.append(child)
    elif order.parent_order_id:
        # Child sale: parent ret_order topib qo'shish (agar mavjud bo'lsa va ham cancelled bo'lsa)
        parent = db.query(Order).filter(Order.id == order.parent_order_id).first()
        if parent and parent.status in ("draft", "cancelled"):
            to_delete.append(parent)
        elif parent:
            return RedirectResponse(
                url="/supervisor/agent-orders?error=obmen_pair&detail=" + quote(
                    f"Bu obmen yangi sotuvi: parent {parent.number} status='{parent.status}'. Avval ikkalasini bekor qiling."
                ),
                status_code=303,
            )

    for o in to_delete:
        if o.status == "cancelled":
            delete_stock_movements_for_document(db, "Sale", o.id)
        db.query(DeliveryModel).filter(DeliveryModel.order_id == o.id).delete()
        db.query(OI).filter(OI.order_id == o.id).delete()

    # Sale child avval o'chirilsin (FK constraint), keyin parent ret_order.
    # SQLAlchemy executemany ni oldini olish uchun flush() har biri uchun chaqiriladi.
    to_delete.sort(key=lambda o: 0 if o.parent_order_id else 1)
    for o in to_delete:
        db.delete(o)
        db.flush()  # Majburiy: SQLAlchemy batch deletes ni qilmasin
    db.commit()
    return RedirectResponse(url="/supervisor/agent-orders", status_code=303)


def _get_yandex_apikey():
    try:
        from app.config.maps_config import YANDEX_MAPS_API_KEY
        return YANDEX_MAPS_API_KEY or ""
    except Exception:
        return ""


def _get_pending_agent_payments(db):
    payments = db.query(AgentPayment).filter(AgentPayment.status == "pending").order_by(AgentPayment.id.desc()).limit(20).all()
    for p in payments:
        p._agent = db.query(Agent).filter(Agent.id == p.agent_id).first()
        p._partner = db.query(Partner).filter(Partner.id == p.partner_id).first()
    return payments


# ==========================================
# SUPERVISOR: AGENT TO'LOVLARI (INKASSATSIYA)
# ==========================================

@router.get("/supervisor/agent-report", response_class=HTMLResponse)
async def supervisor_agent_report(
    request: Request,
    agent_id: int = 0,
    date_from: str = None,
    date_to: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_manager),
):
    """Agent hisoboti: oraliqda savdo, topshirilgan pul, mijozlarda qarz, o'zida qarz.
    agent_id=0 -> barcha agentlar jadvali. agent_id>0 -> tafsilot + mahsulot hisoboti.

    Ta'riflar:
      - Savdo qilgan   = SUM(Order.total) source=agent, type=sale, bekor emas, oraliqda
      - Pul topshirgan = SUM(AgentPayment.amount) confirmed, oraliqda (created_at)
      - Mijozlarda qarz= SUM(Partner.balance>0) agent mijozlari (HOZIRGI, nuqtaviy)
      - O'zida qarz    = SUM(AgentPayment.amount) pending (HOZIRGI, topshirilmagan)
    """
    from datetime import datetime as _dt

    today = _dt.now().replace(hour=0, minute=0, second=0, microsecond=0)
    # oraliq default: shu oy boshidan bugungacha
    try:
        d_from = _dt.strptime(date_from, "%Y-%m-%d") if date_from else today.replace(day=1)
    except (ValueError, TypeError):
        d_from = today.replace(day=1)
    try:
        d_to = (_dt.strptime(date_to, "%Y-%m-%d") if date_to else today).replace(hour=23, minute=59, second=59)
    except (ValueError, TypeError):
        d_to = today.replace(hour=23, minute=59, second=59)
    sel_agent_id = agent_id if (agent_id and agent_id > 0) else None

    agents = db.query(Agent).filter(Agent.is_active == True).order_by(Agent.full_name).all()

    # --- Savdo (oraliq) agent bo'yicha ---
    sales_q = (
        db.query(Order.agent_id, func.coalesce(func.sum(Order.total), 0))
        .filter(Order.source == "agent", Order.type == "sale", Order.status != "cancelled",
                Order.date >= d_from, Order.date <= d_to, Order.agent_id.isnot(None))
        .group_by(Order.agent_id)
    )
    sales_by_agent = {aid: float(t or 0) for aid, t in sales_q.all()}

    # --- Topshirilgan (confirmed, oraliq) ---
    paid_q = (
        db.query(AgentPayment.agent_id, func.coalesce(func.sum(AgentPayment.amount), 0))
        .filter(AgentPayment.status == "confirmed",
                func.date(AgentPayment.created_at) >= d_from.date(),
                func.date(AgentPayment.created_at) <= d_to.date(),
                AgentPayment.agent_id.isnot(None))
        .group_by(AgentPayment.agent_id)
    )
    paid_by_agent = {aid: float(t or 0) for aid, t in paid_q.all()}

    # --- O'zida qarz = pending (HOZIRGI) ---
    pending_q = (
        db.query(AgentPayment.agent_id, func.coalesce(func.sum(AgentPayment.amount), 0))
        .filter(AgentPayment.status == "pending", AgentPayment.agent_id.isnot(None))
        .group_by(AgentPayment.agent_id)
    )
    pending_by_agent = {aid: float(t or 0) for aid, t in pending_q.all()}

    # --- Mijozlarda qarz (HOZIRGI Partner.balance>0) ---
    debt_q = (
        db.query(Partner.agent_id, func.coalesce(func.sum(Partner.balance), 0))
        .filter(Partner.agent_id.isnot(None), Partner.balance > 0)
        .group_by(Partner.agent_id)
    )
    cust_debt_by_agent = {aid: float(t or 0) for aid, t in debt_q.all()}

    rows = []
    for a in agents:
        paid = paid_by_agent.get(a.id, 0.0)
        pct = float(getattr(a, "commission_percent", 0) or 0)
        rows.append({
            "id": a.id, "name": a.full_name, "code": a.code,
            "sales": sales_by_agent.get(a.id, 0.0),
            "paid": paid,
            "cust_debt": cust_debt_by_agent.get(a.id, 0.0),
            "self_debt": pending_by_agent.get(a.id, 0.0),
            "percent": pct,
            "salary": paid * pct / 100.0,  # oylik = foiz% x TOPSHIRGAN pul
        })
    # jami (faqat agentlari bor ko'rsatkichlar)
    totals = {
        "sales": sum(r["sales"] for r in rows),
        "paid": sum(r["paid"] for r in rows),
        "cust_debt": sum(r["cust_debt"] for r in rows),
        "self_debt": sum(r["self_debt"] for r in rows),
        "salary": sum(r["salary"] for r in rows),
    }

    # --- Bitta agent: mahsulot hisoboti (top sotilgan) ---
    sel_agent = None
    sel_row = None
    product_rows = []
    if sel_agent_id:
        sel_agent = db.query(Agent).filter(Agent.id == sel_agent_id).first()
        sel_row = next((r for r in rows if r["id"] == sel_agent_id), None)
        prod_q = (
            db.query(
                Product.name,
                func.coalesce(func.sum(OrderItem.quantity), 0),
                func.coalesce(func.sum(OrderItem.total), 0),
            )
            .join(Order, Order.id == OrderItem.order_id)
            .join(Product, Product.id == OrderItem.product_id)
            .filter(Order.source == "agent", Order.type == "sale", Order.status != "cancelled",
                    Order.agent_id == sel_agent_id, Order.date >= d_from, Order.date <= d_to)
            .group_by(Product.id, Product.name)
            .order_by(func.sum(OrderItem.quantity).desc())
        )
        for nm, qty, tot in prod_q.all():
            product_rows.append({"name": nm, "qty": float(qty or 0), "total": float(tot or 0)})

    is_month = (d_from.date() == today.replace(day=1).date() and d_to.date() == today.date())
    range_label = "Bu oy" if is_month else f"{d_from.strftime('%d.%m.%Y')}–{d_to.strftime('%d.%m.%Y')}"

    return templates.TemplateResponse("supervisor/agent_report.html", {
        "request": request, "current_user": current_user,
        "agents": agents, "rows": rows, "totals": totals,
        "sel_agent_id": sel_agent_id, "sel_agent": sel_agent, "sel_row": sel_row,
        "product_rows": product_rows,
        "date_from": d_from.strftime("%Y-%m-%d"), "date_to": d_to.strftime("%Y-%m-%d"),
        "range_label": range_label,
        "page_title": "Agent hisoboti",
    })


@router.get("/supervisor/agent-payments", response_class=HTMLResponse)
async def supervisor_agent_payments(
    request: Request,
    status: str = "all",
    date_from: str = None,
    date_to: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_manager),
):
    """Agent va Haydovchi to'lovlari ro'yxati — supervisor tasdiqlash uchun."""
    # Sana filtri (default: BUGUN — supervisor har kuni bugungi inkasatsiyani ko'radi,
    # 1 oylik darcha emas). Foydalanuvchi date_from kiritsa, o'sha sana ishlatiladi.
    from datetime import datetime as _dt, date as _date, timedelta as _td
    today = _date.today()
    default_from = today
    try:
        d_from = _dt.strptime(date_from, "%Y-%m-%d").date() if date_from else default_from
    except (ValueError, TypeError):
        d_from = default_from
    try:
        d_to = _dt.strptime(date_to, "%Y-%m-%d").date() if date_to else today
    except (ValueError, TypeError):
        d_to = today
    if d_from > d_to:
        d_from, d_to = d_to, d_from

    q = db.query(AgentPayment).filter(
        func.date(AgentPayment.created_at) >= d_from,
        func.date(AgentPayment.created_at) <= d_to,
    ).order_by(AgentPayment.id.desc())
    if status == "pending":
        q = q.filter(AgentPayment.status == "pending")
    elif status == "confirmed":
        q = q.filter(AgentPayment.status == "confirmed")
    elif status == "cancelled":
        q = q.filter(AgentPayment.status == "cancelled")
    payments = q.limit(QUERY_LIMIT_DEFAULT).all()
    for p in payments:
        p._agent = db.query(Agent).filter(Agent.id == p.agent_id).first()
        p._partner = db.query(Partner).filter(Partner.id == p.partner_id).first()

    # Haydovchi to'lovlari (Payment.category='delivery') — sana filtri bilan
    from app.models.database import Payment as _Payment
    dq = db.query(_Payment).filter(
        _Payment.category == "delivery",
        func.date(_Payment.date) >= d_from,
        func.date(_Payment.date) <= d_to,
    ).order_by(_Payment.id.desc())
    if status == "confirmed":
        dq = dq.filter(_Payment.status == "confirmed")
    driver_payments = dq.limit(QUERY_LIMIT_DEFAULT).all()
    for p in driver_payments:
        p._partner = db.query(Partner).filter(Partner.id == p.partner_id).first() if p.partner_id else None
        # Description'dan delivery raqamini chiqarib olish va Driver topish
        p._driver = None
        if p.user_id:
            from app.models.database import Driver as _Driver
            p._driver = (
                db.query(_Driver)
                .filter((_Driver.employee_id == p.user_id) | (_Driver.id == p.user_id))
                .first()
            )

    # Haydovchi qarz balansi: har haydovchi qo'lida qancha pul (pending)
    from sqlalchemy import func as _func
    from app.models.database import Driver as _Driver
    pending_q = (
        db.query(
            _Payment.user_id,
            _func.sum(_Payment.amount).label("total"),
            _func.count(_Payment.id).label("cnt"),
        )
        .filter(_Payment.category == "delivery", _Payment.status == "pending", _Payment.user_id != None)
        .group_by(_Payment.user_id)
        .all()
    )
    driver_balances = []
    for user_id, total, cnt in pending_q:
        drv = (
            db.query(_Driver)
            .filter((_Driver.employee_id == user_id) | (_Driver.id == user_id))
            .first()
        )
        driver_balances.append({
            "driver_name": drv.full_name if drv else f"user#{user_id}",
            "driver_code": drv.code if drv else "",
            "amount": float(total or 0),
            "count": int(cnt or 0),
        })
    driver_balances.sort(key=lambda x: -x["amount"])
    total_pending = sum(d["amount"] for d in driver_balances)

    # Tanlangan davr jamisi (filtered driver_payments uchun)
    range_total = sum(float(p.amount or 0) for p in driver_payments)
    range_pending = sum(float(p.amount or 0) for p in driver_payments if p.status == "pending")
    range_confirmed = sum(float(p.amount or 0) for p in driver_payments if p.status == "confirmed")

    # Agent va Driver to'lovlari uchun jamlamalar (umumiy + to'lov turi bo'yicha)
    def _breakdown(items):
        bd = {}
        for it in items:
            key = (it.payment_type or "—").lower()
            bd[key] = bd.get(key, 0.0) + float(it.amount or 0)
        return bd

    agent_total = sum(float(p.amount or 0) for p in payments)
    agent_count = len(payments)
    agent_confirmed = sum(float(p.amount or 0) for p in payments if p.status == "confirmed")
    agent_pending = sum(float(p.amount or 0) for p in payments if p.status == "pending")
    agent_by_type = _breakdown(payments)
    driver_by_type = _breakdown(driver_payments)

    return templates.TemplateResponse("supervisor/agent_payments.html", {
        "request": request,
        "payments": payments,
        "driver_payments": driver_payments,
        "driver_balances": driver_balances,
        "total_pending": total_pending,
        "status_filter": status,
        "date_from": d_from.isoformat(),
        "date_to": d_to.isoformat(),
        "today_iso": today.isoformat(),
        "range_total": range_total,
        "range_pending": range_pending,
        "range_confirmed": range_confirmed,
        "agent_total": agent_total,
        "agent_count": agent_count,
        "agent_confirmed": agent_confirmed,
        "agent_pending": agent_pending,
        "agent_by_type": agent_by_type,
        "driver_by_type": driver_by_type,
        "current_user": current_user,
    })


@router.get("/supervisor/agent-payments/export")
async def supervisor_agent_payments_export(
    request: Request,
    status: str = "all",
    date_from: str = None,
    date_to: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_manager),
):
    """Agent va Haydovchi to'lovlarini Excel'ga eksport qilish."""
    import io
    from datetime import datetime as _dt, date as _date, timedelta as _td
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    today = _date.today()
    default_from = today - _td(days=30)
    try:
        d_from = _dt.strptime(date_from, "%Y-%m-%d").date() if date_from else default_from
    except (ValueError, TypeError):
        d_from = default_from
    try:
        d_to = _dt.strptime(date_to, "%Y-%m-%d").date() if date_to else today
    except (ValueError, TypeError):
        d_to = today
    if d_from > d_to:
        d_from, d_to = d_to, d_from

    q = db.query(AgentPayment).filter(
        func.date(AgentPayment.created_at) >= d_from,
        func.date(AgentPayment.created_at) <= d_to,
    ).order_by(AgentPayment.id.desc())
    if status in ("pending", "confirmed", "cancelled"):
        q = q.filter(AgentPayment.status == status)
    ap_rows = q.all()
    for p in ap_rows:
        p._agent = db.query(Agent).filter(Agent.id == p.agent_id).first()
        p._partner = db.query(Partner).filter(Partner.id == p.partner_id).first()

    from app.models.database import Payment as _Payment, Driver as _Driver
    dq = db.query(_Payment).filter(
        _Payment.category == "delivery",
        func.date(_Payment.date) >= d_from,
        func.date(_Payment.date) <= d_to,
    ).order_by(_Payment.id.desc())
    if status == "confirmed":
        dq = dq.filter(_Payment.status == "confirmed")
    dp_rows = dq.all()
    for p in dp_rows:
        p._partner = db.query(Partner).filter(Partner.id == p.partner_id).first() if p.partner_id else None
        p._driver = None
        if p.user_id:
            p._driver = (
                db.query(_Driver)
                .filter((_Driver.employee_id == p.user_id) | (_Driver.id == p.user_id))
                .first()
            )

    status_label_map = {
        "pending": "Kutilmoqda",
        "confirmed": "Tasdiqlangan",
        "cancelled": "Rad etilgan",
    }
    type_label_map = {"naqd": "Naqd", "plastik": "Plastik", "perechisleniye": "Perechisleniye"}

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    # --- Sheet 1: Agent to'lovlari ---
    ws1 = wb.active
    ws1.title = "Agent to'lovlari"
    ws1["A1"] = f"Agent to'lovlari (inkassatsiya) — {d_from} dan {d_to} gacha"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1.merge_cells("A1:H1")
    headers1 = ["#", "Sana", "Agent", "Mijoz", "Kod", "Summa (so'm)", "Tur", "Status", "Izoh"]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ap_total = 0.0
    ap_by_type = {}
    for i, p in enumerate(ap_rows, 1):
        agent_name = p._agent.full_name if p._agent else f"Agent #{p.agent_id}"
        partner_name = p._partner.name if p._partner else ""
        partner_code = p._partner.code if p._partner else ""
        amount = float(p.amount or 0)
        ptype = type_label_map.get(p.payment_type, p.payment_type or "")
        st = status_label_map.get(p.status, p.status or "")
        ws1.append([
            i,
            p.created_at.strftime("%d.%m.%Y %H:%M") if p.created_at else "",
            agent_name,
            partner_name,
            partner_code,
            amount,
            ptype,
            st,
            p.notes or "",
        ])
        ap_total += amount
        ap_by_type[ptype] = ap_by_type.get(ptype, 0.0) + amount

    last_row = 3 + len(ap_rows) + 1
    ws1.cell(row=last_row, column=1, value=f"JAMI: {len(ap_rows)} ta").font = bold_font
    ws1.cell(row=last_row, column=6, value=ap_total).font = bold_font
    for col in range(1, 10):
        ws1.cell(row=last_row, column=col).fill = total_fill
    row_n = last_row + 1
    for ptype, amount in ap_by_type.items():
        ws1.cell(row=row_n, column=5, value=f"{ptype}:").font = bold_font
        ws1.cell(row=row_n, column=6, value=amount).font = bold_font
        row_n += 1

    widths1 = [6, 18, 22, 30, 10, 16, 16, 16, 30]
    for col, w in enumerate(widths1, 1):
        ws1.column_dimensions[chr(64 + col)].width = w

    # --- Sheet 2: Haydovchi to'lovlari ---
    ws2 = wb.create_sheet("Haydovchi to'lovlari")
    ws2["A1"] = f"Haydovchi to'lovlari (yetkazish davomida) — {d_from} dan {d_to} gacha"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.merge_cells("A1:I1")
    headers2 = ["#", "Hujjat raqami", "Sana", "Haydovchi", "Mijoz", "Kod", "Summa (so'm)", "Tur", "Status"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    dp_total = 0.0
    dp_by_type = {}
    for i, p in enumerate(dp_rows, 1):
        driver_name = p._driver.full_name if p._driver else "—"
        partner_name = p._partner.name if p._partner else ""
        partner_code = p._partner.code if p._partner else ""
        amount = float(p.amount or 0)
        ptype = type_label_map.get(p.payment_type, p.payment_type or "")
        st = status_label_map.get(p.status, p.status or "")
        ws2.append([
            i,
            p.number or "",
            p.date.strftime("%d.%m.%Y %H:%M") if p.date else "",
            driver_name,
            partner_name,
            partner_code,
            amount,
            ptype,
            st,
        ])
        dp_total += amount
        dp_by_type[ptype] = dp_by_type.get(ptype, 0.0) + amount

    last_row = 3 + len(dp_rows) + 1
    ws2.cell(row=last_row, column=1, value=f"JAMI: {len(dp_rows)} ta").font = bold_font
    ws2.cell(row=last_row, column=7, value=dp_total).font = bold_font
    for col in range(1, 10):
        ws2.cell(row=last_row, column=col).fill = total_fill
    row_n = last_row + 1
    for ptype, amount in dp_by_type.items():
        ws2.cell(row=row_n, column=6, value=f"{ptype}:").font = bold_font
        ws2.cell(row=row_n, column=7, value=amount).font = bold_font
        row_n += 1

    widths2 = [6, 20, 18, 22, 30, 10, 16, 16, 16]
    for col, w in enumerate(widths2, 1):
        ws2.column_dimensions[chr(64 + col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"agent_payments_{d_from}_{d_to}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _delete_linked_agent_payment(db, ap) -> int:
    """M3: AgentPayment'ga bog'langan kassa Payment'ini o'chiradi.
    Avval FK (ap.payment_id) orqali aniq; FK NULL (eski) bo'lsa category-match fallback.
    Qaytaradi: o'chirilgan Payment soni."""
    deleted = 0
    if ap.payment_id:
        linked = db.query(Payment).filter(Payment.id == ap.payment_id).first()
        if linked:
            db.delete(linked)
            deleted += 1
        return deleted
    for p in db.query(Payment).filter(
        Payment.partner_id == ap.partner_id,
        Payment.amount == float(ap.amount or 0),
        Payment.category == "agent_collection",
    ).all():
        db.delete(p)
        deleted += 1
    return deleted


@router.post("/supervisor/agent-payments/confirm/{payment_id}")
async def supervisor_confirm_agent_payment(
    request: Request,
    payment_id: int,
    received_amount: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_manager),
):
    """Agent to'lovini tasdiqlash — mijoz qarzidan ayirish va kassaga kirim qilish.

    received_amount (ixtiyoriy): agent HAQIQATAN topshirgan summa. Bo'sh bo'lsa = to'liq.
    Agar to'liqdan kam bo'lsa: kassaga faqat topshirilgan kiradi, mijoz qarzi TO'LIQ
    yopiladi (mijoz to'lagan), farq esa agentning o'z qarzi (EmployeeAdvance, ish haqidan
    ushlanadi). Balans: to'liq = kassa + agent_qarzi."""
    ap = db.query(AgentPayment).filter(AgentPayment.id == payment_id).first()
    if not ap:
        return RedirectResponse(url="/supervisor/agent-payments?error=not_found", status_code=303)
    if ap.status != "pending":
        return RedirectResponse(url="/supervisor/agent-payments?error=already_processed", status_code=303)

    # Atomik UPDATE WHERE — double-confirm xavfini oldini olish
    from sqlalchemy import text as _text
    claim = db.execute(
        _text("UPDATE agent_payments SET status='confirmed', confirmed_by=:uid, confirmed_at=:at "
              "WHERE id=:id AND status='pending'"),
        {"id": payment_id, "uid": current_user.id, "at": datetime.now()}
    )
    if claim.rowcount == 0:
        return RedirectResponse(url="/supervisor/agent-payments?error=already_processed", status_code=303)
    db.refresh(ap)

    # Qabul qilingan (agent topshirgan) summa — bo'sh/noto'g'ri bo'lsa to'liq summa
    full_amount = float(ap.amount or 0)
    try:
        accepted = float(received_amount) if received_amount not in (None, "") else full_amount
    except (ValueError, TypeError):
        accepted = full_amount
    # accepted == 0 ruxsat: agent inkassatsiya qildi (mijoz qarzi yopiladi) lekin pulni
    # topshirmadi → kassaga 0 kiradi, butun summa agent qarziga (shortfall = full_amount).
    if accepted < 0 or accepted > full_amount:
        accepted = full_amount
    shortfall = round(full_amount - accepted, 2)  # agent topshirmagan (uning qarzi)

    # 2. Mijoz partner'ini olish (keyinroq recompute + notify uchun)
    partner = db.query(Partner).filter(Partner.id == ap.partner_id).first()

    # 3. Tegishli kassaga kirim — FAQAT agent pul topshirgan bo'lsa (accepted > 0).
    #    accepted == 0: agent umuman topshirmadi → kassaga 0 so'm Payment YARATILMAYDI
    #    (kassa jurnali toza qoladi), butun summa pastda agent qarziga yoziladi.
    payment_number = None
    if accepted > 0:
        pay_type_map = {
            "naqd": "naqd",
            "plastik": "plastik",
            "perechisleniye": "perechisleniye",
            "click": "click",
            "terminal": "terminal",
        }
        ap_pay_type = (ap.payment_type or "").lower().strip()
        cash_type = pay_type_map.get(ap_pay_type)
        if not cash_type:
            db.rollback()
            return RedirectResponse(
                url="/supervisor/agent-payments?error=" + quote(
                    f"Noma'lum to'lov turi: {ap.payment_type!r}. Admin bilan bog'laning."
                ),
                status_code=303,
            )
        cash_register = db.query(CashRegister).filter(
            CashRegister.payment_type == cash_type,
            CashRegister.is_active == True,
            CashRegister.currency == "UZS",
        ).order_by(CashRegister.id.asc()).first()
        if not cash_register:
            db.rollback()
            return RedirectResponse(
                url="/supervisor/agent-payments?error=" + quote(
                    f"'{cash_type}' turidagi faol kassa topilmadi. Avval kassa yarating."
                ),
                status_code=303,
            )

        last_payment = db.query(Payment).order_by(Payment.id.desc()).first()
        next_num = (last_payment.id + 1) if last_payment else 1
        payment_number = f"AGT-{datetime.now().strftime('%Y%m%d')}-{next_num:04d}"

        payment = Payment(
            number=payment_number,
            date=datetime.now(),
            type="income",
            cash_register_id=cash_register.id,
            partner_id=ap.partner_id,
            amount=accepted,  # kassaga FAQAT agent topshirgan summa kiradi
            payment_type=ap.payment_type,
            category="agent_collection",
            description=(
                f"Agent inkassatsiya: {partner.name if partner else ''}"
                + (f" — {ap.notes}" if ap.notes else "")
                + (f" (topshirildi {accepted:,.0f}/{full_amount:,.0f}, qoldi agent qarzi {shortfall:,.0f})" if shortfall > 0 else "")
                + f" [AP#{ap.id}]"
            ),
            user_id=current_user.id,
            status="confirmed",
        )
        db.add(payment)
        db.flush()
        ap.payment_id = payment.id  # M3: FK bog'lash (category-match o'rniga aniq Payment)
        logger.info(
            f"AP#{ap.id} confirmed: payment_type={ap_pay_type!r} -> kassa#{cash_register.id} ({cash_register.name!r}), "
            f"mijoz_to'ladi={full_amount} topshirildi={accepted} agent_qarzi={shortfall}"
        )
    else:
        logger.info(
            f"AP#{ap.id} confirmed: accepted=0 — agent pul topshirmadi, kassaga tegilmadi; "
            f"butun {full_amount:,.0f} agent qarziga yoziladi"
        )

    # Agent topshirmagan summa → uning o'z qarzi (EmployeeAdvance, kassasiz, ish haqidan ushlanadi)
    if shortfall > 0:
        agent = db.query(Agent).filter(Agent.id == ap.agent_id).first()
        if agent and agent.employee_id:
            db.add(EmployeeAdvance(
                employee_id=agent.employee_id,
                amount=shortfall,
                advance_date=datetime.now().date(),
                note=(f"Inkassatsiya topshirilmagan: {partner.name if partner else ''} "
                      f"({accepted:,.0f}/{full_amount:,.0f}) [AP#{ap.id}]"),
                cash_register_id=None,   # kassaga tegmaydi — faqat xodim qarzi
                is_product=False,
                confirmed_at=datetime.now(),
                payment_id=None,
            ))
            logger.info(f"AP#{ap.id}: agent emp#{agent.employee_id} qarziga {shortfall} yozildi (EmployeeAdvance)")
        else:
            logger.warning(f"AP#{ap.id}: shortfall={shortfall} lekin agent#{ap.agent_id} xodimga bog'lanmagan — qarz yozilmadi")

    # 4. Buyurtmalar paid/debt — KANONIK allokatsiya (recompute_partner_order_debts).
    #    Eski qo'lda FIFO (ap.amount'ni to'g'ridan-to'g'ri orderga) OLIB TASHLANDI — 2 ta bug:
    #      (a) mijozning eski/boshlang'ich qarzini (PartnerBalanceDoc) e'tiborga olmasdi →
    #          eski qarz uchun yiqilgan inkassatsiya YANGI orderga noto'g'ri tushardi
    #          (Meva Uz/Al Fajr 2026-06-17);
    #      (b) accepted=0 (agent topshirmagan) holatda Payment YO'Q, lekin order.paid yozardi →
    #          recompute_partner_order_debts (Payment'dan derive) bilan zid, order.paid "uchib"
    #          ketardi.
    #    recompute_partner_order_debts confirmed Payment'lardan derive qiladi va eski qarzni
    #    AVVAL yutadi → mijoz haqiqatan to'lamaguncha (Payment) order qarzdor qoladi.
    if ap.partner_id:
        from app.services.partner_balance_service import recompute_partner_balance, recompute_partner_order_debts
        db.flush()
        recompute_partner_balance(
            db, ap.partner_id,
            reason="agent_payment_confirm",
            ref=payment_number,
            actor=current_user.username if current_user else None,
        )
        recompute_partner_order_debts(db, ap.partner_id)  # eski qarz-aware per-order allokatsiya

    _resync_active_cash(db)  # kassa balansini sync (inkassatsiya kirimi)
    db.commit()
    try:
        if partner is not None:
            from app.bot.customer_bot.notify import notify_customer, msg_agent_payment
            _agent = db.query(Agent).filter(Agent.id == ap.agent_id).first()
            notify_customer(ap.partner_id, msg_agent_payment(
                _agent.code if _agent else "", _agent.full_name if _agent else "",
                ap.amount, partner.balance))
    except Exception:
        pass
    return RedirectResponse(url="/supervisor/agent-payments", status_code=303)


@router.post("/supervisor/agent-payments/reject/{payment_id}")
async def supervisor_reject_agent_payment(
    request: Request,
    payment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_manager),
):
    """Agent to'lovini rad etish."""
    ap = db.query(AgentPayment).filter(AgentPayment.id == payment_id).first()
    if not ap:
        return RedirectResponse(url="/supervisor/agent-payments?error=not_found", status_code=303)
    if ap.status != "pending":
        return RedirectResponse(url="/supervisor/agent-payments?error=already_processed", status_code=303)
    ap.status = "cancelled"
    ap.confirmed_by = current_user.id
    ap.confirmed_at = datetime.now()
    db.commit()
    return RedirectResponse(url="/supervisor/agent-payments", status_code=303)


@router.post("/supervisor/agent-payments/confirm-driver/{payment_id}")
async def supervisor_confirm_driver_payment(
    request: Request,
    payment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_manager),
):
    """Haydovchi yetkazish to'lovini tasdiqlash (inkassatsiya).

    Haydovchi mijozdan pul olib pending status bilan yaratgan. Admin tasdiqlasa:
    - Payment.status='confirmed'
    - Order.paid += amount, Order.debt -= amount (mijoz balansi kamayadi)
    """
    from sqlalchemy import text as _text
    p = db.query(Payment).filter(Payment.id == payment_id, Payment.category == "delivery").first()
    if not p:
        return RedirectResponse(url="/supervisor/agent-payments?error=not_found", status_code=303)
    if p.status != "pending":
        return RedirectResponse(url="/supervisor/agent-payments?error=already_processed", status_code=303)

    # Atomik UPDATE WHERE + Payment.date'ni tasdiqlangan vaqtga ko'chirish.
    # created_at original audit izi sifatida o'zgarmaydi (haydovchi pul olgan vaqt).
    now = datetime.now()
    claim = db.execute(
        _text("UPDATE payments SET status='confirmed', date=:now "
              "WHERE id=:id AND status='pending' AND category='delivery'"),
        {"id": payment_id, "now": now},
    )
    if claim.rowcount == 0:
        return RedirectResponse(url="/supervisor/agent-payments?error=already_processed", status_code=303)
    db.refresh(p)

    # Order.paid yangilash — D3 audit fix: avval Payment.order_id (aniq order),
    # so'ng FIFO fallback (eski yozuvlar yoki order_id NULL bo'lsa)
    remaining = float(p.amount or 0)
    if remaining > 0 and p.partner_id:
        # 1) Payment.order_id bor bo'lsa, avval shu orderga qo'llash
        if p.order_id:
            target = db.query(Order).filter(Order.id == p.order_id).first()
            if target and float(target.debt or 0) > 0:
                target_debt = float(target.debt or 0)
                applied = min(target_debt, remaining)
                target.paid = float(target.paid or 0) + applied
                target.debt = target_debt - applied
                remaining -= applied
        # 2) Qoldiq bo'lsa FIFO bilan boshqa qarz orderlariga
        if remaining > 0:
            debt_orders = (
                db.query(Order)
                .filter(Order.partner_id == p.partner_id, Order.debt > 0, Order.type == "sale")
                .filter(Order.id != (p.order_id or 0))  # asosiy order ikki marta hisoblanmasin
                .order_by(Order.date.asc())
                .all()
            )
            for order in debt_orders:
                if remaining <= 0:
                    break
                order_debt = float(order.debt or 0)
                applied = min(order_debt, remaining)
                order.paid = float(order.paid or 0) + applied
                order.debt = order_debt - applied
                remaining -= applied

    # Site 3: driver payment confirm — recompute after Payment confirmed + order.paid updated
    if p.partner_id:
        from app.services.partner_balance_service import recompute_partner_balance, recompute_partner_order_debts
        db.flush()
        recompute_partner_balance(
            db, p.partner_id,
            reason="agent_payment_confirm",
            ref=getattr(p, "number", None),
            actor=current_user.username if current_user else None,
        )
        recompute_partner_order_debts(db, p.partner_id)  # M2: per-order debt izchillash

    _resync_active_cash(db)  # kassa balansini sync (haydovchi inkassatsiya kirimi)
    db.commit()
    return RedirectResponse(url="/supervisor/agent-payments?info=" + quote("Haydovchi to'lovi tasdiqlandi"), status_code=303)


# ============================================================
# 2026-05-26: Haydovchi (delivery) Payment uchun revert/delete/edit
# AgentPayment uchun mavjud, Payment.delivery uchun yo'q edi.
# ============================================================

@router.post("/supervisor/agent-payments/revert-driver/{payment_id}")
async def supervisor_revert_driver_payment(
    request: Request,
    payment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_manager),
):
    """Tasdiqlangan haydovchi to'lovini bekor qilish — Order.paid, Partner.balance qaytariladi.

    Status pending'ga qaytadi (qaytadan tasdiqlash mumkin).
    """
    p = db.query(Payment).filter(Payment.id == payment_id, Payment.category == "delivery").first()
    if not p:
        return RedirectResponse(url="/supervisor/agent-payments?error=not_found", status_code=303)
    if p.status != "confirmed":
        return RedirectResponse(url="/supervisor/agent-payments?error=not_confirmed", status_code=303)

    amount = float(p.amount or 0)

    # 1. Order.paid kamaytirish, debt qaytarish
    if p.order_id:
        order = db.query(Order).filter(Order.id == p.order_id).first()
        if order:
            order.paid = max(0.0, float(order.paid or 0) - amount)
            order.debt = float(order.debt or 0) + amount

    # 3. Payment status -> pending (payment no longer counted by recompute)
    p.status = "pending"

    # Site 4: driver payment cancel — recompute after Payment set to pending + order.debt restored
    if p.partner_id:
        from app.services.partner_balance_service import recompute_partner_balance, recompute_partner_order_debts
        db.flush()
        recompute_partner_balance(
            db, p.partner_id,
            reason="agent_payment_cancel",
            ref=getattr(p, "number", None),
            actor=current_user.username if current_user else None,
        )
        recompute_partner_order_debts(db, p.partner_id)  # M2: per-order debt izchillash

    _resync_active_cash(db)  # kassa balansini sync (to'lov bekor -> kassadan chiqdi)
    db.commit()
    return RedirectResponse(url="/supervisor/agent-payments?info=" + quote("To'lov bekor qilindi"), status_code=303)


@router.post("/supervisor/agent-payments/delete-driver/{payment_id}")
async def supervisor_delete_driver_payment(
    request: Request,
    payment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_manager),
):
    """Haydovchi to'lovini o'chirish.

    QOIDA (2026-05-26): Tasdiqlangan to'lov to'g'ridan-to'g'ri o'chirilmaydi.
    Avval `revert-driver` (bekor qilish), so'ng pending bo'lganda o'chirish mumkin.
    Bu xavfsizlik qoidasi — tasodifan tasdiqlangan hujjatni o'chirib qo'yish oldini oladi.
    """
    p = db.query(Payment).filter(Payment.id == payment_id, Payment.category == "delivery").first()
    if not p:
        return RedirectResponse(url="/supervisor/agent-payments?error=not_found", status_code=303)

    if p.status == "confirmed":
        return RedirectResponse(
            url="/supervisor/agent-payments?error=" + quote("Tasdiqlangan to'lovni o'chirib bo'lmaydi. Avval 'Bekor qilish' tugmasini bosing."),
            status_code=303,
        )

    db.delete(p)
    db.commit()
    return RedirectResponse(url="/supervisor/agent-payments?info=" + quote("To'lov o'chirildi"), status_code=303)


@router.post("/supervisor/agent-payments/edit-driver/{payment_id}")
async def supervisor_edit_driver_payment(
    request: Request,
    payment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_manager),
):
    """Haydovchi to'lovini tahrirlash — yangi summa va to'lov turi.

    Tasdiqlangan to'lov uchun ham ishlaydi (avval revert qilib eski summani qaytaradi,
    keyin yangi summa bilan qayta tasdiqlaydi).
    """
    form = await request.form()
    new_amount = float(form.get("amount") or 0)
    new_pay_type = (form.get("payment_type") or "naqd").strip().lower()
    if new_amount <= 0:
        return RedirectResponse(url="/supervisor/agent-payments?error=invalid_amount", status_code=303)

    p = db.query(Payment).filter(Payment.id == payment_id, Payment.category == "delivery").first()
    if not p:
        return RedirectResponse(url="/supervisor/agent-payments?error=not_found", status_code=303)

    old_amount = float(p.amount or 0)
    was_confirmed = p.status == "confirmed"

    # Agar confirmed edi: avval eski summa effektni qaytarish (order.paid/debt adjusted)
    if was_confirmed:
        if p.order_id:
            order = db.query(Order).filter(Order.id == p.order_id).first()
            if order:
                order.paid = max(0.0, float(order.paid or 0) - old_amount)
                order.debt = float(order.debt or 0) + old_amount

    # Yangi qiymatlarni yozish
    p.amount = new_amount
    p.payment_type = new_pay_type

    # Cash register'ni yangi pay_type'ga moslash (naqd vs plastik kassa)
    cr = db.query(CashRegister).filter(
        CashRegister.payment_type == new_pay_type,
        CashRegister.is_active == True,
    ).first()
    if cr:
        p.cash_register_id = cr.id

    # Agar confirmed edi: yangi summa bilan qayta qo'llash (order.paid/debt)
    if was_confirmed:
        if p.order_id:
            order = db.query(Order).filter(Order.id == p.order_id).first()
            if order:
                applied = min(float(order.debt or 0), new_amount)
                order.paid = float(order.paid or 0) + applied
                order.debt = max(0.0, float(order.debt or 0) - applied)

    # Site 5: driver payment edit — single recompute after all amount/order updates
    if p.partner_id and was_confirmed:
        from app.services.partner_balance_service import recompute_partner_balance, recompute_partner_order_debts
        db.flush()
        recompute_partner_balance(
            db, p.partner_id,
            reason="agent_payment_edit",
            ref=getattr(p, "number", None),
            actor=current_user.username if current_user else None,
        )
        recompute_partner_order_debts(db, p.partner_id)  # M2: per-order debt izchillash

    _resync_active_cash(db)  # kassa balansini sync (summa/kassa o'zgardi)
    db.commit()
    return RedirectResponse(url="/supervisor/agent-payments?info=" + quote("To'lov tahrirlandi"), status_code=303)


def _delete_shortfall_advance(db, ap) -> float:
    """Inkassatsiya qisman to'lovida yaratilgan agent qarzini (EmployeeAdvance) o'chiradi.
    Kassasiz advance — to'g'ridan o'chirish xavfsiz (kassa/balansga ta'sir yo'q, faqat oylik qarz)."""
    marker = f"[AP#{ap.id}]"
    advs = db.query(EmployeeAdvance).filter(
        EmployeeAdvance.note.like(f"%{marker}%"),
        EmployeeAdvance.cash_register_id.is_(None),
    ).all()
    removed = 0.0
    for a in advs:
        removed += float(a.amount or 0)
        db.delete(a)
    return removed


@router.post("/supervisor/agent-payments/revert/{payment_id}")
async def supervisor_revert_agent_payment(
    request: Request,
    payment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_manager),
):
    """Tasdiqlangan to'lovni bekor qilish — kassadan chiqarish, mijoz qarzini qaytarish."""
    ap = db.query(AgentPayment).filter(AgentPayment.id == payment_id).first()
    if not ap:
        return RedirectResponse(url="/supervisor/agent-payments?error=not_found", status_code=303)
    if ap.status != "confirmed":
        return RedirectResponse(url="/supervisor/agent-payments?error=not_confirmed", status_code=303)

    # 1. Tegishli Payment ni o'chirish (M3 helper: FK orqali aniq — avval category-match
    #    BARCHA bir xil summali agent_collection to'lovni o'chirib, noto'g'ri drift berardi)
    _delete_linked_agent_payment(db, ap)
    ap.payment_id = None
    # Qisman to'lovda yaratilgan agent qarzini ham bekor qilish
    _delete_shortfall_advance(db, ap)

    # 2. Statusni pending ga qaytarish
    ap.status = "pending"
    ap.confirmed_by = None
    ap.confirmed_at = None

    # Site 6: agent payment cancel — recompute after Payment deleted + status pending
    if ap.partner_id:
        from app.services.partner_balance_service import recompute_partner_balance, recompute_partner_order_debts
        db.flush()
        recompute_partner_balance(
            db, ap.partner_id,
            reason="agent_payment_cancel",
            ref=None,
            actor=current_user.username if current_user else None,
        )
        recompute_partner_order_debts(db, ap.partner_id)  # M2: per-order debt izchillash

    _resync_active_cash(db)  # kassa balansini sync (agent to'lov bekor -> kirim qaytarildi)
    db.commit()
    return RedirectResponse(url="/supervisor/agent-payments", status_code=303)


@router.post("/supervisor/agent-payments/delete/{payment_id}")
async def supervisor_delete_agent_payment(
    request: Request,
    payment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_manager),
):
    """Agent to'lovini o'chirish."""
    ap = db.query(AgentPayment).filter(AgentPayment.id == payment_id).first()
    if not ap:
        return RedirectResponse(url="/supervisor/agent-payments?error=not_found", status_code=303)

    # Agar tasdiqlangan bo'lsa — avval Payment'larni o'chirish, so'ng recompute
    partner_id_for_recompute = None
    if ap.status == "confirmed":
        partner_id_for_recompute = ap.partner_id
        _delete_linked_agent_payment(db, ap)  # M3 helper: FK orqali aniq
        _delete_shortfall_advance(db, ap)     # qisman to'lovdagi agent qarzini ham o'chirish

    db.delete(ap)

    # Site 7: agent payment delete — recompute after Payment deleted + ap deleted
    if partner_id_for_recompute:
        from app.services.partner_balance_service import recompute_partner_balance, recompute_partner_order_debts
        db.flush()
        recompute_partner_balance(
            db, partner_id_for_recompute,
            reason="agent_payment_delete",
            ref=None,
            actor=current_user.username if current_user else None,
        )
        recompute_partner_order_debts(db, partner_id_for_recompute)  # M2: per-order debt izchillash

    _resync_active_cash(db)  # kassa balansini sync (agent to'lov o'chirildi)
    db.commit()
    return RedirectResponse(url="/supervisor/agent-payments", status_code=303)


@router.get("/supervisor/agent-payments/edit/{payment_id}", response_class=HTMLResponse)
async def supervisor_edit_agent_payment_form(
    request: Request,
    payment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_manager),
):
    """Agent to'lovini tahrirlash formasi."""
    ap = db.query(AgentPayment).filter(AgentPayment.id == payment_id).first()
    if not ap:
        return RedirectResponse(url="/supervisor/agent-payments?error=not_found", status_code=303)
    ap._agent = db.query(Agent).filter(Agent.id == ap.agent_id).first()
    ap._partner = db.query(Partner).filter(Partner.id == ap.partner_id).first()
    return templates.TemplateResponse("supervisor/agent_payment_edit.html", {
        "request": request,
        "current_user": current_user,
        "payment": ap,
        "page_title": f"To'lov tahrirlash #{ap.id}",
    })


@router.post("/supervisor/agent-payments/edit/{payment_id}")
async def supervisor_edit_agent_payment_save(
    request: Request,
    payment_id: int,
    amount: float = Form(None),
    payment_type: str = Form(None),
    notes: str = Form(""),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_manager),
):
    """Agent to'lovini saqlash."""
    ap = db.query(AgentPayment).filter(AgentPayment.id == payment_id).first()
    if not ap:
        return RedirectResponse(url="/supervisor/agent-payments?error=not_found", status_code=303)

    if ap.status == "pending":
        # Pending — amount, type, notes tahrirlash mumkin
        if amount is not None and amount > 0:
            ap.amount = amount
        if payment_type in ("naqd", "plastik", "perechisleniye"):
            ap.payment_type = payment_type
        ap.notes = notes
    else:
        # Confirmed/cancelled — faqat notes tahrirlash mumkin
        ap.notes = notes

    db.commit()
    return RedirectResponse(url="/supervisor/agent-payments", status_code=303)


@router.post("/delivery/add-driver")
async def add_driver(
    request: Request,
    full_name: str = Form(...),
    phone: str = Form(None),
    vehicle_type: str = Form(None),
    vehicle_number: str = Form(None),
    telegram_id: str = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    last_driver = db.query(Driver).order_by(Driver.id.desc()).first()
    code = f"DR{str((last_driver.id if last_driver else 0) + 1).zfill(3)}"
    driver = Driver(
        code=code,
        full_name=full_name,
        phone=phone,
        vehicle_type=vehicle_type,
        vehicle_number=vehicle_number,
        telegram_id=telegram_id,
        is_active=True,
    )
    db.add(driver)
    db.commit()
    return RedirectResponse(url="/delivery", status_code=303)


@router.post("/delivery/add-order")
async def add_delivery_order(
    request: Request,
    driver_id: int = Form(...),
    order_number: str = Form(...),
    delivery_address: str = Form(...),
    notes: str = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_manager),
):
    # Yetkazish raqami generatsiya: DLV-YYYYMMDD-NNN
    today = datetime.now()
    prefix = f"DLV-{today.strftime('%Y%m%d')}"
    last = db.query(Delivery).filter(Delivery.number.like(f"{prefix}%")).order_by(Delivery.id.desc()).first()
    if last and last.number:
        try:
            seq = int(last.number.split("-")[-1]) + 1
        except Exception:
            seq = 1
    else:
        seq = 1
    delivery_number = f"{prefix}-{seq:03d}"
    # order_id topish (agar mavjud bo'lsa)
    order = db.query(Order).filter(Order.number == order_number).first()
    delivery = Delivery(
        number=delivery_number,
        driver_id=driver_id,
        order_id=order.id if order else None,
        order_number=order_number,
        delivery_address=delivery_address,
        planned_date=today,
        notes=notes,
        status="pending",
    )
    db.add(delivery)
    db.commit()
    return RedirectResponse(url=f"/delivery/{driver_id}", status_code=303)
