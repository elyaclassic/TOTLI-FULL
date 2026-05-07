"""Inventarizatsiya 2026-05-06 uchun xlsx eksport.

Cutoff: 2026-05-05 23:59:59
Tayyor (WH 3), Yarim tayyor (WH 2), Yarim aralash (WH 6) omborlari.

Har xlsx satrida:
- Tizim cutoff (05.05 23:59) — cutoff dagi qoldiq (bugungi harakatlardan
  oldin)
- Bugun ± — bugungi kirim/chiqim natijasi
- Tizim hozir — joriy DB qoldiq
- REAL QOLDIQ — bo'sh, foydalanuvchi qog'ozda sanab to'ldiradi
"""
import sqlite3
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DB = ROOT / "totli_holva.db"
CUTOFF = "2026-05-05 23:59:59"

WAREHOUSES = [
    (3, "tayyor_ombor", "INV-Tayyor-Ombor"),
    (2, "yarim_tayyor", "INV-Yarim-Tayyor"),
    (6, "yarim_aralash", "INV-Yarim-Aralash"),
]


def fetch_rows(conn: sqlite3.Connection, wh_id: int) -> list[dict]:
    cur = conn.cursor()
    # Tizimda mavjud Stock joriy qoldiqlar (faqat shu omborda)
    cur.execute(
        """
        SELECT s.product_id, p.name, COALESCE(u.name, '') AS unit, s.quantity AS now_qty
        FROM stocks s
        JOIN products p ON p.id = s.product_id
        LEFT JOIN units u ON u.id = p.unit_id
        WHERE s.warehouse_id = ?
        ORDER BY p.name COLLATE NOCASE
        """,
        (wh_id,),
    )
    rows = []
    for product_id, name, unit, now_qty in cur.fetchall():
        # Cutoff dan keyingi harakatlar yig'indisi
        cur.execute(
            """
            SELECT COALESCE(SUM(quantity_change), 0)
            FROM stock_movements
            WHERE warehouse_id = ? AND product_id = ?
              AND created_at > ?
            """,
            (wh_id, product_id, CUTOFF),
        )
        delta_after = cur.fetchone()[0] or 0
        cutoff_qty = (now_qty or 0) - delta_after
        rows.append(
            {
                "id": product_id,
                "name": name,
                "unit": unit or "",
                "cutoff_qty": cutoff_qty,
                "delta_today": delta_after,
                "now_qty": now_qty or 0,
            }
        )
    # Cutoff vaqtida bor edi-yu, hozir Stock da yo'q (qoldiq 0 bo'lib o'chirilgan)
    # mahsulotlar — keyingi takomil bo'lsin, hozir Stock jadvali odatda 0 ni
    # ham saqlaydi.
    return rows


def write_xlsx(rows: list[dict], path: Path, sheet_title: str, wh_name: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    headers = [
        "#",
        "ID",
        "Mahsulot",
        "Birlik",
        f"Tizim cutoff (05.05 23:59)",
        "Bugun ±",
        "Tizim hozir (06.05)",
        "REAL QOLDIQ (to'ldiring)",
        "Izoh",
    ]
    bold = Font(bold=True)
    fill_head = PatternFill("solid", fgColor="DDDDDD")
    fill_real = PatternFill("solid", fgColor="FFF2CC")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold
        c.fill = fill_head
        c.alignment = center

    widths = [5, 7, 38, 8, 14, 10, 14, 14, 22]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    for i, r in enumerate(rows, start=1):
        row_idx = i + 1
        ws.cell(row=row_idx, column=1, value=i)
        ws.cell(row=row_idx, column=2, value=r["id"])
        ws.cell(row=row_idx, column=3, value=r["name"])
        ws.cell(row=row_idx, column=4, value=r["unit"])
        ws.cell(row=row_idx, column=5, value=round(r["cutoff_qty"], 3))
        ws.cell(row=row_idx, column=6, value=round(r["delta_today"], 3))
        ws.cell(row=row_idx, column=7, value=round(r["now_qty"], 3))
        ws.cell(row=row_idx, column=8, value="").fill = fill_real
        ws.cell(row=row_idx, column=9, value="")

    ws.freeze_panes = "A2"
    wb.save(path)


def main() -> int:
    if not DB.exists():
        print(f"DB topilmadi: {DB}")
        return 1
    conn = sqlite3.connect(str(DB))
    conn.row_factory = sqlite3.Row

    if len(sys.argv) > 1:
        wanted = {int(x) for x in sys.argv[1:]}
        whs = [w for w in WAREHOUSES if w[0] in wanted]
    else:
        whs = WAREHOUSES

    for wh_id, slug, sheet_title in whs:
        all_rows = fetch_rows(conn, wh_id)
        rows = [r for r in all_rows if r["cutoff_qty"] != 0 or r["now_qty"] != 0]
        out = ROOT / f"inv_{slug}_20260506.xlsx"
        write_xlsx(rows, out, sheet_title, slug)
        print(
            f"WH {wh_id}: {out.name} | {len(rows)} qoldiqli satr "
            f"(jami {len(all_rows)} mahsulot Stock'da)"
        )
    return 0


if __name__ == "__main__":
    sys.exit(main())
