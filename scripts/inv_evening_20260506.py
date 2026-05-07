"""Kechalik inventarizatsiya uchun xlsx — INV-20260506-0001 (cutoff 05.05 23:59:59)
dan keyin nima sodir bo'lganini hisoblab, kutilgan qoldiqni ko'rsatadi.

Foydalanuvchi kechqurun ish tugagach yangidan sanab REAL ga yozadi va
solishtiradi: kutilgan != real bo'lsa qaerdadir hisobsiz harakat bor.
"""
import sqlite3
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DB = ROOT / "totli_holva.db"
INV_DOC = "INV-20260506-0001"
CUTOFF = "2026-05-05 23:59:59"

WAREHOUSES = [
    (3, "tayyor_ombor", "INV-Tayyor-Evening"),
    # WH 2 va 6 uchun ham keyinroq INV bo'lsa shu yerga qo'shiladi
]


def fetch_data(conn: sqlite3.Connection, wh_id: int):
    cur = conn.cursor()

    # 1. Cutoff qoldiqlar — INV ning quantity_after
    cur.execute(
        """
        SELECT product_id, quantity_after, quantity_change
        FROM stock_movements
        WHERE document_number = ? AND warehouse_id = ?
        """,
        (INV_DOC, wh_id),
    )
    cutoff_data = {}
    for pid, qty_after, qty_change in cur.fetchall():
        cutoff_data[pid] = {
            "cutoff_qty": qty_after,
            "inv_change": qty_change,
        }

    # 2. Bugungi (cutoff dan keyin) harakatlar
    cur.execute(
        """
        SELECT product_id, operation_type, document_number, quantity_change, created_at
        FROM stock_movements
        WHERE warehouse_id = ? AND created_at > ? AND document_number != ?
        ORDER BY created_at ASC
        """,
        (wh_id, CUTOFF, INV_DOC),
    )
    today: dict[int, dict] = {}
    for pid, op, doc, change, when in cur.fetchall():
        d = today.setdefault(pid, {
            "kirim": 0.0, "chiqim": 0.0, "ops": [],
        })
        if change > 0:
            d["kirim"] += change
        else:
            d["chiqim"] += -change
        d["ops"].append((when[:16], op, doc, change))

    # 3. Hozirgi Stock
    cur.execute(
        """
        SELECT s.product_id, p.name, COALESCE(u.name, ''), s.quantity
        FROM stocks s
        JOIN products p ON p.id = s.product_id
        LEFT JOIN units u ON u.id = p.unit_id
        WHERE s.warehouse_id = ?
        """,
        (wh_id,),
    )
    stocks = {}
    for pid, name, unit, qty in cur.fetchall():
        stocks[pid] = {"name": name, "unit": unit, "now": qty or 0}

    # Birlashtirish: barcha mahsulotlar (cutoff yoki today da bor)
    all_pids = set(cutoff_data) | set(today) | set(stocks)
    rows = []
    for pid in all_pids:
        cd = cutoff_data.get(pid, {"cutoff_qty": 0, "inv_change": 0})
        td = today.get(pid, {"kirim": 0, "chiqim": 0, "ops": []})
        st = stocks.get(pid, {"name": f"#{pid}", "unit": "", "now": 0})
        cutoff = cd["cutoff_qty"]
        kirim = td["kirim"]
        chiqim = td["chiqim"]
        expected = cutoff + kirim - chiqim
        actual = st["now"]
        rows.append({
            "id": pid,
            "name": st["name"],
            "unit": st["unit"],
            "cutoff": cutoff,
            "kirim": kirim,
            "chiqim": chiqim,
            "expected": expected,
            "actual": actual,
            "ops": td["ops"],
        })
    rows.sort(key=lambda r: (r["name"] or "").lower())
    return rows


def write_xlsx(rows: list[dict], path: Path, sheet_title: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    bold = Font(bold=True)
    fill_head = PatternFill("solid", fgColor="DDDDDD")
    fill_real = PatternFill("solid", fgColor="FFF2CC")
    fill_warn = PatternFill("solid", fgColor="FFCDD2")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "#",
        "ID",
        "Mahsulot",
        "Birlik",
        "Cutoff (05.05 23:59)",
        "Bugun kirim",
        "Bugun chiqim",
        "Kutilgan qoldiq",
        "Tizim hozir",
        "Sistema farq",
        "REAL (sanagandan keyin to'ldiring)",
        "Real-Kutilgan farq",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold
        c.fill = fill_head
        c.alignment = center

    widths = [5, 7, 38, 8, 16, 12, 13, 16, 12, 13, 22, 18]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Filter: faqat bugun harakat bo'lgan yoki cutoff != 0 yoki actual != 0
    visible = [r for r in rows if r["kirim"] or r["chiqim"] or r["cutoff"] or r["actual"]]

    for i, r in enumerate(visible, start=1):
        row_idx = i + 1
        sys_diff = r["actual"] - r["expected"]
        ws.cell(row=row_idx, column=1, value=i)
        ws.cell(row=row_idx, column=2, value=r["id"])
        ws.cell(row=row_idx, column=3, value=r["name"])
        ws.cell(row=row_idx, column=4, value=r["unit"])
        ws.cell(row=row_idx, column=5, value=round(r["cutoff"], 3))
        ws.cell(row=row_idx, column=6, value=round(r["kirim"], 3))
        ws.cell(row=row_idx, column=7, value=round(r["chiqim"], 3))
        ws.cell(row=row_idx, column=8, value=round(r["expected"], 3))
        ws.cell(row=row_idx, column=9, value=round(r["actual"], 3))
        sd = ws.cell(row=row_idx, column=10, value=round(sys_diff, 3))
        if abs(sys_diff) > 0.001:
            sd.fill = fill_warn  # Tizim notog'ri yangilangan
        ws.cell(row=row_idx, column=11, value="").fill = fill_real
        # Real-Kutilgan farq formulasi
        ws.cell(
            row=row_idx,
            column=12,
            value=f'=IF(K{row_idx}="","",K{row_idx}-H{row_idx})',
        )

    ws.freeze_panes = "A2"
    wb.save(path)


def main() -> int:
    if not DB.exists():
        print(f"DB topilmadi: {DB}")
        return 1
    conn = sqlite3.connect(str(DB))

    for wh_id, slug, sheet_title in WAREHOUSES:
        rows = fetch_data(conn, wh_id)
        out = ROOT / f"inv_{slug}_evening_20260506.xlsx"
        write_xlsx(rows, out, sheet_title)
        # Stat
        n_visible = sum(
            1 for r in rows if r["kirim"] or r["chiqim"] or r["cutoff"] or r["actual"]
        )
        n_today = sum(1 for r in rows if r["kirim"] or r["chiqim"])
        n_diff = sum(
            1
            for r in rows
            if abs((r["actual"] or 0) - (r["expected"] or 0)) > 0.001
        )
        print(
            f"WH {wh_id}: {out.name} | {n_visible} satr | bugun harakat: {n_today} | "
            f"sistem farq: {n_diff}"
        )
    return 0


if __name__ == "__main__":
    sys.exit(main())
