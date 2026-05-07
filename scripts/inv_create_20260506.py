"""xlsx dan INV-20260506-NNNN hujjatlarini DB ga yozish (draft holatda).

Foydalanuvchi keyin /qoldiqlar/inv da ko'rib tasdiqlaydi.

Args:
    sys.argv[1]: WH ID (3, 2, yoki 6)
    sys.argv[2]: xlsx fayl nomi (loyiha ildizidan nisbiy)
"""
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DB = ROOT / "totli_holva.db"

NAME_TO_ID_FALLBACK = {
    "YONG'OQ 400gr": 15,
    "Premyum Keshu 400gr": 362,
    "BUTUN PISTA 250gr": 54,
}


def to_float(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def main(wh_id: int, xlsx_path: str) -> int:
    xlsx = ROOT / xlsx_path
    if not xlsx.exists():
        print(f"Fayl topilmadi: {xlsx}")
        return 1

    wb = load_workbook(str(xlsx))
    ws = wb.active

    # Items yig'amiz: (product_id, quantity, previous_quantity, name)
    items = []
    skipped = []
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(row=r, column=2).value
        name = (ws.cell(row=r, column=3).value or "").strip()
        now_qty = to_float(ws.cell(row=r, column=7).value)
        real = to_float(ws.cell(row=r, column=8).value)

        if real is None:
            continue

        # Yangi mahsulot — pid bo'sh, nom orqali topamiz
        if not pid:
            pid = NAME_TO_ID_FALLBACK.get(name)
            if not pid:
                skipped.append((name, real, "ID topilmadi"))
                continue

        items.append({
            "product_id": int(pid),
            "quantity": real,
            "previous_quantity": now_qty or 0,
            "name": name,
        })

    if not items:
        print("Hech qanday item topilmadi")
        return 1

    conn = sqlite3.connect(str(DB))
    cur = conn.cursor()

    # Cost va sale narxlarni olamiz
    pids = [it["product_id"] for it in items]
    placeholders = ",".join("?" * len(pids))
    cur.execute(f"""
        SELECT s.product_id, s.cost_price, p.sale_price
        FROM stocks s
        JOIN products p ON p.id = s.product_id
        WHERE s.warehouse_id = ? AND s.product_id IN ({placeholders})
    """, [wh_id, *pids])
    price_map = {row[0]: (row[1] or 0, row[2] or 0) for row in cur.fetchall()}

    # INV raqami: keyingi seq
    today_str = datetime.now().strftime("%Y%m%d")
    prefix = f"INV-{today_str}"
    cur.execute("SELECT number FROM stock_adjustment_docs WHERE number LIKE ? ORDER BY id DESC LIMIT 1", (f"{prefix}-%",))
    last = cur.fetchone()
    if last:
        try:
            seq = int(last[0].split("-")[-1]) + 1
        except (ValueError, IndexError):
            seq = 1
    else:
        seq = 1
    number = f"{prefix}-{seq:04d}"

    now = datetime.now()
    total_tannarx = 0.0
    total_sotuv = 0.0
    for it in items:
        cost, sale = price_map.get(it["product_id"], (0, 0))
        total_tannarx += it["quantity"] * cost
        total_sotuv += it["quantity"] * sale

    # Doc yaratamiz
    cur.execute("""
        INSERT INTO stock_adjustment_docs (number, date, user_id, status, total_tannarx, total_sotuv, created_at, warehouse_id)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (number, now, 1, "draft", total_tannarx, total_sotuv, now, wh_id))
    doc_id = cur.lastrowid

    # Items
    for it in items:
        cost, sale = price_map.get(it["product_id"], (0, 0))
        cur.execute("""
            INSERT INTO stock_adjustment_doc_items (doc_id, product_id, warehouse_id, quantity, cost_price, sale_price, previous_quantity)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (doc_id, it["product_id"], wh_id, it["quantity"], cost, sale, it["previous_quantity"]))

    conn.commit()
    conn.close()

    print(f"INV hujjat yaratildi: {number} (id={doc_id}, WH={wh_id})")
    print(f"  {len(items)} item | tannarx jami: {total_tannarx:,.0f} | sotuv jami: {total_sotuv:,.0f}")
    if skipped:
        print(f"  O'tkazib yuborildi: {len(skipped)}")
        for n, q, reason in skipped:
            print(f"    - {n} (REAL={q}): {reason}")
    print()
    print(f"Tasdiqlash uchun: /qoldiqlar/inv/{doc_id}")
    return 0


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python inv_create_20260506.py <wh_id> <xlsx_filename>")
        sys.exit(1)
    sys.exit(main(int(sys.argv[1]), sys.argv[2]))
